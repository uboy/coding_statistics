# -*- coding: utf-8 -*-
#import sys                                          # print etc
from jira import JIRA
from configparser import ConfigParser               # able to read configuration file
import argparse                                     # good argument parser
from openpyxl import load_workbook                  # working with Excel files
import logging                                      #
import codecs                                       # used for text encoding in config parser
import csv                                          # working with csv files

CONFIG_POINT_LOCAL = "jira"
CONFIG_POINT_GLOBAL = "global"
CONFIG_FILE = "config.ini"
CONFIG_BASE_URL = "jira-url"
CONFIG_MEMBER_LIST = "member-list"
CONFIG_USERNAME = "username"
CONFIG_PASSWORD = "password"


class DisplayItem:
    def __init__(self, issue_type="", issue_key="", issue_id="", parent_id="", summary="", assignee="",
                 reporter="", priority="", status="", resolution="", created="", updated="",
                 due_date="", resolved="", remaining=0, time_spent=0, original_estimate=0, fix_version="",
                 resolved_by="", reopen_date=""):
        self.issue_type = [issue_type, "Issue Type"]
        self.issue_key = [issue_key, "Issue Key"]
        self.issue_id = [issue_id, "Issue ID"]
        self.parent_id = [parent_id, "Parent ID"]
        self.summary = [summary, "Summary"]
        self.assignee = [assignee.lower(), "Assignee"]
        self.reporter = [reporter.lower(), "Reporter"]
        self.priority = [priority, "Priority"]
        self.status = [status, "Status"]
        self.resolution = [resolution, "Resolution"]
        if len(created) >= 10:
            created_t = created[0:10]
        else:
            created_t = ""
        self.created = [created_t, "Created"]
        if len(updated) >= 10:
            updated_t = updated[0:10]
        else:
            updated_t = ""
        self.updated = [updated_t, "Updated"]
        if len(due_date) >= 10:
            due_date_t = due_date[0:10]
        else:
            due_date_t = ""
        self.due_date = [due_date_t, "Due Date"]
        if len(resolved) >= 10:
            resolved_t = resolved[0:10]
        else:
            resolved_t = ""
        self.resolved = [resolved_t, "Resolved"]
        if remaining is not None:
            remaining_t = remaining/3600
        else:
            remaining_t = 0
        self.remaining = [remaining_t, "Remaining, h"]
        if time_spent is not None:
            time_spent_t = time_spent/3600
        else:
            time_spent_t = 0
        self.time_spent = [time_spent_t, "Time Spent, h"]
        if original_estimate is not None:
            original_estimate_t = original_estimate/3600
        else:
            original_estimate_t = 0
        self.original_estimate = [original_estimate_t, "Original Estimate"]
        self.fix_version = [fix_version, "Fix Version"]
        self.resolved_by = [resolved_by.lower(), "Resolved By"]
        self.reopen_date = [reopen_date, "Reopen Date"]

    # set reopen date
    def update_reopen_date(self, date):
        self.reopen_date[0] = date


def main():
    #   parse arguments and options and get dict
    options = parse_arguments_and_options()
    #   set log level according to options or argument
    level = logging.DEBUG if options['log_level'] else logging.INFO
    logging.basicConfig(format='%(asctime)s %(levelname)s %(message)s', level=level)
    #   read list of members
    #member_list = read_member_list("members.xlsx")
    #   connect to JIRA
    jira = JIRA(basic_auth=(options["username"], options["password"]),
                options={"server": options['base_url'], "verify": "bundle-ca"})  # verify - for self signed certificate
    jira_rawdata = jira_get_data_jql(jira, 'project = myproj ORDER BY created DESC')
    jira_report_data = []
    for jira_issue in jira_rawdata:
        issue_fix_version = ""
        if jira_issue.fields.fixVersions.__len__() > 0:
            for fix_version in jira_issue.fields.fixVersions:
                issue_fix_version = issue_fix_version + fix_version.name + "|"
        duedate = ""
        if hasattr(jira_issue.fields, 'duedate'):
            if jira_issue.fields.duedate is not None:
                duedate = jira_issue.fields.duedate
        resolutiondate = ""
        if jira_issue.fields.resolutiondate is not None:
            resolutiondate = jira_issue.fields.resolutiondate
        parent_id = ""
        if hasattr(jira_issue.fields, 'parent'):
            if jira_issue.fields.parent.id is not None:
                parent_id = jira_issue.fields.parent.id
        resolved_by = ""
        if jira_issue.fields.customfield_10725 is not None:
            resolved_by = jira_issue.fields.customfield_10725.name
        assignee = ""
        if hasattr(jira_issue.fields, 'assignee'):
            if jira_issue.fields.assignee is not None:
                assignee = jira_issue.fields.assignee.name
        jira_report_data.append(DisplayItem(
            jira_issue.fields.issuetype, jira_issue.key, jira_issue.id, parent_id,
            jira_issue.fields.summary, assignee, jira_issue.fields.reporter.name,
            jira_issue.fields.priority, jira_issue.fields.status, jira_issue.fields.resolution,
            jira_issue.fields.created, jira_issue.fields.updated, duedate,
            resolutiondate, jira_issue.fields.timeestimate, jira_issue.fields.timespent,
            jira_issue.fields.timeoriginalestimate, issue_fix_version, resolved_by, ""))
    print("Finished calculating statistics. Total issues {}".format(len(jira_report_data)))
    #   find reopen date
    jira_rawdata = jira_get_data_jql(jira, 'project = myproj AND resolution changed from Done to "" and '
                                           'labels not in (label)')
    reopened_data = {}
    for jira_issue in jira_rawdata:
        issues_jql = jira.issue(jira_issue.key, expand='changelog')  # pass one issue at the time
        changelog = issues_jql.changelog
        for history in changelog.histories:
            for item in history.items:
                if item.field == 'resolution':
                    if item.toString is None:
                        reopened_data[jira_issue.key] = history.created[0:10]
    print("Finished calculating reopen dates. Total reopened issues {}".format(len(reopened_data)))
    #   add reopen date to collected data
    for jira_issue in jira_report_data:
        if jira_issue.issue_key[0] in reopened_data:
            jira_issue.update_reopen_date(reopened_data[jira_issue.issue_key[0]])
    jira.close()
    file_name = "jira_statistics-" + options["date_start"] + ".csv"
    create_and_write_csv_file(file_name, jira_report_data)


def read_member_list(member_list_file):
    wb = load_workbook(member_list_file)
    sheet = wb.active
    has_name = "name" == sheet.cell(row=1, column=1).value
    has_mail = "email" == sheet.cell(row=1, column=2).value
    has_username = "username" == sheet.cell(row=1, column=3).value

    member_list = []
    if has_name and has_mail and has_username:
        for rx in range(2, sheet.max_row + 1):
            name = sheet.cell(row=rx, column=1).value
            mail = sheet.cell(row=rx, column=2).value
            username = sheet.cell(row=rx, column=3).value
            member_list.append({'Name': name, 'Email': mail, 'Username': username})
    else:
        raise ValueError("The table format is incorrect")
    return member_list


def parse_arguments_and_options():
    #   argument parsing section
    tool_description = 'Track comments left in code by using Gerrit API'
    parser = argparse.ArgumentParser(description=tool_description,
                                     formatter_class=argparse.ArgumentDefaultsHelpFormatter)
    parser.add_argument('-u', '--username', dest='username', help='username')
    parser.add_argument('-p', '--password', dest='password', help='password')
    parser.add_argument('-l', '--url-link', dest='base_url', help='the gerrit url')
    parser.add_argument('-f', '--file', dest='member_list_file', help='Group member list file')
    parser.add_argument('-d', '--date', dest='date_start', required=True,
                        help='query date string start, eg. 2019-06 or 2019-06-25')
    parser.add_argument('-v', '--verbose', dest='verbose', action='store_true', help='enable verbose (debug) logging')
    options = parser.parse_args()

    base_url = options.base_url
    username = options.username
    password = options.password
    date_start = options.date_start
    member_list_file = options.member_list_file
    log_level = options.verbose
    config = ConfigParser()
    config.read_file(codecs.open(CONFIG_FILE, 'r', encoding='utf-8-sig'))

    if base_url is None:
        base_url = config.get(CONFIG_POINT_LOCAL, CONFIG_BASE_URL)
    if username is None:
        username = config.get(CONFIG_POINT_LOCAL, CONFIG_USERNAME)
    if password is None:
        password = config.get(CONFIG_POINT_LOCAL, CONFIG_PASSWORD)
    if member_list_file is None:
        member_list_file = config.get(CONFIG_POINT_GLOBAL, CONFIG_MEMBER_LIST)

    if base_url is None or username is None or password is None or member_list_file is None:
        raise ValueError("url or username or password or file is invalid")
    return {'base_url': base_url, 'username': username, 'password': password,
            'member_list_file': member_list_file, 'date_start': date_start, 'log_level': log_level}


def create_and_write_csv_file(file_name, data_to_write):
    with open(file_name, "w", encoding='utf-8-sig', newline='') as csvfile:
        writer = csv.writer(csvfile)
        # Write Title
        empty_class = DisplayItem()
        fields_name = []
        for field in empty_class.__dict__:
            fields_name.append(empty_class.__dict__[field][1])
        writer.writerow(fields_name)
        # Write data
        for item_to_write in data_to_write:
            row_to_write = []
            for field_to_write in item_to_write.__dict__:
                row_to_write.append(item_to_write.__dict__[field_to_write][0])
            writer.writerow(row_to_write)


def jira_get_data_jql(jira, query):
    query_start = 0
    query_limit = 1000
    jira_rawdata = []
    jira_pagination = True
    while jira_pagination:
        jira_query = query
        jira_response = jira.search_issues(jira_query, maxResults=query_limit, startAt=query_start)
        jira_rawdata = jira_rawdata + jira_response
        if jira_response.__len__() + query_start < jira_response.total:
            query_start = query_start + jira_response.__len__()
            jira_pagination = True
        else:
            jira_pagination = False
    return jira_rawdata


#   join fields in structure with delimiter
def field_join(structure, field_name, delimiter):
    joined_fields = ""
    structure_len = len(structure)
    i = 0
    if structure_len > 0:
        for fix_version in structure:
            joined_fields = joined_fields + fix_version.field_name + delimiter
            i = i + 1
    return joined_fields


if __name__ == '__main__':
    main()
