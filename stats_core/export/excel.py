"""
Excel export helpers with optional template support and basic formatting.
"""

from pathlib import Path
from typing import Iterable, Sequence, Optional

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Font


def export_sheet(
    output_path: Path,
    sheet_name: str,
    headers: Sequence[str],
    rows: Iterable[Sequence[str]],
    template: Optional[Path] = None,
    header_bold: bool = True,
) -> None:
    """
    Write tabular data to an Excel worksheet, optionally based on a template.
    """
    if template and template.exists():
        wb = load_workbook(template)
        ws = wb.active
        ws.title = sheet_name
        ws.delete_rows(1, ws.max_row)
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name

    _write_rows(ws, headers, rows, header_bold=header_bold)
    _autosize_columns(ws, max_width=60)
    wb.save(output_path)


def _write_rows(ws: Worksheet, headers: Sequence[str], rows: Iterable[Sequence[str]], header_bold: bool) -> None:
    header_font = Font(bold=header_bold)
    ws.append(list(headers))
    if header_bold:
        for cell in ws[1]:
            cell.font = header_font

    for row in rows:
        ws.append([value if value is not None else "" for value in row])


def _autosize_columns(ws: Worksheet, max_width: int = 60) -> None:
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            value = str(cell.value) if cell.value is not None else ""
            if len(value) > max_length:
                max_length = len(value)
        ws.column_dimensions[column_letter].width = min(max_length + 2, max_width)

