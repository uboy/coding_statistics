"""
Jira comprehensive report - migrated from legacy jira_ranking_report.py.

Generates a multi-sheet Excel workbook:
- Issues (detailed issue export including description/comments)
- Links (URLs extracted from descriptions/comments)
- Engineer/QA/PM performance sheets (requires report_inputs/members.xlsx by default)
"""

from __future__ import annotations

import json
import logging
import os
import re
from configparser import ConfigParser
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any
from urllib.parse import quote, urlsplit, urlunsplit

import pandas as pd
import requests
from pandas.api.types import is_object_dtype, is_string_dtype
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.styles import Alignment, Font, PatternFill

from ..pathing import resolve_member_list_path
from ..sources.jira import JiraSource
from ..utils.ai_retry import retry_ai_call
from ..utils.parallel import parallel_map
from ..utils.progress import NoopProgressManager
from . import registry

logger = logging.getLogger(__name__)


_DONE_STATUSES = {"done", "resolved", "closed"}
_SUMMARY_MAX_SENTENCES = 2
_SUBTASK_TYPES = {"sub-task", "subtask", "sub task"}

_TT_COUNTER_PATTERNS: dict[str, re.Pattern[str]] = {
    # Accept counters written as:
    # - "TT_tdev_APIs = 2" (any spaces around "=")
    # - "TT_tdev_APIs: 2"
    # - "TT_tdev_APIs - 2"
    # - "TT_tdev_APIs - some explanation = 2"
    "TT_tdev_APIs": re.compile(r"\bTT_tdev_APIs\b[^\n\r]*?(?:[:=]\s*|\s-\s*)(\d+)", re.IGNORECASE),
    "TT_tested_APIs": re.compile(r"\bTT_tested_APIs\b[^\n\r]*?(?:[:=]\s*|\s-\s*)(\d+)", re.IGNORECASE),
    "TT_tested_perf": re.compile(r"\bTT_tested_perf\b[^\n\r]*?(?:[:=]\s*|\s-\s*)(\d+)", re.IGNORECASE),
    "TT_tdev_perf": re.compile(r"\bTT_tdev_perf\b[^\n\r]*?(?:[:=]\s*|\s-\s*)(\d+)", re.IGNORECASE),
}

_OUTSTANDING_CONTRIBUTION_PATTERN = re.compile(r"outstanding[_ -]?contribution", re.IGNORECASE)
_RESULT_HEADING_PREFIX_PATTERN = re.compile(r"^\s*(?:h[1-6]\.\s+|#{1,6}\s+)", re.IGNORECASE)
_RESULT_TAG_LINE_PATTERN = re.compile(
    r"^\s*(?:[*_+]+\s*)?(results?)(?:\s*[*_+]+)?(?:\s*[:\-–—]\s*(?:[*_+]+\s*)?(.*))?\s*$",
    re.IGNORECASE,
)
_ATTACHMENT_MARKER_PATTERN = re.compile(r"\[\^([^\]\r\n]+)\]")


def _normalize_text(value: Any) -> str:
    if value is None:
        return ""
    try:
        if pd.isna(value):
            return ""
    except Exception:
        pass
    return " ".join(str(value).strip().split()).casefold()


def _compact_text(value: Any) -> str:
    if value is None:
        return ""
    try:
        if pd.isna(value):
            return ""
    except Exception:
        pass
    return " ".join(str(value).strip().split())


def _bool_value(value: Any, default: bool) -> bool:
    if value is None:
        return default
    if isinstance(value, bool):
        return value
    return str(value).strip().lower() in {"1", "true", "yes", "y", "on"}


def _strip_wrapping_quotes(value: str) -> str:
    cleaned = _compact_text(value)
    if len(cleaned) >= 2 and ((cleaned[0] == '"' and cleaned[-1] == '"') or (cleaned[0] == "'" and cleaned[-1] == "'")):
        return cleaned[1:-1].strip()
    return cleaned


def _to_plain_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, (dict, list)):
        return _compact_text(_comment_body_to_text(value))
    return _compact_text(value)


def _summarize_detail(detail: str) -> str:
    text = _compact_text(detail)
    if not text:
        return ""
    sentences = [
        token.strip(" -,:;")
        for token in re.split(r"(?<=[.!?])\s+", text)
        if token.strip(" -,:;")
    ]
    if not sentences:
        return ""
    limited = " ".join(sentences[:_SUMMARY_MAX_SENTENCES]).strip()
    if limited and limited[-1] not in ".!?":
        limited += "."
    words = limited.split()
    if len(words) > 45:
        limited = " ".join(words[:45]).rstrip(" ,;:-")
        if limited and limited[-1] not in ".!?":
            limited += "..."
    return limited


def _fallback_issue_achievement(summary: str, description: str, last_comment: str) -> str:
    base = _compact_text(summary) or "Task completed"
    detail = _summarize_detail(last_comment) or _summarize_detail(description)
    if detail:
        if detail.casefold().startswith(base.casefold()):
            return detail
        return f"{base}. {detail}"
    return f"{base}. Delivered within the reporting period."


def _extract_json_object(text: str) -> dict[str, Any] | None:
    raw = str(text or "").strip()
    if not raw:
        return None
    try:
        value = json.loads(raw)
        return value if isinstance(value, dict) else None
    except Exception:
        pass
    start_idx = raw.find("{")
    end_idx = raw.rfind("}")
    if start_idx == -1 or end_idx == -1 or end_idx <= start_idx:
        return None
    try:
        value = json.loads(raw[start_idx : end_idx + 1])
        return value if isinstance(value, dict) else None
    except Exception:
        return None


def _sanitize_summary_ai_text(text: Any) -> str:
    cleaned = _compact_text(text)
    if not cleaned:
        return ""
    cleaned = re.sub(r"(?:https?://|ftp://|file://|www\.)\S+", "", cleaned, flags=re.IGNORECASE)
    cleaned = re.sub(r"\b[A-Z]+-\d+\b", "", cleaned)
    cleaned = re.sub(r"\b[0-9a-f]{7,40}\b", "", cleaned, flags=re.IGNORECASE)
    cleaned = re.sub(r"(?i)\b(?:pr|mr|pull request|merge request|commit)\b\s*[:#-]?\s*[A-Za-z0-9/_-]*", "", cleaned)
    cleaned = _compact_text(cleaned.strip(" -,:;"))
    if not cleaned:
        return ""
    return _summarize_detail(cleaned)


def _build_summary_prompt(items: list[dict[str, str]], *, start_index: int = 1) -> tuple[dict[str, str], str]:
    prompt_lines = [
        "You are preparing a monthly software delivery summary for ArkUI OpenHarmony framework development.",
        "Task: rewrite EACH issue into concise achievement text.",
        "Strict rules:",
        "1) Output language: English.",
        "2) For each item, produce exactly 1-2 short complete sentences.",
        "3) Focus on RESULT and delivered value, not process details.",
        "4) Use only provided fields: issue title, description, latest comment, grouping hint.",
        "5) Exclude all links, repository references, file paths, PR/MR mentions, commit hashes, and Jira keys.",
        "6) Do not invent facts; if details are weak, state a safe concrete completion outcome.",
        "7) Keep phrasing clear for monthly management summary.",
        "8) Input item can represent a feature group (parent task + subtasks); summarize the whole feature in one statement.",
        "9) Return ONLY one valid JSON object mapping id to rewritten text.",
        "JSON example: {\"t1\":\"...\", \"t2\":\"...\"}",
        "---",
        "Input items:",
    ]
    target_map: dict[str, str] = {}
    for idx, item in enumerate(items, start=start_index):
        target_id = f"t{idx}"
        target_map[target_id] = item["id"]
        prompt_lines.append(
            f'ID={target_id}; title="{item.get("summary", "")}"; description="{item.get("description", "")}"; latest_comment="{item.get("last_comment", "")}"; grouping_hint="{item.get("grouping_hint", "")}"'
        )
    return target_map, "\n".join(prompt_lines)


def _strip_links_and_markup(text: str) -> str:
    cleaned = str(text or "")
    cleaned = re.sub(r"\b[A-Z]+-\d+\b", "", cleaned)
    cleaned = re.sub(r"(?:https?://|ftp://|file://|www\.)\S+", "", cleaned, flags=re.IGNORECASE)
    cleaned = re.sub(r"\[[^\]]+\]\([^)]+\)", "", cleaned)  # markdown links
    cleaned = re.sub(r"[`*_>#]+", " ", cleaned)
    cleaned = re.sub(r"\s+", " ", cleaned).strip(" -,:;")
    return cleaned


def _build_comment_summary_prompt(
    items: list[dict[str, str]],
    *,
    start_index: int = 1,
) -> tuple[dict[str, str], str]:
    prompt_lines = [
        "Ты анализируешь комментарии по задачам Jira за период.",
        "Цель: составить краткую и структурированную сводку прогресса по задаче.",
        "Используй ТОЛЬКО текст комментариев за период, не выдумывай факты.",
        "Удали ссылки, URL, упоминания PR/MR/репозиториев, ключи Jira и разметку.",
        "Ответ должен быть строго в JSON, без пояснений, без кода.",
        "Формат JSON: {\"t1\": {\"done\":\"...\", \"planned\":\"...\", \"risks\":\"...\", \"dependencies\":\"...\", \"notes\":\"...\"}}",
        "Если данных недостаточно, заполни notes='Недостаточно данных', остальные поля оставь пустыми.",
        "Если фрагмент неочевиден или двусмысленен, перенеси его в notes.",
        "Извлекай факты: done=сделано, planned=планируется, risks=риски/проблемы, dependencies=зависимости.",
        "Вывод на русском языке. Поля всегда строки.",
        "Пример: {\"t1\": {\"done\":\"Исправлен дефект X.\", \"planned\":\"Доработать тесты.\", "
        "\"risks\":\"Нет данных.\", \"dependencies\":\"Нет данных.\", \"notes\":\"\"}}",
        "---",
        "Input items:",
    ]
    target_map: dict[str, str] = {}
    for idx, item in enumerate(items, start=start_index):
        target_id = f"t{idx}"
        target_map[target_id] = item["id"]
        prompt_lines.append(f'ID={target_id}; comments="{item.get("comments", "")}"')
    return target_map, "\n".join(prompt_lines)


def _format_ai_comment_summary(value: dict[str, Any] | None) -> str:
    if not isinstance(value, dict):
        return "Недостаточно данных."

    labels = {
        "done": "Сделано",
        "planned": "Планы",
        "risks": "Риски",
        "dependencies": "Зависимости",
        "notes": "Примечания",
    }
    parts: list[str] = []
    empty_fields = 0
    for key, label in labels.items():
        raw = value.get(key, "")
        cleaned = _strip_links_and_markup(raw)
        if not cleaned:
            empty_fields += 1
            cleaned = "нет данных"
        parts.append(f"{label}: {cleaned}")
    if empty_fields == len(labels):
        return "Недостаточно данных."
    return ". ".join(parts).strip()


def _parse_label_set(value: Any) -> set[str]:
    text = _compact_text(value).casefold()
    if not text:
        return set()
    return {token for token in re.split(r"[\s,;|]+", text) if token}


def _parse_iso_date(value: Any) -> date | None:
    text = _compact_text(value)
    if not text:
        return None
    try:
        return datetime.strptime(text[:10], "%Y-%m-%d").date()
    except Exception:
        return None


def _parse_jira_date(value: Any) -> date | None:
    if value is None:
        return None
    try:
        return datetime.strptime(str(value)[:10], "%Y-%m-%d").date()
    except Exception:
        return None


def _parallel_workers(extra_params: dict[str, Any], default: int = 4) -> int:
    raw = extra_params.get("parallel_workers")
    try:
        value = int(str(raw))
        return max(value, 1)
    except Exception:
        return default


def _get_progress(extra_params: dict[str, Any], report_name: str, total_steps: int):
    progress = extra_params.get("progress_manager")
    if progress is None:
        progress = NoopProgressManager()
    progress.set_total(total_steps)
    return progress


def _summary_period(extra_params: dict[str, Any]) -> tuple[date | None, date | None]:
    start_raw = _first_value(
        [
            extra_params.get("start"),
            extra_params.get("start_date"),
            extra_params.get("start-date"),
        ]
    )
    end_raw = _first_value(
        [
            extra_params.get("end"),
            extra_params.get("end_date"),
            extra_params.get("end-date"),
        ]
    )
    return _parse_iso_date(start_raw), _parse_iso_date(end_raw)


def _resolved_in_period(resolved_date: date | None, start_dt: date | None, end_dt: date | None) -> bool:
    if resolved_date is None:
        return False
    if start_dt and end_dt:
        return start_dt <= resolved_date <= end_dt
    return True


def _build_epic_metadata_map(issues_df: pd.DataFrame) -> dict[str, dict[str, str]]:
    if issues_df.empty:
        return {}
    if "Type" not in issues_df.columns or "Issue_Key" not in issues_df.columns:
        return {}

    epic_rows = issues_df[issues_df["Type"].fillna("").astype(str).map(_normalize_text) == "epic"]
    if epic_rows.empty:
        return {}

    metadata: dict[str, dict[str, str]] = {}
    for _, row in epic_rows.iterrows():
        epic_key = _compact_text(row.get("Issue_Key"))
        if not epic_key:
            continue
        metadata[epic_key] = {
            "labels": _compact_text(row.get("Labels")),
            "status": _compact_text(row.get("Status")),
            "resolved": _compact_text(row.get("Resolved")),
        }
    return metadata


def _compose_grouped_summary_inputs(
    planned_df: pd.DataFrame,
    epic_link: str,
) -> tuple[list[dict[str, str]], list[dict[str, str]]]:
    parent_key_map: dict[str, str] = {}
    for _, row in planned_df.iterrows():
        issue_key = _compact_text(row.get("Issue_Key"))
        if issue_key:
            parent_key_map[issue_key] = _compact_text(row.get("Parent_Summary")) or _to_plain_text(row.get("Summary"))

    subtask_parent_keys: set[str] = set()
    for _, row in planned_df.iterrows():
        type_norm = _normalize_text(row.get("Type"))
        parent_key = _compact_text(row.get("Parent")) or _compact_text(row.get("Parent_Key"))
        if type_norm in _SUBTASK_TYPES and parent_key:
            subtask_parent_keys.add(parent_key)

    grouped: dict[str, dict[str, Any]] = {}
    ordered_group_keys: list[str] = []
    for _, row in planned_df.iterrows():
        issue_key = _compact_text(row.get("Issue_Key"))
        parent_key = _compact_text(row.get("Parent")) or _compact_text(row.get("Parent_Key"))
        parent_summary = _compact_text(row.get("Parent_Summary")) or parent_key_map.get(parent_key, "")
        type_norm = _normalize_text(row.get("Type"))
        is_subtask = type_norm in _SUBTASK_TYPES

        if is_subtask and parent_key:
            group_key = parent_key
        elif issue_key and issue_key in subtask_parent_keys:
            group_key = issue_key
        else:
            group_key = issue_key or f"row-{len(ordered_group_keys) + 1}"

        if group_key not in grouped:
            ordered_group_keys.append(group_key)
            grouped[group_key] = {
                "group_key": group_key,
                "parent_summary": parent_key_map.get(group_key, "") or parent_summary,
                "items": [],
            }
        grouped[group_key]["items"].append(
            {
                "issue_key": issue_key,
                "summary": _to_plain_text(row.get("Summary")),
                "description": _to_plain_text(row.get("Description")),
                "last_comment": _to_plain_text(row.get("Last_Comment")) or _to_plain_text(row.get("Comments")),
                "is_subtask": is_subtask,
            }
        )

    grouped_items: list[dict[str, str]] = []
    ai_inputs: list[dict[str, str]] = []
    for idx, group_key in enumerate(ordered_group_keys, start=1):
        group = grouped[group_key]
        components = group.get("items", [])
        if not components:
            continue

        main_summary = _compact_text(group.get("parent_summary")) or _compact_text(components[0].get("summary")) or "Task"
        description_parts: list[str] = []
        comment_parts: list[str] = []
        grouping_hint = ""
        if len(components) > 1:
            subtask_labels: list[str] = []
            for component in components:
                label = _compact_text(component.get("summary"))
                if component.get("is_subtask"):
                    subtask_labels.append(label)
                if _compact_text(component.get("description")):
                    description_parts.append(_compact_text(component.get("description")))
                if _compact_text(component.get("last_comment")):
                    comment_parts.append(_compact_text(component.get("last_comment")))
            if subtask_labels:
                grouping_hint = (
                    f"Feature group: parent task '{main_summary}' includes subtasks: "
                    + "; ".join(subtask_labels)
                )
                description_parts.append(grouping_hint)
        else:
            only = components[0]
            if _compact_text(only.get("description")):
                description_parts.append(_compact_text(only.get("description")))
            if _compact_text(only.get("last_comment")):
                comment_parts.append(_compact_text(only.get("last_comment")))

        merged_description = " ".join(description_parts).strip()
        merged_comment = " ".join(comment_parts).strip()
        fallback = _fallback_issue_achievement(main_summary, merged_description, merged_comment)
        item_id = f"{epic_link or 'NOEPIC'}::{group_key or str(idx)}"
        grouped_item = {
            "id": item_id,
            "summary": main_summary,
            "description": merged_description,
            "last_comment": merged_comment,
            "grouping_hint": grouping_hint,
            "fallback": fallback,
        }
        grouped_items.append(grouped_item)
        ai_inputs.append(
            {
                "id": item_id,
                "summary": main_summary,
                "description": merged_description,
                "last_comment": merged_comment,
                "grouping_hint": grouping_hint,
            }
        )
    return grouped_items, ai_inputs


def _build_webui_api_url(base_url: str, endpoint: str) -> str:
    base = _compact_text(base_url).rstrip("/")
    ep = _compact_text(endpoint)
    if not ep:
        return base
    if ep.startswith("http://") or ep.startswith("https://"):
        return ep.rstrip("/")
    ep_path = "/" + ep.lstrip("/")
    if base.endswith(ep_path):
        return base
    parts = urlsplit(base)
    if not parts.scheme or not parts.netloc:
        if not base:
            return ep_path
        return f"{base}/{ep.lstrip('/')}"
    base_path = parts.path.rstrip("/")
    merged_path = f"{base_path}{ep_path}" if base_path else ep_path
    return urlunsplit((parts.scheme, parts.netloc, merged_path, "", ""))


def _jira_user_identifier(user: Any | None) -> str:
    """
    Jira user identifiers differ between Server/DC and Cloud.

    Prefer the legacy username when available, then fall back to other stable IDs.
    """
    if not user:
        return ""
    for attr in ("name", "key", "accountId"):
        candidate = getattr(user, attr, None)
        if candidate:
            return str(candidate)
    return ""


def _resolved_mask(issues_df: pd.DataFrame) -> pd.Series:
    resolved_value = issues_df.get("Resolved")
    if resolved_value is None:
        resolved_value = pd.Series([""] * len(issues_df), index=issues_df.index)
    resolved_by_date = (
        resolved_value.fillna("").astype(str).str.strip().ne("")
    )

    status_value = issues_df.get("Status")
    if status_value is None:
        status_value = pd.Series([""] * len(issues_df), index=issues_df.index)
    status_norm = status_value.fillna("").astype(str).map(_normalize_text)
    resolved_by_status = status_norm.isin(_DONE_STATUSES)

    return resolved_by_date | resolved_by_status


def _countable_mask(issues_df: pd.DataFrame) -> pd.Series:
    resolution_value = issues_df.get("Resolution")
    if resolution_value is None:
        return pd.Series([True] * len(issues_df), index=issues_df.index)

    resolution_norm = resolution_value.fillna("").astype(str).map(_normalize_text)
    excluded = (
        resolution_norm.str.contains(r"won['’]t do", regex=True, na=False)
        | resolution_norm.str.contains("wont do", regex=False, na=False)
        | resolution_norm.str.contains("invalid", regex=False, na=False)
    )
    return ~excluded


def _extract_tt_counters(text: Any) -> dict[str, int]:
    if text is None:
        return {key: 0 for key in _TT_COUNTER_PATTERNS}
    try:
        if pd.isna(text):
            return {key: 0 for key in _TT_COUNTER_PATTERNS}
    except Exception:
        pass

    payload = str(text)
    payload = (
        payload.replace("\u00a0", " ")
        .replace("\uff1a", ":")  # fullwidth colon
        .replace("\uff1d", "=")  # fullwidth equals
        .replace("\u2013", "-")  # en-dash
        .replace("\u2014", "-")  # em-dash
        .replace("\u2212", "-")  # minus sign
    )
    payload = re.sub(r"\s+", " ", payload)
    counters: dict[str, int] = {}
    for key, pattern in _TT_COUNTER_PATTERNS.items():
        matches = pattern.findall(payload)
        counters[key] = sum(int(match) for match in matches) if matches else 0
    return counters


def build_jql_query(params: dict[str, Any]) -> str:
    """Build JQL query based on provided parameters."""
    if params.get("jql"):
        return str(params["jql"])

    conditions: list[str] = []

    project = params.get("project")
    if project:
        conditions.append(f"project = {project}")

    start_date = params.get("start_date")
    end_date = params.get("end_date")
    if start_date and end_date:
        try:
            end_exclusive = (
                datetime.strptime(str(end_date), "%Y-%m-%d").date() + timedelta(days=1)
            ).strftime("%Y-%m-%d")
            conditions.append(f"resolved >= '{start_date}' AND resolved < '{end_exclusive}'")
        except ValueError:
            conditions.append(f"resolved >= '{start_date}' AND resolved <= '{end_date}'")

    version = params.get("version")
    if version:
        conditions.append(f"fixVersion = '{version}'")

    epic = params.get("epic")
    if epic:
        conditions.append(f"'Epic Link' = {epic}")

    if not conditions:
        raise ValueError("Must specify at least one of: project+dates, version, epic, or jql")

    return " AND ".join(conditions) + " ORDER BY created DESC"


def build_comments_period_jql(params: dict[str, Any]) -> str:
    """Build JQL query for comments-period sheet (updated-date based)."""
    if params.get("jql"):
        return str(params["jql"])

    conditions: list[str] = []

    project = params.get("project")
    if project:
        conditions.append(f"project = {project}")

    start_date = params.get("start_date")
    end_date = params.get("end_date")
    if start_date and end_date:
        try:
            end_exclusive = (
                datetime.strptime(str(end_date), "%Y-%m-%d").date() + timedelta(days=1)
            ).strftime("%Y-%m-%d")
            conditions.append(f"updated >= '{start_date}' AND updated < '{end_exclusive}'")
        except ValueError:
            conditions.append(f"updated >= '{start_date}' AND updated <= '{end_date}'")

    version = params.get("version")
    if version:
        conditions.append(f"fixVersion = '{version}'")

    epic = params.get("epic")
    if epic:
        conditions.append(f"'Epic Link' = {epic}")

    if not conditions:
        raise ValueError("Must specify at least one of: project+dates, version, epic, or jql")

    return " AND ".join(conditions) + " ORDER BY created DESC"


def extract_urls_from_text(text: str | None) -> list[str]:
    """Extract all URLs from text."""
    if not text:
        return []
    url_pattern = r'https?://[^\s<>"{}|\\^`\[\]]+'
    return re.findall(url_pattern, text)


def _comment_body_to_text(body: Any) -> str:
    if body is None:
        return ""
    if isinstance(body, str):
        return body
    if isinstance(body, list):
        return " ".join(_comment_body_to_text(item) for item in body).strip()
    if isinstance(body, dict):
        parts: list[str] = []
        stack = [body]
        while stack:
            node = stack.pop()
            if isinstance(node, dict):
                text_value = node.get("text")
                if isinstance(text_value, str) and text_value:
                    parts.append(text_value)
                content = node.get("content")
                if isinstance(content, list):
                    stack.extend(reversed(content))
            elif isinstance(node, list):
                stack.extend(reversed(node))
        return " ".join(parts).strip()
    return str(body)


def _extract_result_text(comment_body: Any) -> str | None:
    text = "" if comment_body is None else str(comment_body)
    stripped = text.lstrip()
    if not stripped:
        return None

    lines = stripped.splitlines()
    first_line = lines[0] if lines else ""
    remaining_lines = lines[1:] if len(lines) > 1 else []

    first_line_no_heading = _RESULT_HEADING_PREFIX_PATTERN.sub("", first_line, count=1)
    match = _RESULT_TAG_LINE_PATTERN.match(first_line_no_heading)
    if not match:
        return None

    # Strip only the leading result tag/heading; keep remaining content intact.
    same_line_tail = match.group(2) or ""
    if same_line_tail:
        if remaining_lines:
            return (same_line_tail + "\n" + "\n".join(remaining_lines)).lstrip()
        return same_line_tail.lstrip()
    return "\n".join(remaining_lines).lstrip()


def _issue_attachment_links(issue: Any, jira_url: str) -> dict[str, str]:
    attachments = getattr(getattr(issue, "fields", None), "attachment", None) or []
    result: dict[str, str] = {}
    for attachment in attachments:
        filename = str(getattr(attachment, "filename", "") or "").strip()
        if not filename:
            continue
        url = str(getattr(attachment, "content", "") or "").strip()
        if not url:
            attachment_id = str(getattr(attachment, "id", "") or "").strip()
            if attachment_id and jira_url:
                url = f"{jira_url}/secure/attachment/{attachment_id}/{quote(filename)}"
        if not url:
            continue
        result[filename.casefold()] = url
    return result


def _extract_attachment_links(text: str, attachment_links: dict[str, str]) -> list[str]:
    links: list[str] = []
    for match in _ATTACHMENT_MARKER_PATTERN.finditer(text or ""):
        filename = match.group(1).strip().casefold()
        if not filename:
            continue
        url = attachment_links.get(filename)
        if url and url not in links:
            links.append(url)
    return links


def _replace_attachment_markers_with_links(text: str, attachment_links: dict[str, str]) -> str:
    if not text:
        return text

    def _replace(match: re.Match[str]) -> str:
        filename = match.group(1).strip().casefold()
        if not filename:
            return match.group(0)
        return attachment_links.get(filename) or match.group(0)

    return _ATTACHMENT_MARKER_PATTERN.sub(_replace, text)


def _sort_by_epic_and_resolved(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return df
    if "Epic_Name" not in df.columns or "Resolved" not in df.columns:
        return df

    sorted_df = df.copy()
    sorted_df["_sort_epic_name"] = sorted_df["Epic_Name"].fillna("").astype(str).str.casefold()
    sorted_df["_sort_resolved"] = pd.to_datetime(sorted_df["Resolved"], errors="coerce")
    sorted_df = sorted_df.sort_values(
        by=["_sort_epic_name", "_sort_resolved"],
        ascending=[True, True],
        na_position="last",
        kind="mergesort",
    )
    return sorted_df.drop(columns=["_sort_epic_name", "_sort_resolved"])


def _fetch_epic_metadata(jira, epic_keys: list[str]) -> dict[str, dict[str, str]]:
    if not epic_keys:
        return {}
    epic_keys = [key for key in epic_keys if key]
    if not epic_keys:
        return {}

    epic_metadata: dict[str, dict[str, str]] = {}
    chunk_size = 50
    for i in range(0, len(epic_keys), chunk_size):
        chunk = epic_keys[i:i + chunk_size]
        epics = jira.search_issues(
            f"issuekey in ({', '.join(chunk)})",
            maxResults=1000,
            fields=["key", "summary", "status", "resolutiondate", "labels"],
        )
        for epic in epics:
            labels_raw = getattr(epic.fields, "labels", None) or []
            labels = ", ".join(str(label) for label in labels_raw if str(label).strip())
            status_name = getattr(getattr(epic.fields, "status", None), "name", "") or ""
            resolved_raw = getattr(epic.fields, "resolutiondate", "") or ""
            epic_metadata[epic.key] = {
                "name": _compact_text(getattr(epic.fields, "summary", "")),
                "status": _compact_text(status_name),
                "resolved": _compact_text(str(resolved_raw)[:10]),
                "labels": _compact_text(labels),
            }
    return epic_metadata


def _build_comment_link(jira_url: str, issue_key: str, comment_id: str | None) -> str:
    if not comment_id:
        return f"{jira_url}/browse/{issue_key}"
    return f"{jira_url}/browse/{issue_key}?focusedCommentId={comment_id}#comment-{comment_id}"


def fetch_jira_data(jira, jql_query: str) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    """
    Fetch Jira issues with all details including comments.

    Returns:
        (issues_df, links_df, results_df)
    """
    start_at = 0
    max_results = 100
    all_issues: list[Any] = []

    logger.info("Executing JQL: %s", jql_query)

    while True:
        issues = jira.search_issues(
            jql_query,
            startAt=start_at,
            maxResults=max_results,
            fields=[
                "key",
                "summary",
                "assignee",
                "reporter",
                "resolutiondate",
                "created",
                "updated",
                "description",
                "comment",
                "labels",
                "priority",
                "status",
                "resolution",
                "issuetype",
                "timeestimate",
                "timespent",
                "timeoriginalestimate",
                "customfield_10000",  # Epic Link
                "parent",
                "attachment",
            ],
            expand="changelog",
        )

        all_issues.extend(issues)

        if len(issues) < max_results:
            break
        start_at += max_results

    logger.info("Fetched %s issues", len(all_issues))

    data: list[dict[str, Any]] = []
    all_links: list[dict[str, str]] = []
    all_comments: list[dict[str, Any]] = []

    issue_epic_map: dict[str, str] = {}
    parent_keys_needed: set[str] = set()
    for issue in all_issues:
        epic_link = getattr(issue.fields, "customfield_10000", "") or ""
        parent = getattr(issue.fields, "parent", None)
        parent_key = parent.key if parent else ""
        if not epic_link and parent_key:
            parent_keys_needed.add(parent_key)
        issue_epic_map[issue.key] = epic_link or ""

    if parent_keys_needed:
        missing_parent_keys = [
            key for key in parent_keys_needed if not issue_epic_map.get(key)
        ]
        if missing_parent_keys:
            chunk_size = 50
            for i in range(0, len(missing_parent_keys), chunk_size):
                chunk = missing_parent_keys[i:i + chunk_size]
                parent_issues = jira.search_issues(
                    f"issuekey in ({', '.join(chunk)})",
                    maxResults=1000,
                    fields=["key", "customfield_10000"],
                )
                for parent_issue in parent_issues:
                    parent_epic = getattr(parent_issue.fields, "customfield_10000", "") or ""
                    issue_epic_map[parent_issue.key] = parent_epic

    epic_metadata_map = _fetch_epic_metadata(jira, list({value for value in issue_epic_map.values() if value}))
    epic_name_map = {key: value.get("name", "") for key, value in epic_metadata_map.items()}
    jira_url = jira._options.get("server", "")

    for issue in all_issues:
        key = issue.key
        summary = issue.fields.summary
        assignee = issue.fields.assignee.displayName if issue.fields.assignee else "Unassigned"
        assignee_name = _jira_user_identifier(issue.fields.assignee)
        reporter = issue.fields.reporter.displayName if issue.fields.reporter else ""
        reporter_name = _jira_user_identifier(issue.fields.reporter)

        created = issue.fields.created[:10] if issue.fields.created else ""
        resolved = issue.fields.resolutiondate[:10] if issue.fields.resolutiondate else ""

        original_estimate = issue.fields.timeoriginalestimate / 3600 if issue.fields.timeoriginalestimate else 0
        time_spent = issue.fields.timespent / 3600 if issue.fields.timespent else 0
        remaining = issue.fields.timeestimate / 3600 if issue.fields.timeestimate else 0

        description = issue.fields.description or ""
        attachment_links = _issue_attachment_links(issue, jira_url)

        for link in extract_urls_from_text(description):
            all_links.append({"Issue_Key": key, "Source": "Description", "URL": link})

        comments: list[str] = []
        latest_comment_text = ""
        latest_comment_meta: tuple[str, int] | None = None
        if hasattr(issue.fields, "comment") and issue.fields.comment.comments:
            for comment_idx, comment in enumerate(issue.fields.comment.comments):
                comment_text = _comment_body_to_text(comment.body)
                comment_author = comment.author.displayName if comment.author else "Unknown"
                comment_created = comment.created[:10] if comment.created else ""
                comment_id = getattr(comment, "id", "") or ""
                comments.append(f"[{comment_created}] {comment_author}: {comment_text}")
                created_raw = str(comment.created or "")
                if comment_text:
                    marker = (created_raw, comment_idx)
                    if latest_comment_meta is None or marker >= latest_comment_meta:
                        latest_comment_meta = marker
                        latest_comment_text = comment_text

                for link in extract_urls_from_text(comment_text):
                    all_links.append(
                        {
                            "Issue_Key": key,
                            "Source": f"Comment by {comment_author}",
                            "URL": link,
                        }
                    )
                for attachment_link in _extract_attachment_links(comment_text, attachment_links):
                    all_links.append(
                        {
                            "Issue_Key": key,
                            "Source": f"Comment by {comment_author}",
                            "URL": attachment_link,
                        }
                    )
                all_comments.append(
                    {
                        "Issue_Key": key,
                        "Summary": summary,
                        "Assignee": assignee,
                        "Comment_Created": comment.created or "",
                        "Comment_Id": str(comment_id) if comment_id is not None else "",
                        "Comment_Body": comment_text,
                        "Comment_Link": _build_comment_link(jira_url, key, str(comment_id) if comment_id else None),
                        "Attachment_Links": attachment_links,
                    }
                )

        all_comments_text = "\n---\n".join(comments)

        labels = ", ".join(issue.fields.labels) if issue.fields.labels else ""

        priority = issue.fields.priority.name if issue.fields.priority else ""
        status = issue.fields.status.name if issue.fields.status else ""
        resolution = issue.fields.resolution.name if issue.fields.resolution else ""
        issue_type = issue.fields.issuetype.name if issue.fields.issuetype else ""
        epic_link = issue_epic_map.get(issue.key, "")
        parent = getattr(issue.fields, "parent", None)
        parent_key = parent.key if parent else ""
        if not epic_link and parent_key:
            epic_link = issue_epic_map.get(parent_key, "")
        epic_name = epic_name_map.get(epic_link, "Unknown Epic") if epic_link else "Unknown Epic"
        epic_meta = epic_metadata_map.get(epic_link, {}) if epic_link else {}

        data.append(
            {
                "Issue_Key": key,
                "Summary": summary,
                "Type": issue_type,
                "Status": status,
                "Resolution": resolution,
                "Priority": priority,
                "Assignee": assignee,
                "Assignee_Username": assignee_name,
                "Reporter": reporter,
                "Reporter_Username": reporter_name,
                "Created": created,
                "Resolved": resolved,
                "Original_Estimate_Hours": original_estimate,
                "Time_Spent_Hours": time_spent,
                "Remaining_Hours": remaining,
                "Description": description,
                "Comments": all_comments_text,
                "Last_Comment": latest_comment_text,
                "Labels": labels,
                "Epic_Link": epic_link,
                "Epic_Name": epic_name,
                "Epic_Status": epic_meta.get("status", ""),
                "Epic_Resolved": epic_meta.get("resolved", ""),
                "Epic_Labels": epic_meta.get("labels", ""),
                "Parent": parent_key,
            }
        )

    issues_df = pd.DataFrame(data)
    links_df = pd.DataFrame(all_links)
    results_df = pd.DataFrame(
        columns=["Issue_Key", "Summary", "Epic_Name", "Resolved", "Assignee", "Result", "Result_Links"]
    )

    if not issues_df.empty:
        issues_df = _sort_by_epic_and_resolved(issues_df)

        resolution_value = issues_df.get("Resolution")
        if resolution_value is not None:
            resolution_norm = resolution_value.fillna("").astype(str).str.strip().str.casefold()
            resolved_keys = set(
                issues_df.loc[resolution_norm.isin({"done", "resolved"}), "Issue_Key"]
                .dropna()
                .astype(str)
            )
        else:
            resolved_keys = set()

        result_rows: list[dict[str, Any]] = []
        results_by_issue: set[str] = set()
        latest_comment_links: dict[str, str] = {}
        if all_comments:
            latest_comment_meta: dict[str, tuple[str, int]] = {}
            for idx, entry in enumerate(all_comments):
                issue_key = str(entry.get("Issue_Key", "") or "")
                if not issue_key:
                    continue
                created = str(entry.get("Comment_Created", "") or "")
                prev_meta = latest_comment_meta.get(issue_key)
                if prev_meta is None or (created, idx) >= prev_meta:
                    latest_comment_meta[issue_key] = (created, idx)
                    latest_comment_links[issue_key] = str(entry.get("Comment_Link", "") or "")

        if resolved_keys and all_comments:
            epic_name_map = issues_df.set_index("Issue_Key")["Epic_Name"].to_dict()
            resolved_date_map = issues_df.set_index("Issue_Key")["Resolved"].to_dict()
            for entry in all_comments:
                issue_key = str(entry.get("Issue_Key", "") or "")
                if issue_key not in resolved_keys:
                    continue
                comment_body = entry.get("Comment_Body") or ""
                parsed_result = _extract_result_text(comment_body)
                if parsed_result is None:
                    continue

                attachment_links = entry.get("Attachment_Links")
                if not isinstance(attachment_links, dict):
                    attachment_links = {}
                parsed_result = _replace_attachment_markers_with_links(parsed_result, attachment_links)

                links: list[str] = []
                for link in extract_urls_from_text(comment_body):
                    if link not in links:
                        links.append(link)
                for link in _extract_attachment_links(comment_body, attachment_links):
                    if link not in links:
                        links.append(link)

                if links:
                    result_links = "\n".join(links)
                else:
                    result_links = entry.get("Comment_Link") or ""

                result_rows.append(
                    {
                        "Issue_Key": issue_key,
                        "Summary": entry.get("Summary", ""),
                        "Epic_Name": epic_name_map.get(issue_key, "Unknown Epic"),
                        "Resolved": resolved_date_map.get(issue_key, ""),
                        "Assignee": entry.get("Assignee", "") or "Unassigned",
                        "Result": parsed_result,
                        "Result_Links": result_links,
                    }
                )
                results_by_issue.add(issue_key)

        if resolved_keys:
            summary_map = issues_df.set_index("Issue_Key")["Summary"].to_dict()
            assignee_map = issues_df.set_index("Issue_Key")["Assignee"].to_dict()
            epic_name_map = issues_df.set_index("Issue_Key")["Epic_Name"].to_dict()
            resolved_date_map = issues_df.set_index("Issue_Key")["Resolved"].to_dict()
            for issue_key in sorted(resolved_keys):
                if issue_key in results_by_issue:
                    continue
                result_rows.append(
                    {
                        "Issue_Key": issue_key,
                        "Summary": summary_map.get(issue_key, ""),
                        "Epic_Name": epic_name_map.get(issue_key, "Unknown Epic"),
                        "Resolved": resolved_date_map.get(issue_key, ""),
                        "Assignee": assignee_map.get(issue_key, "") or "Unassigned",
                        "Result": "no results",
                        "Result_Links": latest_comment_links.get(issue_key)
                        or (f"{jira_url}/browse/{issue_key}" if jira_url else ""),
                    }
                )

        if result_rows:
            results_df = pd.DataFrame(result_rows)
            results_df = _sort_by_epic_and_resolved(results_df)

    return issues_df, links_df, results_df


def _comment_activity_date(
    created_dt: date | None,
    updated_dt: date | None,
    start_dt: date | None,
    end_dt: date | None,
) -> tuple[date | None, bool]:
    if start_dt and end_dt:
        if updated_dt and start_dt <= updated_dt <= end_dt:
            return updated_dt, True
        if created_dt and start_dt <= created_dt <= end_dt:
            return created_dt, False
        return None, False
    return (updated_dt or created_dt), bool(updated_dt and updated_dt != created_dt)


def build_comments_period_df(
    jira,
    jql_query: str,
    start_date: str | None,
    end_date: str | None,
    config: ConfigParser,
    extra_params: dict[str, Any],
) -> pd.DataFrame:
    start_dt = _parse_iso_date(start_date) if start_date else None
    end_dt = _parse_iso_date(end_date) if end_date else None

    start_at = 0
    max_results = 100
    all_issues: list[Any] = []
    while True:
        issues = jira.search_issues(
            jql_query,
            startAt=start_at,
            maxResults=max_results,
            fields=[
                "key",
                "summary",
                "assignee",
                "created",
                "description",
                "comment",
                "priority",
                "status",
                "issuetype",
                "customfield_10000",
                "parent",
            ],
        )
        all_issues.extend(issues)
        if len(issues) < max_results:
            break
        start_at += max_results

    if not all_issues:
        columns = [
            "Issue_Key",
            "Summary",
            "Type",
            "Status",
            "Priority",
            "Assignee",
            "Created",
            "Epic_Name",
            "Parent",
            "Description",
            "Comments",
            "AI_Comments",
            "Comments_In_Period",
        ]
        return pd.DataFrame(columns=columns)

    issue_epic_map: dict[str, str] = {}
    parent_keys_needed: set[str] = set()
    for issue in all_issues:
        epic_link = getattr(issue.fields, "customfield_10000", "") or ""
        parent = getattr(issue.fields, "parent", None)
        parent_key = parent.key if parent else ""
        if not epic_link and parent_key:
            parent_keys_needed.add(parent_key)
        issue_epic_map[issue.key] = epic_link or ""

    if parent_keys_needed:
        missing_parent_keys = [
            key for key in parent_keys_needed if not issue_epic_map.get(key)
        ]
        if missing_parent_keys:
            chunk_size = 50
            for i in range(0, len(missing_parent_keys), chunk_size):
                chunk = missing_parent_keys[i:i + chunk_size]
                parent_issues = jira.search_issues(
                    f"issuekey in ({', '.join(chunk)})",
                    maxResults=1000,
                    fields=["key", "customfield_10000"],
                )
                for parent_issue in parent_issues:
                    parent_epic = getattr(parent_issue.fields, "customfield_10000", "") or ""
                    issue_epic_map[parent_issue.key] = parent_epic

    epic_metadata_map = _fetch_epic_metadata(jira, list({value for value in issue_epic_map.values() if value}))
    epic_name_map = {key: value.get("name", "") for key, value in epic_metadata_map.items()}

    rows: list[dict[str, Any]] = []
    ai_inputs: list[dict[str, str]] = []

    for issue in all_issues:
        key = issue.key
        summary = issue.fields.summary
        issue_type = issue.fields.issuetype.name if issue.fields.issuetype else ""
        status = issue.fields.status.name if issue.fields.status else ""
        priority = issue.fields.priority.name if issue.fields.priority else ""
        assignee = issue.fields.assignee.displayName if issue.fields.assignee else "Unassigned"
        created = issue.fields.created[:10] if issue.fields.created else ""
        description = issue.fields.description or ""
        parent = getattr(issue.fields, "parent", None)
        parent_key = parent.key if parent else ""
        epic_link = issue_epic_map.get(issue.key, "")
        if not epic_link and parent_key:
            epic_link = issue_epic_map.get(parent_key, "")
        epic_name = epic_name_map.get(epic_link, "Unknown Epic") if epic_link else "Unknown Epic"

        comments: list[str] = []
        comments_in_period: list[str] = []
        if hasattr(issue.fields, "comment") and issue.fields.comment.comments:
            for comment in issue.fields.comment.comments:
                comment_text = _compact_text(_comment_body_to_text(comment.body))
                if not comment_text:
                    continue
                comment_author = comment.author.displayName if comment.author else "Unknown"
                created_dt = _parse_jira_date(getattr(comment, "created", None))
                updated_dt = _parse_jira_date(getattr(comment, "updated", None))
                created_str = created_dt.strftime("%Y-%m-%d") if created_dt else ""
                comments.append(f"[{created_str}] {comment_author}: {comment_text}")

                activity_dt, used_updated = _comment_activity_date(created_dt, updated_dt, start_dt, end_dt)
                if activity_dt:
                    activity_str = activity_dt.strftime("%Y-%m-%d")
                    tag = f"{activity_str} updated" if used_updated else activity_str
                    comments_in_period.append(f"[{tag}] {comment_author}: {comment_text}")

        if start_dt and end_dt and not comments_in_period:
            continue

        all_comments_text = "\n---\n".join(comments)
        comments_in_period_text = "\n---\n".join(comments_in_period)
        rows.append(
            {
                "Issue_Key": key,
                "Summary": summary,
                "Type": issue_type,
                "Status": status,
                "Priority": priority,
                "Assignee": assignee,
                "Created": created,
                "Epic_Name": epic_name,
                "Parent": parent_key,
                "Description": description,
                "Comments": all_comments_text,
                "AI_Comments": "",
                "Comments_In_Period": comments_in_period_text,
            }
        )
        ai_inputs.append({"id": key, "comments": comments_in_period_text})

    ai_map = rewrite_comment_items_with_ai(ai_inputs, config, extra_params) if ai_inputs else {}
    for row in rows:
        ai_value = ai_map.get(row["Issue_Key"])
        row["AI_Comments"] = _format_ai_comment_summary(ai_value)

    return pd.DataFrame(rows)


def _rewrite_summary_items_with_ollama(
    items: list[dict[str, str]],
    config: ConfigParser,
    extra_params: dict[str, Any],
) -> dict[str, str]:
    ollama_enabled = _bool_value(
        extra_params.get("ollama_enabled", config.get("ollama", "enabled", fallback="true")),
        True,
    )
    if not ollama_enabled:
        return {}
    model = _compact_text(extra_params.get("ollama_model") or config.get("ollama", "model", fallback=""))
    if not model:
        logger.warning("Summary AI: Ollama model is not configured; using deterministic summary text.")
        return {}
    if not items:
        return {}

    ollama_url = _compact_text(extra_params.get("ollama_url") or config.get("ollama", "url", fallback="http://localhost:11434"))
    ollama_api_key = _strip_wrapping_quotes(
        _compact_text(extra_params.get("ollama_api_key") or config.get("ollama", "api_key", fallback=""))
    )
    timeout_seconds = int(
        _compact_text(extra_params.get("ollama_timeout_seconds") or config.get("ollama", "timeout_seconds", fallback="60"))
        or "60"
    )
    temperature = float(
        _compact_text(extra_params.get("ollama_temperature") or config.get("ollama", "temperature", fallback="0.2"))
        or "0.2"
    )

    headers: dict[str, str] = {"Content-Type": "application/json"}
    if ollama_api_key:
        headers["Authorization"] = f"Bearer {ollama_api_key}"

    rewritten: dict[str, str] = {}
    batch_size = 8
    batches = [(i, items[i : i + batch_size]) for i in range(0, len(items), batch_size)]
    max_workers = _parallel_workers(extra_params)

    def _run_batch(batch_info: tuple[int, list[dict[str, str]]]) -> dict[str, str]:
        start_index, batch = batch_info
        target_map, prompt = _build_summary_prompt(batch, start_index=start_index + 1)

        def _request():
            return requests.post(
                f"{ollama_url.rstrip('/')}/api/generate",
                headers=headers,
                json={
                    "model": model,
                    "prompt": prompt,
                    "stream": False,
                    "options": {"temperature": temperature},
                },
                timeout=timeout_seconds,
            )

        try:
            response = retry_ai_call(_request, logger=logger)
            response.raise_for_status()
            response_json = response.json()
            response_text = str(response_json.get("response", "") or "")
            rewrite_map = _extract_json_object(response_text) or {}
            batch_result: dict[str, str] = {}
            for target_id, issue_id in target_map.items():
                candidate = _sanitize_summary_ai_text(rewrite_map.get(target_id))
                if candidate:
                    batch_result[issue_id] = candidate
            return batch_result
        except Exception as exc:
            logger.warning("Summary AI (Ollama) batch %s failed: %s", start_index // batch_size + 1, exc)
            return {}

    for batch_result in parallel_map(_run_batch, batches, max_workers=max_workers):
        rewritten.update(batch_result)

    return rewritten


def _rewrite_summary_items_with_webui(
    items: list[dict[str, str]],
    config: ConfigParser,
    extra_params: dict[str, Any],
) -> dict[str, str]:
    webui_section = config["webui"] if config.has_section("webui") else {}
    webui_enabled = _bool_value(
        extra_params.get("webui_enabled")
        or webui_section.get("enabled")
        or config.get("webui", "enabled", fallback="false"),
        False,
    )
    if not webui_enabled:
        return {}
    model = _compact_text(
        extra_params.get("webui_model")
        or webui_section.get("model")
        or config.get("webui", "model", fallback="")
    )
    if not model:
        logger.warning("Summary AI: WebUI model is not configured; using deterministic summary text.")
        return {}
    if not items:
        return {}

    base_url = _compact_text(
        extra_params.get("webui_url")
        or webui_section.get("url")
        or config.get("webui", "url", fallback="http://localhost:3000")
    )
    endpoint = _compact_text(
        extra_params.get("webui_endpoint")
        or webui_section.get("endpoint")
        or config.get("webui", "endpoint", fallback="/api/chat/completions")
    )
    api_url = _build_webui_api_url(base_url, endpoint)
    webui_api_key = _strip_wrapping_quotes(
        _compact_text(
            extra_params.get("webui_api_key")
            or webui_section.get("api_key")
            or config.get("webui", "api_key", fallback="")
        )
    )
    timeout_seconds = int(
        _compact_text(
            extra_params.get("webui_timeout_seconds")
            or webui_section.get("timeout_seconds")
            or config.get("webui", "timeout_seconds", fallback="120")
        )
        or "120"
    )
    connect_timeout_seconds = int(
        _compact_text(
            extra_params.get("webui_connect_timeout_seconds")
            or webui_section.get("connect_timeout_seconds")
            or config.get("webui", "connect_timeout_seconds", fallback="10")
        )
        or "10"
    )
    temperature = float(
        _compact_text(
            extra_params.get("webui_temperature")
            or webui_section.get("temperature")
            or config.get("webui", "temperature", fallback="0.2")
        )
        or "0.2"
    )

    headers: dict[str, str] = {"Content-Type": "application/json"}
    if webui_api_key:
        headers["Authorization"] = f"Bearer {webui_api_key}"

    rewritten: dict[str, str] = {}
    batch_size = 8
    batches = [(i, items[i : i + batch_size]) for i in range(0, len(items), batch_size)]
    max_workers = _parallel_workers(extra_params)

    def _run_batch(batch_info: tuple[int, list[dict[str, str]]]) -> dict[str, str]:
        start_index, batch = batch_info
        target_map, prompt = _build_summary_prompt(batch, start_index=start_index + 1)

        def _request():
            return requests.post(
                api_url,
                headers=headers,
                json={
                    "model": model,
                    "messages": [
                        {
                            "role": "system",
                            "content": (
                                "You rewrite software task evidence into short, business-facing achievement statements. "
                                "Return only strict JSON."
                            ),
                        },
                        {"role": "user", "content": prompt},
                    ],
                    "stream": False,
                    "temperature": temperature,
                },
                timeout=(connect_timeout_seconds, timeout_seconds),
            )

        try:
            response = retry_ai_call(_request, logger=logger)
            response.raise_for_status()
            response_json = response.json()
            response_text = ""
            choices = response_json.get("choices")
            if isinstance(choices, list) and choices:
                first_choice = choices[0] or {}
                message = first_choice.get("message") or {}
                response_text = str(message.get("content", "") or "")
            if not response_text:
                response_text = str(response_json.get("response", "") or "")
            rewrite_map = _extract_json_object(response_text) or {}
            batch_result: dict[str, str] = {}
            for target_id, issue_id in target_map.items():
                candidate = _sanitize_summary_ai_text(rewrite_map.get(target_id))
                if candidate:
                    batch_result[issue_id] = candidate
            return batch_result
        except Exception as exc:
            logger.warning("Summary AI (WebUI) batch %s failed: %s", start_index // batch_size + 1, exc)
            return {}

    for batch_result in parallel_map(_run_batch, batches, max_workers=max_workers):
        rewritten.update(batch_result)

    return rewritten


def rewrite_summary_items_with_ai(
    items: list[dict[str, str]],
    config: ConfigParser,
    extra_params: dict[str, Any],
) -> dict[str, str]:
    section = config["jira_comprehensive"] if config.has_section("jira_comprehensive") else {}
    provider_raw = _compact_text(extra_params.get("ai_provider") or section.get("ai_provider"))
    if provider_raw:
        provider = provider_raw.casefold()
    else:
        webui_enabled = _bool_value(
            extra_params.get("webui_enabled", config.get("webui", "enabled", fallback="false")),
            False,
        )
        provider = "webui" if webui_enabled else "ollama"
    if provider == "webui":
        return _rewrite_summary_items_with_webui(items, config, extra_params)
    if provider not in {"", "ollama"}:
        logger.warning("Summary AI: unknown ai_provider=%s, falling back to ollama.", provider)
    return _rewrite_summary_items_with_ollama(items, config, extra_params)


def _rewrite_comment_items_with_ollama(
    items: list[dict[str, str]],
    config: ConfigParser,
    extra_params: dict[str, Any],
) -> dict[str, dict[str, Any]]:
    ollama_enabled = _bool_value(
        extra_params.get("ollama_enabled", config.get("ollama", "enabled", fallback="true")),
        True,
    )
    if not ollama_enabled:
        return {}
    model = _compact_text(extra_params.get("ollama_model") or config.get("ollama", "model", fallback=""))
    if not model:
        logger.warning("Comments AI: Ollama model is not configured; using empty summaries.")
        return {}
    if not items:
        return {}

    ollama_url = _compact_text(extra_params.get("ollama_url") or config.get("ollama", "url", fallback="http://localhost:11434"))
    ollama_api_key = _strip_wrapping_quotes(
        _compact_text(extra_params.get("ollama_api_key") or config.get("ollama", "api_key", fallback=""))
    )
    timeout_seconds = int(
        _compact_text(extra_params.get("ollama_timeout_seconds") or config.get("ollama", "timeout_seconds", fallback="60"))
        or "60"
    )
    temperature = float(
        _compact_text(extra_params.get("ollama_temperature") or config.get("ollama", "temperature", fallback="0.2"))
        or "0.2"
    )

    headers: dict[str, str] = {"Content-Type": "application/json"}
    if ollama_api_key:
        headers["Authorization"] = f"Bearer {ollama_api_key}"

    rewritten: dict[str, dict[str, Any]] = {}
    batch_size = 8
    batches = [(i, items[i : i + batch_size]) for i in range(0, len(items), batch_size)]
    max_workers = _parallel_workers(extra_params)

    def _run_batch(batch_info: tuple[int, list[dict[str, str]]]) -> dict[str, dict[str, Any]]:
        start_index, batch = batch_info
        target_map, prompt = _build_comment_summary_prompt(batch, start_index=start_index + 1)

        def _request():
            return requests.post(
                f"{ollama_url.rstrip('/')}/api/generate",
                headers=headers,
                json={
                    "model": model,
                    "prompt": prompt,
                    "stream": False,
                    "options": {"temperature": temperature},
                },
                timeout=timeout_seconds,
            )

        try:
            response = retry_ai_call(_request, logger=logger)
            response.raise_for_status()
            response_json = response.json()
            response_text = str(response_json.get("response", "") or "")
            rewrite_map = _extract_json_object(response_text) or {}
            batch_result: dict[str, dict[str, Any]] = {}
            for target_id, issue_id in target_map.items():
                candidate = rewrite_map.get(target_id)
                if isinstance(candidate, dict):
                    batch_result[issue_id] = candidate
            return batch_result
        except Exception as exc:
            logger.warning("Comments AI (Ollama) batch %s failed: %s", start_index // batch_size + 1, exc)
            return {}

    for batch_result in parallel_map(_run_batch, batches, max_workers=max_workers):
        rewritten.update(batch_result)

    return rewritten


def _rewrite_comment_items_with_webui(
    items: list[dict[str, str]],
    config: ConfigParser,
    extra_params: dict[str, Any],
) -> dict[str, dict[str, Any]]:
    webui_section = config["webui"] if config.has_section("webui") else {}
    webui_enabled = _bool_value(
        extra_params.get("webui_enabled")
        or webui_section.get("enabled")
        or config.get("webui", "enabled", fallback="false"),
        False,
    )
    if not webui_enabled:
        return {}
    model = _compact_text(
        extra_params.get("webui_model")
        or webui_section.get("model")
        or config.get("webui", "model", fallback="")
    )
    if not model:
        logger.warning("Comments AI: WebUI model is not configured; using empty summaries.")
        return {}
    if not items:
        return {}

    base_url = _compact_text(
        extra_params.get("webui_url")
        or webui_section.get("url")
        or config.get("webui", "url", fallback="http://localhost:3000")
    )
    endpoint = _compact_text(
        extra_params.get("webui_endpoint")
        or webui_section.get("endpoint")
        or config.get("webui", "endpoint", fallback="/api/chat/completions")
    )
    api_url = _build_webui_api_url(base_url, endpoint)
    webui_api_key = _strip_wrapping_quotes(
        _compact_text(
            extra_params.get("webui_api_key")
            or webui_section.get("api_key")
            or config.get("webui", "api_key", fallback="")
        )
    )
    timeout_seconds = int(
        _compact_text(
            extra_params.get("webui_timeout_seconds")
            or webui_section.get("timeout_seconds")
            or config.get("webui", "timeout_seconds", fallback="120")
        )
        or "120"
    )
    connect_timeout_seconds = int(
        _compact_text(
            extra_params.get("webui_connect_timeout_seconds")
            or webui_section.get("connect_timeout_seconds")
            or config.get("webui", "connect_timeout_seconds", fallback="10")
        )
        or "10"
    )
    temperature = float(
        _compact_text(
            extra_params.get("webui_temperature")
            or webui_section.get("temperature")
            or config.get("webui", "temperature", fallback="0.2")
        )
        or "0.2"
    )

    headers: dict[str, str] = {"Content-Type": "application/json"}
    if webui_api_key:
        headers["Authorization"] = f"Bearer {webui_api_key}"

    rewritten: dict[str, dict[str, Any]] = {}
    batch_size = 8
    batches = [(i, items[i : i + batch_size]) for i in range(0, len(items), batch_size)]
    max_workers = _parallel_workers(extra_params)

    def _run_batch(batch_info: tuple[int, list[dict[str, str]]]) -> dict[str, dict[str, Any]]:
        start_index, batch = batch_info
        target_map, prompt = _build_comment_summary_prompt(batch, start_index=start_index + 1)

        def _request():
            return requests.post(
                api_url,
                headers=headers,
                json={
                    "model": model,
                    "messages": [
                        {
                            "role": "system",
                            "content": (
                                "You summarize Jira comment activity into structured progress JSON. "
                                "Return only strict JSON."
                            ),
                        },
                        {"role": "user", "content": prompt},
                    ],
                    "stream": False,
                    "temperature": temperature,
                },
                timeout=(connect_timeout_seconds, timeout_seconds),
            )

        try:
            response = retry_ai_call(_request, logger=logger)
            response.raise_for_status()
            response_json = response.json()
            response_text = ""
            choices = response_json.get("choices")
            if isinstance(choices, list) and choices:
                first_choice = choices[0] or {}
                message = first_choice.get("message") or {}
                response_text = str(message.get("content", "") or "")
            if not response_text:
                response_text = str(response_json.get("response", "") or "")
            rewrite_map = _extract_json_object(response_text) or {}
            batch_result: dict[str, dict[str, Any]] = {}
            for target_id, issue_id in target_map.items():
                candidate = rewrite_map.get(target_id)
                if isinstance(candidate, dict):
                    batch_result[issue_id] = candidate
            return batch_result
        except Exception as exc:
            logger.warning("Comments AI (WebUI) batch %s failed: %s", start_index // batch_size + 1, exc)
            return {}

    for batch_result in parallel_map(_run_batch, batches, max_workers=max_workers):
        rewritten.update(batch_result)

    return rewritten


def rewrite_comment_items_with_ai(
    items: list[dict[str, str]],
    config: ConfigParser,
    extra_params: dict[str, Any],
) -> dict[str, dict[str, Any]]:
    if not _bool_value(extra_params.get("ai_comments_enabled"), False):
        return {}
    section = config["jira_comprehensive"] if config.has_section("jira_comprehensive") else {}
    provider_raw = _compact_text(extra_params.get("ai_provider") or section.get("ai_provider"))
    if provider_raw:
        provider = provider_raw.casefold()
    else:
        webui_enabled = _bool_value(
            extra_params.get("webui_enabled", config.get("webui", "enabled", fallback="false")),
            False,
        )
        provider = "webui" if webui_enabled else "ollama"
    if provider == "webui":
        return _rewrite_comment_items_with_webui(items, config, extra_params)
    if provider not in {"", "ollama"}:
        logger.warning("Comments AI: unknown ai_provider=%s, falling back to ollama.", provider)
    return _rewrite_comment_items_with_ollama(items, config, extra_params)


def build_monthly_summary_df(
    issues_df: pd.DataFrame,
    config: ConfigParser,
    extra_params: dict[str, Any],
) -> pd.DataFrame:
    columns = [
        "Epic_Link",
        "Epic_Name",
        "Summary",
        "Planned_Tasks_Resolved",
        "Reported_Issues_Resolved",
    ]
    if issues_df.empty:
        return pd.DataFrame(columns=columns)

    resolved_mask = _resolved_mask(issues_df) & _countable_mask(issues_df)
    resolved_df = issues_df[resolved_mask].copy()
    if resolved_df.empty:
        return pd.DataFrame(columns=columns)

    for required_col in ("Type", "Issue_Key", "Epic_Name", "Epic_Link"):
        if required_col not in resolved_df.columns:
            resolved_df[required_col] = ""
    for optional_col in (
        "Resolved",
        "Last_Comment",
        "Description",
        "Summary",
        "Labels",
        "Status",
        "Parent",
        "Parent_Key",
        "Parent_Summary",
        "Epic_Labels",
        "Epic_Status",
        "Epic_Resolved",
    ):
        if optional_col not in resolved_df.columns:
            resolved_df[optional_col] = ""
    if "Parent_Key" in resolved_df.columns:
        resolved_df["Parent"] = resolved_df["Parent"].fillna("").astype(str)
        parent_keys = resolved_df["Parent_Key"].fillna("").astype(str)
        empty_parent_mask = resolved_df["Parent"].str.strip().eq("")
        resolved_df.loc[empty_parent_mask, "Parent"] = parent_keys.loc[empty_parent_mask]

    epic_metadata_map = _build_epic_metadata_map(issues_df)
    if epic_metadata_map:
        for col_name, meta_key in (
            ("Epic_Labels", "labels"),
            ("Epic_Status", "status"),
            ("Epic_Resolved", "resolved"),
        ):
            values: list[str] = []
            for _, row in resolved_df.iterrows():
                current_value = _compact_text(row.get(col_name))
                if current_value:
                    values.append(current_value)
                    continue
                epic_key = _compact_text(row.get("Epic_Link"))
                values.append(_compact_text(epic_metadata_map.get(epic_key, {}).get(meta_key)))
            resolved_df[col_name] = values

    resolved_df["_type_norm"] = resolved_df["Type"].fillna("").astype(str).map(_normalize_text)
    resolved_df["_epic_norm"] = resolved_df["Epic_Name"].fillna("").astype(str).map(_normalize_text)
    resolved_df["_resolved_sort"] = pd.to_datetime(resolved_df["Resolved"], errors="coerce")
    resolved_df = resolved_df.sort_values(
        by=["_epic_norm", "_resolved_sort", "Issue_Key"],
        ascending=[True, True, True],
        na_position="last",
        kind="mergesort",
    )

    summary_start_dt, summary_end_dt = _summary_period(extra_params)
    epic_payloads: list[dict[str, Any]] = []
    ai_inputs: list[dict[str, str]] = []
    for (epic_link_raw, epic_name_raw), epic_df in resolved_df.groupby(["Epic_Link", "Epic_Name"], dropna=False, sort=False):
        epic_link = _compact_text(epic_link_raw)
        if not epic_link:
            continue
        epic_name = _compact_text(epic_name_raw) or "Unknown Epic"

        epic_meta = epic_metadata_map.get(epic_link, {})
        epic_labels_raw = _first_value(list(epic_df.get("Epic_Labels", pd.Series([], dtype=str)).tolist())) or epic_meta.get("labels", "")
        epic_status_raw = _first_value(list(epic_df.get("Epic_Status", pd.Series([], dtype=str)).tolist())) or epic_meta.get("status", "")
        epic_resolved_raw = _first_value(list(epic_df.get("Epic_Resolved", pd.Series([], dtype=str)).tolist())) or epic_meta.get("resolved", "")

        has_report_label = "report" in _parse_label_set(epic_labels_raw)
        epic_resolved_dt = _parse_iso_date(epic_resolved_raw)
        epic_status_norm = _normalize_text(epic_status_raw)
        epic_is_open = epic_resolved_dt is None and epic_status_norm not in _DONE_STATUSES
        epic_closed_in_period = _resolved_in_period(epic_resolved_dt, summary_start_dt, summary_end_dt)
        if not (has_report_label and (epic_is_open or epic_closed_in_period)):
            continue

        bug_count = int((epic_df["_type_norm"] == "bug").sum())

        planned_df = epic_df[~epic_df["_type_norm"].isin({"bug", "epic"})]
        planned_issue_count = int(planned_df.shape[0])
        grouped_items, grouped_ai_inputs = _compose_grouped_summary_inputs(planned_df, epic_link)
        ai_inputs.extend(grouped_ai_inputs)

        epic_payloads.append(
            {
                "epic_link": epic_link,
                "epic_name": epic_name,
                "planned_items": grouped_items,
                "planned_issue_count": planned_issue_count,
                "bug_count": bug_count,
            }
        )

    rewrite_map = rewrite_summary_items_with_ai(ai_inputs, config, extra_params) if ai_inputs else {}

    rows: list[dict[str, Any]] = []
    for epic in epic_payloads:
        planned_items = epic["planned_items"]
        bug_count = int(epic["bug_count"])
        planned_count = int(epic.get("planned_issue_count", len(planned_items)))

        summary_lines: list[str] = []
        for item in planned_items:
            rewritten = _sanitize_summary_ai_text(rewrite_map.get(item["id"]))
            if not rewritten:
                rewritten = _sanitize_summary_ai_text(item["fallback"]) or item["fallback"]
            summary_lines.append(f"- {rewritten}")

        summary_lines.append(f"Resolved {planned_count} planned tasks on time.")
        if bug_count > 0:
            summary_lines.append(f"Resolved {bug_count} reported issues.")

        rows.append(
            {
                "Epic_Link": epic["epic_link"],
                "Epic_Name": epic["epic_name"],
                "Summary": "\n".join(summary_lines),
                "Planned_Tasks_Resolved": planned_count,
                "Reported_Issues_Resolved": bug_count,
            }
        )

    return pd.DataFrame(rows, columns=columns)


def fetch_worklog_activity(
    jira_source: JiraSource,
    issues_df: pd.DataFrame,
    start_date: str | None = None,
    end_date: str | None = None,
) -> pd.DataFrame:
    """
    Fetch aggregated worklog activity per issue and author.

    Returns:
        DataFrame with Issue_Key, Summary, Assignee, Worklog_Author, Total_Hours,
        First_Log_Date, Last_Log_Date
    """
    if issues_df.empty:
        return pd.DataFrame(
            columns=["Issue_Key", "Summary", "Assignee", "Worklog_Author", "Total_Hours", "First_Log_Date", "Last_Log_Date"]
        )

    start_dt = datetime.strptime(start_date, "%Y-%m-%d").date() if start_date else None
    end_dt = datetime.strptime(end_date, "%Y-%m-%d").date() if end_date else None

    summary_map = issues_df.set_index("Issue_Key")["Summary"].to_dict()
    assignee_map = issues_df.set_index("Issue_Key")["Assignee"].to_dict()

    agg: dict[tuple[str, str], dict[str, Any]] = {}

    for issue_key in issues_df["Issue_Key"].dropna().unique():
        worklogs = jira_source.get_all_worklogs(issue_key)
        for log in worklogs:
            try:
                author = log.get("author", {}).get("displayName", "") or "Unknown"
                log_date = datetime.strptime(log["started"].split("T")[0], "%Y-%m-%d").date()
            except Exception:
                continue

            if start_dt and log_date < start_dt:
                continue
            if end_dt and log_date > end_dt:
                continue

            time_spent = int(log.get("timeSpentSeconds") or 0)
            key = (issue_key, author)
            entry = agg.setdefault(
                key,
                {
                    "Issue_Key": issue_key,
                    "Summary": summary_map.get(issue_key, ""),
                    "Assignee": assignee_map.get(issue_key, ""),
                    "Worklog_Author": author,
                    "Total_Seconds": 0,
                    "First_Log_Date": log_date,
                    "Last_Log_Date": log_date,
                },
            )
            entry["Total_Seconds"] += time_spent
            if log_date < entry["First_Log_Date"]:
                entry["First_Log_Date"] = log_date
            if log_date > entry["Last_Log_Date"]:
                entry["Last_Log_Date"] = log_date

    rows = []
    for entry in agg.values():
        rows.append(
            {
                "Issue_Key": entry["Issue_Key"],
                "Summary": entry["Summary"],
                "Assignee": entry["Assignee"],
                "Worklog_Author": entry["Worklog_Author"],
                "Total_Hours": round(entry["Total_Seconds"] / 3600, 2),
                "First_Log_Date": entry["First_Log_Date"].strftime("%Y-%m-%d"),
                "Last_Log_Date": entry["Last_Log_Date"].strftime("%Y-%m-%d"),
            }
        )

    activity_df = pd.DataFrame(rows)
    if activity_df.empty:
        return activity_df

    multi_author_issues = (
        activity_df.groupby("Issue_Key")["Worklog_Author"]
        .nunique()
        .reset_index()
    )
    multi_author_issues = multi_author_issues[multi_author_issues["Worklog_Author"] > 1]["Issue_Key"]
    activity_df = activity_df[activity_df["Issue_Key"].isin(multi_author_issues)]
    activity_df = activity_df.sort_values(by=["Issue_Key", "Worklog_Author"])
    return activity_df


def fetch_worklog_entries(
    jira_source: JiraSource,
    issues_df: pd.DataFrame,
    start_date: str | None = None,
    end_date: str | None = None,
) -> pd.DataFrame:
    """
    Fetch raw worklog entries for the requested period.

    Returns:
        DataFrame with Issue_Key, Summary, Assignee, Worklog_Author, Date, Time_Spent_Hours, Comment
    """
    if issues_df.empty:
        return pd.DataFrame(
            columns=["Issue_Key", "Summary", "Assignee", "Worklog_Author", "Date", "Time_Spent_Hours", "Comment"]
        )

    start_dt = datetime.strptime(start_date, "%Y-%m-%d").date() if start_date else None
    end_dt = datetime.strptime(end_date, "%Y-%m-%d").date() if end_date else None

    summary_map = issues_df.set_index("Issue_Key")["Summary"].to_dict()
    assignee_map = issues_df.set_index("Issue_Key")["Assignee"].to_dict()

    rows: list[dict[str, Any]] = []

    for issue_key in issues_df["Issue_Key"].dropna().unique():
        worklogs = jira_source.get_all_worklogs(issue_key)
        for log in worklogs:
            try:
                author = log.get("author", {}).get("displayName", "") or "Unknown"
                log_date = datetime.strptime(log["started"].split("T")[0], "%Y-%m-%d").date()
            except Exception:
                continue

            if start_dt and log_date < start_dt:
                continue
            if end_dt and log_date > end_dt:
                continue

            time_spent = int(log.get("timeSpentSeconds") or 0)
            rows.append(
                {
                    "Issue_Key": issue_key,
                    "Summary": summary_map.get(issue_key, ""),
                    "Assignee": assignee_map.get(issue_key, ""),
                    "Worklog_Author": author,
                    "Date": log_date.strftime("%Y-%m-%d"),
                    "Time_Spent_Hours": round(time_spent / 3600, 2),
                    "Comment": log.get("comment", "") or "",
                }
            )

    entries_df = pd.DataFrame(rows)
    if not entries_df.empty:
        entries_df = entries_df.sort_values(by=["Date", "Issue_Key"])
    return entries_df


def read_member_list(member_list_file: str) -> pd.DataFrame:
    """Read team member details from Excel file."""
    if not os.path.exists(member_list_file):
        logger.warning("Member list file %r not found", member_list_file)
        return pd.DataFrame(columns=["name", "email", "username", "role"])

    df = pd.read_excel(member_list_file)
    required_columns = ["name", "username", "role"]
    for col in required_columns:
        if col not in df.columns:
            raise ValueError(f"Member list file must contain '{col}' column")
    return df


def read_code_volume(code_volume_file: str | None) -> pd.DataFrame:
    """Read code volume data from Excel file."""
    if not code_volume_file or not os.path.exists(code_volume_file):
        return pd.DataFrame(columns=["username", "code_volume"])
    return pd.read_excel(code_volume_file)


def calculate_engineer_metrics(
    issues_df: pd.DataFrame,
    members_df: pd.DataFrame,
    code_volume_df: pd.DataFrame,
    worklog_entries_df: pd.DataFrame,
) -> pd.DataFrame:
    """Calculate metrics for Engineers."""
    metrics: list[dict[str, Any]] = []
    member_role_norm = members_df.get(
        "role", pd.Series([""] * len(members_df), index=members_df.index)
    ).map(_normalize_text)
    engineers = members_df[member_role_norm.isin({"engineer", "huawei"})]

    jira_column = next(
        (
            col
            for col in members_df.columns
            if _normalize_text(col) in {"jira", "jira username", "jira_user", "jira account"}
        ),
        None,
    )

    assignee_username_value = issues_df.get(
        "Assignee_Username", pd.Series([""] * len(issues_df), index=issues_df.index)
    )
    assignee_username_norm = assignee_username_value.fillna("").astype(str).map(_normalize_text)

    assignee_value = issues_df.get("Assignee", pd.Series([""] * len(issues_df), index=issues_df.index))
    assignee_name_norm = assignee_value.fillna("").astype(str).map(_normalize_text)
    status_resolved_mask = _resolved_mask(issues_df)
    countable_mask = _countable_mask(issues_df)
    labels_value = issues_df.get("Labels")
    labels_norm = (
        labels_value.fillna("").astype(str)
        if labels_value is not None
        else pd.Series([""] * len(issues_df), index=issues_df.index)
    )
    issue_assignee_username_by_key: dict[str, str] = {}
    issue_assignee_name_by_key: dict[str, str] = {}
    if "Issue_Key" in issues_df.columns:
        issue_keys_value = issues_df.get("Issue_Key", pd.Series([""] * len(issues_df), index=issues_df.index))
        for idx in issues_df.index:
            issue_key_norm = _normalize_text(issue_keys_value.loc[idx])
            if not issue_key_norm:
                continue
            if issue_key_norm in issue_assignee_username_by_key:
                continue
            issue_assignee_username_by_key[issue_key_norm] = _normalize_text(assignee_username_value.loc[idx])
            issue_assignee_name_by_key[issue_key_norm] = _normalize_text(assignee_value.loc[idx])

    for _, engineer in engineers.iterrows():
        username = str(engineer.get("username", "")).strip()
        username_norm = _normalize_text(username)
        name = engineer.get("name", "")
        name_norm = _normalize_text(name)

        jira_username_raw = engineer.get(jira_column, "") if jira_column else ""
        jira_username = _first_value([jira_username_raw, username]) or ""
        jira_username_norm = _normalize_text(jira_username)

        identifier_candidates = {jira_username_norm, username_norm} - {""}
        assistance_identity = set(identifier_candidates)
        if name_norm:
            assistance_identity.add(name_norm)

        user_mask = assignee_username_norm.isin(identifier_candidates)
        if name_norm:
            user_mask = user_mask | (assignee_name_norm == name_norm)

        user_issues = issues_df[user_mask & countable_mask]
        resolved_issues = user_issues[status_resolved_mask.loc[user_issues.index]]

        code_volume = 0
        if not code_volume_df.empty and "username" in code_volume_df.columns:
            cv_row = code_volume_df[code_volume_df["username"].fillna("").astype(str).map(_normalize_text) == username_norm]
            if not cv_row.empty:
                code_volume = cv_row.iloc[0].get("code_volume", 0)

        bugs = resolved_issues[resolved_issues["Type"] == "Bug"].shape[0]
        features = resolved_issues[
            resolved_issues["Type"].isin(["Story", "New Feature", "Improvement"])
        ].shape[0]
        code_quality = bugs / features if features > 0 else 0

        doc_tasks = labels_norm.loc[resolved_issues.index].str.contains("documentation", case=False, na=False).sum()
        assistance_issue_keys: set[str] = set()
        if (
            not worklog_entries_df.empty
            and "Issue_Key" in worklog_entries_df.columns
            and "Worklog_Author" in worklog_entries_df.columns
            and assistance_identity
        ):
            worklog_issue_keys = worklog_entries_df["Issue_Key"].fillna("").astype(str).map(_normalize_text)
            worklog_authors = worklog_entries_df["Worklog_Author"].fillna("").astype(str).map(_normalize_text)
            worklog_assignees = (
                worklog_entries_df.get(
                    "Assignee",
                    pd.Series([""] * len(worklog_entries_df), index=worklog_entries_df.index),
                )
                .fillna("")
                .astype(str)
                .map(_normalize_text)
            )
            author_mask = worklog_authors.isin(assistance_identity)
            for idx in worklog_entries_df[author_mask].index:
                issue_key_norm = worklog_issue_keys.loc[idx]
                if not issue_key_norm:
                    continue
                issue_assignee_candidates = {
                    issue_assignee_username_by_key.get(issue_key_norm, ""),
                    issue_assignee_name_by_key.get(issue_key_norm, ""),
                    worklog_assignees.loc[idx],
                } - {""}
                if issue_assignee_candidates & assistance_identity:
                    continue
                assistance_issue_keys.add(issue_key_norm)
        assistance_provided = len(assistance_issue_keys)

        metrics.append(
            {
                "Name": name,
                "Role": "Engineer",
                "Code_Volume": code_volume,
                "Code_Quality_Score": round(1 / (1 + code_quality), 2) if features > 0 else 1.0,
                "Bugs": bugs,
                "Features": features,
                "Documentation_Tasks": doc_tasks,
                "Outstanding_Contribution": 0,
                "Assistance_Provided": int(assistance_provided),
                "Total_Resolved_Issues": resolved_issues.shape[0],
            }
        )

    return pd.DataFrame(metrics)


def calculate_qa_metrics(issues_df: pd.DataFrame, members_df: pd.DataFrame) -> pd.DataFrame:
    """Calculate metrics for QA Engineers."""
    metrics: list[dict[str, Any]] = []
    member_role_norm = members_df.get(
        "role", pd.Series([""] * len(members_df), index=members_df.index)
    ).map(_normalize_text)
    qa_engineers = members_df[member_role_norm.isin({"qa engineer", "test engineer", "tester", "qa"})]

    jira_column = next(
        (
            col
            for col in members_df.columns
            if _normalize_text(col) in {"jira", "jira username", "jira_user", "jira account"}
        ),
        None,
    )

    assignee_username_value = issues_df.get(
        "Assignee_Username", pd.Series([""] * len(issues_df), index=issues_df.index)
    )
    assignee_username_norm = assignee_username_value.fillna("").astype(str).map(_normalize_text)

    assignee_value = issues_df.get("Assignee", pd.Series([""] * len(issues_df), index=issues_df.index))
    assignee_name_norm = assignee_value.fillna("").astype(str).map(_normalize_text)

    reporter_username_value = issues_df.get(
        "Reporter_Username", pd.Series([""] * len(issues_df), index=issues_df.index)
    )
    reporter_username_norm = reporter_username_value.fillna("").astype(str).map(_normalize_text)

    reporter_value = issues_df.get("Reporter", pd.Series([""] * len(issues_df), index=issues_df.index))
    reporter_name_norm = reporter_value.fillna("").astype(str).map(_normalize_text)
    status_resolved_mask = _resolved_mask(issues_df)
    countable_mask = _countable_mask(issues_df)
    labels_value = issues_df.get("Labels")
    labels_norm = (
        labels_value.fillna("").astype(str)
        if labels_value is not None
        else pd.Series([""] * len(issues_df), index=issues_df.index)
    )
    summary_value = issues_df.get("Summary", pd.Series([""] * len(issues_df), index=issues_df.index))
    summary_norm = summary_value.fillna("").astype(str)

    for _, qa in qa_engineers.iterrows():
        username = str(qa.get("username", "")).strip()
        username_norm = _normalize_text(username)
        name = qa.get("name", "")
        name_norm = _normalize_text(name)

        jira_username_raw = qa.get(jira_column, "") if jira_column else ""
        jira_username = _first_value([jira_username_raw, username]) or ""
        jira_username_norm = _normalize_text(jira_username)
        identifier_candidates = {jira_username_norm, username_norm} - {""}

        user_mask = assignee_username_norm.isin(identifier_candidates)
        if name_norm:
            user_mask = user_mask | (assignee_name_norm == name_norm)

        user_issues = issues_df[user_mask & countable_mask]
        resolved_issues = user_issues[status_resolved_mask.loc[user_issues.index]]
        tt_totals = {key: 0 for key in _TT_COUNTER_PATTERNS}
        for payload in resolved_issues.get("Comments", pd.Series([], dtype=object)).fillna("").astype(str):
            counters = _extract_tt_counters(payload)
            for key in tt_totals:
                tt_totals[key] += counters.get(key, 0)

        test_scenarios = tt_totals["TT_tdev_APIs"] + tt_totals["TT_tested_APIs"]
        perf_tasks = tt_totals["TT_tested_perf"] + tt_totals["TT_tdev_perf"]

        reporter_mask = reporter_username_norm.isin(identifier_candidates)
        if name_norm:
            reporter_mask = reporter_mask | (reporter_name_norm == name_norm)
        bugs_created = issues_df[(reporter_mask & countable_mask) & (issues_df["Type"] == "Bug")].shape[0]

        doc_tasks = labels_norm.loc[resolved_issues.index].str.contains("documentation", case=False, na=False).sum()
        outstanding_tasks = (
            labels_norm.loc[resolved_issues.index].str.contains(
                _OUTSTANDING_CONTRIBUTION_PATTERN, regex=True, na=False
            )
            | summary_norm.loc[resolved_issues.index].str.contains(
                _OUTSTANDING_CONTRIBUTION_PATTERN, regex=True, na=False
            )
        ).sum()

        metrics.append(
            {
                "Name": name,
                "Role": "QA Engineer",
                "Test_Scenarios_Executed": test_scenarios,
                "Issues_Raised": bugs_created,
                "Performance_Benchmarks": perf_tasks,
                "Documentation_Tasks": doc_tasks,
                "TT_tdev_APIs": tt_totals["TT_tdev_APIs"],
                "TT_tested_APIs": tt_totals["TT_tested_APIs"],
                "TT_tested_perf": tt_totals["TT_tested_perf"],
                "TT_tdev_perf": tt_totals["TT_tdev_perf"],
                "Outstanding_Contribution": int(outstanding_tasks),
                "Total_Resolved_Issues": resolved_issues.shape[0],
            }
        )

    return pd.DataFrame(metrics)


def calculate_pm_metrics(
    issues_df: pd.DataFrame, members_df: pd.DataFrame, jira, jql_query: str
) -> pd.DataFrame:
    """Calculate metrics for Project Managers."""
    metrics: list[dict[str, Any]] = []
    member_role_norm = members_df.get(
        "role", pd.Series([""] * len(members_df), index=members_df.index)
    ).map(_normalize_text)
    pms = members_df[member_role_norm.isin({"project manager", "pm"})]

    epic_count = issues_df[issues_df["Type"] == "Epic"].shape[0]
    resolved_issues_mask = _resolved_mask(issues_df)
    countable_mask = _countable_mask(issues_df)
    resolved_countable_mask = resolved_issues_mask & countable_mask

    for _, pm in pms.iterrows():
        name = pm["name"]

        total_closed = int(resolved_countable_mask.sum())

        labels_value = issues_df.get("Labels")
        labels_norm = (
            labels_value.fillna("").astype(str)
            if labels_value is not None
            else pd.Series([""] * len(issues_df), index=issues_df.index)
        )
        doc_tasks = labels_norm[resolved_countable_mask].str.contains("documentation", case=False, na=False).sum()

        metrics.append(
            {
                "Name": name,
                "Role": "Project Manager",
                "Epics_Created": epic_count,
                "Total_Closed_Tasks": total_closed,
                "Documentation_Tasks": doc_tasks,
            }
        )

    return pd.DataFrame(metrics)


def export_to_excel(
    issues_df: pd.DataFrame,
    links_df: pd.DataFrame,
    results_df: pd.DataFrame,
    summary_df: pd.DataFrame,
    comments_period_df: pd.DataFrame,
    engineer_metrics: pd.DataFrame,
    qa_metrics: pd.DataFrame,
    pm_metrics: pd.DataFrame,
    worklog_activity_df: pd.DataFrame,
    worklog_entries_df: pd.DataFrame,
    output_file: str | Path,
) -> None:
    """Export all data to Excel file with multiple sheets."""
    issues_df = _sanitize_dataframe_for_excel(issues_df)
    links_df = _sanitize_dataframe_for_excel(links_df)
    results_df = _sanitize_dataframe_for_excel(results_df)
    summary_df = _sanitize_dataframe_for_excel(summary_df)
    comments_period_df = _sanitize_dataframe_for_excel(comments_period_df)
    engineer_metrics = _sanitize_dataframe_for_excel(engineer_metrics)
    qa_metrics = _sanitize_dataframe_for_excel(qa_metrics)
    pm_metrics = _sanitize_dataframe_for_excel(pm_metrics)
    worklog_activity_df = _sanitize_dataframe_for_excel(worklog_activity_df)
    worklog_entries_df = _sanitize_dataframe_for_excel(worklog_entries_df)

    output_path = Path(output_file)
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        issues_df.to_excel(writer, sheet_name="Issues", index=False)

        if not links_df.empty:
            links_df.to_excel(writer, sheet_name="Links", index=False)

        if not results_df.empty:
            results_df.to_excel(writer, sheet_name="Results", index=False)

        if not summary_df.empty:
            summary_df.to_excel(writer, sheet_name="Summary", index=False)

        comments_period_df.to_excel(writer, sheet_name="Comments_Period", index=False)

        if not engineer_metrics.empty:
            engineer_metrics.to_excel(writer, sheet_name="Engineer_Performance", index=False)

        if not qa_metrics.empty:
            qa_metrics.to_excel(writer, sheet_name="QA_Performance", index=False)

        if not pm_metrics.empty:
            pm_metrics.to_excel(writer, sheet_name="PM_Performance", index=False)

        if not worklog_activity_df.empty:
            worklog_activity_df.to_excel(writer, sheet_name="Worklog_Activity", index=False)

        if not worklog_entries_df.empty:
            worklog_entries_df.to_excel(writer, sheet_name="Worklog_Entries", index=False)

        workbook = writer.book
        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]

            for cell in worksheet[1]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        max_length = max(max_length, len(str(cell.value)))
                    except Exception:
                        continue
                worksheet.column_dimensions[column_letter].width = min(max_length + 2, 50)

    logger.info("Excel report created: %s", output_path)


def _sanitize_dataframe_for_excel(df: pd.DataFrame) -> pd.DataFrame:
    """
    Strip illegal control characters that cause openpyxl IllegalCharacterError.
    """
    if df.empty:
        return df
    result = df.copy()
    for column in result.columns:
        series = result[column]
        if not (is_object_dtype(series.dtype) or is_string_dtype(series.dtype)):
            continue
        result[column] = series.map(_sanitize_excel_value)
    return result


def _sanitize_excel_value(value: Any) -> Any:
    if isinstance(value, str):
        return ILLEGAL_CHARACTERS_RE.sub("", value)
    return value


def _first_value(values: list[str | None]) -> str | None:
    for value in values:
        if value is None:
            continue
        try:
            if pd.isna(value):
                continue
        except Exception:
            continue
        normalized = str(value).strip()
        if normalized:
            return normalized
    return None


def _extra_param(extra_params: dict[str, Any], *names: str) -> str | None:
    return _first_value([extra_params.get(name) for name in names])


@registry.register
class JiraComprehensiveReport:
    name = "jira_comprehensive"

    def run(
        self,
        dataset: dict,
        config: ConfigParser,
        output_formats: list[str],
        extra_params: dict | None = None,
    ) -> None:
        extra_params = extra_params or {}
        progress = _get_progress(extra_params, self.name, total_steps=5)

        if "excel" not in output_formats:
            logger.warning("jira_comprehensive supports only Excel output. Skipping.")
            return

        params: dict[str, Any] = {}
        params["project"] = _extra_param(extra_params, "project")
        params["start_date"] = _extra_param(extra_params, "start_date", "start-date", "start")
        params["end_date"] = _extra_param(extra_params, "end_date", "end-date", "end")
        params["version"] = _extra_param(extra_params, "version")
        params["epic"] = _extra_param(extra_params, "epic")
        params["jql"] = _extra_param(extra_params, "jql")
        params["member_list_file"] = _extra_param(
            extra_params, "member_list_file", "member-list-file"
        )
        params["member_list_file"] = str(resolve_member_list_path(params["member_list_file"]))
        params["code_volume_file"] = _extra_param(
            extra_params, "code_volume_file", "code-volume-file"
        )

        jql_query = build_jql_query(params)
        comments_jql = build_comments_period_jql(params)

        jira_source = JiraSource(config["jira"])
        jira = jira_source.jira

        with progress.step("Fetch issues"):
            issues_df, links_df, results_df = fetch_jira_data(jira, jql_query)
        if issues_df.empty:
            logger.warning("No issues found matching the query.")
            return
        with progress.step("Build summaries"):
            summary_df = build_monthly_summary_df(issues_df, config, extra_params)
            comments_period_df = build_comments_period_df(
                jira,
                comments_jql,
                params.get("start_date"),
                params.get("end_date"),
                config,
                extra_params,
            )

        with progress.step("Fetch worklogs"):
            max_workers = _parallel_workers(extra_params)
            if max_workers > 1:
                def _activity():
                    return fetch_worklog_activity(
                        jira_source,
                        issues_df,
                        params.get("start_date"),
                        params.get("end_date"),
                    )

                def _entries():
                    return fetch_worklog_entries(
                        jira_source,
                        issues_df,
                        params.get("start_date"),
                        params.get("end_date"),
                    )

                worklog_activity_df, worklog_entries_df = parallel_map(
                    lambda fn: fn(),
                    [_activity, _entries],
                    max_workers=min(max_workers, 2),
                )
            else:
                worklog_activity_df = fetch_worklog_activity(
                    jira_source,
                    issues_df,
                    params.get("start_date"),
                    params.get("end_date"),
                )
                worklog_entries_df = fetch_worklog_entries(
                    jira_source,
                    issues_df,
                    params.get("start_date"),
                    params.get("end_date"),
                )

        members_df = read_member_list(params["member_list_file"])
        code_volume_df = read_code_volume(params["code_volume_file"])

        engineer_metrics = pd.DataFrame()
        qa_metrics = pd.DataFrame()
        pm_metrics = pd.DataFrame()

        with progress.step("Compute metrics"):
            if not members_df.empty:
                engineer_metrics = calculate_engineer_metrics(
                    issues_df,
                    members_df,
                    code_volume_df,
                    worklog_entries_df,
                )
                qa_metrics = calculate_qa_metrics(issues_df, members_df)
                pm_metrics = calculate_pm_metrics(issues_df, members_df, jira, jql_query)
            else:
                logger.warning("No member list found, skipping team performance calculations.")

        output_dir = _extra_param(extra_params, "output_dir") or config.get(
            "reporting", "output_dir", fallback="reports"
        )
        output_base = Path(str(output_dir))
        output_base.mkdir(parents=True, exist_ok=True)

        output_name = _extra_param(extra_params, "output", "output_file")
        if output_name:
            output_path = Path(output_name)
            if not output_path.is_absolute():
                output_path = output_base / output_path
            if output_path.suffix.lower() != ".xlsx":
                output_path = output_path.with_suffix(".xlsx")
        else:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_path = output_base / f"jira_comprehensive_report_{timestamp}.xlsx"

        with progress.step("Export Excel"):
            export_to_excel(
                issues_df,
                links_df,
                results_df,
                summary_df,
                comments_period_df,
                engineer_metrics,
                qa_metrics,
                pm_metrics,
                worklog_activity_df,
                worklog_entries_df,
                output_path,
            )

        resolution_value = issues_df.get("Resolution")
        if resolution_value is not None:
            resolution_norm = resolution_value.fillna("").astype(str).str.strip().str.casefold()
            resolved_count = int(resolution_norm.isin({"done", "resolved"}).sum())
        else:
            resolved_count = 0

        logger.info(
            "REPORT SUMMARY: issues=%s links=%s results=%s summary_epics=%s resolved=%s comments_period=%s",
            len(issues_df),
            len(links_df),
            len(results_df),
            len(summary_df),
            resolved_count,
            len(comments_period_df),
        )
