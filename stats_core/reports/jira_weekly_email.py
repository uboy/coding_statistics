"""
Jira weekly HTML email report with optional Ollama text polishing.
"""

from __future__ import annotations

import difflib
import html
import json
import logging
import platform
import re
import shutil
import subprocess
import sys
import tempfile
from configparser import ConfigParser
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from pathlib import Path
from typing import Any
from urllib.parse import urlsplit, urlunsplit

import requests
from tqdm import tqdm
from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel

from ..sources.jira import JiraSource
from . import registry
from ..utils.ai_retry import retry_ai_call
from ..utils.progress import NoopProgressManager

logger = logging.getLogger(__name__)


_DONE_VALUES = {"done", "resolved", "closed"}
_REPORT_CLOSED_RESOLUTION_VALUES = {"done", "resolved"}
_IN_PROGRESS_VALUES = {"in progress", "in-progress"}
_RISK_LABELS = frozenset({"risk", "issue"})

_SOFFICE_DOWNLOAD_URL = "https://www.libreoffice.org/download/download-libreoffice/"
_OUTLOOK_INFO_URL = "https://www.microsoft.com/en-us/microsoft-365/outlook/email-and-calendar-software-microsoft-outlook"


@dataclass(frozen=True)
class WeekWindow:
    year: int
    week: int
    start: date
    end: date
    key: str


def _bool_value(value: Any, default: bool) -> bool:
    if value is None:
        return default
    if isinstance(value, bool):
        return value
    return str(value).strip().lower() in {"1", "true", "yes", "y", "on"}


def _get_progress(extra_params: dict[str, Any], total_steps: int):
    progress = extra_params.get("progress_manager")
    if progress is None:
        progress = NoopProgressManager()
    progress.set_total(total_steps)
    return progress


def _split_csv(value: str | None, default: list[str]) -> list[str]:
    if not value:
        return default
    raw = str(value).strip()
    if len(raw) >= 2 and ((raw[0] == '"' and raw[-1] == '"') or (raw[0] == "'" and raw[-1] == "'")):
        raw = raw[1:-1].strip()
    items: list[str] = []
    for token in raw.split(","):
        cleaned = token.strip()
        if len(cleaned) >= 2 and (
            (cleaned[0] == '"' and cleaned[-1] == '"') or (cleaned[0] == "'" and cleaned[-1] == "'")
        ):
            cleaned = cleaned[1:-1].strip()
        if cleaned:
            items.append(cleaned)
    return [item for item in items if item]


def _normalize_text(value: Any) -> str:
    if value is None:
        return ""
    return " ".join(str(value).strip().split())


def _normalize_html_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _normalize_key(value: Any) -> str:
    return _normalize_text(value).casefold()


def _contains_cyrillic(text: str) -> bool:
    return bool(re.search(r"[А-Яа-яЁё]", text or ""))


def _clean_comment_for_report(value: Any) -> str:
    cleaned = _normalize_text(value)
    if not cleaned:
        return ""
    # Remove Jira/Confluence-style block markers and headings.
    cleaned = re.sub(r"\{code(?::[^}]*)?\}|\{noformat\}|\{quote\}|\{panel(?::[^}]*)?\}|\{color(?::[^}]*)?\}|\{\/color\}", " ", cleaned, flags=re.IGNORECASE)
    cleaned = re.sub(r"(?<!\w)h[1-6]\.\s*", " ", cleaned, flags=re.IGNORECASE)
    cleaned = re.sub(r"\{[^}]{1,80}\}", " ", cleaned)
    # Remove inline JSON-like fragments and long key/value blobs.
    cleaned = re.sub(r"\[[\{\[][^][]{20,}[\}\]]\]", " ", cleaned)
    cleaned = re.sub(r"\{[^{}]{20,}\}", " ", cleaned)
    cleaned = re.sub(r"\[[^\]]+\]\([^)]+\)", "", cleaned)
    cleaned = re.sub(r"(?:https?://|ftp://|file://|www\.)\S+", "", cleaned, flags=re.IGNORECASE)
    cleaned = re.sub(r"\b[A-Z]+-\d+\b", "", cleaned)
    cleaned = re.sub(r"@\w+", " ", cleaned)
    cleaned = re.sub(r"\b\d{6,}\b", " ", cleaned)
    cleaned = re.sub(r"\b[A-Za-z0-9+/=_-]{20,}\b", " ", cleaned)
    cleaned = re.sub(r"[`*_>#]+", " ", cleaned)
    cleaned = re.sub(r"\s*[\[\]{}|]+\s*", " ", cleaned)
    cleaned = _normalize_text(cleaned.strip(" -,:;"))
    if not cleaned:
        return ""
    # Keep final weekly email report English-only even when source comments are non-English.
    if _contains_cyrillic(cleaned) and not re.search(r"[A-Za-z]", cleaned):
        return "Progress update recorded in Jira comments."
    return cleaned


def _strip_wrapping_quotes(value: str) -> str:
    cleaned = _normalize_text(value)
    if len(cleaned) >= 2 and ((cleaned[0] == '"' and cleaned[-1] == '"') or (cleaned[0] == "'" and cleaned[-1] == "'")):
        return cleaned[1:-1].strip()
    return cleaned


def _parse_positive_int_with_fallback(value: Any, default: int, *, name: str) -> int:
    text = _normalize_text(value)
    if not text:
        return default
    try:
        parsed = int(text)
    except ValueError:
        logger.error(
            "Invalid %s=%r. Expected a positive integer. Using default=%s.",
            name,
            value,
            default,
        )
        return default
    if parsed <= 0:
        logger.error(
            "Invalid %s=%r. Value must be > 0. Using default=%s.",
            name,
            value,
            default,
        )
        return default
    return parsed


def _week_key(year: int, week: int) -> str:
    return f"{str(year)[-2:]}\'w{week:02d}"


def resolve_week_window(params: dict[str, Any], now: date | None = None) -> WeekWindow:
    now = now or date.today()
    week_date_value = params.get("week_date") or params.get("date")
    week_value = params.get("week")
    year_value = params.get("year")
    start_value = params.get("start") or params.get("start_date")
    end_value = params.get("end") or params.get("end_date")

    if week_date_value:
        dt = datetime.strptime(str(week_date_value), "%Y-%m-%d").date()
        iso = dt.isocalendar()
        week_start = dt - timedelta(days=dt.weekday())
        week_end = week_start + timedelta(days=6)
        return WeekWindow(
            year=iso.year,
            week=iso.week,
            start=week_start,
            end=week_end,
            key=_week_key(iso.year, iso.week),
        )

    if week_value:
        week_raw = _normalize_text(week_value).replace(" ", "")
        week = 0
        year = now.year

        week_with_year_match = re.fullmatch(r"(?i)(\d{1,2})w(\d{2,4})", week_raw)
        if week_with_year_match:
            week = int(week_with_year_match.group(1))
            year_token = week_with_year_match.group(2)
            if len(year_token) == 2:
                year = 2000 + int(year_token)
            else:
                year = int(year_token)
        else:
            week = int(week_raw)
            year = int(str(year_value)) if year_value is not None else now.year

        week_start = date.fromisocalendar(year, week, 1)
        week_end = week_start + timedelta(days=6)
        return WeekWindow(
            year=year,
            week=week,
            start=week_start,
            end=week_end,
            key=_week_key(year, week),
        )

    if start_value and end_value:
        start_dt = datetime.strptime(str(start_value), "%Y-%m-%d").date()
        end_dt = datetime.strptime(str(end_value), "%Y-%m-%d").date()
        start_iso = start_dt.isocalendar()
        end_iso = end_dt.isocalendar()
        if (start_iso.year, start_iso.week) != (end_iso.year, end_iso.week):
            raise ValueError("start/end must be in the same ISO week for jira_weekly_email.")
        week_start = start_dt - timedelta(days=start_dt.weekday())
        week_end = week_start + timedelta(days=6)
        return WeekWindow(
            year=start_iso.year,
            week=start_iso.week,
            start=week_start,
            end=week_end,
            key=_week_key(start_iso.year, start_iso.week),
        )

    raise ValueError(
        "Provide one selector: week_date=YYYY-MM-DD, or week=<WWwYY|WWwYYYY|WW>, "
        "or start/end in one ISO week."
    )


def _parse_jira_date(value: Any) -> date | None:
    if not value:
        return None
    raw = str(value)
    token = raw.split("T")[0]
    try:
        return datetime.strptime(token, "%Y-%m-%d").date()
    except ValueError:
        return None


def _comment_body_to_text(body: Any) -> str:
    if body is None:
        return ""
    if isinstance(body, str):
        return body
    if isinstance(body, list):
        return " ".join(_comment_body_to_text(item) for item in body).strip()
    if isinstance(body, dict):
        node_type = _normalize_key(body.get("type"))
        content = body.get("content")
        if node_type in {"bulletlist", "orderedlist"} and isinstance(content, list):
            items: list[str] = []
            for item in content:
                item_text = _normalize_text(_comment_body_to_text(item))
                item_text = re.sub(r"^-+\s*", "", item_text)
                if item_text:
                    items.append(f"- {item_text}")
            return "; ".join(items)
        if node_type == "listitem" and isinstance(content, list):
            item_text = _normalize_text(" ".join(_comment_body_to_text(part) for part in content))
            return f"- {item_text}" if item_text else ""
        if node_type in {"hardbreak", "hard_break"}:
            return "; "

        parts: list[str] = []
        text_value = body.get("text")
        if isinstance(text_value, str) and text_value:
            parts.append(text_value)
        if isinstance(content, list):
            for item in content:
                part_text = _comment_body_to_text(item)
                if part_text:
                    parts.append(part_text)
        return " ".join(parts).strip()
    return str(body)


def _is_finished(status: str, resolution: str) -> bool:
    return _normalize_key(status) in _DONE_VALUES or _normalize_key(resolution) in _DONE_VALUES


def _is_in_progress_status(status: Any) -> bool:
    return _normalize_key(status) in _IN_PROGRESS_VALUES


def _safe_project_key(project: str) -> str:
    normalized = _normalize_text(project)
    if not re.fullmatch(r"[A-Za-z0-9_-]+", normalized):
        raise ValueError("Invalid project key format.")
    return normalized


def _fetch_epic_names(
    jira_source: JiraSource,
    jira,
    all_issues: list[Any],
) -> tuple[dict[str, str], dict[str, str], dict[str, str], dict[str, list[str]], dict[str, dict[str, str]]]:
    issue_epic_map: dict[str, str] = {}
    issue_parent_map: dict[str, str] = {}
    parent_keys_needed: set[str] = set()
    issue_details: dict[str, dict[str, str]] = {}

    for issue in all_issues:
        issue_type = _normalize_key(getattr(getattr(issue.fields, "issuetype", None), "name", ""))
        epic_link = getattr(issue.fields, "customfield_10000", "") or ""
        if not epic_link and issue_type == "epic":
            epic_link = issue.key
        parent = getattr(issue.fields, "parent", None)
        parent_key = parent.key if parent else ""
        issue_epic_map[issue.key] = epic_link
        issue_parent_map[issue.key] = parent_key
        if not epic_link and parent_key:
            parent_keys_needed.add(parent_key)
        issue_details[issue.key] = {
            "summary": _normalize_text(getattr(issue.fields, "summary", "")),
            "status": _normalize_text(getattr(getattr(issue.fields, "status", None), "name", "")),
            "resolution": _normalize_text(getattr(getattr(issue.fields, "resolution", None), "name", "")),
            "labels": [str(label) for label in (getattr(issue.fields, "labels", []) or [])],
        }

    pending_parent_keys = {key for key in parent_keys_needed if key}
    fetched_parent_keys: set[str] = set()
    chunk_size = 50
    while pending_parent_keys:
        chunk = [key for key in list(pending_parent_keys)[:chunk_size] if key not in fetched_parent_keys]
        if not chunk:
            break
        fetched_parent_keys.update(chunk)
        parent_issues = jira.search_issues(
            f"issuekey in ({', '.join(chunk)})",
            maxResults=1000,
            fields=["key", "customfield_10000", "parent", "issuetype", "summary", "status", "resolution", "labels"],
        )
        pending_parent_keys.difference_update(chunk)
        for parent_issue in parent_issues:
            parent_type = _normalize_key(getattr(getattr(parent_issue.fields, "issuetype", None), "name", ""))
            parent_parent = getattr(parent_issue.fields, "parent", None)
            parent_parent_key = parent_parent.key if parent_parent else ""
            parent_epic = getattr(parent_issue.fields, "customfield_10000", "") or ""
            if not parent_epic and parent_type == "epic":
                parent_epic = parent_issue.key

            issue_epic_map[parent_issue.key] = parent_epic
            issue_parent_map[parent_issue.key] = parent_parent_key
            issue_details[parent_issue.key] = {
                "summary": _normalize_text(getattr(parent_issue.fields, "summary", "")),
                "status": _normalize_text(getattr(getattr(parent_issue.fields, "status", None), "name", "")),
                "resolution": _normalize_text(getattr(getattr(parent_issue.fields, "resolution", None), "name", "")),
                "labels": [str(label) for label in (getattr(parent_issue.fields, "labels", []) or [])],
            }
            if not parent_epic and parent_parent_key and parent_parent_key not in fetched_parent_keys:
                pending_parent_keys.add(parent_parent_key)

    def _resolve_epic_for_issue(issue_key: str) -> str:
        visited: set[str] = set()
        current_key = issue_key
        while current_key and current_key not in visited:
            visited.add(current_key)
            epic_key = _normalize_text(issue_epic_map.get(current_key, ""))
            if epic_key:
                return epic_key
            current_key = _normalize_text(issue_parent_map.get(current_key, ""))
        return ""

    for key in list(issue_epic_map.keys()):
        resolved_epic = _resolve_epic_for_issue(key)
        if resolved_epic:
            issue_epic_map[key] = resolved_epic

    epic_keys = list({epic for epic in issue_epic_map.values() if epic})
    epic_names = jira_source.fetch_epic_names(epic_keys)
    epic_labels: dict[str, list[str]] = {}
    if epic_keys:
        chunk_size = 50
        for idx in range(0, len(epic_keys), chunk_size):
            chunk = epic_keys[idx : idx + chunk_size]
            try:
                epic_issues = jira.search_issues(
                    f"issuekey in ({', '.join(chunk)})",
                    maxResults=1000,
                    fields=["key", "labels"],
                )
            except Exception:
                logger.warning("Failed to fetch epic labels for chunk size=%s.", len(chunk))
                continue
            for epic_issue in epic_issues:
                if epic_issue.key not in chunk:
                    continue
                epic_labels[epic_issue.key] = [str(label) for label in (epic_issue.fields.labels or [])]
    return issue_epic_map, issue_parent_map, epic_names, epic_labels, issue_details


def collect_weekly_comment_evidence(
    jira_source: JiraSource,
    project: str,
    week: WeekWindow,
) -> list[dict[str, Any]]:
    jira = jira_source.jira
    project_key = _safe_project_key(project)
    start_value = week.start.strftime("%Y-%m-%d")
    end_exclusive = (week.end + timedelta(days=1)).strftime("%Y-%m-%d")
    jql_query = (
        f"project = {project_key} "
        f"AND updated >= '{start_value}' "
        f"AND updated < '{end_exclusive}' "
        "ORDER BY created DESC"
    )
    logger.info(
        "JIRA QUERY: project=%s week=%s range=[%s..%s] jql=%s fields=%s page_size=%s",
        project,
        week.key,
        start_value,
        week.end.strftime("%Y-%m-%d"),
        jql_query,
        "key,summary,status,resolution,issuetype,labels,priority,customfield_10000,parent,comment",
        100,
    )

    start_at = 0
    max_results = 100
    all_issues: list[Any] = []
    pages = 0
    while True:
        issues = jira.search_issues(
            jql_query,
            startAt=start_at,
            maxResults=max_results,
            fields=[
                "key",
                "summary",
                "status",
                "resolution",
                "issuetype",
                "labels",
                "priority",
                "customfield_10000",
                "parent",
                "comment",
            ],
        )
        pages += 1
        all_issues.extend(issues)
        if len(issues) < max_results:
            break
        start_at += max_results

    if not all_issues:
        logger.info("JIRA FETCH RESULT: project=%s week=%s pages=%s raw_issues=0", project, week.key, pages)
        return []

    issue_epic_map, issue_parent_map, epic_names, epic_labels_map, issue_details = _fetch_epic_names(
        jira_source, jira, all_issues
    )
    summary_map: dict[str, str] = {
        key: _normalize_text((details or {}).get("summary", ""))
        for key, details in issue_details.items()
    }
    status_map: dict[str, str] = {
        key: _normalize_text((details or {}).get("status", ""))
        for key, details in issue_details.items()
    }
    resolution_map: dict[str, str] = {
        key: _normalize_text((details or {}).get("resolution", ""))
        for key, details in issue_details.items()
    }
    labels_map: dict[str, list[str]] = {
        key: [str(label) for label in ((details or {}).get("labels") or [])]
        for key, details in issue_details.items()
    }

    evidence: list[dict[str, Any]] = []
    total_comments_in_week = 0

    for issue in all_issues:
        issue_key = issue.key
        labels = [str(label) for label in (issue.fields.labels or [])]
        priority = issue.fields.priority.name if issue.fields.priority else ""
        status = issue.fields.status.name if issue.fields.status else ""
        resolution = issue.fields.resolution.name if issue.fields.resolution else ""
        issue_type_obj = issue.fields.issuetype if issue.fields.issuetype else None
        issue_type = issue_type_obj.name if issue_type_obj else ""
        issue_is_subtask = bool(getattr(issue_type_obj, "subtask", False))

        epic_link = issue_epic_map.get(issue_key, "")
        parent_key = issue_parent_map.get(issue_key, "")
        if not epic_link and parent_key:
            epic_link = issue_epic_map.get(parent_key, "")
        epic_name = epic_names.get(epic_link, "Unknown Epic") if epic_link else "Unknown Epic"

        comments_in_week_rows: list[tuple[date, int, str]] = []
        comment_block = getattr(getattr(issue.fields, "comment", None), "comments", []) or []
        for comment_idx, comment in enumerate(comment_block):
            created_dt = _parse_jira_date(getattr(comment, "created", ""))
            if not created_dt or not (week.start <= created_dt <= week.end):
                continue
            body_text = _normalize_text(_comment_body_to_text(getattr(comment, "body", "")))
            if body_text:
                comments_in_week_rows.append((created_dt, comment_idx, body_text))

        comments_in_week_rows.sort(key=lambda item: (item[0], item[1]))
        comments_in_week = [item[2] for item in comments_in_week_rows]

        total_comments_in_week += len(comments_in_week)

        parent_finished = False
        if parent_key:
            parent_finished = _is_finished(status_map.get(parent_key, ""), resolution_map.get(parent_key, ""))

        evidence.append(
            {
                "Issue_Key": issue_key,
                "Summary": _normalize_text(issue.fields.summary),
                "Epic_Key": epic_link,
                "Epic_Name": epic_name,
                "Parent_Key": parent_key,
                "Type": issue_type,
                "Status": status,
                "Resolution": resolution,
                "Priority": priority,
                "Labels": labels,
                "Epic_Labels": epic_labels_map.get(epic_link, []),
                "Epic_Labels_Known": epic_link in epic_labels_map,
                "Comments": comments_in_week,
                "Finished": _is_finished(status, resolution),
                "Bug": _normalize_key(issue_type) == "bug",
                "Subtask": issue_is_subtask or _normalize_key(issue_type) in {"sub-task", "subtask"},
                "Parent_Finished": parent_finished,
                "Parent_Summary": summary_map.get(parent_key, ""),
                "Parent_Status": status_map.get(parent_key, ""),
                "Parent_Resolution": resolution_map.get(parent_key, ""),
                "Parent_Labels": labels_map.get(parent_key, []),
            }
        )

    logger.info(
        "JIRA FETCH RESULT: project=%s week=%s pages=%s raw_issues=%s evidence_issues=%s comments_in_week=%s unique_epics=%s",
        project,
        week.key,
        pages,
        len(all_issues),
        len(evidence),
        total_comments_in_week,
        len({item.get('Epic_Key') for item in evidence if _normalize_text(item.get('Epic_Key'))}),
    )
    return evidence


def _issue_type_name(issue: Any) -> str:
    fields = getattr(issue, "fields", None)
    issue_type = getattr(fields, "issuetype", None)
    return _normalize_text(getattr(issue_type, "name", ""))


def _issue_status_name(issue: Any) -> str:
    fields = getattr(issue, "fields", None)
    status = getattr(fields, "status", None)
    return _normalize_text(getattr(status, "name", ""))


def _issue_resolution_name(issue: Any) -> str:
    fields = getattr(issue, "fields", None)
    resolution = getattr(fields, "resolution", None)
    if resolution is None:
        return ""
    return _normalize_text(getattr(resolution, "name", ""))


def _is_bug_issue(issue: Any) -> bool:
    return _normalize_key(_issue_type_name(issue)) == "bug"


def _is_open_bug_issue(issue: Any) -> bool:
    if not _is_bug_issue(issue):
        return False
    resolution_key = _normalize_key(_issue_resolution_name(issue))
    return not resolution_key or resolution_key == "unresolved"


def _is_in_progress_bug_issue(issue: Any) -> bool:
    if not _is_bug_issue(issue):
        return False
    return _is_in_progress_status(_issue_status_name(issue))


def _count_issues_for_jql(jira: Any, jql: str, fallback_match: Any) -> int:
    try:
        result = jira.search_issues(
            jql,
            maxResults=0,
            fields=["key", "issuetype", "status", "resolution"],
        )
    except Exception as exc:
        logger.warning("Bug counters query failed: jql=%s error=%s", jql, exc)
        return 0
    total = getattr(result, "total", None)
    if isinstance(total, int):
        return total
    if isinstance(result, list):
        return sum(1 for issue in result if fallback_match(issue))
    try:
        materialized = list(result)
    except Exception:
        return 0
    return sum(1 for issue in materialized if fallback_match(issue))


def collect_priority_always_evidence(
    jira_source: JiraSource,
    project: str,
    week: WeekWindow,
    priority_values: set[str],
) -> list[dict[str, Any]]:
    """Fetch all non-Epic issues with the given priority values that are open or closed in the week period.

    Returns evidence in the same structure as collect_weekly_comment_evidence.
    Comments are filtered to the week window; issues without week comments will have Comments=[].
    """
    if not priority_values:
        return []
    jira = jira_source.jira
    project_key = _safe_project_key(project)
    start_value = week.start.strftime("%Y-%m-%d")
    end_exclusive = (week.end + timedelta(days=1)).strftime("%Y-%m-%d")
    priority_list = ", ".join(f'"{p}"' for p in sorted(priority_values))

    open_jql = (
        f"project = {project_key} "
        f"AND priority in ({priority_list}) "
        "AND resolution = Unresolved "
        "AND issuetype not in (Epic) "
        "ORDER BY created DESC"
    )
    closed_jql = (
        f"project = {project_key} "
        f"AND priority in ({priority_list}) "
        f"AND resolutiondate >= '{start_value}' "
        f"AND resolutiondate < '{end_exclusive}' "
        "AND issuetype not in (Epic) "
        "ORDER BY created DESC"
    )

    fields = [
        "key", "summary", "status", "resolution", "issuetype",
        "labels", "priority", "customfield_10000", "parent", "comment",
    ]
    all_issues: list[Any] = []
    seen_keys: set[str] = set()

    for jql_query in [open_jql, closed_jql]:
        start_at = 0
        max_results = 100
        while True:
            issues = jira.search_issues(jql_query, startAt=start_at, maxResults=max_results, fields=fields)
            for issue in issues:
                if issue.key not in seen_keys:
                    seen_keys.add(issue.key)
                    all_issues.append(issue)
            if len(issues) < max_results:
                break
            start_at += max_results

    if not all_issues:
        return []

    issue_epic_map, issue_parent_map, epic_names, epic_labels_map, issue_details = _fetch_epic_names(
        jira_source, jira, all_issues
    )
    summary_map: dict[str, str] = {
        key: _normalize_text((d or {}).get("summary", "")) for key, d in issue_details.items()
    }
    status_map: dict[str, str] = {
        key: _normalize_text((d or {}).get("status", "")) for key, d in issue_details.items()
    }
    resolution_map: dict[str, str] = {
        key: _normalize_text((d or {}).get("resolution", "")) for key, d in issue_details.items()
    }
    labels_map: dict[str, list[str]] = {
        key: [str(label) for label in ((d or {}).get("labels") or [])]
        for key, d in issue_details.items()
    }

    evidence: list[dict[str, Any]] = []
    for issue in all_issues:
        issue_key = issue.key
        labels = [str(label) for label in (issue.fields.labels or [])]
        priority = issue.fields.priority.name if issue.fields.priority else ""
        status = issue.fields.status.name if issue.fields.status else ""
        resolution = issue.fields.resolution.name if issue.fields.resolution else ""
        issue_type_obj = issue.fields.issuetype if issue.fields.issuetype else None
        issue_type = issue_type_obj.name if issue_type_obj else ""
        issue_is_subtask = bool(getattr(issue_type_obj, "subtask", False))

        epic_link = issue_epic_map.get(issue_key, "")
        parent_key = issue_parent_map.get(issue_key, "")
        if not epic_link and parent_key:
            epic_link = issue_epic_map.get(parent_key, "")
        epic_name = epic_names.get(epic_link, "Unknown Epic") if epic_link else "Unknown Epic"

        comments_rows: list[tuple[date, int, str]] = []
        comment_block = getattr(getattr(issue.fields, "comment", None), "comments", []) or []
        for comment_idx, comment in enumerate(comment_block):
            created_dt = _parse_jira_date(getattr(comment, "created", ""))
            if not created_dt or not (week.start <= created_dt <= week.end):
                continue
            body_text = _normalize_text(_comment_body_to_text(getattr(comment, "body", "")))
            if body_text:
                comments_rows.append((created_dt, comment_idx, body_text))
        comments_rows.sort(key=lambda item: (item[0], item[1]))
        comments_in_week = [item[2] for item in comments_rows]

        parent_finished = False
        if parent_key:
            parent_finished = _is_finished(status_map.get(parent_key, ""), resolution_map.get(parent_key, ""))

        evidence.append(
            {
                "Issue_Key": issue_key,
                "Summary": _normalize_text(issue.fields.summary),
                "Epic_Key": epic_link,
                "Epic_Name": epic_name,
                "Parent_Key": parent_key,
                "Type": issue_type,
                "Status": status,
                "Resolution": resolution,
                "Priority": priority,
                "Labels": labels,
                "Epic_Labels": epic_labels_map.get(epic_link, []),
                "Epic_Labels_Known": epic_link in epic_labels_map,
                "Comments": comments_in_week,
                "Finished": _is_finished(status, resolution),
                "Bug": _normalize_key(issue_type) == "bug",
                "Subtask": issue_is_subtask or _normalize_key(issue_type) in {"sub-task", "subtask"},
                "Parent_Finished": parent_finished,
                "Parent_Summary": summary_map.get(parent_key, ""),
                "Parent_Status": status_map.get(parent_key, ""),
                "Parent_Resolution": resolution_map.get(parent_key, ""),
                "Parent_Labels": labels_map.get(parent_key, []),
                "AlwaysShow": True,
            }
        )

    logger.info(
        "PRIORITY ALWAYS EVIDENCE: project=%s week=%s priority=%s issues=%s",
        project,
        week.key,
        ",".join(sorted(priority_values)),
        len(evidence),
    )
    return evidence


def collect_risk_evidence(
    jira_source: JiraSource,
    project: str,
    week: WeekWindow,
) -> list[dict[str, Any]]:
    """Fetch all risk/issue-labeled Jira issues: open ones and those updated within the week period.

    Returns a simplified evidence list with key, summary, status, assignee,
    reporter, created date, and comments filtered to the week window.
    """
    jira = jira_source.jira
    project_key = _safe_project_key(project)
    start_value = week.start.strftime("%Y-%m-%d")
    end_exclusive = (week.end + timedelta(days=1)).strftime("%Y-%m-%d")
    jql = (
        f"project = {project_key} "
        'AND labels in ("risk", "issue") '
        "AND (statusCategory != Done "
        f"    OR (updated >= '{start_value}' AND updated < '{end_exclusive}')) "
        "ORDER BY created ASC"
    )
    fields = ["key", "summary", "status", "labels",
              "assignee", "reporter", "created", "comment"]
    all_issues: list[Any] = []
    start_at = 0
    max_results = 100
    while True:
        issues = jira.search_issues(jql, startAt=start_at, maxResults=max_results, fields=fields)
        all_issues.extend(issues)
        if len(issues) < max_results:
            break
        start_at += max_results

    results: list[dict[str, Any]] = []
    for issue in all_issues:
        assignee_obj = getattr(issue.fields, "assignee", None)
        assignee = _normalize_text(getattr(assignee_obj, "displayName", "") or "") if assignee_obj else ""
        reporter_obj = getattr(issue.fields, "reporter", None)
        reporter = _normalize_text(getattr(reporter_obj, "displayName", "") or "") if reporter_obj else ""
        created_raw = getattr(issue.fields, "created", None)
        created_str = str(created_raw)[:10] if created_raw else ""
        comments_in_week: list[str] = []
        comment_block = getattr(getattr(issue.fields, "comment", None), "comments", []) or []
        for comment in comment_block:
            created_dt = _parse_jira_date(getattr(comment, "created", ""))
            if not created_dt or not (week.start <= created_dt <= week.end):
                continue
            body_text = _normalize_text(_comment_body_to_text(getattr(comment, "body", "")))
            if body_text:
                comments_in_week.append(body_text)
        results.append({
            "Issue_Key": issue.key,
            "Summary":   getattr(issue.fields, "summary", "") or "",
            "Status":    getattr(issue.fields.status, "name", "") if issue.fields.status else "",
            "Assignee":  assignee,
            "Reporter":  reporter,
            "Created":   created_str,
            "Comments":  comments_in_week,
        })
    logger.info(
        "RISK EVIDENCE: project=%s week=%s issues=%s",
        project, week.key, len(results),
    )
    return results


def collect_project_bug_stats(jira_source: JiraSource, project: str) -> dict[str, int]:
    jira = jira_source.jira
    project_key = _safe_project_key(project)
    in_progress_jql = (
        f"project = {project_key} "
        "AND issuetype = Bug "
        "AND statusCategory = 'In Progress'"
    )
    open_jql = (
        f"project = {project_key} "
        "AND issuetype = Bug "
        "AND resolution = Unresolved"
    )
    in_progress = _count_issues_for_jql(jira, in_progress_jql, _is_in_progress_bug_issue)
    open_count = _count_issues_for_jql(jira, open_jql, _is_open_bug_issue)
    logger.info(
        "PROJECT BUG COUNTS: project=%s in_progress=%s open=%s",
        project,
        in_progress,
        open_count,
    )
    return {
        "in_progress": in_progress,
        "open": open_count,
    }


def _first_sentence(text: str) -> str:
    value = _normalize_text(text)
    if not value:
        return ""
    parts = re.split(r"(?<=[.!?])\s+", value)
    sentence = parts[0].strip() if parts else value
    if len(sentence) > 180:
        sentence = sentence[:177].rstrip() + "..."
    if sentence and sentence[-1] not in ".!?":
        sentence += "."
    return sentence


def _comment_hints_joined(comments: Any) -> str:
    values = comments if isinstance(comments, list) else [comments]
    points: list[str] = []
    for value in values:
        cleaned = _clean_comment_for_report(value)
        if not cleaned:
            continue
        split_points = _split_progress_points(cleaned)
        if split_points:
            points.extend(split_points)
        else:
            points.append(cleaned)
    if not points:
        return ""
    unique_points: list[str] = []
    seen: set[str] = set()
    for point in points:
        marker = _normalize_key(point)
        if not marker or marker in seen:
            continue
        seen.add(marker)
        unique_points.append(point)
    if not unique_points:
        return ""
    joined = "; ".join(unique_points[:6])
    words = joined.split()
    if len(words) > 80:
        joined = " ".join(words[:80]).rstrip(" ,;:-")
        if joined and joined[-1] not in ".!?":
            joined += "..."
    return joined


def _split_progress_points(text: Any) -> list[str]:
    raw = _normalize_text(text)
    if not raw:
        return []
    normalized = re.sub(r"[•▪◦]+", "; ", raw)
    normalized = re.sub(r"(?:(?<=^)|(?<=[;:]))\s*\d+[.)]\s+", "; ", normalized)
    normalized = re.sub(r"(?:(?<=^)|(?<=[;:]))\s*-\s+", "; ", normalized)
    parts = [_normalize_text(part) for part in normalized.split(";")]
    return [part for part in parts if part]


def _collect_comment_points(comments: Any) -> list[str]:
    values = comments if isinstance(comments, list) else [comments]
    raw_points: list[str] = []
    for value in values:
        cleaned = _clean_comment_for_report(value)
        if not cleaned:
            continue
        split_points = _split_progress_points(cleaned)
        if split_points:
            raw_points.extend(split_points)
        else:
            raw_points.append(cleaned)
    unique_points: list[str] = []
    seen: set[str] = set()
    for point in raw_points:
        point = _first_sentence(point)
        marker = _normalize_key(point)
        if not marker or marker in seen:
            continue
        if re.search(r"\b(weekly report|h2|h3|chapter|results summary)\b", marker):
            continue
        if len(point.split()) < 2:
            continue
        if len(re.findall(r"[^A-Za-z0-9\s.,;:!?()/-]", point)) > 4:
            continue
        seen.add(marker)
        unique_points.append(point)
    return unique_points


def _truncate_words(text: str, max_words: int = 12) -> str:
    words = _normalize_text(text).split()
    if not words:
        return ""
    if len(words) <= max_words:
        return " ".join(words)
    truncated = " ".join(words[:max_words]).rstrip(" ,;:-")
    if truncated and truncated[-1] not in ".!?":
        truncated += "..."
    return truncated


def _limit_points(points: list[str], *, max_items: int = 2, max_words_per_item: int = 12) -> str:
    if not points:
        return ""
    compact = [_truncate_words(item, max_words=max_words_per_item) for item in points[:max_items]]
    compact = [item for item in compact if item]
    return "; ".join(compact)


def _classify_progress_points(points: list[str]) -> dict[str, list[str]]:
    done: list[str] = []
    plan: list[str] = []
    risk: list[str] = []
    dependency: list[str] = []
    misc: list[str] = []
    for point in points:
        marker = _normalize_key(point)
        if not marker:
            continue
        if re.search(
            r"\b(done|completed|fixed|merged|implemented|released|resolved|verified|tested|closed|prepared|delivered)\b",
            marker,
        ):
            done.append(point)
            continue
        if re.search(
            r"\b(next|plan|will|todo|to do|continue|prepare|scheduled|pending|need to|going to|target)\b",
            marker,
        ):
            plan.append(point)
            continue
        if re.search(r"\b(block|blocked|issue|problem|risk|fail|failed|unstable|delay|stuck|timeout|hold)\b", marker):
            risk.append(point)
            continue
        if re.search(r"\b(depend|dependency|waiting|await|requires|need from|external|review|approval)\b", marker):
            dependency.append(point)
            continue
        misc.append(point)
    return {"done": done, "plan": plan, "risk": risk, "dependency": dependency, "misc": misc}


def _build_compact_feature_status(feature: dict[str, Any]) -> str:
    classified = _classify_progress_points(feature.get("points") or [])
    parts: list[str] = []

    done_part = _limit_points(classified["done"], max_items=2, max_words_per_item=12)
    plan_part = _limit_points(classified["plan"], max_items=2, max_words_per_item=12)
    risk_part = _limit_points(classified["risk"], max_items=1, max_words_per_item=14)
    dependency_part = _limit_points(classified["dependency"], max_items=1, max_words_per_item=14)
    misc_part = _limit_points(classified["misc"], max_items=1, max_words_per_item=14)

    has_comments = bool((feature.get("points") or []))

    if int(feature.get("blocked_tasks") or 0) > 0:
        parts.append("Blocked")
    elif int(feature.get("in_progress_tasks") or 0) > 0 and int(feature.get("closed_tasks") or 0) <= 0:
        parts.append("In progress")
    elif int(feature.get("closed_tasks") or 0) > 0 and int(feature.get("in_progress_tasks") or 0) <= 0:
        parts.append("Completed")

    if done_part:
        parts.append(f"Done: {done_part}")

    if misc_part:
        parts.append(misc_part)

    if plan_part:
        parts.append(f"Next: {plan_part}")

    if risk_part:
        parts.append(f"Risk: {risk_part}")
    elif int(feature.get("blocked_tasks") or 0) > 0 and has_comments:
        parts.append("Blocked; requires follow-up")

    if dependency_part:
        parts.append(f"Depends on: {dependency_part}")

    if not parts:
        if int(feature.get("blocked_tasks") or 0) > 0:
            return "Blocked."
        if int(feature.get("closed_tasks") or 0) > 0:
            return "Completed."
        if int(feature.get("in_progress_tasks") or 0) > 0:
            return "In progress."
        return ""

    compact = "; ".join(parts)
    if compact and compact[-1] not in ".!?":
        compact += "."
    return compact


def _build_compact_plan_status(feature: dict[str, Any]) -> str:
    classified = _classify_progress_points(feature.get("points") or [])
    plan_part = _limit_points(classified["plan"], max_items=2, max_words_per_item=12)
    misc_part = _limit_points(classified["misc"], max_items=1, max_words_per_item=14)
    risk_part = _limit_points(classified["risk"], max_items=1, max_words_per_item=14)
    dependency_part = _limit_points(classified["dependency"], max_items=1, max_words_per_item=14)

    parts: list[str] = []
    if plan_part:
        parts.append(plan_part)
    elif misc_part:
        parts.append(misc_part)

    if risk_part:
        parts.append(f"Risk: {risk_part}")
    if dependency_part:
        parts.append(f"Depends on: {dependency_part}")

    if not parts:
        return ""

    compact = "; ".join(parts)
    if compact and compact[-1] not in ".!?":
        compact += "."
    return compact


def _build_aggregate_input(feature: dict[str, Any], mode: str = "result") -> str:
    """Build a text input for AI aggregation of multiple subtask contributions."""
    parent_name = _normalize_text(feature.get("feature_name"))
    subtask_keys = list(feature.get("subtask_issue_keys") or [])
    subtask_summaries = feature.get("subtask_summaries") or {}

    names = [_normalize_text(subtask_summaries.get(k, k)) for k in subtask_keys[:5]]
    names_str = "; ".join(n for n in names if n)
    if len(subtask_keys) > 5:
        names_str += f"; +{len(subtask_keys) - 5} more"

    points = feature.get("points") or []
    classified = _classify_progress_points(points)

    if mode == "plan":
        action = (classified["plan"] + classified["misc"])[:4]
        notes = "; ".join(action) if action else "Work in progress"
        text = f"Feature: {parent_name}. Tasks: {names_str}. Next week: {notes}."
    else:
        done = (classified["done"] + classified["misc"])[:4]
        notes = "; ".join(done)
        if not notes:
            closed = int(feature.get("closed_tasks") or 0)
            notes = f"{closed} task(s) completed" if closed else "In progress"
        text = f"Feature: {parent_name}. Tasks: {names_str}. Done: {notes}."

    words = text.split()
    if len(words) > 60:
        text = " ".join(words[:60]).rstrip(".,;:") + "."
    return text


def _build_highlight_progress(entry: dict[str, Any], subtasks: list[dict[str, Any]]) -> str:
    if entry.get("Finished"):
        return "Finished this week."

    progress_parts: list[str] = []
    issue_comment = _comment_hints_joined(entry.get("Comments") or [])
    if issue_comment:
        issue_points = _split_progress_points(issue_comment)
        progress_parts.extend(issue_points or [issue_comment])

    if not entry.get("Subtask"):
        ordered_subtasks = sorted(
            list(subtasks or []),
            key=lambda item: _normalize_key(item.get("Issue_Key")),
        )
        for subtask in ordered_subtasks:
            subtask_title = _normalize_text(subtask.get("Summary")) or _normalize_text(subtask.get("Issue_Key")) or "Subtask"
            if subtask.get("Finished"):
                progress_parts.append(subtask_title)
                continue

            subtask_comment = _comment_hints_joined(subtask.get("Comments") or [])
            if subtask_comment:
                subtask_points = _split_progress_points(subtask_comment)
                if subtask_points:
                    progress_parts.append(f"{subtask_title}: {'; '.join(subtask_points)}")
                else:
                    progress_parts.append(f"{subtask_title}: {subtask_comment}")
            else:
                progress_parts.append(subtask_title)

    if progress_parts:
        return f"Progress: {'; '.join(progress_parts)}"
    return "No progress this week."


def _build_item_text(entry: dict[str, Any], *, mode: str) -> str:
    summary = _normalize_text(entry.get("Summary"))
    issue_key = _normalize_text(entry.get("Issue_Key"))
    comment_hint = _comment_hints_joined(entry.get("Comments") or [])

    if mode == "highlight":
        headline = summary or issue_key or "Task"
        if entry.get("Finished"):
            return f"{headline} - Finished this week."
        if comment_hint:
            return f"{headline} - Progress: {comment_hint}"
        return f"{headline} - No progress this week."

    if mode == "completed":
        return summary or issue_key or "Task"

    if mode == "subtask":
        if summary:
            return summary
        return comment_hint or "Subtask update"

    if mode == "plan":
        return comment_hint

    if mode == "result_progress":
        return summary or issue_key or "Task"

    if mode == "high":
        return summary or issue_key or "High priority item"

    return comment_hint or summary


def _epic_id(entry: dict[str, Any]) -> str:
    epic_key = _normalize_text(entry.get("Epic_Key"))
    if epic_key:
        return epic_key
    return f"name::{_normalize_text(entry.get('Epic_Name'))}"


def _parse_label_set(value: str | None, default: list[str]) -> set[str]:
    return {_normalize_key(item) for item in _split_csv(value, default)}


def build_report_payload(
    evidence: list[dict[str, Any]],
    week: WeekWindow,
    config: ConfigParser,
    project: str,
    *,
    labels_highlights: set[str],
    labels_report: set[str],
    priority_high_values: set[str],
    priority_always_show_values: set[str] = frozenset(),
    hp_always_evidence: list[dict[str, Any]] | None = None,
    always_show_evidence: list[dict[str, Any]] | None = None,
    project_bug_stats: dict[str, int] | None = None,
    risk_evidence: list[dict[str, Any]] | None = None,
) -> dict[str, Any]:
    highlights: list[dict[str, str]] = []
    epics: dict[str, dict[str, Any]] = {}
    subtasks_by_parent: dict[str, list[dict[str, Any]]] = {}
    report_epic_ids: set[str] = set()
    report_all_labels = "@all" in labels_report
    blocked_status_values = {"blocked", "on hold", "hold", "waiting", "stalled"}
    project_bug_stats = project_bug_stats or {"in_progress": 0, "open": 0}

    for item in evidence:
        if not item.get("Subtask"):
            continue
        parent_key = _normalize_text(item.get("Parent_Key"))
        if not parent_key:
            continue
        subtasks_by_parent.setdefault(parent_key, []).append(item)

    for entry in evidence:
        issue_key = _normalize_text(entry.get("Issue_Key"))
        labels_norm = {_normalize_key(label) for label in (entry.get("Labels") or [])}
        if labels_norm & _RISK_LABELS:
            continue  # handled exclusively by the Top Issues/Risks section
        epic_name = _normalize_text(entry.get("Epic_Name")) or "No Epic"
        epic_key = _normalize_text(entry.get("Epic_Key"))

        if labels_norm & labels_highlights:
            issue_summary = _normalize_text(entry.get("Summary"))
            highlights.append(
                {
                    "issue_key": issue_key,
                    "headline": issue_summary or issue_key or "Task",
                    "comment": _build_highlight_progress(entry, subtasks_by_parent.get(issue_key, [])),
                    "epic_key": epic_key,
                    "epic_name": epic_name,
                }
            )

        is_bug = bool(entry.get("Bug"))
        priority_key = _normalize_key(entry.get("Priority"))
        if is_bug and priority_key not in (priority_high_values | priority_always_show_values):
            continue

        resolution_key = _normalize_key(entry.get("Resolution"))
        if entry.get("Finished") and resolution_key and resolution_key not in _REPORT_CLOSED_RESOLUTION_VALUES:
            # HP tasks bypass the resolution filter — they go to high_priority_items regardless of resolution
            if priority_key not in priority_high_values:
                continue

        epic_identifier = _epic_id(entry)
        epic_labels_norm = {_normalize_key(label) for label in (entry.get("Epic_Labels") or [])}
        epic_labels_known = bool(entry.get("Epic_Labels_Known"))
        issue_report_scope = report_all_labels or bool(labels_norm & labels_report)
        issue_in_report_scope = report_all_labels or bool(epic_labels_norm & labels_report)
        if not issue_in_report_scope and epic_key and not epic_labels_known:
            issue_in_report_scope = bool(labels_norm & labels_report)

        parent_labels_norm = {_normalize_key(label) for label in (entry.get("Parent_Labels") or [])}
        parent_report_scope = report_all_labels or bool(parent_labels_norm & labels_report)
        subtask_parent_key = _normalize_text(entry.get("Parent_Key"))
        subtask_parent_scope = bool(entry.get("Subtask") and subtask_parent_key and parent_report_scope)

        if not (issue_report_scope or issue_in_report_scope or subtask_parent_scope):
            continue
        report_epic_ids.add(epic_identifier)

        epic_bucket = epics.setdefault(
            epic_identifier,
            {
                "epic_key": epic_key,
                "epic_name": epic_name,
                "report_items": [],
                "completed_items": [],
                "progress_items": [],
                "parent_subtasks": [],
                "high_priority_items": [],
                "feature_statuses": [],
                "next_week_items": [],
                "closed_tasks": 0,
                "bugs": {"closed": 0, "in_progress": 0},
                "_feature_map": {},
            },
        )

        if is_bug:
            if entry.get("Finished"):
                epic_bucket["bugs"]["closed"] += 1
            elif _is_in_progress_status(entry.get("Status")):
                epic_bucket["bugs"]["in_progress"] += 1

        if priority_key in priority_high_values and not entry.get("Subtask"):
            logger.debug(
                "HP_ITEM: issue_key=%s priority=%s is_bug=%s resolution=%s finished=%s",
                issue_key, priority_key, is_bug, resolution_key, entry.get("Finished"),
            )
            high_status = "Finished" if entry.get("Finished") else (
                _normalize_text(entry.get("Status")) or _normalize_text(entry.get("Resolution"))
            )
            all_hp_comments = list(entry.get("Comments") or [])
            for subtask_entry in (subtasks_by_parent.get(issue_key) or []):
                all_hp_comments.extend(subtask_entry.get("Comments") or [])
            epic_bucket["high_priority_items"].append(
                {
                    "issue_key": issue_key,
                    "text": _normalize_text(entry.get("Summary")) or issue_key,
                    "status": high_status,
                    "comment": _comment_hints_joined(all_hp_comments),
                }
            )

        feature_key = _normalize_text(entry.get("Parent_Key")) if entry.get("Subtask") else issue_key
        if not feature_key:
            feature_key = issue_key or _normalize_text(entry.get("Summary")) or "feature"
        feature_name = (
            _normalize_text(entry.get("Parent_Summary"))
            if entry.get("Subtask")
            else _normalize_text(entry.get("Summary"))
        )
        if not feature_name:
            feature_name = feature_key

        feature = epic_bucket["_feature_map"].setdefault(
            feature_key,
            {
                "feature_key": feature_key,
                "feature_name": feature_name,
                "issue_keys": set(),
                "points": [],
                "comments_count": 0,
                "closed_tasks": 0,
                "in_progress_tasks": 0,
                "blocked_tasks": 0,
                "subtask_issue_keys": [],
                "subtask_summaries": {},
            },
        )

        if issue_key:
            feature["issue_keys"].add(issue_key)
        if entry.get("Subtask") and issue_key and issue_key not in feature["subtask_summaries"]:
            feature["subtask_issue_keys"].append(issue_key)
            feature["subtask_summaries"][issue_key] = _normalize_text(entry.get("Summary"))
        comment_points = _collect_comment_points(entry.get("Comments") or [])
        if comment_points:
            feature["points"].extend(comment_points)
        feature["comments_count"] += len(entry.get("Comments") or [])

        status_key = _normalize_key(entry.get("Status"))
        if entry.get("Finished"):
            feature["closed_tasks"] += 1
            epic_bucket["closed_tasks"] += 1
        elif _is_in_progress_status(entry.get("Status")):
            feature["in_progress_tasks"] += 1
        if status_key in blocked_status_values:
            feature["blocked_tasks"] += 1

    # Post-process hp_always_evidence: add HP always-show items not yet in high_priority_items
    _existing_hp_keys: set[str] = {
        _normalize_key(item.get("issue_key"))
        for epic in epics.values()
        for item in (epic.get("high_priority_items") or [])
        if item.get("issue_key")
    }
    for _hp_entry in (hp_always_evidence or []):
        _hp_ik = _normalize_text(_hp_entry.get("Issue_Key"))
        _hp_ik_norm = _normalize_key(_hp_ik)
        if _normalize_key(_hp_entry.get("Priority")) not in priority_high_values:
            continue
        if bool(_hp_entry.get("Subtask")):
            continue
        if _hp_ik_norm in _existing_hp_keys:
            # Already present — set "No updates this week." if comment is empty
            for _epic in epics.values():
                for _item in (_epic.get("high_priority_items") or []):
                    if _normalize_key(_item.get("issue_key")) == _hp_ik_norm and not _item.get("comment"):
                        _item["comment"] = "No updates this week."
            continue
        _hp_epic_id = _epic_id(_hp_entry)
        _hp_epic_key = _normalize_text(_hp_entry.get("Epic_Key"))
        _hp_epic_name = _normalize_text(_hp_entry.get("Epic_Name")) or "No Epic"
        _hp_bucket = epics.setdefault(
            _hp_epic_id,
            {
                "epic_key": _hp_epic_key,
                "epic_name": _hp_epic_name,
                "report_items": [],
                "completed_items": [],
                "progress_items": [],
                "parent_subtasks": [],
                "high_priority_items": [],
                "feature_statuses": [],
                "next_week_items": [],
                "closed_tasks": 0,
                "bugs": {"closed": 0, "in_progress": 0},
                "_feature_map": {},
            },
        )
        report_epic_ids.add(_hp_epic_id)
        _hp_comments = list(_hp_entry.get("Comments") or [])
        _hp_comment_text = _comment_hints_joined(_hp_comments) if _hp_comments else "No updates this week."
        _hp_status = "Finished" if _hp_entry.get("Finished") else (
            _normalize_text(_hp_entry.get("Status")) or _normalize_text(_hp_entry.get("Resolution"))
        )
        _hp_bucket["high_priority_items"].append(
            {
                "issue_key": _hp_ik,
                "text": _normalize_text(_hp_entry.get("Summary")) or _hp_ik,
                "status": _hp_status,
                "comment": _hp_comment_text,
            }
        )
        _existing_hp_keys.add(_hp_ik_norm)

    # Post-process always_show_evidence: add Highest always-show tasks to _feature_map
    for _as_entry in (always_show_evidence or []):
        _as_ik = _normalize_text(_as_entry.get("Issue_Key"))
        _as_priority_key = _normalize_key(_as_entry.get("Priority"))
        if _as_priority_key not in priority_always_show_values:
            continue
        if bool(_as_entry.get("Bug")):
            continue
        _as_epic_id = _epic_id(_as_entry)
        _as_epic_key = _normalize_text(_as_entry.get("Epic_Key"))
        _as_epic_name = _normalize_text(_as_entry.get("Epic_Name")) or "No Epic"
        _as_bucket = epics.setdefault(
            _as_epic_id,
            {
                "epic_key": _as_epic_key,
                "epic_name": _as_epic_name,
                "report_items": [],
                "completed_items": [],
                "progress_items": [],
                "parent_subtasks": [],
                "high_priority_items": [],
                "feature_statuses": [],
                "next_week_items": [],
                "closed_tasks": 0,
                "bugs": {"closed": 0, "in_progress": 0},
                "_feature_map": {},
            },
        )
        report_epic_ids.add(_as_epic_id)
        _as_fmap = _as_bucket["_feature_map"]
        _as_is_subtask = bool(_as_entry.get("Subtask"))
        _as_parent_key = _normalize_text(_as_entry.get("Parent_Key"))
        _as_feat_key = (
            (_as_parent_key if _as_parent_key else _as_ik)
            if _as_is_subtask
            else _as_ik
        ) or _normalize_text(_as_entry.get("Summary")) or "feature"
        _as_feat_name = (
            (_normalize_text(_as_entry.get("Parent_Summary")) or _as_parent_key or _as_feat_key)
            if _as_is_subtask and _as_parent_key
            else (_normalize_text(_as_entry.get("Summary")) or _as_feat_key)
        )
        if _as_feat_key in _as_fmap:
            _as_fmap[_as_feat_key]["always_show"] = True
            if _as_is_subtask and _as_ik:
                _as_fmap[_as_feat_key]["issue_keys"].add(_as_ik)
                _as_subtask_summaries = _as_fmap[_as_feat_key].setdefault("subtask_summaries", {})
                if _as_ik not in _as_subtask_summaries:
                    _as_fmap[_as_feat_key].setdefault("subtask_issue_keys", []).append(_as_ik)
                    _as_subtask_summaries[_as_ik] = _normalize_text(_as_entry.get("Summary"))
        else:
            _as_comment_points = _collect_comment_points(_as_entry.get("Comments") or [])
            _as_fmap[_as_feat_key] = {
                "feature_key": _as_feat_key,
                "feature_name": _as_feat_name,
                "issue_keys": {_as_ik} if _as_ik else set(),
                "points": _as_comment_points,
                "comments_count": len(_as_entry.get("Comments") or []),
                "closed_tasks": 1 if _as_entry.get("Finished") else 0,
                "in_progress_tasks": 1 if not _as_entry.get("Finished") and _is_in_progress_status(_as_entry.get("Status")) else 0,
                "blocked_tasks": 0,
                "always_show": True,
                "subtask_issue_keys": [_as_ik] if _as_is_subtask and _as_ik else [],
                "subtask_summaries": {_as_ik: _normalize_text(_as_entry.get("Summary"))} if _as_is_subtask and _as_ik else {},
            }

    epic_entries: list[dict[str, Any]] = []
    for epic_id, epic in epics.items():
        if epic_id not in report_epic_ids:
            continue

        hp_issue_keys: set[str] = {
            _normalize_key(item.get("issue_key"))
            for item in (epic.get("high_priority_items") or [])
            if item.get("issue_key")
        }
        feature_statuses: list[dict[str, Any]] = []
        next_week_items: list[dict[str, Any]] = []
        for feature_key, feature in sorted(
            (epic.get("_feature_map") or {}).items(),
            key=lambda item: (_normalize_key(item[0]), _normalize_key(item[1].get("feature_name"))),
        ):
            raw_points = list(feature.get("points") or [])
            dedup_points: list[str] = []
            seen: set[str] = set()
            for point in raw_points:
                marker = _normalize_key(point)
                if not marker or marker in seen:
                    continue
                seen.add(marker)
                dedup_points.append(point)
            feature["points"] = dedup_points

            status_text = _build_compact_feature_status(feature)
            plan_text = _build_compact_plan_status(feature)

            feature_item = {
                "issue_key": feature_key,
                "text": _normalize_text(feature.get("feature_name")) or feature_key,
                "status": status_text,
                "comment": status_text,
                "closed_tasks": int(feature.get("closed_tasks") or 0),
                "issue_keys": sorted(list(feature.get("issue_keys") or []), key=_normalize_key),
            }
            subtask_keys = list(feature.get("subtask_issue_keys") or [])
            has_results = (
                int(feature.get("closed_tasks") or 0) > 0
                or len(feature.get("points") or []) > 0
            )
            is_high_priority = _normalize_key(feature_key) in hp_issue_keys
            is_always_show = bool(feature.get("always_show"))
            logger.debug(
                "FEATURE_CLASSIFY: feature_key=%s has_results=%s is_high_priority=%s is_always_show=%s hp_issue_keys=%s",
                feature_key, has_results, is_high_priority, is_always_show, hp_issue_keys,
            )
            if (has_results or is_always_show) and not is_high_priority:
                if is_always_show and not has_results:
                    feature_item["status"] = "No updates this week."
                    feature_item["comment"] = "No updates this week."
                if len(subtask_keys) > 2:
                    feature_item["aggregate_input"] = _build_aggregate_input(feature, mode="result")
                    feature_item["subtask_keys_in_report"] = subtask_keys
                    feature_item["aggregate_status"] = ""
                feature_statuses.append(feature_item)

            if int(feature.get("in_progress_tasks") or 0) > 0:
                plan_item: dict[str, Any] = {
                    "issue_key": feature_key,
                    "text": feature_item["text"],
                    "status": plan_text,
                    "comment": plan_text,
                    "subtasks": [],
                }
                if len(subtask_keys) > 2:
                    plan_item["aggregate_input"] = _build_aggregate_input(feature, mode="plan")
                    plan_item["subtask_keys_in_report"] = subtask_keys
                    plan_item["aggregate_status"] = ""
                next_week_items.append(plan_item)

        epic["feature_statuses"] = feature_statuses
        epic["next_week_items"] = next_week_items
        epic["report_items"] = []
        epic["completed_items"] = []
        epic["progress_items"] = []
        epic["parent_subtasks"] = []
        # high_priority_items already populated during evidence loop — do not overwrite
        epic.pop("_feature_map", None)
        epic_entries.append(epic)

    if not report_all_labels:
        epic_entries = [epic for epic in epic_entries if _normalize_text(epic.get("epic_key"))]

    epic_entries = sorted(
        epic_entries,
        key=lambda item: (_normalize_key(item.get("epic_name")), _normalize_key(item.get("epic_key"))),
    )

    next_week_plans: list[dict[str, Any]] = []
    for epic in epic_entries:
        items = list(epic.get("next_week_items") or [])
        if not items:
            continue
        next_week_plans.append(
            {
                "epic_key": epic["epic_key"],
                "epic_name": epic["epic_name"],
                "items": items,
            }
        )

    summary_rows_map: dict[str, dict[str, Any]] = {}
    for epic in epic_entries:
        epic_key = _normalize_text(epic.get("epic_key"))
        epic_name = _normalize_text(epic.get("epic_name")) or "No Epic"
        epic_id = epic_key if epic_key else f"name::{epic_name}"
        summary_rows_map[epic_id] = {
            "epic_key": epic_key,
            "epic_name": epic_name,
            "highlights": 0,
            "this_week": len(epic.get("feature_statuses") or []),
            "next_week": len(epic.get("next_week_items") or []),
            "closed_tasks": int(epic.get("closed_tasks") or 0),
        }

    for item in highlights:
        highlight_epic_key = _normalize_text(item.get("epic_key"))
        highlight_epic_name = _normalize_text(item.get("epic_name")) or "No Epic"
        highlight_epic_id = highlight_epic_key if highlight_epic_key else f"name::{highlight_epic_name}"
        row = summary_rows_map.setdefault(
            highlight_epic_id,
            {
                "epic_key": highlight_epic_key,
                "epic_name": highlight_epic_name,
                "highlights": 0,
                "this_week": 0,
                "next_week": 0,
                "closed_tasks": 0,
            },
        )
        row["highlights"] += 1

    summary_rows = sorted(
        summary_rows_map.values(),
        key=lambda item: (_normalize_key(item.get("epic_name")), _normalize_key(item.get("epic_key"))),
    )
    summary_totals = {
        "highlights": sum(int(row.get("highlights") or 0) for row in summary_rows),
        "this_week": sum(int(row.get("this_week") or 0) for row in summary_rows),
        "next_week": sum(int(row.get("next_week") or 0) for row in summary_rows),
        "closed_tasks": sum(int(row.get("closed_tasks") or 0) for row in summary_rows),
        "epics_covered": len(summary_rows),
    }

    risk_items: list[dict[str, Any]] = []
    for _re in (risk_evidence or []):
        _re_comments = _re.get("Comments") or []
        _re_action = "; ".join(_normalize_text(c) for c in _re_comments[:3] if c) or ""
        risk_items.append({
            "issue_key":     _normalize_text(_re.get("Issue_Key")),
            "text":          _normalize_text(_re.get("Summary")),
            "assignee":      _normalize_text(_re.get("Assignee")),
            "created":       _normalize_text(_re.get("Created")),
            "status":        _normalize_text(_re.get("Status")),
            "action_points": _re_action,
            "reporter":      _normalize_text(_re.get("Reporter")),
        })

    logger.info(
        "PAYLOAD SUMMARY: project=%s week=%s evidence=%s epics_total=%s epics_in_report=%s highlights=%s this_week_features=%s next_week_features=%s closed_tasks=%s bugs_closed=%s bugs_in_progress=%s project_bugs_in_progress=%s project_bugs_open=%s",
        project,
        week.key,
        len(evidence),
        len(epics),
        len(epic_entries),
        len(highlights),
        sum(len(epic.get("feature_statuses") or []) for epic in epic_entries),
        sum(len(epic.get("next_week_items") or []) for epic in epic_entries),
        sum(int(epic.get("closed_tasks") or 0) for epic in epic_entries),
        sum(int((epic.get("bugs") or {}).get("closed") or 0) for epic in epic_entries),
        sum(int((epic.get("bugs") or {}).get("in_progress") or 0) for epic in epic_entries),
        int(project_bug_stats.get("in_progress") or 0),
        int(project_bug_stats.get("open") or 0),
    )
    return {
        "meta": {
            "project": project,
            "week_key": week.key,
            "week_start": week.start.strftime("%Y-%m-%d"),
            "week_end": week.end.strftime("%Y-%m-%d"),
        },
        "highlights": highlights,
        "epics": epic_entries,
        "next_week_plans": next_week_plans,
        "summary_table": {
            "rows": summary_rows,
            "totals": summary_totals,
        },
        "project_bugs": {
            "in_progress": int(project_bug_stats.get("in_progress") or 0),
            "open": int(project_bug_stats.get("open") or 0),
        },
        "vacations": [],
        "risk_items": risk_items,
        "titles": {
            "main": config.get("jira_weekly_email", "title_main", fallback="Weekly Report"),
            "highlights": config.get("jira_weekly_email", "chapter_highlights_title", fallback="Highlights"),
            "results": config.get(
                "jira_weekly_email",
                "chapter_results_title",
                fallback="Key Results and Achievements",
            ),
            "plans": config.get("jira_weekly_email", "chapter_next_week_title", fallback="Next Week Plans"),
            "summary": config.get("jira_weekly_email", "chapter_summary_title", fallback="Summary"),
            "vacations": config.get("jira_weekly_email", "chapter_vacations_title", fallback="Vacations (next 60 days)"),
            "high_priority_subtitle": config.get(
                "jira_weekly_email", "chapter_results_high_priority_subtitle", fallback="High priority items"
            ),
            "bugs_subtitle": config.get(
                "jira_weekly_email", "chapter_results_bugs_subtitle", fallback="Bugs summary"
            ),
            "bugs_summary_template": config.get(
                "jira_weekly_email",
                "bugs_summary_template_closed_in_progress",
                fallback=(
                    "{closed} trouble reports/issues are analyzed and closed, "
                    "{in_progress} currently in progress, {open} open in project."
                ),
            ),
            "header_project_info": config.get(
                "jira_weekly_email", "header_project_info_title", fallback="Weekly execution summary"
            ),
            "header_banner_bg_color": config.get(
                "jira_weekly_email", "header_banner_bg_color", fallback="rgb(63,78,0)"
            ),
            "meta_report_period": config.get(
                "jira_weekly_email", "meta_report_period_label", fallback="Report Period"
            ),
            "meta_active_iteration": config.get(
                "jira_weekly_email", "meta_active_iteration_label", fallback="Active iteration"
            ),
            "meta_active_iteration_value": config.get(
                "jira_weekly_email", "meta_active_iteration_value", fallback=""
            ),
            "meta_report_owner": config.get(
                "jira_weekly_email",
                "meta_report_owner_label",
                fallback=config.get("jira_weekly_email", "meta_project_label", fallback="Report Owner"),
            ),
            "meta_report_owner_value": config.get(
                "jira_weekly_email", "meta_report_owner_value", fallback=project
            ),
            "meta_team_member": config.get(
                "jira_weekly_email",
                "meta_team_member_label",
                fallback=config.get("jira_weekly_email", "meta_generated_label", fallback="Team Member"),
            ),
            "meta_team_member_value": config.get(
                "jira_weekly_email", "meta_team_member_value", fallback=""
            ),
            "footer_html": config.get(
                "jira_weekly_email", "footer_html", fallback=""
            ),
        },
    }


def _apply_order_for_items(items: list[dict[str, Any]], order_map: list[str]) -> list[dict[str, Any]]:
    rank = {key: idx for idx, key in enumerate(order_map)}
    return sorted(items, key=lambda item: (rank.get(item.get("issue_key", ""), 10**9), item.get("issue_key", "")))


def apply_previous_order(payload: dict[str, Any], previous_snapshot: dict[str, Any] | None) -> dict[str, Any]:
    if not previous_snapshot:
        return payload

    order = previous_snapshot.get("order") or {}
    epic_order = [str(item) for item in (order.get("epic_order") or [])]
    issue_order_by_epic = order.get("issue_order_by_epic") or {}

    current = dict(payload)
    epics = list(current.get("epics") or [])
    if epic_order:
        epic_rank = {epic_id: idx for idx, epic_id in enumerate(epic_order)}

        def _epic_identifier(epic_entry: dict[str, Any]) -> str:
            epic_key = _normalize_text(epic_entry.get("epic_key"))
            if epic_key:
                return epic_key
            return f"name::{_normalize_text(epic_entry.get('epic_name'))}"

        epics = sorted(
            epics,
            key=lambda item: (
                epic_rank.get(_epic_identifier(item), 10**9),
                _normalize_key(item.get("epic_name")),
            ),
        )

        for epic in epics:
            epic_id = _epic_identifier(epic)
            issue_order = [str(item) for item in (issue_order_by_epic.get(epic_id) or [])]
            if issue_order:
                rank = {key: idx for idx, key in enumerate(issue_order)}
                epic["feature_statuses"] = _apply_order_for_items(list(epic.get("feature_statuses") or []), issue_order)
                epic["next_week_items"] = _apply_order_for_items(list(epic.get("next_week_items") or []), issue_order)
                epic["report_items"] = _apply_order_for_items(list(epic.get("report_items") or []), issue_order)
                epic["completed_items"] = _apply_order_for_items(list(epic.get("completed_items") or []), issue_order)
                epic["progress_items"] = _apply_order_for_items(list(epic.get("progress_items") or []), issue_order)
                epic["high_priority_items"] = _apply_order_for_items(
                    list(epic.get("high_priority_items") or []), issue_order
                )
                parent_subtasks = list(epic.get("parent_subtasks") or [])
                parent_subtasks = sorted(
                    parent_subtasks,
                    key=lambda item: (
                        rank.get(_normalize_text(item.get("parent_issue_key")), 10**9),
                        _normalize_key(item.get("parent_issue_key")),
                    ),
                )
                for group in parent_subtasks:
                    group["subtasks"] = sorted(
                        list(group.get("subtasks") or []),
                        key=lambda item: (
                            rank.get(_normalize_text(item.get("issue_key")), 10**9),
                            _normalize_key(item.get("issue_key")),
                        ),
                    )
                epic["parent_subtasks"] = parent_subtasks

    current["epics"] = epics
    return current


def _collect_text_targets(payload: dict[str, Any]) -> list[tuple[str, str]]:
    targets: list[tuple[str, str]] = []

    # SECTION: Highlights -> only progress/comment (headline/title is never AI-processed)
    for idx, item in enumerate(payload.get("highlights") or []):
        comment = _normalize_text(item.get("comment"))
        if comment:
            targets.append((f"highlights.{idx}.comment", comment))

    # SECTION: Key Results (epics) -> compact feature statuses
    for epic_idx, epic in enumerate(payload.get("epics") or []):
        for item_idx, item in enumerate(epic.get("feature_statuses") or []):
            aggregate_input = _normalize_text(item.get("aggregate_input"))
            if aggregate_input:
                targets.append((f"epics.{epic_idx}.feature_statuses.{item_idx}.aggregate_status", aggregate_input))
            else:
                status_text = _normalize_text(item.get("status"))
                if status_text and not _is_ai_skip_text(status_text):
                    targets.append((f"epics.{epic_idx}.feature_statuses.{item_idx}.status", status_text))

        for item_idx, item in enumerate(epic.get("next_week_items") or []):
            if item.get("aggregate_input"):
                continue  # handled via next_week_plans path
            status_text = _normalize_text(item.get("status"))
            if status_text and not _is_ai_skip_text(status_text):
                targets.append((f"epics.{epic_idx}.next_week_items.{item_idx}.status", status_text))

    # Backward-compatible section paths (legacy payload shape)
    for epic_idx, epic in enumerate(payload.get("epics") or []):
        for section in ("report_items", "completed_items", "progress_items", "high_priority_items"):
            for item_idx, item in enumerate(epic.get(section) or []):
                comment = _normalize_text(item.get("comment"))
                if comment:
                    targets.append((f"epics.{epic_idx}.{section}.{item_idx}.comment", comment))

        for parent_idx, parent_group in enumerate(epic.get("parent_subtasks") or []):
            for subtask_idx, subtask in enumerate(parent_group.get("subtasks") or []):
                comment = _normalize_text(subtask.get("comment"))
                if comment:
                    targets.append(
                        (
                            f"epics.{epic_idx}.parent_subtasks.{parent_idx}.subtasks.{subtask_idx}.comment",
                            comment,
                        )
                    )

    # SECTION: Plans (next_week_plans) -> only comments
    for epic_idx, plan_epic in enumerate(payload.get("next_week_plans") or []):
        for item_idx, item in enumerate(plan_epic.get("items") or []):
            aggregate_input = _normalize_text(item.get("aggregate_input"))
            if aggregate_input:
                targets.append((f"next_week_plans.{epic_idx}.items.{item_idx}.aggregate_status", aggregate_input))
            else:
                status_text = _normalize_text(item.get("status"))
                if status_text and not _is_ai_skip_text(status_text):
                    targets.append((f"next_week_plans.{epic_idx}.items.{item_idx}.status", status_text))
                comment = _normalize_text(item.get("comment"))
                if comment and not _is_ai_skip_text(comment):
                    targets.append((f"next_week_plans.{epic_idx}.items.{item_idx}.comment", comment))

            for subtask_idx, subtask in enumerate(item.get("subtasks") or []):
                subtask_comment = _normalize_text(subtask.get("comment"))
                if subtask_comment:
                    targets.append(
                        (
                            f"next_week_plans.{epic_idx}.items.{item_idx}.subtasks.{subtask_idx}.comment",
                            subtask_comment,
                        )
                    )
    return targets


def _extract_json_object(text: str) -> dict[str, Any] | None:
    raw = text.strip()
    if not raw:
        return None
    try:
        value = json.loads(raw)
        return value if isinstance(value, dict) else None
    except Exception:
        pass

    match = re.search(r"\{.*\}", raw, re.DOTALL)
    if not match:
        return None
    try:
        value = json.loads(match.group(0))
    except Exception:
        return None
    return value if isinstance(value, dict) else None


def _json_dict_or_raise(response: requests.Response) -> dict[str, Any]:
    parsed = response.json()
    if isinstance(parsed, dict):
        return parsed
    raise ValueError(f"Invalid JSON response type: {type(parsed).__name__}")


def _log_ollama_check_commands(ollama_url: str, model: str, has_api_key: bool) -> None:
    base = ollama_url.rstrip("/")
    logger.error("Ollama API health checks (run in console):")
    logger.error('curl -i "%s/api/tags"', base)
    if has_api_key:
        logger.error('curl -i -H "Authorization: Bearer <OLLAMA_API_KEY>" "%s/api/tags"', base)
        logger.error(
            'curl -i -X POST -H "Content-Type: application/json" -H "Authorization: Bearer <OLLAMA_API_KEY>" "%s/api/generate" -d "{\\"model\\":\\"%s\\",\\"prompt\\":\\"ping\\",\\"stream\\":false}"',
            base,
            model or "<MODEL>",
        )
    else:
        logger.error(
            'curl -i -X POST -H "Content-Type: application/json" "%s/api/generate" -d "{\\"model\\":\\"%s\\",\\"prompt\\":\\"ping\\",\\"stream\\":false}"',
            base,
            model or "<MODEL>",
        )


_AI_SKIP_PHRASES: frozenset[str] = frozenset({
    "in progress.",
    "in progress",
    "completed.",
    "completed",
    "blocked.",
    "blocked",
})


def _is_ai_skip_text(text: str) -> bool:
    """Return True if the text is a trivial status word that AI cannot improve."""
    return _normalize_key(text).rstrip(". ") in _AI_SKIP_PHRASES


def _build_rewrite_prompt(targets: list[tuple[str, str]], start_index: int = 1) -> tuple[dict[str, str], str]:
    if not targets:
        return {}, ""

    def _intent_from_path(path: str) -> str:
        if path.endswith(".aggregate_status"):
            if "next_week" in path:
                return "AGGREGATE_PLAN"
            return "AGGREGATE"
        if path.startswith("next_week_plans.") or ".next_week_items." in path:
            return "PLAN"
        if path.startswith("highlights."):
            return "HIGHLIGHT"
        return "RESULT"

    # NOTE: model range is 32b–120b; keep prompt short and unambiguous.
    prompt_lines = [
        "You are a technical writer producing a formal weekly engineering status report.",
        "Task: rewrite each numbered input into one clear English status line for management.",
        "",
        "Rules (follow exactly):",
        "- Output language: English only.",
        "- Length: 1 compact line, up to 35 words.",
        "- No links, no URLs, no commit/PR/MR hashes, no Jira ticket numbers (e.g. PROJ-123).",
        "- No markdown, no bullet markers, no code blocks.",
        "- RESULT item: state what was accomplished; add next step only if explicitly mentioned in input.",
        "- PLAN item: state what will be done next week based on input; use future tense.",
        "- HIGHLIGHT item: state progress only (title is shown separately, do not repeat it).",
        "- AGGREGATE item: one English line (max 35 words) summarizing COLLECTIVE progress of all listed tasks. Do NOT list individual task names — describe the outcome.",
        "- AGGREGATE_PLAN item: one English line (max 35 words) on what will be done COLLECTIVELY next week. Future tense. Do NOT list individual task names.",
        "- If input has real content: rewrite it concisely.",
        "- If input has no real content (e.g. only status words): return it unchanged.",
        "",
        "Output: return ONLY a valid JSON object. Example:",
        '{"t1":"Merged authentication cache fix into main branch.","t2":"Continue integration testing for payment module."}',
        "",
        "Inputs:",
    ]
    target_map: dict[str, str] = {}
    for idx, (path, text_value) in enumerate(targets, start=start_index):
        target_id = f"t{idx}"
        target_map[target_id] = path
        intent = _intent_from_path(path)
        prompt_lines.append(f'{target_id} [{intent}]: "{text_value}"')
    return target_map, "\n".join(prompt_lines)


def _sanitize_ai_text(text: str) -> str:
    cleaned = _normalize_text(text)
    if not cleaned:
        return ""

    # Remove links and repository/file paths regardless of formatting.
    cleaned = re.sub(r"\[[^\]]+\]\([^)]+\)", "", cleaned, flags=re.IGNORECASE)
    cleaned = re.sub(r"<(?:https?://|www\.|file://)[^>]+>", "", cleaned, flags=re.IGNORECASE)
    cleaned = re.sub(r"(?:https?://|ftp://|file://|www\.)\S+", "", cleaned, flags=re.IGNORECASE)
    cleaned = re.sub(r"\b[a-z0-9.-]+\.[a-z]{2,}(?:/[^\s)\],;]+)+", "", cleaned, flags=re.IGNORECASE)
    cleaned = re.sub(r"\\\\[A-Za-z0-9._$ -]+\\[^\s,;)\]]+", "", cleaned)
    cleaned = re.sub(r"\b[A-Za-z]:\\(?:[^\\/:*?\"<>|\r\n]+\\)*[^\\/:*?\"<>|\s,;)\]]*", "", cleaned)
    cleaned = re.sub(r"(?:(?<=\s)|^)(?:\.\.?/|/)[^\s,;)\]]+", " ", cleaned)
    cleaned = re.sub(r"\b(?:[A-Za-z0-9_.-]+/){1,}[A-Za-z0-9_.-]+\.[A-Za-z0-9]{1,8}\b", "", cleaned)
    cleaned = re.sub(r"\b(?:[A-Za-z0-9_.-]+\\){1,}[A-Za-z0-9_.-]+\.[A-Za-z0-9]{1,8}\b", "", cleaned)
    cleaned = re.sub(r"\b[0-9a-f]{7,40}\b", "", cleaned, flags=re.IGNORECASE)
    cleaned = re.sub(
        r"(?i)\b(?:pr|mr|pull request|merge request|commit)\b\s*[:#-]?\s*[A-Za-z0-9/_-]*",
        "",
        cleaned,
    )
    cleaned = re.sub(r"(?i)\b(?:results?|status|plan|update|details)\s*:\s*", "", cleaned)
    cleaned = re.sub(r"(?i)\b(?:see|ref(?:erence)?)\s+(?:commit|pr|mr|pull request|merge request|link|url)\b", "", cleaned)
    cleaned = re.sub(
        r"\(\s*(?:https?://|www\.|file://|\\\\|[A-Za-z]:\\|/|\.\./|\./)[^)]*\)",
        "",
        cleaned,
        flags=re.IGNORECASE,
    )
    cleaned = re.sub(
        r"\[\s*(?:https?://|www\.|file://|\\\\|[A-Za-z]:\\|/|\.\./|\./)[^\]]*\]",
        "",
        cleaned,
        flags=re.IGNORECASE,
    )
    cleaned = re.sub(r"\s+([,.;:!?])", r"\1", cleaned)
    cleaned = re.sub(r"[ ]{2,}", " ", cleaned)
    cleaned = _normalize_text(cleaned.strip(" -,:;"))
    if not cleaned:
        return ""
    if _contains_cyrillic(cleaned):
        return ""
    words = cleaned.split()
    if not words:
        return ""
    if len(words) > 70:
        cleaned = " ".join(words[:70]).rstrip(" ,;:-")
        if cleaned and cleaned[-1] not in ".!?":
            cleaned += "..."
    return cleaned


def _apply_rewrite_map(payload: dict[str, Any], target_map: dict[str, str], rewrite_map: dict[str, Any]) -> dict[str, Any]:
    updated = json.loads(json.dumps(payload))
    for target_id, target_path in target_map.items():
        rewritten = _sanitize_ai_text(_normalize_text(rewrite_map.get(target_id)))
        if not rewritten:
            continue
        path_tokens = target_path.split(".")
        cursor: Any = updated
        try:
            for token in path_tokens[:-1]:
                if token.isdigit():
                    cursor = cursor[int(token)]
                else:
                    cursor = cursor[token]
            leaf = path_tokens[-1]
            if leaf.isdigit():
                cursor[int(leaf)] = rewritten
            else:
                cursor[leaf] = rewritten
        except Exception:
            continue
    return updated


def _log_webui_check_commands(api_url: str, model: str, has_api_key: bool, prompt: str = "ping") -> None:
    logger.error("WebUI API health checks (run in console):")
    # Use a truncated prompt to show it's not just 'ping', but avoid massive logs
    display_prompt = prompt if len(prompt) < 200 else (prompt[:200] + "... (truncated)")
    
    json_body = json.dumps({
        "model": model or "<MODEL>",
        "messages": [{"role": "user", "content": display_prompt}],
        "stream": False
    })
    # Escape for shell (simple approach)
    json_body_sh = json_body.replace("'", "'\\''")
    json_body_ps = json_body.replace('"', '\\"')

    if has_api_key:
        logger.error(
            f"curl -i -X POST '{api_url}' -H 'Content-Type: application/json' -H 'Authorization: Bearer <WEBUI_API_KEY>' -d '{json_body_sh}'"
        )
        logger.error(
            f'PowerShell: curl.exe -sS -i -X POST "{api_url}" -H "Content-Type: application/json" -H "Authorization: Bearer <WEBUI_API_KEY>" --data-raw "{json_body_ps}"'
        )
    else:
        logger.error(
            f"curl -i -X POST '{api_url}' -H 'Content-Type: application/json' -d '{json_body_sh}'"
        )
        logger.error(
            f'PowerShell: curl.exe -sS -i -X POST "{api_url}" -H "Content-Type: application/json" --data-raw "{json_body_ps}"'
        )


def _build_webui_api_url(base_url: str, endpoint: str) -> str:
    base = _normalize_text(base_url).rstrip("/")
    ep = _normalize_text(endpoint)
    if not ep:
        return base
    if ep.startswith("http://") or ep.startswith("https://"):
        return ep.rstrip("/")

    ep_path = "/" + ep.lstrip("/")
    if base.endswith(ep_path):
        return base

    parts = urlsplit(base)
    if not parts.scheme or not parts.netloc:
        if not base:
            return ep_path
        return f"{base}/{ep.lstrip('/')}"

    base_path = parts.path.rstrip("/")
    if base_path and ep_path.startswith(base_path + "/"):
        merged_path = ep_path
    elif base_path:
        merged_path = f"{base_path}{ep_path}"
    else:
        merged_path = ep_path
    return urlunsplit((parts.scheme, parts.netloc, merged_path, "", ""))


def rewrite_payload_with_ollama(payload: dict[str, Any], config: ConfigParser, extra_params: dict[str, Any]) -> dict[str, Any]:
    ollama_enabled = _bool_value(
        extra_params.get("ollama_enabled", config.get("ollama", "enabled", fallback="true")),
        True,
    )
    if not ollama_enabled:
        return payload

    model = _normalize_text(extra_params.get("ollama_model") or config.get("ollama", "model", fallback=""))
    if not model:
        logger.warning("Ollama model is not configured; using deterministic text.")
        return payload

    all_targets = _collect_text_targets(payload)
    if not all_targets:
        return payload

    ollama_url = _normalize_text(extra_params.get("ollama_url") or config.get("ollama", "url", fallback="http://localhost:11434"))
    ollama_api_key = _strip_wrapping_quotes(
        _normalize_text(extra_params.get("ollama_api_key") or config.get("ollama", "api_key", fallback=""))
    )
    timeout_seconds = int(
        _normalize_text(extra_params.get("ollama_timeout_seconds") or config.get("ollama", "timeout_seconds", fallback="60"))
        or "60"
    )
    headers: dict[str, str] = {"Content-Type": "application/json"}
    if ollama_api_key:
        headers["Authorization"] = f"Bearer {ollama_api_key}"

    full_rewrite_map: dict[str, Any] = {}
    full_target_map: dict[str, str] = {}
    batch_size = 5

    for i in range(0, len(all_targets), batch_size):
        batch = all_targets[i : i + batch_size]
        target_map, prompt = _build_rewrite_prompt(batch, start_index=i + 1)

        try:
            def _request():
                response = requests.post(
                    f"{ollama_url.rstrip('/')}/api/generate",
                    headers=headers,
                    json={
                        "model": model,
                        "prompt": prompt,
                        "stream": False,
                        "options": {
                            "temperature": float(
                                _normalize_text(
                                    extra_params.get("ollama_temperature") or config.get("ollama", "temperature", fallback="0.2")
                                )
                                or "0.2"
                            )
                        },
                    },
                    timeout=timeout_seconds,
                )
                response.raise_for_status()
                return _json_dict_or_raise(response)

            response_json = retry_ai_call(
                _request,
                logger=logger,
                retry_exceptions=(
                    requests.Timeout,
                    requests.exceptions.ReadTimeout,
                    requests.exceptions.ConnectTimeout,
                    ValueError,
                ),
            )
            response_text = _normalize_text(response_json.get("response", ""))
            rewrite_map = _extract_json_object(response_text)
            if rewrite_map:
                full_rewrite_map.update(rewrite_map)
                full_target_map.update(target_map)
            else:
                logger.warning("Ollama response is not valid JSON map for batch %s; skipping.", i)
        except requests.HTTPError as exc:
            status = exc.response.status_code if exc.response is not None else "n/a"
            response_text = _normalize_text(exc.response.text if exc.response is not None else "")
            if len(response_text) > 500:
                response_text = response_text[:500] + "..."
            logger.error("Ollama HTTP error: status=%s body=%s", status, response_text or "<empty>")
            _log_ollama_check_commands(ollama_url, model, bool(ollama_api_key))
        except (
            requests.Timeout,
            requests.exceptions.ReadTimeout,
            requests.exceptions.ConnectTimeout,
            ValueError,
        ) as exc:
            logger.warning("Ollama transient failure (batch %s): %s", (i // batch_size) + 1, exc)
        except Exception as exc:
            logger.error("Ollama call failed: %s", exc)
            _log_ollama_check_commands(ollama_url, model, bool(ollama_api_key))

    return _apply_rewrite_map(payload, full_target_map, full_rewrite_map)


def rewrite_payload_with_webui(payload: dict[str, Any], config: ConfigParser, extra_params: dict[str, Any]) -> dict[str, Any]:
    webui_section = config["webui"] if config.has_section("webui") else {}
    webui_enabled = _bool_value(
        extra_params.get("webui_enabled")
        or webui_section.get("enabled")
        or config.get("webui", "enabled", fallback="true"),
        True,
    )
    if not webui_enabled:
        return payload

    model = _normalize_text(
        extra_params.get("webui_model")
        or webui_section.get("model")
        or config.get("webui", "model", fallback="")
    )
    if not model:
        logger.warning("WebUI model is not configured; using deterministic text.")
        return payload

    all_targets = _collect_text_targets(payload)
    if not all_targets:
        return payload

    base_url = _normalize_text(
        extra_params.get("webui_url")
        or webui_section.get("url")
        or config.get("webui", "url", fallback="http://localhost:3000")
    )
    endpoint = _normalize_text(
        extra_params.get("webui_endpoint")
        or webui_section.get("endpoint")
        or config.get("webui", "endpoint", fallback="/api/chat/completions")
    )
    api_url = _build_webui_api_url(base_url, endpoint)
    configured_webui_api_key = (
        webui_section.get("api_key")
        if webui_section
        else config.get("webui", "api_key", fallback="")
    )
    webui_api_key = _strip_wrapping_quotes(
        _normalize_text(extra_params.get("webui_api_key") or configured_webui_api_key)
    )
    timeout_seconds = int(
        _normalize_text(
            extra_params.get("webui_timeout_seconds")
            or webui_section.get("timeout_seconds")
            or config.get("webui", "timeout_seconds", fallback="120")
        )
        or "120"
    )
    connect_timeout_seconds = int(
        _normalize_text(
            extra_params.get("webui_connect_timeout_seconds")
            or webui_section.get("connect_timeout_seconds")
            or config.get("webui", "connect_timeout_seconds", fallback="10")
        )
        or "10"
    )
    temperature = float(
        _normalize_text(
            extra_params.get("webui_temperature")
            or webui_section.get("temperature")
            or config.get("webui", "temperature", fallback="0.2")
        )
        or "0.2"
    )
    logger.info(
        "WEBUI CONFIG: enabled=%s url=%s endpoint=%s api_url=%s model=%s api_key_set=%s timeout(connect/read)=%s/%s",
        webui_enabled,
        base_url,
        endpoint,
        api_url,
        model,
        bool(webui_api_key),
        connect_timeout_seconds,
        timeout_seconds,
    )

    headers: dict[str, str] = {"Content-Type": "application/json"}
    if webui_api_key:
        headers["Authorization"] = f"Bearer {webui_api_key}"

    full_rewrite_map: dict[str, Any] = {}
    full_target_map: dict[str, str] = {}
    batch_size = 5

    logger.info("WEBUI REQUEST: api_url=%s total_targets=%s batch_size=%s", api_url, len(all_targets), batch_size)

    for i in range(0, len(all_targets), batch_size):
        batch = all_targets[i : i + batch_size]
        target_map, prompt = _build_rewrite_prompt(batch, start_index=i + 1)

        try:
            def _request():
                response = requests.post(
                    api_url,
                    headers=headers,
                    json={
                        "model": model,
                        "messages": [
                            {"role": "system", "content": "You are an AI assistant that rewrites raw text snippets into formal report entries, returning only a single, valid JSON object with the results."},
                            {"role": "user", "content": prompt},
                        ],
                        "stream": False,
                        "temperature": temperature,
                    },
                    timeout=(connect_timeout_seconds, timeout_seconds),
                )
                response.raise_for_status()
                return _json_dict_or_raise(response)

            response_json = retry_ai_call(
                _request,
                logger=logger,
                retry_exceptions=(
                    requests.Timeout,
                    requests.exceptions.ReadTimeout,
                    requests.exceptions.ConnectTimeout,
                    ValueError,
                ),
            )

            response_text = ""
            choices = response_json.get("choices")
            if isinstance(choices, list) and choices:
                first_choice = choices[0]
                if isinstance(first_choice, dict):
                    message = first_choice.get("message")
                    if isinstance(message, dict):
                        response_text = _normalize_text(message.get("content"))
            if not response_text:
                response_text = _normalize_text(response_json.get("response", ""))

            rewrite_map = _extract_json_object(response_text)
            if rewrite_map:
                full_rewrite_map.update(rewrite_map)
                full_target_map.update(target_map)
            else:
                logger.warning("WebUI response is not valid JSON map for batch %s; skipping.", i)
        except requests.HTTPError as exc:
            status = exc.response.status_code if exc.response is not None else "n/a"
            response_text = _normalize_text(exc.response.text if exc.response is not None else "")
            if len(response_text) > 500:
                response_text = response_text[:500] + "..."
            logger.error("WebUI HTTP error: status=%s body=%s", status, response_text or "<empty>")
            _log_webui_check_commands(api_url, model, bool(webui_api_key), prompt)
        except (
            requests.Timeout,
            requests.exceptions.ReadTimeout,
            requests.exceptions.ConnectTimeout,
            ValueError,
        ) as exc:
            logger.warning("WebUI transient failure (batch %s): %s", (i // batch_size) + 1, exc)
        except Exception as exc:
            logger.error("WebUI call failed: %s", exc)

    return _apply_rewrite_map(payload, full_target_map, full_rewrite_map)


def rewrite_payload_with_ai(payload: dict[str, Any], config: ConfigParser, extra_params: dict[str, Any]) -> dict[str, Any]:
    section = config["jira_weekly_email"] if config.has_section("jira_weekly_email") else {}
    provider_raw = _normalize_text(extra_params.get("ai_provider") or section.get("ai_provider"))
    if provider_raw:
        provider = _normalize_key(provider_raw)
    else:
        webui_enabled = _bool_value(
            extra_params.get("webui_enabled", config.get("webui", "enabled", fallback="false")),
            False,
        )
        provider = "webui" if webui_enabled else "ollama"
    if provider == "webui":
        return rewrite_payload_with_webui(payload, config, extra_params)
    if provider not in {"", "ollama"}:
        logger.warning("Unknown ai_provider=%s, falling back to ollama.", provider)
    return rewrite_payload_with_ollama(payload, config, extra_params)


def _snapshot_week_tuple(snapshot: dict[str, Any]) -> tuple[int, int] | None:
    meta = snapshot.get("meta") or {}
    try:
        year = int(meta.get("year"))
        week = int(meta.get("week"))
        return year, week
    except Exception:
        pass

    week_key = _normalize_text(meta.get("week_key"))
    match = re.fullmatch(r"(\d{2,4})'?w(\d{1,2})", week_key, flags=re.IGNORECASE)
    if not match:
        return None
    year_val = int(match.group(1))
    if year_val < 100:
        year_val += 2000
    return year_val, int(match.group(2))


def _previous_week_window(current_week: WeekWindow) -> WeekWindow:
    previous_date = current_week.start - timedelta(days=7)
    previous_iso = previous_date.isocalendar()
    previous_start = previous_date - timedelta(days=previous_date.weekday())
    previous_end = previous_start + timedelta(days=6)
    return WeekWindow(
        year=previous_iso.year,
        week=previous_iso.week,
        start=previous_start,
        end=previous_end,
        key=_week_key(previous_iso.year, previous_iso.week),
    )


def _read_snapshot_json(path: Path) -> dict[str, Any] | None:
    for encoding in ("utf-8", "utf-8-sig"):
        try:
            value = json.loads(path.read_text(encoding=encoding))
        except Exception:
            continue
        if isinstance(value, dict):
            return value
    return None


def load_previous_snapshot(snapshot_dir: Path, project: str, current_week: WeekWindow) -> dict[str, Any] | None:
    previous_week = _previous_week_window(current_week)
    previous_week_compact = previous_week.key.replace("'", "")
    legacy_base = snapshot_dir / "snapshots" / "jira_weekly_email"
    search_dirs = [
        snapshot_dir,
        snapshot_dir / project,
        legacy_base,
        legacy_base / project,
    ]
    previous_candidates = [
        f"jira_weekly_email_{project}_{previous_week.key}.json",
        f"jira_weekly_email_{project}_{previous_week_compact}.json",
        f"{previous_week.key}.json",
        f"{previous_week_compact}.json",
    ]
    previous_candidates_folded = {name.casefold() for name in previous_candidates}
    logger.info(
        "SNAPSHOT SEARCH: previous_week=%s project=%s candidates=%s dirs=%s abs_dirs=%s",
        previous_week.key,
        project,
        ",".join(previous_candidates),
        ",".join(str(path) for path in search_dirs),
        ",".join(str(path.resolve()) for path in search_dirs),
    )

    for directory in search_dirs:
        if not directory.exists():
            continue
        for previous_path in directory.glob("*.json"):
            if previous_path.name.casefold() not in previous_candidates_folded:
                continue
            logger.info("Checking candidate file: %s", previous_path)
            payload = _read_snapshot_json(previous_path)
            if payload is None:
                logger.warning("Failed to read JSON from %s", previous_path)
                continue
            week_tuple = _snapshot_week_tuple(payload)
            target_tuple = (previous_week.year, previous_week.week)
            if week_tuple == target_tuple or week_tuple is None:
                logger.info("Snapshot accepted: %s", previous_path)
                return payload
            logger.warning("Snapshot rejected: %s (week=%s, expected=%s)", previous_path.name, week_tuple, target_tuple)

    if snapshot_dir.exists():
        expected_names = {
            f"jira_weekly_email_{project}_{previous_week.key}.json".casefold(),
            f"jira_weekly_email_{project}_{previous_week_compact}.json".casefold(),
        }
        for candidate_path in snapshot_dir.rglob("*.json"):
            if candidate_path.name.casefold() not in expected_names:
                continue
            payload = _read_snapshot_json(candidate_path)
            if payload is None:
                continue
            week_tuple = _snapshot_week_tuple(payload)
            if week_tuple == (previous_week.year, previous_week.week) or week_tuple is None:
                return payload

    latest_candidate: tuple[tuple[int, int], dict[str, Any]] | None = None
    current_tuple = (current_week.year, current_week.week)
    for directory in search_dirs:
        if not directory.exists():
            continue
        for candidate_path in directory.glob("*.json"):
            payload = _read_snapshot_json(candidate_path)
            if payload is None:
                continue
            week_tuple = _snapshot_week_tuple(payload)
            if not week_tuple or week_tuple >= current_tuple:
                continue
            if latest_candidate is None or week_tuple > latest_candidate[0]:
                latest_candidate = (week_tuple, payload)
    if latest_candidate:
        return latest_candidate[1]
    return None


def _extract_order(payload: dict[str, Any]) -> dict[str, Any]:
    epic_order: list[str] = []
    issue_order_by_epic: dict[str, list[str]] = {}
    for epic in payload.get("epics") or []:
        epic_key = _normalize_text(epic.get("epic_key"))
        epic_name = _normalize_text(epic.get("epic_name"))
        epic_id = epic_key if epic_key else f"name::{epic_name}"
        epic_order.append(epic_id)
        keys: list[str] = []
        for section in (
            "feature_statuses",
            "next_week_items",
            "report_items",
            "completed_items",
            "progress_items",
            "high_priority_items",
        ):
            for item in epic.get(section) or []:
                key = _normalize_text(item.get("issue_key"))
                if key and key not in keys:
                    keys.append(key)
        for group in epic.get("parent_subtasks") or []:
            parent_key = _normalize_text(group.get("parent_issue_key"))
            if parent_key and parent_key not in keys:
                keys.append(parent_key)
            for subtask in group.get("subtasks") or []:
                subtask_key = _normalize_text(subtask.get("issue_key"))
                if subtask_key and subtask_key not in keys:
                    keys.append(subtask_key)
        issue_order_by_epic[epic_id] = keys
    return {"epic_order": epic_order, "issue_order_by_epic": issue_order_by_epic}


def _payload_to_lines(payload: dict[str, Any]) -> list[str]:
    lines: list[str] = []
    for item in payload.get("highlights") or []:
        headline = _normalize_text(item.get("headline"))
        comment = _normalize_text(item.get("comment"))
        highlight_line = headline
        if comment:
            highlight_line = f"{headline} - {comment}" if headline else comment
        lines.append(f"HIGHLIGHT {highlight_line} ({item.get('issue_key')})")

    for epic in payload.get("epics") or []:
        lines.append(f"EPIC {epic.get('epic_name')} ({epic.get('epic_key')})")
        for item in epic.get("feature_statuses") or []:
            lines.append(f"feature:{item.get('issue_key')} {item.get('text')} status={item.get('status')}")
        for item in epic.get("next_week_items") or []:
            lines.append(f"next_week:{item.get('issue_key')} {item.get('text')} status={item.get('status')}")
        for section in ("report_items", "completed_items", "progress_items", "high_priority_items"):
            for item in epic.get(section) or []:
                lines.append(f"{section}:{item.get('issue_key')} {item.get('text')}")
        for group in epic.get("parent_subtasks") or []:
            lines.append(f"parent:{group.get('parent_issue_key')} {group.get('parent_text')}")
            for subtask in group.get("subtasks") or []:
                lines.append(
                    f"subtask:{subtask.get('issue_key')} {subtask.get('text')} status={subtask.get('status')}"
                )
                if _normalize_text(subtask.get("comment")):
                    lines.append(f"subtask_comment:{subtask.get('issue_key')} {subtask.get('comment')}")
        bugs = epic.get("bugs") or {}
        lines.append(f"bugs closed={bugs.get('closed', 0)} in_progress={bugs.get('in_progress', 0)}")

    for epic in payload.get("next_week_plans") or []:
        lines.append(f"PLAN {epic.get('epic_name')} ({epic.get('epic_key')})")
        for item in epic.get("items") or []:
            lines.append(f"plan:{item.get('issue_key')} {item.get('text')}")
            if _normalize_text(item.get("comment")):
                lines.append(f"plan_comment:{item.get('issue_key')} {item.get('comment')}")
            for subtask in item.get("subtasks") or []:
                lines.append(
                    f"plan_subtask:{subtask.get('issue_key')} {subtask.get('text')} status={subtask.get('status')}"
                )
                if _normalize_text(subtask.get("comment")):
                    lines.append(f"plan_subtask_comment:{subtask.get('issue_key')} {subtask.get('comment')}")

    summary_table = payload.get("summary_table") or {}
    for row in summary_table.get("rows") or []:
        lines.append(
            "SUMMARY "
            f"{row.get('epic_name')} ({row.get('epic_key')}) "
            f"H={row.get('highlights', 0)} W={row.get('this_week', 0)} "
            f"N={row.get('next_week', 0)} C={row.get('closed_tasks', 0)}"
        )

    for vacation in payload.get("vacations") or []:
        lines.append(f"VACATION {vacation}")
    return lines


def _render_summary_table_console(summary_table: dict[str, Any]) -> str:
    rows = list((summary_table or {}).get("rows") or [])
    totals = (summary_table or {}).get("totals") or {}
    if not rows:
        return "Section summary: no data."

    headers = ["Epic", "Highlights", "ThisWeek", "NextWeek", "Closed"]
    table_rows: list[list[str]] = []
    for row in rows:
        epic_name = _normalize_text(row.get("epic_name")) or "No Epic"
        epic_key = _normalize_text(row.get("epic_key"))
        epic_label = f"{epic_name} ({epic_key})" if epic_key else epic_name
        table_rows.append(
            [
                epic_label,
                str(int(row.get("highlights") or 0)),
                str(int(row.get("this_week") or 0)),
                str(int(row.get("next_week") or 0)),
                str(int(row.get("closed_tasks") or 0)),
            ]
        )

    table_rows.append(
        [
            "TOTAL",
            str(int(totals.get("highlights") or 0)),
            str(int(totals.get("this_week") or 0)),
            str(int(totals.get("next_week") or 0)),
            str(int(totals.get("closed_tasks") or 0)),
        ]
    )
    widths = [len(h) for h in headers]
    for row in table_rows:
        for idx, cell in enumerate(row):
            widths[idx] = max(widths[idx], len(cell))

    def _fmt(cells: list[str]) -> str:
        return " | ".join(cell.ljust(widths[idx]) for idx, cell in enumerate(cells))

    separator = "-+-".join("-" * width for width in widths)
    lines = [
        f"Section summary (epics covered: {int(totals.get('epics_covered') or len(rows))})",
        _fmt(headers),
        separator,
    ]
    for row in table_rows:
        lines.append(_fmt(row))
    return "\n".join(lines)


def compute_payload_diff(previous_payload: dict[str, Any] | None, current_payload: dict[str, Any]) -> list[str]:
    if not previous_payload:
        return []
    prev_lines = _payload_to_lines(previous_payload)
    curr_lines = _payload_to_lines(current_payload)
    return list(difflib.ndiff(prev_lines, curr_lines))


def _diff_stats(diff_lines: list[str]) -> dict[str, int]:
    stats = {"added": 0, "removed": 0, "unchanged": 0}
    for line in diff_lines:
        if line.startswith("+ "):
            stats["added"] += 1
        elif line.startswith("- "):
            stats["removed"] += 1
        elif line.startswith("  "):
            stats["unchanged"] += 1
    return stats


def _strikethrough(text: str) -> str:
    return "".join(f"{char}\u0336" for char in text)


def render_console_diff(
    diff_lines: list[str],
    *,
    project: str,
    current_week_key: str,
    previous_week_key: str,
    use_color: bool = True,
) -> None:
    if not diff_lines:
        return

    tqdm.write(f"[DIFF] {project} {current_week_key} vs {previous_week_key}")
    for line in diff_lines:
        if line.startswith("? "):
            continue
        payload = line[2:]
        if line.startswith("- "):
            old_text = _strikethrough(payload)
            if use_color:
                tqdm.write(f"  - \x1b[31m{old_text}\x1b[0m")
            else:
                tqdm.write(f"  - {old_text}")
        elif line.startswith("+ "):
            if use_color:
                tqdm.write(f"  + \x1b[32m{payload}\x1b[0m")
            else:
                tqdm.write(f"  + {payload}")
        else:
            if use_color:
                tqdm.write(f"    \x1b[37m{payload}\x1b[0m")
            else:
                tqdm.write(f"    {payload}")


def parse_vacations_excel(
    path: Path,
    *,
    sheet: str,
    markers: set[str],
    horizon_start: date,
    horizon_days: int,
) -> list[str]:
    if not path.exists():
        logger.warning("Vacation file not found: %s", path)
        return []

    try:
        workbook = load_workbook(path, data_only=True)
    except Exception as exc:
        logger.error("Vacation file cannot be read: file=%s error=%s", path, exc)
        return []
    if sheet not in workbook.sheetnames:
        logger.warning("Vacation sheet not found: file=%s sheet=%s", path, sheet)
        return []
    ws = workbook[sheet]
    max_col = ws.max_column
    max_row = ws.max_row
    if max_col < 6 or max_row < 5:
        logger.warning("Vacation sheet has unexpected shape: file=%s sheet=%s rows=%s cols=%s", path, sheet, max_row, max_col)
        return []

    marker_set = {_normalize_key(value) for value in markers}
    horizon_end = horizon_start + timedelta(days=horizon_days)
    logger.info(
        "VACATION INPUT: file=%s sheet=%s markers=%s horizon=[%s..%s]",
        path,
        sheet,
        ",".join(sorted(marker_set)),
        horizon_start.strftime("%Y-%m-%d"),
        horizon_end.strftime("%Y-%m-%d"),
    )

    def _coerce_excel_day(raw: Any) -> date | None:
        if isinstance(raw, datetime):
            return raw.date()
        if isinstance(raw, date):
            return raw
        if isinstance(raw, (int, float)):
            try:
                converted = from_excel(raw)
            except Exception:
                return None
            if isinstance(converted, datetime):
                return converted.date()
            if isinstance(converted, date):
                return converted
            return None
        if isinstance(raw, str):
            text = _normalize_text(raw)
            if not text:
                return None
            for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%d/%m/%Y", "%m/%d/%Y"):
                try:
                    return datetime.strptime(text, fmt).date()
                except ValueError:
                    continue
        return None

    date_by_col: dict[int, date] = {}
    for col in range(6, max_col + 1):
        raw_date = ws.cell(row=3, column=col).value
        parsed_day = _coerce_excel_day(raw_date)
        if parsed_day:
            date_by_col[col] = parsed_day

    vacation_lines: list[str] = []
    marker_hits_total = 0
    marker_hits_in_horizon = 0
    for row in range(5, max_row + 1):
        name = _normalize_text(ws.cell(row=row, column=2).value)
        if not name:
            continue

        marker_dates: list[date] = []
        for col in range(6, max_col + 1):
            marker_raw = ws.cell(row=row, column=col).value
            marker_text = _normalize_key(marker_raw)
            if not marker_text:
                continue
            marker_tokens = [token for token in re.split(r"[,;/\s]+", marker_text) if token]
            if not marker_tokens:
                continue
            if not any(token in marker_set for token in marker_tokens):
                continue
            marker_hits_total += 1
            day = date_by_col.get(col)
            if not day:
                continue
            marker_dates.append(day)
            if horizon_start <= day <= horizon_end:
                marker_hits_in_horizon += 1

        if not marker_dates:
            continue

        marker_dates = sorted(set(marker_dates))
        ranges: list[tuple[date, date]] = []
        range_start = marker_dates[0]
        range_end = marker_dates[0]
        for day in marker_dates[1:]:
            if day == range_end + timedelta(days=1):
                range_end = day
                continue
            ranges.append((range_start, range_end))
            range_start = day
            range_end = day
        ranges.append((range_start, range_end))

        for range_start, range_end in ranges:
            # Include any vacation range intersecting horizon, but keep original full boundaries.
            if range_end < horizon_start or range_start > horizon_end:
                continue
            vacation_lines.append(
                f"{name} vacation {range_start.strftime('%d.%m.%Y')} - {range_end.strftime('%d.%m.%Y')}"
            )

    logger.info(
        "VACATION PARSE RESULT: marker_hits=%s marker_hits_in_horizon=%s entries=%s",
        marker_hits_total,
        marker_hits_in_horizon,
        len(vacation_lines),
    )
    return vacation_lines

def render_outlook_html(payload: dict[str, Any]) -> str:
    meta = payload.get("meta") or {}
    titles = payload.get("titles") or {}
    project = html.escape(_normalize_text(meta.get("project")))
    week_key = html.escape(_normalize_text(meta.get("week_key")))

    def _cfg_html(key: str, fallback: str) -> str:
        value = _normalize_html_text(titles.get(key, fallback))
        return value if value else fallback

    report_title = _cfg_html("main", "Weekly Report")
    highlights_title = _cfg_html("highlights", "Highlights")
    results_title = _cfg_html("results", "Key Results and Achievements")
    plans_title = _cfg_html("plans", "Next Week Plans")
    vacations_title = _cfg_html("vacations", "Vacations (next 60 days)")
    high_priority_title = _cfg_html("high_priority_subtitle", "High priority items")
    bugs_title = _cfg_html("bugs_subtitle", "Bugs summary")
    bugs_summary_template = _normalize_text(
        titles.get(
            "bugs_summary_template",
            (
                "{closed} trouble reports/issues are analyzed and closed, "
                "{in_progress} currently in progress, {open} open in project."
            ),
        )
    )
    header_project_info = _cfg_html("header_project_info", "Weekly execution summary")
    header_banner_bg_color = html.escape(_normalize_text(titles.get("header_banner_bg_color", "rgb(63,78,0)")))
    meta_report_period = _cfg_html("meta_report_period", "Report Period")
    meta_active_iteration = _cfg_html("meta_active_iteration", "Active iteration")
    meta_active_iteration_value = _cfg_html("meta_active_iteration_value", "")
    meta_report_owner = _cfg_html("meta_report_owner", "Report Owner")
    meta_report_owner_value = _cfg_html(
        "meta_report_owner_value",
        html.escape(_normalize_text(meta.get("project"))),
    )
    meta_team_member = _cfg_html("meta_team_member", "Team Member")
    meta_team_member_value = _cfg_html("meta_team_member_value", "")
    footer_html = str(titles.get("footer_html") or "")

    def _header_date(value: Any) -> str:
        text = _normalize_text(value)
        if not text:
            return ""
        for fmt in ("%Y-%m-%d", "%Y/%m/%d"):
            try:
                parsed = datetime.strptime(text, fmt)
                return parsed.strftime("%Y/%m/%d")
            except ValueError:
                continue
        return text.replace("-", "/")

    period_value = html.escape(f"{_header_date(meta.get('week_start'))} - {_header_date(meta.get('week_end'))}")

    rows: list[str] = []
    rows.append("<!doctype html>")
    rows.append("<html lang='en'>")
    rows.append("<head>")
    rows.append("<meta charset='utf-8' />")
    rows.append("<meta name='viewport' content='width=device-width, initial-scale=1' />")
    rows.append(f"<title>{report_title}</title>")
    rows.append("<style>")
    rows.append("html, body { height:100%; }")
    rows.append("body{margin:0;padding:24px;background:#0b0b0b;font-family:Calibri,'Segoe UI',Arial,sans-serif;color:#ffffff;}")
    rows.append(".sheet,.sheet *{color:#ffffff;}")
    rows.append(".sheet{width:1040px;max-width:100%;margin:0 auto;border:2px solid #ffffff;background:#141414;box-shadow:0 14px 40px rgba(0,0,0,.45);}")
    rows.append(".title{text-align:center;font-weight:700;font-size:20px;letter-spacing:.2px;padding:12px 14px;background:#141414;background-image:linear-gradient(#1f1f1f,#141414);border-bottom:2px solid #ffffff;}")
    rows.append("table{border-collapse:collapse;width:100%;}td{vertical-align:top;}")
    rows.append(".subhead td{border-bottom:1px solid #ffffff;font-size:13px;font-weight:700;padding:8px 10px;text-align:center;}")
    rows.append(".meta td{border-bottom:1px solid #ffffff;border-right:1px solid #ffffff;padding:8px 10px;font-size:13px;line-height:1.2;}")
    rows.append(".meta tr td:last-child{border-right:none;}.meta{border-bottom:2px solid #ffffff;}")
    rows.append(".label{background:rgb(63,78,0);background-image:linear-gradient(rgb(76,92,10),rgb(63,78,0));font-weight:700;width:190px;white-space:nowrap;}")
    rows.append(".value{background:#202020;background-image:linear-gradient(#2a2a2a,#202020);}.label.small{width:140px;}")
    rows.append(".blue-panel{padding:14px 14px 18px;background-color:rgb(23,88,98);background-image:radial-gradient(circle at 38% 20%,rgba(255,255,255,.18) 0%,rgba(255,255,255,0) 35%),linear-gradient(135deg,rgb(32,110,123) 0%,rgb(23,88,98) 48%,rgb(18,70,78) 100%);}")
    rows.append(".content td{padding:8px 12px;font-size:12.6px;line-height:1.25;}")
    rows.append(".content .sec-label{width:190px;font-weight:700;font-size:14pt;padding:8px 10px;background:rgba(0,0,0,.2);}")
    rows.append(".divider{height:1px;background:rgba(255,255,255,.35);margin:10px 0;}")
    rows.append(".muted{color:rgba(255,255,255,.78);}")
    rows.append("ul{margin:6px 0;padding:0;}")
    rows.append(".lvl1 li,.lvl2 li,.lvl3 li,.lvl4 li{list-style:none;margin:3px 0;position:relative;padding-left:18px;}")
    rows.append(".lvl1 li:before{content:'\\25A0';position:absolute;left:0;top:0;font-size:10px;line-height:1.2;color:#ffffff;}")
    rows.append(".lvl2{margin-left:18px;}.lvl2 li:before{content:'\\25C6';position:absolute;left:0;top:0;font-size:12px;line-height:1.2;color:#ffffff;}")
    rows.append(".lvl3{margin-left:36px;}.lvl3 li:before{content:'\\2022';position:absolute;left:0;top:0;font-size:12px;line-height:1.2;color:#ffffff;}")
    rows.append(".lvl4{margin-left:20px;}.lvl4 li:before{content:'\\25E6';position:absolute;left:0;top:0;font-size:12px;line-height:1.2;color:#ffffff;}")
    rows.append("@media print{body{background:#ffffff;padding:0;}.sheet{width:100%;box-shadow:none;}}")
    rows.append("</style>")
    rows.append("</head>")
    rows.append("<body>")
    rows.append("<div class='sheet'>")
    rows.append(f"<div class='title'>{report_title} - {project} - {week_key}</div>")
    rows.append(
        f"<table class='subhead' cellspacing='0' cellpadding='0'><tr><td class='sub-banner' style='background:{header_banner_bg_color};'>{header_project_info}</td></tr></table>"
    )
    rows.append("<table class='meta' cellspacing='0' cellpadding='0'>")
    rows.append(
        "<tr>"
        f"<td class='label'>{meta_active_iteration}</td>"
        f"<td class='value'>{meta_active_iteration_value}</td>"
        f"<td class='label small'>{meta_report_owner}</td>"
        f"<td class='value' style='border-right:none;'>{meta_report_owner_value}</td>"
        "</tr>"
    )
    rows.append(
        "<tr>"
        f"<td class='label'>{meta_report_period}</td>"
        f"<td class='value'>{period_value}</td>"
        f"<td class='label small'>{meta_team_member}</td>"
        f"<td class='value' style='border-right:none;'>{meta_team_member_value}</td>"
        "</tr>"
    )
    rows.append("</table>")
    rows.append("<div class='blue-panel'>")
    rows.append("<table class='content' cellspacing='0' cellpadding='0'>")

    _SEP_ROW = "<tr><td colspan='2' style='height:3px;background:rgba(255,255,255,.45);padding:0;'></td></tr>"

    rows.append("<tr>")
    rows.append(f"<td class='sec-label'>{highlights_title}</td><td class='sec-body'><ul class='lvl1'>")
    for item in payload.get("highlights") or []:
        headline = _normalize_text(item.get("headline"))
        comment = _normalize_text(item.get("comment"))
        highlight_text = headline
        if comment:
            highlight_text = f"{headline} - {comment}" if headline else comment
        headline_html = html.escape(highlight_text)
        issue_key = html.escape(_normalize_text(item.get("issue_key")))
        rows.append(f"<li>{headline_html}{f' ({issue_key})' if issue_key else ''}</li>")
    if not (payload.get("highlights") or []):
        rows.append("<li>No highlight updates in this week.</li>")
    rows.append("</ul></td></tr>")

    rows.append(_SEP_ROW)
    rows.append("<tr>")
    rows.append(f"<td class='sec-label'>{results_title}</td><td class='sec-body'>")
    for epic_idx, epic in enumerate(payload.get("epics") or []):
        epic_name = html.escape(_normalize_text(epic.get("epic_name")))
        epic_key = html.escape(_normalize_text(epic.get("epic_key")))
        rows.append("<ul class='lvl1'>")
        rows.append(f"<li><b>{epic_name} ({epic_key})</b></li>")
        rows.append("</ul>")
        feature_items = list(epic.get("feature_statuses") or [])
        if feature_items:
            rows.append("<ul class='lvl2'>")
            for item in feature_items:
                text = html.escape(_normalize_text(item.get("text")))
                issue_key = html.escape(_normalize_text(item.get("issue_key")))
                aggregate_status = _normalize_text(item.get("aggregate_status"))
                regular_status = _normalize_text(item.get("status"))
                status = html.escape(aggregate_status or regular_status)
                subtask_keys_r = list(item.get("subtask_keys_in_report") or [])
                rows.append(f"<li>{text}{f' ({issue_key})' if issue_key else ''}</li>")
                if status:
                    rows.append("</ul><ul class='lvl3'>")
                    if aggregate_status and subtask_keys_r:
                        debug = html.escape(f" ({', '.join(subtask_keys_r)})")
                        rows.append(f"<li>{status}{debug}</li>")
                    else:
                        rows.append(f"<li>{status}</li>")
                    rows.append("</ul><ul class='lvl2'>")
            rows.append("</ul>")
        else:
            rows.append("<p class='muted'>No feature updates in selected period.</p>")
        high_priority_items = list(epic.get("high_priority_items") or [])
        if high_priority_items:
            rows.append("<ul class='lvl2'>")
            rows.append(f"<li><b>{high_priority_title}</b></li>")
            rows.append("</ul>")
            seen_hp: set[str] = set()
            rows.append("<ul class='lvl3'>")
            for item in high_priority_items:
                item_key_norm = _normalize_key(item.get("issue_key"))
                if item_key_norm and item_key_norm in seen_hp:
                    continue
                if item_key_norm:
                    seen_hp.add(item_key_norm)
                text = html.escape(_normalize_text(item.get("text")))
                issue_key_hp = html.escape(_normalize_text(item.get("issue_key")))
                status_hp = html.escape(_normalize_text(item.get("status")))
                comment_hp = html.escape(_normalize_text(item.get("comment")))
                status_suffix = f" — {status_hp}" if status_hp else ""
                rows.append(f"<li>{text}{f' ({issue_key_hp})' if issue_key_hp else ''}{status_suffix}</li>")
                if comment_hp:
                    rows.append("<ul class='lvl4'>")
                    rows.append(f"<li>{comment_hp}</li>")
                    rows.append("</ul>")
            rows.append("</ul>")
        if epic_idx < len(payload.get("epics") or []) - 1:
            rows.append("<div class='divider'></div>")
    if not (payload.get("epics") or []):
        rows.append("<p class='muted'>No epic feature updates for selected scope.</p>")
    total_closed_bugs = sum(
        int((epic.get("bugs") or {}).get("closed", 0))
        for epic in (payload.get("epics") or [])
    )
    project_bugs = payload.get("project_bugs") or {}
    in_progress_bugs = int(project_bugs.get("in_progress", 0) or 0)
    open_bugs = int(project_bugs.get("open", 0) or 0)
    if total_closed_bugs or in_progress_bugs or open_bugs:
        try:
            bugs_summary_line = bugs_summary_template.format(
                closed=total_closed_bugs,
                in_progress=in_progress_bugs,
                open=open_bugs,
            )
        except Exception:
            bugs_summary_line = (
                f"{total_closed_bugs} trouble reports/issues are analyzed and closed, "
                f"{in_progress_bugs} currently in progress, {open_bugs} open in project."
            )
        rows.append(
            f"<ul class='lvl1'><li><b>{bugs_title}</b>: "
            f"{html.escape(_normalize_text(bugs_summary_line))}</li></ul>"
        )
    rows.append("</td></tr>")

    rows.append(_SEP_ROW)
    rows.append("<tr>")
    rows.append(f"<td class='sec-label'>{plans_title}</td><td class='sec-body'>")
    for epic in payload.get("next_week_plans") or []:
        epic_name = html.escape(_normalize_text(epic.get("epic_name")))
        epic_key = html.escape(_normalize_text(epic.get("epic_key")))
        rows.append("<ul class='lvl1'>")
        rows.append(f"<li><b>{epic_name} ({epic_key})</b></li>")
        rows.append("</ul>")
        rows.append("<ul class='lvl2'>")
        for item in epic.get("items") or []:
            text = html.escape(_normalize_text(item.get("text")))
            issue_key = html.escape(_normalize_text(item.get("issue_key")))
            aggregate_status = _normalize_text(item.get("aggregate_status"))
            regular_status = _normalize_text(item.get("status"))
            status = html.escape(aggregate_status or regular_status)
            subtask_keys_p = list(item.get("subtask_keys_in_report") or [])
            rows.append(f"<li>{text}{f' ({issue_key})' if issue_key else ''}</li>")
            if status:
                rows.append("</ul><ul class='lvl3'>")
                if aggregate_status and subtask_keys_p:
                    debug = html.escape(f" ({', '.join(subtask_keys_p)})")
                    rows.append(f"<li>{status}{debug}</li>")
                else:
                    rows.append(f"<li>{status}</li>")
                rows.append("</ul><ul class='lvl2'>")
        rows.append("</ul>")
    if not (payload.get("next_week_plans") or []):
        rows.append("<p class='muted'>No in-progress plans collected for next week.</p>")
    rows.append("</td></tr>")

    rows.append(_SEP_ROW)
    rows.append("<tr>")
    rows.append(f"<td class='sec-label'>{vacations_title}</td><td class='sec-body'><ul class='lvl1'>")
    for item in payload.get("vacations") or []:
        rows.append(f"<li>{html.escape(_normalize_text(item))}</li>")
    if not (payload.get("vacations") or []):
        rows.append("<li>No vacations found for the configured horizon.</li>")
    rows.append("</ul></td></tr>")

    # footer_html as a content-table row (same blue-panel background)
    if footer_html.strip():
        rows.append(
            "<tr><td colspan='2'"
            " style='padding:10px 14px;background-color:rgb(23,88,98);'>"
            f"{footer_html}"
            "</td></tr>"
        )

    rows.append("</table>")
    rows.append("</div>")  # close blue-panel

    # Separator between blue-panel and risks block
    rows.append(
        "<div class='risk-separator'"
        " style='background-color:#9C5600;height:24px;font-size:1px;line-height:1px;'>"
        "</div>"
    )

    risk_title = titles.get("risk_title", "Top Issues / Risks / For Help")
    risk_items = payload.get("risk_items") or []
    rows.append("<div class='risk-section' style='background-color:#706721;padding:10px 14px 14px;'>")
    rows.append(
        f"<table style='width:100%;border-collapse:collapse;font-size:11px;color:#ffffff;'>"
        f"<thead>"
        f"<tr><th colspan='6' style='padding:6px 8px 10px;font-size:14pt;font-weight:700;"
        f"text-align:left;background:none;border-bottom:1px solid rgba(255,255,255,.3);'>"
        f"{html.escape(risk_title)}</th></tr>"
        f"<tr style='background:rgba(0,0,0,.3);font-weight:700;'>"
        f"<th style='padding:4px 8px;text-align:left;'>Risk/Issue</th>"
        f"<th style='padding:4px 8px;text-align:left;'>Assignee</th>"
        f"<th style='padding:4px 8px;text-align:left;'>Created</th>"
        f"<th style='padding:4px 8px;text-align:left;'>Status</th>"
        f"<th style='padding:4px 8px;text-align:left;'>Action points / Comments</th>"
        f"<th style='padding:4px 8px;text-align:left;'>Created by</th>"
        f"</tr></thead><tbody>"
    )
    if risk_items:
        for i, item in enumerate(risk_items):
            bg = "rgba(0,0,0,.12)" if i % 2 == 0 else "rgba(0,0,0,.04)"
            key_esc  = html.escape(item.get("issue_key") or "")
            text_esc = html.escape(item.get("text") or "")
            rows.append(
                f"<tr style='background:{bg};'>"
                f"<td style='padding:4px 8px;white-space:nowrap;'>{key_esc}</td>"
                f"<td style='padding:4px 8px;'>{html.escape(item.get('assignee') or '')}</td>"
                f"<td style='padding:4px 8px;white-space:nowrap;'>{html.escape(item.get('created') or '')}</td>"
                f"<td style='padding:4px 8px;'>{html.escape(item.get('status') or '')}</td>"
                f"<td style='padding:4px 8px;'>{html.escape(item.get('action_points') or '')} "
                f"<span class='muted'>{text_esc}</span></td>"
                f"<td style='padding:4px 8px;'>{html.escape(item.get('reporter') or '')}</td>"
                f"</tr>"
            )
    else:
        rows.append(
            "<tr><td colspan='6' style='padding:8px;color:rgba(255,255,255,.7);'>"
            "No open risks or issues.</td></tr>"
        )
    rows.append("</tbody></table>")
    rows.append("</div>")  # close risks section div
    rows.append("</div></body></html>")  # close sheet only
    return "\n".join(rows)


def save_snapshot(path: Path, payload: dict[str, Any], week: WeekWindow) -> None:
    snapshot = {
        "meta": {
            "project": payload.get("meta", {}).get("project", ""),
            "week_key": week.key,
            "year": week.year,
            "week": week.week,
            "generated_at": datetime.now().strftime("%Y-%m-%dT%H:%M:%S"),
        },
        "order": _extract_order(payload),
        "payload": payload,
    }
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(snapshot, ensure_ascii=False, indent=2), encoding="utf-8")


# ---------------------------------------------------------------------------
# Output format dependency checks and format-specific renderers
# ---------------------------------------------------------------------------

_SOFFICE_WINDOWS_PATHS = [
    r"C:\Program Files\LibreOffice\program\soffice.exe",
    r"C:\Program Files (x86)\LibreOffice\program\soffice.exe",
]


def _find_soffice() -> str | None:
    """Return the full path to the soffice binary, or None if not found."""
    found = shutil.which("soffice")
    if found:
        return found
    for candidate in _SOFFICE_WINDOWS_PATHS:
        if Path(candidate).is_file():
            return candidate
    return None


def _check_soffice() -> bool:
    """Return True if the soffice binary is findable."""
    return _find_soffice() is not None


def _check_outlook_available() -> bool:
    """Return True when Classic Outlook is installed and win32com is available.

    Uses two lightweight checks without launching Outlook:
    1. Windows registry — is the ``Outlook.Application`` COM class registered?
    2. Python import  — is ``win32com`` (pywin32) installed?
    """
    if platform.system() != "Windows":
        return False
    # Registry check: Classic Outlook registers its COM class here on install.
    try:
        import winreg
        winreg.OpenKey(winreg.HKEY_CLASSES_ROOT, "Outlook.Application")
    except ImportError:
        print("[jira_weekly_email] outlook_draft: winreg not available (not Windows?)")
        return False
    except (FileNotFoundError, OSError) as exc:
        print(
            "[jira_weekly_email] outlook_draft: Outlook.Application COM class not found in registry.\n"
            "  This means Classic Outlook (Office 2016/2019/2021/2024 LTSC) is not installed.\n"
            f"  Registry error: {exc}"
        )
        return False
    # Dependency check: win32com must be available to create the draft.
    try:
        import win32com.client  # noqa: F401
    except ImportError:
        print(
            "[jira_weekly_email] outlook_draft: pywin32 (win32com) is not installed in this Python environment.\n"
            f"  Python: {__import__('sys').executable}\n"
            "  Fix: pip install pywin32"
        )
        return False
    return True


def _prompt_user(message: str) -> bool:
    """Print *message* and ask the user to confirm (y/N).  Returns True for 'y'."""
    try:
        answer = input(f"{message} [y/N] ").strip().lower()
        return answer == "y"
    except (EOFError, KeyboardInterrupt):
        return False


def _resolve_output_format_deps(output_formats: list[str]) -> list[str]:
    """Validate runtime dependencies for requested output formats.

    * ``docx``  — requires LibreOffice ``soffice`` binary.
    * ``outlook_draft`` — requires Windows + Microsoft Outlook (COM).

    Interactive prompts guide the user when a dependency is missing.
    If the user refuses to continue, the process exits immediately.
    Adjusted format list is returned (e.g. ``outlook_draft`` replaced by ``eml``).
    """
    # Normalize: split comma-separated values so both "html docx" and "html,docx" work.
    formats: list[str] = [
        part.strip()
        for token in output_formats
        for part in token.split(",")
        if part.strip()
    ]

    if "docx" in formats:
        if not _check_soffice():
            print(
                "\n[jira_weekly_email] 'docx' output requires LibreOffice (soffice).\n"
                f"  Download and install from: {_SOFFICE_DOWNLOAD_URL}\n"
                "  After installation, restart the script to enable DOCX export.\n"
            )
            if _prompt_user("Continue without DOCX conversion?"):
                formats.remove("docx")
                logger.info("DOCX conversion skipped: soffice not found.")
            else:
                logger.error("Aborted by user: soffice not available.")
                sys.exit(1)

    if "outlook_draft" in formats:
        if platform.system() != "Windows":
            print(
                "\n[jira_weekly_email] 'outlook_draft' requires Windows + Microsoft Outlook.\n"
                f"  More info: {_OUTLOOK_INFO_URL}\n"
                "  An EML file can be opened in any email client instead.\n"
            )
            if _prompt_user("Generate EML file instead?"):
                formats[formats.index("outlook_draft")] = "eml"
                logger.info("Falling back to EML: not running on Windows.")
            else:
                logger.error("Aborted by user: Outlook not available on this platform.")
                sys.exit(1)
        elif not _check_outlook_available():
            print(
                "\n[jira_weekly_email] 'outlook_draft' requires Microsoft Outlook to be installed.\n"
                f"  Download from: {_OUTLOOK_INFO_URL}\n"
                "  An EML file can be opened in Outlook once installed.\n"
            )
            if _prompt_user("Generate EML file instead?"):
                formats[formats.index("outlook_draft")] = "eml"
                logger.info("Falling back to EML: Outlook COM not accessible.")
            else:
                logger.error("Aborted by user: Outlook COM not accessible.")
                sys.exit(1)

    return formats


# CSS injected into the HTML before DOCX conversion to produce a light,
# print-friendly document instead of the dark-themed email layout.
_DOCX_STYLE_OVERRIDE = """
<style id='docx-override'>
  @page { size: A4 landscape; margin: 1.5cm; }
  body {
    background: #ffffff !important;
    color: #000000 !important;
    padding: 0 !important;
    width: 100% !important;
    max-width: 100% !important;
  }
  /* Reset all dark backgrounds and white text to black-on-white */
  body *, .sheet, .sheet * {
    color: #000000 !important;
    background-color: transparent !important;
    background-image: none !important;
    box-shadow: none !important;
  }
  /* Restore light structural backgrounds */
  .sheet {
    width: 100% !important;
    max-width: 100% !important;
    border: 1px solid #cccccc !important;
  }
  .label, .sec-label { background-color: #e8e8e8 !important; }
  .value            { background-color: #f5f5f5 !important; }
  .blue-panel       { background-color: #ddeef2 !important; }
  .divider          { background-color: #cccccc !important; }
  .muted            { color: #555555 !important; }
  .title            { border-bottom: 1px solid #cccccc !important; }
  .risk-separator   { background-color: #9C5600 !important; height: 24px !important; }
  .risk-section     { background-color: #706721 !important; }
  table { width: 100% !important; }
</style>
"""


def _prepare_html_for_docx(html_text: str) -> str:
    """Return a light-themed variant of *html_text* suitable for LibreOffice DOCX conversion.

    LibreOffice's HTML importer does not implement CSS cascade like a browser.
    ``!important`` overrides in an injected stylesheet are ignored for page-layout
    decisions.  We therefore patch the HTML source directly:
      - replace the fixed ``.sheet`` pixel width with 100 %
      - strip ``margin:0 auto`` centering (irrelevant / harmful in DOCX)
      - remove the ``class='sheet'`` attribute so LibreOffice does not create a
        fixed-width text frame for that div (the main cause of the narrow column)
      - inject the light-theme CSS block (colours)
    Page size / orientation is handled separately via python-docx post-processing.
    """
    # 1. Remove the fixed pixel width from the CSS rule.
    html_text = re.sub(r'\bwidth\s*:\s*1040px\b', 'width:100%', html_text)
    # 2. Remove centering that is meaningless in DOCX layout.
    html_text = re.sub(r';\s*margin\s*:\s*0\s+auto\b', '', html_text)
    # 3. Strip class='sheet' from the outer container div so LibreOffice does not
    #    treat it as a named, width-constrained text frame.
    html_text = re.sub(r"<div\s+class=['\"]sheet['\"]>", "<div>", html_text)
    # 4. Patch <div class='blue-panel'>: light-blue inline background for DOCX.
    html_text = re.sub(
        r"<div\s+class=['\"]blue-panel['\"](\s[^>]*)?>",
        "<div class='blue-panel'\\1 style='background-color:#ddeef2;padding:14px;'>",
        html_text,
    )
    # 5. Convert .divider divs to <hr> — LibreOffice renders <hr> natively as a line.
    html_text = re.sub(
        r"<div\s+class=['\"]divider['\"][^>]*>\s*</div>",
        "<hr style='border:none;border-top:1px solid #aaaaaa;margin:6px 0;'>",
        html_text,
    )
    # 6. Add bgcolor to the content table so LibreOffice shows the blue-panel background.
    html_text = html_text.replace(
        "<table class='content' cellspacing='0' cellpadding='0'>",
        "<table class='content' cellspacing='0' cellpadding='0'"
        " bgcolor='#ddeef2' style='background-color:#ddeef2;'>",
        1,
    )
    # 7. Patch risk-separator and risk-section with explicit inline backgrounds for LibreOffice.
    #    The _DOCX_STYLE_OVERRIDE has `body * { background-color: transparent !important }`
    #    which can suppress these divs — direct inline styles are more reliable.
    html_text = re.sub(
        r"<div\s+class='risk-separator'[^>]*>",
        "<div class='risk-separator'"
        " style='background-color:#9C5600;height:24px;font-size:1px;line-height:1px;'>",
        html_text,
    )
    html_text = re.sub(
        r"<div\s+class='risk-section'[^>]*>",
        "<div class='risk-section'"
        " style='background-color:#706721;padding:10px 14px 14px;'>",
        html_text,
    )
    # 8. Inject light-theme CSS block right before </head>.
    if "</head>" in html_text:
        return html_text.replace("</head>", f"{_DOCX_STYLE_OVERRIDE}</head>", 1)
    return _DOCX_STYLE_OVERRIDE + html_text


def _postprocess_docx_page(docx_path: Path) -> None:
    """Set every section to A4 landscape and stretch all tables to the full text width.

    LibreOffice computes table widths at its own internal DPI/page assumption during
    HTML import, then freezes those widths.  python-docx re-writing the page dimensions
    afterwards does NOT reflow the table columns — they stay at the narrow widths
    LibreOffice computed.  We therefore also patch every table's total width and
    redistribute its column widths proportionally to fill the new text area.
    """
    try:
        from docx import Document  # type: ignore[import]
        from docx.shared import Mm  # type: ignore[import]
        from docx.enum.section import WD_ORIENT  # type: ignore[import]
        from docx.oxml.ns import qn  # type: ignore[import]
        from docx.oxml import OxmlElement  # type: ignore[import]

        doc = Document(str(docx_path))

        # --- 1. Fix page size -------------------------------------------------
        margin_mm = 15
        for section in doc.sections:
            section.orientation = WD_ORIENT.LANDSCAPE
            section.page_width = Mm(297)   # A4 landscape
            section.page_height = Mm(210)
            section.left_margin = Mm(margin_mm)
            section.right_margin = Mm(margin_mm)
            section.top_margin = Mm(margin_mm)
            section.bottom_margin = Mm(margin_mm)

        # Text width in twips (dxa): A4 landscape minus margins on both sides.
        # 1 mm = 1440/25.4 twips
        page_w_dxa = round(297 / 25.4 * 1440)          # 16838 dxa
        margin_dxa = round(margin_mm / 25.4 * 1440)    # 851 dxa each side
        text_w_dxa = page_w_dxa - 2 * margin_dxa       # usable text width

        # --- 2. Stretch every table to the full text width --------------------
        def _set_tbl_width(tbl_elem, width_dxa: int) -> None:
            tblPr = tbl_elem.find(qn("w:tblPr"))
            if tblPr is None:
                tblPr = OxmlElement("w:tblPr")
                tbl_elem.insert(0, tblPr)
            tblW = tblPr.find(qn("w:tblW"))
            if tblW is None:
                tblW = OxmlElement("w:tblW")
                tblPr.append(tblW)
            tblW.set(qn("w:w"), str(width_dxa))
            tblW.set(qn("w:type"), "dxa")

        def _stretch_grid(tbl_elem, text_w_dxa: int) -> None:
            """Scale the tblGrid column definitions proportionally to fill *text_w_dxa*."""
            tblGrid = tbl_elem.find(qn("w:tblGrid"))
            if tblGrid is None:
                return
            cols = tblGrid.findall(qn("w:gridCol"))
            if not cols:
                return
            old_total = sum(int(c.get(qn("w:w"), 0)) for c in cols)
            if old_total == 0:
                # Distribute evenly
                per_col = text_w_dxa // len(cols)
                for c in cols:
                    c.set(qn("w:w"), str(per_col))
                return
            scale = text_w_dxa / old_total
            # Scale each column and track the remainder to avoid rounding drift
            new_total = 0
            for i, c in enumerate(cols):
                new_w = round(int(c.get(qn("w:w"), 0)) * scale)
                c.set(qn("w:w"), str(new_w))
                new_total += new_w
            # Give any leftover twips to the last column
            remainder = text_w_dxa - new_total
            if remainder and cols:
                last = cols[-1]
                last.set(qn("w:w"), str(int(last.get(qn("w:w"), 0)) + remainder))

        def _stretch_cells(tbl_elem, text_w_dxa: int) -> None:
            """Scale every cell's tcW so rows add up to *text_w_dxa*."""
            for tr in tbl_elem.findall(qn("w:tr")):
                cells = tr.findall(qn("w:tc"))
                old_row_w = 0
                cell_widths = []
                for tc in cells:
                    tcPr = tc.find(qn("w:tcPr"))
                    tcW = tcPr.find(qn("w:tcW")) if tcPr is not None else None
                    w = int(tcW.get(qn("w:w"), 0)) if tcW is not None else 0
                    cell_widths.append((tc, tcPr, tcW, w))
                    old_row_w += w
                if old_row_w == 0:
                    continue
                scale = text_w_dxa / old_row_w
                new_total = 0
                for i, (tc, tcPr, tcW, old_w) in enumerate(cell_widths):
                    new_w = round(old_w * scale)
                    new_total += new_w
                    if tcW is not None:
                        tcW.set(qn("w:w"), str(new_w))
                        tcW.set(qn("w:type"), "dxa")
                # Fix rounding drift on the last cell
                if cell_widths:
                    tc, tcPr, tcW, _ = cell_widths[-1]
                    if tcW is not None:
                        cur = int(tcW.get(qn("w:w"), 0))
                        tcW.set(qn("w:w"), str(cur + (text_w_dxa - new_total)))

        for table in doc.tables:
            tbl = table._tbl
            _set_tbl_width(tbl, text_w_dxa)
            _stretch_grid(tbl, text_w_dxa)
            _stretch_cells(tbl, text_w_dxa)

        doc.save(str(docx_path))
    except Exception as exc:
        logger.warning("Could not post-process DOCX page settings: %s", exc)


def _convert_html_to_docx(html_path: Path, output_dir: Path) -> Path | None:
    """Convert *html_path* to DOCX via LibreOffice headless.

    A light-themed version of the HTML is written to a temp file first so the
    dark email stylesheet does not carry over into the document.
    Returns the resulting ``.docx`` path on success, or ``None`` on failure.
    """
    soffice_bin = _find_soffice()
    if not soffice_bin:
        logger.error("soffice binary not found; cannot convert to DOCX.")
        return None

    # Write a light-themed temp copy of the HTML so the dark email CSS
    # does not produce invisible white-on-white text in the DOCX.
    original_html = html_path.read_text(encoding="utf-8")
    prepared_html = _prepare_html_for_docx(original_html)
    tmp_file = tempfile.NamedTemporaryFile(
        suffix=".html", dir=output_dir, delete=False, mode="w", encoding="utf-8"
    )
    tmp_file.write(prepared_html)
    tmp_file.close()
    tmp_path = Path(tmp_file.name)
    try:
        result = subprocess.run(
            [
                soffice_bin,
                "--headless",
                "--norestore",          # skip crash-recovery dialog
                "--nocrashreport",      # suppress crash reporter
                "--infilter=HTML (StarWriter)",          # explicit HTML import filter
                "--convert-to", "docx:MS Word 2007 XML",  # explicit export filter
                "--outdir", str(output_dir),
                str(tmp_path),
            ],
            capture_output=True,
            text=True,
            timeout=120,
        )
        # LibreOffice writes informational lines to stderr even on success;
        # treat rc != 0 as failure only when output mentions a real error.
        stderr_lc = result.stderr.strip().lower()
        has_error = result.returncode != 0 and (
            "error" in stderr_lc or "failed" in stderr_lc or not result.stderr.strip()
        )
        if has_error:
            logger.error(
                "soffice conversion failed (rc=%s): %s",
                result.returncode,
                result.stderr.strip() or result.stdout.strip(),
            )
            return None
        # soffice names the output after the *input* stem (i.e. the temp file stem).
        tmp_docx = output_dir / (tmp_path.stem + ".docx")
        desired_docx = output_dir / (html_path.stem + ".docx")
        if tmp_docx.exists():
            # Path.replace() overwrites the destination on Windows (unlike Path.rename()).
            tmp_docx.replace(desired_docx)
        elif not desired_docx.exists():
            logger.error("soffice ran successfully but DOCX output not found at %s", tmp_docx)
            return None
        _postprocess_docx_page(desired_docx)
        return desired_docx
    except subprocess.TimeoutExpired:
        logger.error("soffice conversion timed out after 120 s.")
        return None
    except Exception as exc:
        logger.error("soffice conversion error: %s", exc)
        return None
    finally:
        # Always remove the temporary light-themed HTML file.
        try:
            tmp_path.unlink(missing_ok=True)
        except Exception:
            pass


def _prepare_html_for_eml(html_text: str) -> str:
    """Patch HTML for email-client rendering.

    Email clients (Thunderbird, Outlook, Windows Mail) strip ``<style>`` tags
    and ignore CSS background-color on div elements and CSS width on divs.
    We therefore use HTML attributes and table-based layout that all clients respect:

    * ``<body>``             → bgcolor + no padding (outer table handles spacing)
    * Outer centering table  → width=100%, bgcolor=#0b0b0b
    * Sheet table (1040px)   → replaces the div, provides width constraint + border + bgcolor
    * ``<div class='blue-panel'>`` → inline background-color
    * ``<table class='content'>``  → bgcolor so cells show blue, not white
    """
    # 1. Patch <body>: dark bgcolor, no padding (outer table handles spacing).
    html_text = re.sub(
        r"<body(\s[^>]*)?>",
        (
            "<body\\1"
            " bgcolor='#0b0b0b'"
            " style='background:#0b0b0b;margin:0;padding:0;"
            "font-family:Calibri,\"Segoe UI\",Arial,sans-serif;color:#ffffff;'>"
            # Outer 100%-wide centering table with dark background
            "<table role='presentation' border='0' cellspacing='0' cellpadding='0'"
            " width='100%' bgcolor='#0b0b0b'"
            " style='background:#0b0b0b;'><tr>"
            "<td align='center' valign='top'"
            " style='padding:24px 12px;background:#0b0b0b;' bgcolor='#0b0b0b'>"
        ),
        html_text,
        count=1,
    )
    # 2. Close the outer centering table before </body>.
    html_text = html_text.replace("</body>", "</td></tr></table></body>", 1)
    # 3. Replace <div class='sheet'> with a width-1040 table (reliable in all email clients).
    html_text = re.sub(
        r"<div\s+class=['\"]sheet['\"](\s[^>]*)?>",
        (
            "<table role='presentation' border='0' cellspacing='0' cellpadding='0'"
            " width='1040' align='center' bgcolor='#141414'"
            " style='max-width:1040px;width:1040px;background:#141414;"
            "border:2px solid #ffffff;'>"
            "<tr><td bgcolor='#141414' style='background:#141414;'>"
        ),
        html_text,
        count=1,
    )
    # 4. Replace the sheet div's closing tag with a table close.
    #    After restructuring, the HTML ends with: ...risks-div... </div></body></html>
    #    where the final </div> is ONLY the sheet closing div (blue-panel and risks-div
    #    both close before it).
    html_text = html_text.replace(
        "</div></body></html>",
        "</td></tr></table></body></html>",
        1,
    )
    # 5. Patch <div class='blue-panel'>: inline background for email clients.
    html_text = re.sub(
        r"<div\s+class=['\"]blue-panel['\"](\s[^>]*)?>",
        (
            "<div class='blue-panel'\\1"
            " style='background-color:rgb(23,88,98);padding:14px 14px 18px;color:#ffffff;'>"
        ),
        html_text,
        count=1,
    )
    # 6. Add bgcolor to the content table so email clients show the blue background
    #    (table cells don't inherit background from parent div in email clients).
    html_text = html_text.replace(
        "<table class='content' cellspacing='0' cellpadding='0'>",
        "<table class='content' cellspacing='0' cellpadding='0'"
        " bgcolor='#175862' style='background-color:rgb(23,88,98);'>",
        1,
    )
    # 7. Convert risk-separator div to a table row — Outlook's Word engine ignores
    #    background-color on div elements entirely.
    html_text = re.sub(
        r"<div\s+class='risk-separator'[^>]*>\s*</div>",
        (
            "<table role='presentation' width='100%' border='0' cellspacing='0' cellpadding='0'"
            " bgcolor='#9C5600' style='background-color:#9C5600;'>"
            "<tr><td height='24' bgcolor='#9C5600'"
            " style='background-color:#9C5600;height:24px;font-size:1px;line-height:1px;'>"
            "&nbsp;</td></tr></table>"
        ),
        html_text,
        count=1,
    )
    # 8. Convert risk-section div opening to a bgcolor table for Outlook.
    html_text = re.sub(
        r"<div\s+class='risk-section'[^>]*>",
        (
            "<table role='presentation' width='100%' border='0' cellspacing='0' cellpadding='0'"
            " bgcolor='#706721' style='background-color:#706721;'>"
            "<tr><td bgcolor='#706721'"
            " style='background-color:#706721;padding:10px 14px 14px;'>"
        ),
        html_text,
        count=1,
    )
    # 9. Convert risk-section closing </div> to table close.
    #    After step 4, the sheet is already a table, so the HTML ends:
    #      ...risk-content...</div>\n</td></tr></table></body></html>
    #    The leading </div> is now the risk-section cell close (step 8 made it a <td>).
    html_text = re.sub(
        r"</div>\s*</td></tr></table></body></html>",
        "</td></tr></table>\n</td></tr></table></body></html>",
        html_text,
        count=1,
    )
    return html_text


def _save_as_eml(html_text: str, subject: str, output_path: Path) -> None:
    """Write *html_text* as a standards-compliant ``.eml`` file."""
    prepared = _prepare_html_for_eml(html_text)
    msg = MIMEMultipart("alternative")
    msg["Subject"] = subject
    msg["From"] = "noreply@weekly-report"
    msg.attach(MIMEText(prepared, "html", "utf-8"))
    output_path.write_bytes(msg.as_bytes())
    logger.info("EML saved: %s", output_path)


def _open_outlook_draft(html_text: str, subject: str) -> None:
    """Open a new Outlook MailItem window pre-filled with *html_text* (Windows only)."""
    import pythoncom  # part of pywin32
    import win32com.client
    # CoInitialize is required when COM is called from a non-main thread
    # (e.g. inside tqdm context managers or progress wrappers).
    pythoncom.CoInitialize()
    try:
        outlook = win32com.client.Dispatch("Outlook.Application")
        mail = outlook.CreateItem(0)  # 0 = olMailItem
        mail.Subject = subject
        mail.HTMLBody = html_text
        mail.Display()  # opens the composer window; user reviews and sends manually
        logger.info("Outlook draft opened: subject=%r", subject)
    finally:
        pythoncom.CoUninitialize()


@registry.register
class JiraWeeklyEmailReport:
    name = "jira_weekly_email"

    def run(
        self,
        dataset: dict,
        config: ConfigParser,
        output_formats: list[str],
        extra_params: dict | None = None,
    ) -> None:
        extra_params = extra_params or {}
        # Also accept output_formats from --params (e.g. output_formats=html,docx).
        # When passed via --params it arrives as a string in extra_params instead of the
        # output_formats list argument — merge both sources before dependency checks.
        params_formats_raw = str(extra_params.get("output_formats") or "")
        if params_formats_raw:
            output_formats = list(output_formats) + [params_formats_raw]
        # Check output-format dependencies first — before progress bar, config parsing, or Jira I/O.
        output_formats = _resolve_output_format_deps(output_formats)
        progress = _get_progress(extra_params, total_steps=4)

        if not config.has_section("jira"):
            logger.error(
                "jira_weekly_email: missing [jira] section in config. "
                "Add jira-url/username/password to configs/local/config.ini or pass --config <path>."
            )
            return
        jira_section = config["jira"]
        missing_jira_keys = [
            key for key in ("jira-url", "username", "password") if not _normalize_text(jira_section.get(key))
        ]
        if missing_jira_keys:
            logger.error(
                "jira_weekly_email: missing required [jira] keys: %s. "
                "Fill them in configs/local/config.ini or pass --config <path>.",
                ", ".join(missing_jira_keys),
            )
            return

        section = config["jira_weekly_email"] if config.has_section("jira_weekly_email") else {}
        project = _normalize_text(
            extra_params.get("project")
            or section.get("project")
            or config.get("jira", "project", fallback="")
        )
        if not project:
            logger.error(
                "jira_weekly_email: project key is required. "
                "Pass --params project=ABC or set [jira_weekly_email].project in config."
            )
            return

        try:
            week = resolve_week_window(extra_params)
        except Exception as exc:
            logger.error(
                "jira_weekly_email: failed to parse week parameters (%s). "
                "Use one of: week_date=YYYY-MM-DD, week=WW (optional year=YYYY), or start/end as ISO dates.",
                exc,
            )
            return

        labels_highlights = _parse_label_set(
            extra_params.get("labels_highlights") or section.get("labels_highlights"),
            ["highlights"],
        )
        labels_report = _parse_label_set(
            extra_params.get("labels_report") or section.get("labels_report"),
            ["report"],
        )
        priority_high_values = _parse_label_set(
            extra_params.get("priority_high_values") or section.get("priority_high_values"),
            ["High"],
        )
        priority_always_show_values = _parse_label_set(
            extra_params.get("priority_always_show_values") or section.get("priority_always_show_values"),
            ["Highest"],
        )
        ai_enabled_value = extra_params.get("ai_enabled")
        if ai_enabled_value is None:
            ai_enabled_value = extra_params.get("enable_ai")
        if ai_enabled_value is None:
            ai_enabled_value = section.get("ai_enabled")
        if ai_enabled_value is None:
            ai_enabled_value = config.get("jira_weekly_email", "ai_enabled", fallback="false")
        ai_enabled = _bool_value(ai_enabled_value, False)
        logger.info(
            "REPORT PARAMS: project=%s week=%s range=[%s..%s] labels_highlights=%s labels_report=%s priority_high_values=%s priority_always_show_values=%s ai_enabled=%s",
            project,
            week.key,
            week.start.strftime("%Y-%m-%d"),
            week.end.strftime("%Y-%m-%d"),
            ",".join(sorted(labels_highlights)),
            ",".join(sorted(labels_report)),
            ",".join(sorted(priority_high_values)),
            ",".join(sorted(priority_always_show_values)),
            ai_enabled,
        )

        try:
            jira_source = JiraSource(jira_section)
        except Exception as exc:
            logger.error(
                "jira_weekly_email: failed to initialize Jira client (%s). "
                "Check [jira] jira-url/username/password and network access.",
                exc,
            )
            return

        with progress.step("Fetch Jira data"):
            try:
                evidence = collect_weekly_comment_evidence(jira_source, project, week)
                project_bug_stats = collect_project_bug_stats(jira_source, project)
                hp_always_evidence = collect_priority_always_evidence(
                    jira_source, project, week, priority_high_values
                )
                always_show_evidence = collect_priority_always_evidence(
                    jira_source, project, week, priority_always_show_values
                )
                risk_evidence = collect_risk_evidence(jira_source, project, week)
            except Exception as exc:
                logger.error(
                    "jira_weekly_email: failed to fetch Jira data (%s). "
                    "Verify project key, Jira connectivity, and credentials.",
                    exc,
                )
                return

        with progress.step("Build payload"):
            payload = build_report_payload(
                evidence,
                week,
                config,
                project,
                labels_highlights=labels_highlights,
                labels_report=labels_report,
                priority_high_values=priority_high_values,
                priority_always_show_values=priority_always_show_values,
                hp_always_evidence=hp_always_evidence,
                always_show_evidence=always_show_evidence,
                project_bug_stats=project_bug_stats,
                risk_evidence=risk_evidence,
            )

        vacation_file = _strip_wrapping_quotes(
            _normalize_text(extra_params.get("vacation_file") or section.get("vacation_file"))
        )
        if vacation_file:
            vacation_sheet = _normalize_text(extra_params.get("vacation_sheet") or section.get("vacation_sheet")) or "Vacations2026"
            vacation_markers = {
                item.strip()
                for item in _split_csv(
                    extra_params.get("vacation_marker_values") or section.get("vacation_marker_values"),
                    ["p", "P"],
                )
            }
            vacation_days_value = (
                _normalize_text(extra_params.get("vacation_horizon_days") or section.get("vacation_horizon_days")) or "60"
            )
            vacation_horizon_days = _parse_positive_int_with_fallback(
                vacation_days_value,
                60,
                name="vacation_horizon_days",
            )
            vacation_horizon_anchor = _normalize_key(
                extra_params.get("vacation_horizon_anchor") or section.get("vacation_horizon_anchor") or "today"
            )
            if vacation_horizon_anchor in {"week", "week_start", "report_week_start"}:
                vacation_horizon_start = week.start
            elif vacation_horizon_anchor in {"today", ""}:
                vacation_horizon_start = date.today()
            else:
                logger.error(
                    "Invalid vacation_horizon_anchor=%r. Allowed values: today, week_start. Using today.",
                    vacation_horizon_anchor,
                )
                vacation_horizon_start = date.today()
            alternate_horizon_start = week.start if vacation_horizon_start != week.start else date.today()
            vacation_path = Path(vacation_file)
            if not vacation_path.is_absolute():
                parent_candidate = (Path.cwd().parent / vacation_path).resolve()
                cwd_candidate = (Path.cwd() / vacation_path).resolve()
                vacation_path = parent_candidate if parent_candidate.exists() else cwd_candidate
            if not vacation_path.exists():
                payload["vacations"] = []
                logger.error(
                    "Vacation file not found: %s. "
                    "Set jira_weekly_email.vacation_file to an existing .xlsx path "
                    "(recommended location: report_inputs/<file>.xlsx).",
                    vacation_path,
                )
            else:
                try:
                    payload["vacations"] = parse_vacations_excel(
                        vacation_path,
                        sheet=vacation_sheet,
                        markers=vacation_markers,
                        horizon_start=vacation_horizon_start,
                        horizon_days=vacation_horizon_days,
                    )
                    if not payload.get("vacations"):
                        alternate_vacations = parse_vacations_excel(
                            vacation_path,
                            sheet=vacation_sheet,
                            markers=vacation_markers,
                            horizon_start=alternate_horizon_start,
                            horizon_days=vacation_horizon_days,
                        )
                        if alternate_vacations:
                            payload["vacations"] = alternate_vacations
                            logger.info(
                                "VACATION FALLBACK APPLIED: original_start=%s alternate_start=%s entries=%s",
                                vacation_horizon_start.strftime("%Y-%m-%d"),
                                alternate_horizon_start.strftime("%Y-%m-%d"),
                                len(payload.get("vacations") or []),
                            )
                    logger.info(
                        "VACATION RESULT: entries=%s file=%s anchor=%s horizon_start=%s",
                        len(payload.get("vacations") or []),
                        vacation_path,
                        vacation_horizon_anchor,
                        vacation_horizon_start.strftime("%Y-%m-%d"),
                    )
                except Exception as exc:
                    payload["vacations"] = []
                    logger.error(
                        "Vacation data read failed: file=%s error=%s. "
                        "Check workbook path/sheet name and markers configuration.",
                        vacation_path,
                        exc,
                    )
        else:
            logger.info("VACATION RESULT: skipped (vacation_file is empty).")

        output_dir = _normalize_text(extra_params.get("output_dir") or section.get("output_dir") or config.get("reporting", "output_dir", fallback="reports"))
        output_base = Path(output_dir)
        try:
            output_base.mkdir(parents=True, exist_ok=True)
        except Exception as exc:
            logger.error(
                "jira_weekly_email: failed to prepare output directory %s (%s). "
                "Set writable output_dir in config or --params output_dir=<path>.",
                output_base,
                exc,
            )
            return
        snapshot_dir = _normalize_text(extra_params.get("snapshot_dir") or section.get("snapshot_dir")) or str(output_base)
        snapshot_base = Path(snapshot_dir).resolve()
        previous_week = _previous_week_window(week)
        logger.info(
            "SNAPSHOT INPUT: dir=%s project=%s current_week=%s previous_week=%s",
            snapshot_base,
            project,
            week.key,
            previous_week.key,
        )

        with progress.step("Apply snapshot + AI"):
            previous_snapshot = load_previous_snapshot(snapshot_base, project, week)
            previous_payload = previous_snapshot.get("payload") if previous_snapshot else None
            previous_week_key = _normalize_text((previous_snapshot.get("meta") or {}).get("week_key")) if previous_snapshot else ""
            if previous_snapshot:
                logger.info("SNAPSHOT FOUND: previous_week=%s", previous_week_key or previous_week.key)
            else:
                logger.info("SNAPSHOT NOT FOUND: expected_previous_week=%s", previous_week.key)
            payload = apply_previous_order(payload, previous_snapshot)
            if ai_enabled:
                payload = rewrite_payload_with_ai(payload, config, extra_params)
            else:
                logger.info("AI rewrite skipped: ai_enabled=false")

        with progress.step("Export report"):
            html_text = render_outlook_html(payload)
        output_name = _normalize_text(extra_params.get("output") or extra_params.get("output_file"))
        if output_name:
            output_path = Path(output_name)
            if not output_path.is_absolute():
                output_path = output_base / output_path
            if output_path.suffix.lower() != ".html":
                output_path = output_path.with_suffix(".html")
        else:
            output_path = output_base / f"jira_weekly_email_{project}_{week.key}.html"
        try:
            output_path.write_text(html_text, encoding="utf-8")
        except Exception as exc:
            logger.error(
                "jira_weekly_email: failed to write HTML output %s (%s). "
                "Check output path permissions and disk space.",
                output_path,
                exc,
            )
            return

        # --- Optional extra output formats ---
        email_subject = f"Weekly Report {project} {week.key}"

        if "docx" in output_formats:
            docx_result = _convert_html_to_docx(output_path, output_path.parent)
            if docx_result:
                logger.info("DOCX saved: %s", docx_result)
            else:
                logger.warning("DOCX conversion failed; HTML output is still available at %s", output_path)

        if "eml" in output_formats:
            eml_path = output_path.with_suffix(".eml")
            try:
                _save_as_eml(html_text, email_subject, eml_path)
            except Exception as exc:
                logger.error("jira_weekly_email: failed to write EML output %s (%s).", eml_path, exc)

        if "outlook_draft" in output_formats:
            try:
                _open_outlook_draft(_prepare_html_for_eml(html_text), email_subject)
            except Exception as exc:
                logger.error("jira_weekly_email: failed to open Outlook draft (%s).", exc)

        snapshot_path = snapshot_base / f"jira_weekly_email_{project}_{week.key}.json"
        try:
            save_snapshot(snapshot_path, payload, week)
        except Exception as exc:
            logger.error(
                "jira_weekly_email: failed to write snapshot %s (%s). "
                "Check snapshot_dir permissions.",
                snapshot_path,
                exc,
            )
            return

        diff_lines = compute_payload_diff(previous_payload, payload)
        diff_stats = _diff_stats(diff_lines)
        logger.info(
            "DIFF SUMMARY: previous_week=%s added=%s removed=%s unchanged=%s",
            previous_week_key or "none",
            diff_stats["added"],
            diff_stats["removed"],
            diff_stats["unchanged"],
        )
        if diff_lines and previous_week_key:
            render_console_diff(
                diff_lines,
                project=project,
                current_week_key=week.key,
                previous_week_key=previous_week_key,
                use_color=True,
            )
        elif not previous_week_key:
            logger.info("DIFF SKIPPED: previous week snapshot is missing for week=%s", previous_week.key)

        logger.info(
            "REPORT SUMMARY: project=%s week=%s issues=%s highlights=%s epics=%s plans=%s vacations=%s output=%s",
            project,
            week.key,
            len(evidence),
            len(payload.get("highlights") or []),
            len(payload.get("epics") or []),
            sum(len(item.get("items") or []) for item in (payload.get("next_week_plans") or [])),
            len(payload.get("vacations") or []),
            output_path,
        )
        logger.info("\n%s", _render_summary_table_console(payload.get("summary_table") or {}))
