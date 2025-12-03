"""
Utilities for working with engineer/member lists shared across reports.
"""

from __future__ import annotations

from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook
from openpyxl.utils.cell import column_index_from_string

NAME_HEADERS = {
    "name",
    "display name",
    "assignee",
    "engineer",
    "member",
    "full name",
}

LOGIN_HEADERS = {
    "login",
    "username",
    "user name",
    "account",
    "account name",
}


def read_member_list(
    member_list_file: str | Path,
    *,
    prefer: str = "name",
    fallback_column: str = "E",
) -> list[str]:
    """
    Read distinct member identifiers from an Excel file.

    Args:
        member_list_file: Path to the Excel file with members.
        prefer: Which column group to use ("name" or "login").
        fallback_column: Excel column letter to fall back to when no headers match.

    Returns:
        List of unique member values (preserving file order).
    """

    member_path = Path(member_list_file)
    if not member_path.exists():
        raise FileNotFoundError(f"Member list file not found: {member_list_file}")

    workbook = load_workbook(member_path, data_only=True)
    sheet = workbook.active

    header_map = _build_header_map(sheet)
    prefer = prefer.lower()

    column_idx = None
    if prefer == "name":
        column_idx = _find_column_index(header_map, NAME_HEADERS)
        if column_idx is None:
            # fall back to login headers if names missing
            column_idx = _find_column_index(header_map, LOGIN_HEADERS)
    elif prefer == "login":
        column_idx = _find_column_index(header_map, LOGIN_HEADERS)
    else:
        raise ValueError("prefer must be either 'name' or 'login'")

    if column_idx is None:
        # use fallback column letter (default E from legacy file)
        column_idx = column_index_from_string(fallback_column)

    values: list[str] = []
    seen: set[str] = set()

    for row in range(2, sheet.max_row + 1):
        raw_value = sheet.cell(row=row, column=column_idx).value
        cleaned = _clean_value(raw_value)
        if not cleaned:
            continue
        marker = _normalize(cleaned)
        if marker in seen:
            continue
        seen.add(marker)
        values.append(cleaned)

    return values


def _build_header_map(sheet) -> dict[str, int]:
    """Map normalized header titles to column indices."""
    headers: dict[str, int] = {}
    for cell in sheet[1]:
        if not cell.value:
            continue
        normalized = _normalize(str(cell.value))
        if normalized:
            headers[normalized] = cell.col_idx
    return headers


def _find_column_index(header_map: dict[str, int], candidates: Iterable[str]) -> int | None:
    """Locate column index by matching header titles."""
    for candidate in candidates:
        normalized = _normalize(candidate)
        if normalized in header_map:
            return header_map[normalized]
    return None


def _clean_value(value: object) -> str | None:
    """Convert cell content to a stripped string."""
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _normalize(value: str) -> str:
    """Normalize strings for comparison."""
    return " ".join(value.split()).casefold()

