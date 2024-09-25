# -*- coding: utf-8 -*-
import sys                             #
import requests                        #
import csv                             #
import codecs                          # used for text encoding in config parser
import argparse                        # good argument parser
import logging                         #
import re                              # for parsing ticket no from description
from openpyxl import load_workbook     #
from configparser import ConfigParser  # able to read configuration file
from datetime import date              # used for manipulations with dates
from datetime import timedelta

CONFIG_POINT_LOCAL = "codehub"
CONFIG_POINT_GLOBAL = "global"
CONFIG_FILE = "config.ini"
CONFIG_BASE_URL = "codehub-url"
CONFIG_TOKEN = "token"
CONFIG_MEMBER_LIST = "member-list"
CONFIG_BRANCH = "branch"
CONFIG_PROJECT = "project"


def main():
    #   parse arguments and options and get dict
    options = parse_arguments_and_options()
    #   set log level according to options or argument
    level = logging.DEBUG if options['log_level'] else logging.INFO
    logging.basicConfig(format='%(asctime)s %(levelname)s %(message)s', level=level)
    #   read list of members
    member_list = read_member_list("members.xlsx")
    s = requests.Session()
    s.headers = {'Private-Token': options['token']}
    # bundle-ca is a text file with certificate in Base64 format of intermediate CA and root CA.
    # Used for self-signed certificates which does not exist in certifi
    s.verify = 'bundle-ca'
    since = '2021-12-01'
    until = '2021-12-31'
    if options['project'] is None or len(options['project']) == 0:
        ids = get_contribute_project_ids(options['base_url'], s, since, until)
    else:
        ids = get_project_id_by_name(options['base_url'], s, options['project'])
    project_report = []
    project_num = len(ids)
    # we got id's of projects where events exists and will iterate through users from the list to collect statistics for these users
    for user in member_list:
        author = user['Username']
        idx = 1

        for project_id in ids:
            project_info = get_project_info(options['base_url'], s, project_id)
            print('(' + str(idx) + '/' + str(project_num) + ')Analyzing project: ' + project_info['path_with_namespace'] + ' For Username: ' + author)
            for branch in options['branch'].split(','):
                commits = get_all_commits(options['base_url'], s, project_id, branch, since, until, author)
                for c in commits:
                    detail = get_commit_detail(options['base_url'], s, project_id, c['id'])
                    stats = detail['stats']
                    name = detail['author_name']  # TODO Gitlab returns ID instead of a name
                    title = detail['title']
                    commit_url = '{}/{}/{}/{}'.format(
                        options['base_url'], project_info['path_with_namespace'], "files/commit", c['id'])
                    commit_date = detail['committed_date']
                    description = detail['message']
                    #additions = stats['additions']
                    #deletions = stats['deletions']
                    total = stats['total']
                    comments = get_commit_comments_count(options['base_url'], s, project_id, c['id'], member_list, name)
                    ticket_no = get_ticketno(description)
                    # combining all data into array
                    project_report.append({ 'Name': user['Name'],
                                            'Project': project_info['path_with_namespace'],
                                            'Date': commit_date,
                                            'Gitlab id': commit_url,
                                            'Title': title,
                                            'branch': branch,
                                            'description': description,
                                            'LOC': total,
                                            'Reviewed By': ';'.join(comments[2]),
                                            'AR': ticket_no,
                                            'change_inner_group_comments_count': comments[0],
                                            'change_outer_group_comments_count': comments[1]
                                            })
                idx += 1
        # TODO comments in other team member's code and KLOCs reviewed
        #get_users_comments
    # Create a report file with headlines
    file_name = CONFIG_POINT_LOCAL + "-commits-" + since + ".csv"
    create_csv_file(file_name)

    with open(file_name, "a", encoding='utf-8-sig', newline='') as csvfile:
        writer = csv.writer(csvfile)
        for commit in project_report:
            writer.writerow([commit['Name'],
                             commit['Project'],
                             str(commit['Date']),
                             str(commit['Gitlab id']),
                             commit['branch'],
                             commit['description'],
                             str(commit['LOC']),
                             commit['Reviewed By'],
                             commit['AR'],
                             str(commit['change_inner_group_comments_count']),
                             str(commit['change_outer_group_comments_count'])
                             ])
    print("All done!")
    return 0


def parse_arguments_and_options():
    #   argument parsing section
    tool_description = 'Track comments left in code by using Gerrit API'
    parser = argparse.ArgumentParser(description=tool_description,
                                     formatter_class=argparse.ArgumentDefaultsHelpFormatter)
    parser.add_argument('-t', '--token', dest='token', help='token')
    parser.add_argument('-l', '--url-link', dest='base_url', help='the gerrit url')
    parser.add_argument('-f', '--file', dest='member_list_file', help='Group member list file')
    parser.add_argument('-p', '--project', dest='project', help='Project where calculate statistics')
    parser.add_argument('-b', '--branch', dest='branch', help='Branch for statistics calculation')
    parser.add_argument('-d', '--date', dest='date_start', required=True,
                        help='query date string start, eg. 2019-06 or 2019-06-25')
    parser.add_argument('-v', '--verbose', dest='verbose', action='store_true', help='enable verbose (debug) logging')
    options = parser.parse_args()

    base_url = options.base_url
    token = options.token
    branch = options.branch
    project = options.project
    date_start = options.date_start
    member_list_file = options.member_list_file
    log_level = options.verbose
    config = ConfigParser()
    config.read_file(codecs.open(CONFIG_FILE, 'r', encoding='utf-8-sig'))

    if base_url is None:
        base_url = config.get(CONFIG_POINT_LOCAL, CONFIG_BASE_URL)
    if token is None:
        token = config.get(CONFIG_POINT_LOCAL, CONFIG_TOKEN)
    if branch is None:
        branch = config.get(CONFIG_POINT_LOCAL, CONFIG_BRANCH)
    if project is None:
        project = config.get(CONFIG_POINT_LOCAL, CONFIG_PROJECT)
    if member_list_file is None:
        member_list_file = config.get(CONFIG_POINT_GLOBAL, CONFIG_MEMBER_LIST)

    if base_url is None or token is None or member_list_file is None:
        raise ValueError("url or username or password or file is invalid")
    return {'base_url': base_url, 'token': token, 'branch': branch, 'project': project,
            'member_list_file': member_list_file, 'date_start': date_start, 'log_level': log_level}


def get_contribute_project_ids(base_url, session, since, until):
    res = []
    next_page = 1
    # make date period bigger as API does not include borders
    since_d = date.fromisoformat(since)
    until_d = date.fromisoformat(until)
    while next_page != '':
        url = '{}/api/v4/events?per_page=100&page={}&action=pushed&after={}&before={}'\
            .format(base_url, next_page, (since_d - timedelta(days=1)).isoformat(), (until_d + timedelta(days=1)).isoformat())
        resp = session.get(url)
        next_page = resp.headers.get('X-Next-Page')
        for event_date in resp.json():
            for event in resp.json()[event_date]:
                project_id = event['project_id']
                res.append(project_id)
    res = set(res)
    return res


def get_project_info(base_url, session, project_id):
    url = '{}{}/{}'.format(base_url, '/api/v4/projects', project_id)
    resp = session.get(url)
    return resp.json()


def get_all_commits(base_url, session, project_id, branch, since, until, filter_author=''):
    res = []
    next_page = 1
    url_format = '{}/api/v4/projects/{}/repository/commits?ref={}&since={}&until={}&per_page=100&page={}'
    while next_page != '':
        url = url_format.format(base_url, project_id, branch, since + "T00:00:00Z", until + "T23:59:59Z", next_page)
        resp = session.get(url)
        next_page = resp.headers.get('X-Next-Page')
        if filter_author == '':
            res.extend(resp.json())
        else:
            for commit in resp.json():
                if commit['author_name'] == filter_author:
                    res.append(commit)
    return res


def get_commit_detail(base_url, session, project_id, commit_id):
    url = '{}/api/v4/projects/{}/repository/commits/{}'.format(base_url, project_id, commit_id)
    resp = session.get(url)
    return resp.json()


def get_commit_comments_count(base_url, session, project_id, commit_id, member_list, commit_author):
    url = '{}/api/v4/projects/{}/repository/commits/{}/comments'.format(base_url, project_id, commit_id)
    resp = session.get(url)
    inner = 0
    outer = 0
    reviewers = []
    for comment in resp.json():
        if commit_author != comment['author']['username']:
            filter_result = \
                list(filter(lambda author: author['Username'] == comment['author']['username'], member_list))
            if filter_result:
                inner += 1
                reviewers.append(comment['author']['name'])
            else:
                outer += 1
    return inner, outer, reviewers


def read_member_list(member_list_file):
    wb = load_workbook(member_list_file)
    sheet = wb.active
    has_name = "name" == sheet.cell(row=1, column=1).value
    has_mail = "email" == sheet.cell(row=1, column=2).value
    has_username = "username" == sheet.cell(row=1, column=3).value

    member_list = []
    if has_name and has_mail and has_username:
        for rx in range(2, sheet.max_row + 1):
            name = sheet.cell(row=rx, column=1).value
            mail = sheet.cell(row=rx, column=2).value
            username = sheet.cell(row=rx, column=3).value
            member_list.append({'Name': name, 'Email': mail, 'Username': username})
    else:
        raise ValueError("The table format is incorrect")
    return member_list


def get_project_id_by_name(base_url, session, project_name):
    projects = project_name.split(",")
    project_id = []
    for project in projects:
        url = '{}/api/v4/projects?search={}'.format(base_url, project)
        resp = session.get(url)
        if len(resp.json()) != 0:
            project_id.append(resp.json()[0]['id'])
        else:
            print("Empty response, project not found or you do not have permissions to it")
    return project_id


def create_csv_file(file_name):
    # Write Title
    with open(file_name, "w", encoding='utf-8-sig', newline='') as csvfile:
        writer = csv.writer(csvfile)
        writer.writerow(["Name", "Project", "Date", "Gitlab id", "branch", "description", "LOC", "Reviewed By", "AR",
                         "change_inner_group_comments_count", "change_outer_group_comments_count"])


def get_ticketno(description):
    ticket_pattern = re.compile(r'(?<=Description:)(.*?)(?=TicketNo:|Team:|Change-Id)')
    ticket_number = re.findall(ticket_pattern, description)
    if ticket_number:
        ticket_number = ticket_number[0]
    else:
        ticket_number = ""
    return ticket_number


if __name__ == '__main__':
    sys.exit(main())  # TODO check exit code
