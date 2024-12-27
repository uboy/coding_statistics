# -*- coding: utf-8 -*-
import os.path
import sys  ###
import requests  ###
import csv  ###
import codecs  ### used for text encoding in config parser
import argparse  ### good argument parser
import logging  ###
from configparser import ConfigParser  ### able to read configuration file
from datetime import datetime, time, timedelta         ### used for manipulations with dates
from openpyxl import load_workbook  ###
import openpyxl
from tqdm import tqdm  # show progress

#from pkg_resources import empty_provider

CONFIG_POINT_LOCAL = "gitee"
CONFIG_POINT_GLOBAL = "global"
CONFIG_FILE = "config.ini"
CONFIG_BASE_URL = "gitee-url"
CONFIG_TOKEN = "token"
CONFIG_MEMBER_LIST = "member-list"
CONFIG_BRANCH = "branch"
CONFIG_REPOSITORY = "repository"
CONFIG_UNTIL = "date_until"
# Default values, can be overridden by config
PER_PAGE = 50
PAUSE_AFTER_REQUESTS = 10
PAUSE_DURATION = 2  # seconds
# see https://gitee.com/api/v5/swagger#/getV5ReposOwnerRepoPulls
GET_LIST_PR = '{}/api/v5/repos/{}/{}/pulls?base={}&state=all&since={}&per_page={}}&page={}'
#  see https://gitee.com/api/v5/swagger#/getV5ReposOwnerRepoPullsNumberFiles
GET_PR_FILES = '{}/api/v5/repos/{}/{}/pulls/{}/files'

def main():
    ### argument parsing section
    tool_description = 'Track commit by using Gitee API'
    parser = argparse.ArgumentParser(description=tool_description,
                                     formatter_class=argparse.ArgumentDefaultsHelpFormatter)
    parser.add_argument('-t', '--token', dest='token', help='token')
    parser.add_argument('-l', '--url-link', dest='base_url', help='the gitee url')
    parser.add_argument('-f', '--file', dest='member_list_file', help='Group member list file')
    parser.add_argument('-b', '--branch', dest='branch',
                        help='Branch to get report for all projects. Can be comma separated list like "master,dev,main"')
    parser.add_argument('-p', '--project', dest='project',
                        help='Specify project for report path with namespace. Possible to specify several project separated by comma"')
    parser.add_argument('-d', '--date', dest='date_since', required=True,
                        help='query date string since, eg. 2019-06-25')
    parser.add_argument('-u', '--until', dest='date_until', required=False,
                        help='query date string until, eg. 2019-12-25')
    parser.add_argument('-v', '--verbose', dest='verbose', action='store_true', help='enable verbose (debug) logging')
    options = parser.parse_args()
    level = logging.DEBUG if options.verbose else logging.INFO
    logging.basicConfig(format='%(asctime)s %(levelname)s %(message)s', level=level)

    base_url = options.base_url
    token = options.token
    since = (options.date_since).strip()
    branch_list = options.branch

    member_list_file = options.member_list_file
    ###
    ### config file section
    config = ConfigParser(allow_no_value=False, comment_prefixes=('#', ';'),inline_comment_prefixes='#')
    config.read_file(codecs.open(CONFIG_FILE, 'r', encoding='utf-8-sig'))

    if base_url is None:
        base_url = config.get(CONFIG_POINT_LOCAL, CONFIG_BASE_URL)
    if token is None:
        token = config.get(CONFIG_POINT_LOCAL, CONFIG_TOKEN)
    if branch_list is None:
        branch_list = config.get(CONFIG_POINT_LOCAL, CONFIG_BRANCH).split(",")
    if member_list_file is None:
        member_list_file = config.get(CONFIG_POINT_GLOBAL, CONFIG_MEMBER_LIST)
    ###

    if config.has_option(CONFIG_POINT_LOCAL, 'per_page'):
        PER_PAGE = int(config.get(CONFIG_POINT_LOCAL, 'per_page'))
    if config.has_option(CONFIG_POINT_LOCAL, 'pause_after_requests'):
        PAUSE_AFTER_REQUESTS = int(config.get(CONFIG_POINT_LOCAL, 'pause_after_requests'))
    if config.has_option(CONFIG_POINT_LOCAL, 'pause_duration'):
        PAUSE_DURATION = int(config.get(CONFIG_POINT_LOCAL, 'pause_duration'))

    if base_url is None or token is None or member_list_file is None or branch_list is None:
        raise ValueError("url or token or file is invalid")
    until = options.date_until or config.get(CONFIG_POINT_LOCAL, CONFIG_UNTIL, fallback=None) or datetime.today().strftime('%Y-%m-%d')
    repositories = config.get(CONFIG_POINT_LOCAL, CONFIG_REPOSITORY).split(",")
    print(f"Starting to prepare report from Gitee for branch: {branch_list}, repositories {repositories}, date since {since}, date until {until}")
    # member_list = read_member_list("members.xlsx") ### TODO add report for set members only
    s = requests.Session()
    s.headers = {'Private-Token': token}
    ### bundle-ca is a text file with certificate in Base64 format of intermediate CA and root CA. Used for self-signed certificates which does not exist in certifi
    s.verify = 'bundle-ca' if os.path.exists("bundle-ca") else True
    project_report = []
    # get comma-separated repositories with projects
    
    for repository in repositories:
        branches = config.get(CONFIG_POINT_LOCAL, CONFIG_BRANCH).split(",")
        repo_string = repository.split('/')
        repo_owner = repo_string[0]
        repo = repo_string[1]

    # Прогресс-бар для веток
    for branch in tqdm(branches, desc=f"Processing Branches in {repository}", leave=False):
        prs = get_all_prs(s, base_url, repo_owner, repo, branch, since)

        # Прогресс-бар для страниц (в get_all_prs)
        for c in tqdm(prs, desc=f"Processing PRs in {branch}", leave=False):
                user_name = c['user']['name']
                user_login = c['user']['login']
                pr_title = c['title']
                pr_url = c['html_url']
                pr_state = c['state']
                pr_date = c['created_at']
                pr_merged_date = c['merged_at']
                description = c['body']
                size = get_pr_size(s, base_url, repo_owner, repo, c['number'])
                ### combining all data into array
                project_report.append({
                    'Name': user_name,
                    'Login': user_login,
                    'PR_Name': pr_title,
                    'PR_URL': pr_url,
                    'PR_State': pr_state,
                    'PR_Created_Date': pr_date,
                    'PR_Merged_Date': pr_merged_date,
                    'PR_Description': description,
                    'branch': branch,
                    'Repo': repo_owner + "/" + repo,
                    'additions': size[0],
                    'deletions': size[1]
                })
    ### TODO comments in other team member's code and KLOCs reviewed
    # get_users_comments
    ### Create a report file with headlines
    file_name = "gitee-prs-since-" + since + "-until-" + until + ".csv"
    create_csv_file(file_name)

    with open(file_name, "a", encoding='utf-8-sig', newline='') as csvfile:
        writer = csv.writer(csvfile)
        for commit in project_report:
            writer.writerow([commit['Name'],
                             commit['Login'],
                             commit['PR_Name'],
                             commit['PR_URL'],
                             commit['PR_State'],
                             str(commit['PR_Created_Date']),
                             commit['PR_Merged_Date'],
                             #commit['PR_Description'],
                             commit['branch'],
                             commit['Repo'],
                             commit['additions'],
                             commit['deletions']
                             ])

    # Create an Excel report file with headlines
    excel_file_name = "gitee-prs-since-" + since + "-until-" + until + ".xlsx"
    create_excel_file(excel_file_name, project_report)  # Call the new function

    print("All done!")
    return 0


def get_all_prs(session, base_url, project_id, repository, branch, since):
    res = []
    next_page = 1
    url_format = GET_LIST_PR
    request_count = 0

    # Предварительное получение общего количества страниц
    url = url_format.format(base_url, project_id, repository, branch, since + "T00:00:00Z", PER_PAGE, next_page)
    initial_resp = session.get(url)
    total_pages = int(initial_resp.headers.get('total_page', 1))  # Дефолтное значение - 1

    # Прогресс-бар для страниц
    with tqdm(total=total_pages, desc=f"Processing Pages for {repository}/{branch}", leave=False) as pbar:
        while next_page <= total_pages:
            url = url_format.format(base_url, project_id, repository, branch, since + "T00:00:00Z", PER_PAGE, next_page)
            resp = session.get(url)
            request_count += 1

            if resp.status_code == 429:
                logging.warning("Rate limit exceeded. Waiting for 1 second...")
                time.sleep(1)
                continue
            elif resp.status_code != 200:
                logging.error(f"Failed to fetch data for page {next_page}. Status code: {resp.status_code}")
                break

            res.extend(resp.json())
            pbar.update(1)
            next_page += 1

            # Пауза после заданного количества запросов
            if request_count >= PAUSE_AFTER_REQUESTS:
                logging.info(f"Pausing for {PAUSE_DURATION} seconds after {PAUSE_AFTER_REQUESTS} requests.")
                time.sleep(PAUSE_DURATION)
                request_count = 0

    return res


def get_pr_files(session, base_url, project_id, repository, pr):
    url_format = GET_PR_FILES
    url = url_format.format(base_url, project_id, repository, pr)
    resp = session.get(url)
    return resp.json()

def get_pr_size(session, base_url, project_id, repository, pr):
    additions = 0
    deletions = 0
    files_changed = get_pr_files(session, base_url, project_id, repository, pr)
    for file in files_changed:
        additions += int(file['additions'])
        deletions += int(file['deletions'])
    return [additions,deletions]


def read_member_list(member_list_file):
    wb = load_workbook(member_list_file)
    sheet = wb.active
    has_name = "name" == sheet.cell(row=1, column=1).value
    has_mail = "email" == sheet.cell(row=1, column=2).value
    has_username = "username" == sheet.cell(row=1, column=3).value
    has_giteeaccount = "gitee_account" == sheet.cell(row=1, column=4).value

    member_list = []
    if has_name and has_mail and has_username and has_giteeaccount:
        for rx in range(2, sheet.max_row + 1):
            name = sheet.cell(row=rx, column=1).value
            mail = sheet.cell(row=rx, column=2).value
            username = sheet.cell(row=rx, column=3).value
            gitee_account = sheet.cell(row=rx, column=4).value
            member_list.append({'Name': name, 'Email': mail, 'Username': username, 'GiteeAccount': gitee_account})
    else:
        raise ValueError("The table format is incorrect")
    return member_list


def create_csv_file(file_name):
    # Write Title
    with open(file_name, "w", encoding='utf-8-sig', newline='') as csvfile:
        writer = csv.writer(csvfile)
        writer.writerow(["Name", "Login", "PR_Name", "PR_URL", "PR_State", "PR_Created_Date",
                         "PR_Merged_Date", "branch", "Repo", "Additions", "Deletions"])


def create_excel_file(file_name, project_report):
    # Create a workbook and select the active worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    
    # Write headers
    headers = ["Name", "Login", "PR_Name", "PR_URL", "PR_State", 
               "PR_Created_Date", "PR_Merged_Date", "branch", "Repo", 
               "Additions", "Deletions"]
    ws.append(headers)
    
    # Write data rows
    for commit in project_report:
        row_data = [commit['Name'],
                    commit['Login'],
                    commit['PR_Name'],
                    commit['PR_URL'],
                    commit['PR_State'],
                    str(commit['PR_Created_Date']),
                    commit['PR_Merged_Date'],
                    commit['branch'],
                    commit['Repo'],
                    commit['additions'],
                    commit['deletions']
                   ]
        ws.append(row_data)
    
    # Save the workbook
    wb.save(file_name)

if __name__ == '__main__':
    sys.exit(main())  ### TODO check exit code
