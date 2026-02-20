from __future__ import annotations

import json
from configparser import ConfigParser
from datetime import date, datetime, timedelta
from pathlib import Path
from types import SimpleNamespace
from unittest.mock import Mock, patch

from openpyxl import Workbook

from stats_core.reports.jira_weekly_email import (
    JiraWeeklyEmailReport,
    load_previous_snapshot,
    parse_vacations_excel,
    render_outlook_html,
    rewrite_payload_with_ai,
    resolve_week_window,
)


def _make_issue(
    key: str,
    *,
    summary: str,
    issue_type: str,
    status: str,
    resolution: str,
    labels: list[str],
    priority: str,
    epic_link: str = "EPIC-1",
    parent_key: str | None = None,
    comment_created: str = "2026-03-03T10:00:00.000+0000",
    comment_body: str = "Weekly update",
    issue_is_subtask: bool = False,
):
    parent = SimpleNamespace(key=parent_key) if parent_key else None
    comment = SimpleNamespace(
        comments=[
            SimpleNamespace(
                body=comment_body,
                created=comment_created,
                id=f"c-{key}",
            )
        ]
    )
    fields = SimpleNamespace(
        summary=summary,
        status=SimpleNamespace(name=status),
        resolution=SimpleNamespace(name=resolution) if resolution else None,
        issuetype=SimpleNamespace(name=issue_type, subtask=issue_is_subtask),
        labels=labels,
        priority=SimpleNamespace(name=priority) if priority else None,
        customfield_10000=epic_link,
        parent=parent,
        comment=comment,
    )
    return SimpleNamespace(key=key, fields=fields)


def _make_epic_issue(key: str, summary: str, labels: list[str]):
    fields = SimpleNamespace(
        summary=summary,
        labels=labels,
        issuetype=SimpleNamespace(name="Epic"),
        customfield_10000="",
        parent=None,
    )
    return SimpleNamespace(key=key, fields=fields)


def _make_parent_issue(
    key: str,
    *,
    summary: str,
    status: str = "In Progress",
    resolution: str = "",
    epic_link: str = "EPIC-1",
    parent_key: str | None = None,
    issue_type: str = "Task",
    labels: list[str] | None = None,
):
    fields = SimpleNamespace(
        summary=summary,
        status=SimpleNamespace(name=status) if status else None,
        resolution=SimpleNamespace(name=resolution) if resolution else None,
        customfield_10000=epic_link,
        parent=SimpleNamespace(key=parent_key) if parent_key else None,
        issuetype=SimpleNamespace(name=issue_type) if issue_type else None,
        labels=labels or [],
    )
    return SimpleNamespace(key=key, fields=fields)


def test_resolve_week_window_from_week_defaults_year():
    week = resolve_week_window({"week": "8"}, now=date(2026, 2, 18))
    assert week.year == 2026
    assert week.week == 8
    assert week.key == "26'w08"
    assert week.start == date(2026, 2, 16)
    assert week.end == date(2026, 2, 22)


def test_resolve_week_window_from_compact_week_and_year():
    week = resolve_week_window({"week": "08w26"}, now=date(2025, 1, 1))
    assert week.year == 2026
    assert week.week == 8
    assert week.key == "26'w08"
    assert week.start == date(2026, 2, 16)
    assert week.end == date(2026, 2, 22)


def test_jira_weekly_email_missing_project_does_not_raise(tmp_path: Path):
    config = ConfigParser()
    config.read_dict(
        {
            "jira": {"jira-url": "https://jira.example.com", "username": "u", "password": "p"},
            "reporting": {"output_dir": str(tmp_path)},
            "ollama": {"enabled": "false"},
            "jira_weekly_email": {},
        }
    )
    report = JiraWeeklyEmailReport()
    report.run(
        dataset={},
        config=config,
        output_formats=["html"],
        extra_params={
            "week_date": "2026-03-03",
            "output_dir": str(tmp_path),
        },
    )
    assert not (tmp_path / "jira_weekly_email__26'w10.html").exists()


@patch("stats_core.reports.jira_weekly_email.JiraSource")
def test_jira_weekly_email_invalid_vacation_horizon_days_does_not_raise(mock_jira_source_cls, tmp_path: Path):
    issues = [
        _make_issue(
            "ABC-10",
            summary="Report task done",
            issue_type="Task",
            status="Done",
            resolution="Done",
            labels=["reportx"],
            priority="Medium",
            comment_body="Done this week.",
        )
    ]
    fake_jira = Mock()

    def _fake_search_issues(jql, *args, **kwargs):
        if "issuekey in (EPIC-1)" in str(jql):
            return [_make_epic_issue("EPIC-1", "Epic One", ["reportx"])]
        return issues

    fake_jira.search_issues.side_effect = _fake_search_issues
    fake_source = Mock()
    fake_source.jira = fake_jira
    fake_source.fetch_epic_names.return_value = {"EPIC-1": "Epic One"}
    mock_jira_source_cls.return_value = fake_source

    config = ConfigParser()
    config.read_dict(
        {
            "jira": {"jira-url": "https://jira.example.com", "username": "u", "password": "p"},
            "reporting": {"output_dir": str(tmp_path)},
            "ollama": {"enabled": "false"},
            "jira_weekly_email": {"labels_report": "reportx"},
        }
    )

    report = JiraWeeklyEmailReport()
    report.run(
        dataset={},
        config=config,
        output_formats=["html"],
        extra_params={
            "project": "ABC",
            "week_date": "2026-03-03",
            "labels_report": "reportx",
            "vacation_file": str(tmp_path / "missing-vacations.xlsx"),
            "vacation_horizon_days": "abc",
            "output_dir": str(tmp_path),
        },
    )

    html_path = tmp_path / "jira_weekly_email_ABC_26'w10.html"
    assert html_path.exists()


@patch("stats_core.reports.jira_weekly_email.requests.post")
def test_rewrite_payload_with_webui_provider(mock_post):
    response = Mock()
    response.raise_for_status.return_value = None
    response.json.return_value = {
        "choices": [
            {
                "message": {
                    "content": "{\"t1\":\"Updated highlight progress\"}",
                }
            }
        ]
    }
    mock_post.return_value = response

    payload = {
        "highlights": [{"issue_key": "ABC-1", "headline": "Old headline", "comment": "Old comment"}],
        "epics": [],
        "next_week_plans": [],
    }
    config = ConfigParser()
    config.read_dict(
        {
            "jira_weekly_email": {},
            "webui": {
                "enabled": "true",
                "url": "http://localhost:3000",
                "endpoint": "/api/chat/completions",
                "api_key": "cfg-key",
                "model": "qwen",
                "timeout_seconds": "30",
                "temperature": "0.2",
            },
        }
    )

    rewritten = rewrite_payload_with_ai(payload, config, {})
    assert rewritten["highlights"][0]["headline"] == "Old headline"
    assert rewritten["highlights"][0]["comment"] == "Updated highlight progress"
    called_url = mock_post.call_args.kwargs["url"] if "url" in mock_post.call_args.kwargs else mock_post.call_args.args[0]
    assert called_url.endswith("/api/chat/completions")
    assert mock_post.call_args.kwargs["headers"]["Authorization"] == "Bearer cfg-key"


@patch("stats_core.reports.jira_weekly_email.requests.post")
def test_rewrite_payload_with_webui_provider_avoids_duplicate_api_path(mock_post):
    response = Mock()
    response.raise_for_status.return_value = None
    response.json.return_value = {
        "choices": [
            {
                "message": {
                    "content": "{\"t1\":\"Updated highlight progress\"}",
                }
            }
        ]
    }
    mock_post.return_value = response

    payload = {
        "highlights": [{"issue_key": "ABC-1", "headline": "Old headline", "comment": "Old comment"}],
        "epics": [],
        "next_week_plans": [],
    }
    config = ConfigParser()
    config.read_dict(
        {
            "jira_weekly_email": {},
            "webui": {
                "enabled": "true",
                "url": "http://localhost:3000/api",
                "endpoint": "/api/chat/completions",
                "api_key": "cfg-key",
                "model": "qwen",
                "timeout_seconds": "30",
            },
        }
    )

    rewrite_payload_with_ai(payload, config, {})
    called_url = mock_post.call_args.kwargs["url"] if "url" in mock_post.call_args.kwargs else mock_post.call_args.args[0]
    assert called_url == "http://localhost:3000/api/chat/completions"
    assert mock_post.call_args.kwargs["timeout"] == (10, 30)


@patch("stats_core.reports.jira_weekly_email.requests.post")
def test_rewrite_payload_with_webui_logs_powershell_curl_on_timeout(mock_post, caplog):
    mock_post.side_effect = Exception("read timeout")
    payload = {
        "highlights": [{"issue_key": "ABC-1", "headline": "Old headline", "comment": "Old comment"}],
        "epics": [],
        "next_week_plans": [],
    }
    config = ConfigParser()
    config.read_dict(
        {
            "jira_weekly_email": {},
            "webui": {
                "enabled": "true",
                "url": "http://localhost:3000",
                "endpoint": "/api/chat/completions",
                "api_key": "cfg-key",
                "model": "qwen",
                "timeout_seconds": "30",
            },
        }
    )

    rewrite_payload_with_ai(payload, config, {})
    assert any("PowerShell: curl.exe" in record.message for record in caplog.records)


@patch("stats_core.reports.jira_weekly_email.requests.post")
def test_rewrite_payload_with_webui_sanitizes_links_and_limits_to_two_sentences(mock_post):
    response = Mock()
    response.raise_for_status.return_value = None
    response_payload = {
        "t1": (
            "Results: Implemented API endpoint (\\\\srv\\share\\team\\notes.md). "
            "See details at [build log](https://git.local/path/to/build). "
            "Added tests for weekly report processing. Coordinated handover."
        ),
        "t2": (
            "Plan: Continue rollout; see commit a1b2c3d4 and PR #123. "
            "Next step in C:\\repo\\stats_core\\reports\\jira_weekly_email.py. "
            "Prepare validation with QA."
        ),
        "t3": (
            "Update: Expand coverage. "
            "See git.local/pr/123 and /mnt/data/tmp/run.log. "
            "Track regressions and align monitoring."
        ),
    }
    response.json.return_value = {
        "choices": [
            {
                "message": {
                    "content": json.dumps(response_payload),
                }
            }
        ]
    }
    mock_post.return_value = response

    payload = {
        "highlights": [{"issue_key": "ABC-1", "headline": "Old headline", "comment": "Old comment"}],
        "epics": [],
        "next_week_plans": [
            {
                "epic_key": "EPIC-1",
                "epic_name": "Epic One",
                "items": [
                    {"issue_key": "ABC-2", "text": "Old plan", "comment": "Old comment", "subtasks": []},
                ],
            }
        ],
    }
    config = ConfigParser()
    config.read_dict(
        {
            "jira_weekly_email": {"ai_provider": "webui"},
            "webui": {
                "enabled": "true",
                "url": "http://localhost:3000",
                "endpoint": "/api/chat/completions",
                "api_key": "cfg-key",
                "model": "qwen",
                "timeout_seconds": "30",
            },
        }
    )

    rewritten = rewrite_payload_with_ai(payload, config, {})
    prompt = mock_post.call_args.kwargs["json"]["messages"][1]["content"]
    assert "Maximum is 2 sentences." in prompt
    assert rewritten["highlights"][0]["headline"] == "Old headline"
    assert rewritten["highlights"][0]["comment"] != "Old comment"
    assert rewritten["next_week_plans"][0]["items"][0]["text"] == "Old plan"
    assert rewritten["next_week_plans"][0]["items"][0]["comment"] != "Old comment"
    values = [
        rewritten["highlights"][0]["comment"],
        rewritten["next_week_plans"][0]["items"][0]["comment"],
    ]
    for value in values:
        assert "http://" not in value and "https://" not in value
        assert "git.local" not in value
        assert "\\\\srv\\share" not in value
        assert "C:\\repo" not in value
        assert "/mnt/data/tmp" not in value
        assert "see commit" not in value.lower()
        assert "see pr" not in value.lower()
        assert "results:" not in value.lower()
        assert "plan:" not in value.lower()
        assert "update:" not in value.lower()
        sentences = [s for s in value.replace("?", ".").replace("!", ".").split(".") if s.strip()]
        assert len(sentences) <= 2
        assert len(value.split()) <= 48


@patch("stats_core.reports.jira_weekly_email.requests.post")
def test_rewrite_payload_with_ollama_sanitizes_links_and_limits_to_two_sentences(mock_post):
    response = Mock()
    response.raise_for_status.return_value = None
    response_payload = {
        "t1": (
            "Results: Fixed weekly report parser in src/stats_core/reports/parser.py. "
            "See commit 9f3e4ab and https://git.local/c/12345. Added regression coverage."
        ),
        "t2": (
            "Plan: Continue rollout in \\\\server\\share\\roadmap.md and C:\\repo\\next_steps.md; "
            "coordinate remaining checks with QA and product owners this week."
        ),
        "t3": (
            "Update: Collect final sign-off from /mnt/data/release-notes.md and git.local/pr/999. "
            "Prepare deployment checklist."
        ),
    }
    response.json.return_value = {"response": json.dumps(response_payload)}
    mock_post.return_value = response

    payload = {
        "highlights": [{"issue_key": "ABC-1", "headline": "Old headline", "comment": "Old comment"}],
        "epics": [],
        "next_week_plans": [
            {
                "epic_key": "EPIC-1",
                "epic_name": "Epic One",
                "items": [
                    {"issue_key": "ABC-2", "text": "Old plan", "comment": "Old comment", "subtasks": []},
                ],
            }
        ],
    }
    config = ConfigParser()
    config.read_dict(
        {
            "jira_weekly_email": {"ai_provider": "ollama"},
            "ollama": {
                "enabled": "true",
                "url": "http://localhost:11434",
                "model": "qwen",
                "timeout_seconds": "30",
                "temperature": "0.2",
            },
        }
    )

    rewritten = rewrite_payload_with_ai(payload, config, {})
    prompt = mock_post.call_args.kwargs["json"]["prompt"]
    assert "Maximum is 2 sentences." in prompt
    assert rewritten["highlights"][0]["headline"] == "Old headline"
    assert rewritten["highlights"][0]["comment"] != "Old comment"
    assert rewritten["next_week_plans"][0]["items"][0]["text"] == "Old plan"
    assert rewritten["next_week_plans"][0]["items"][0]["comment"] != "Old comment"
    values = [
        rewritten["highlights"][0]["comment"],
        rewritten["next_week_plans"][0]["items"][0]["comment"],
    ]
    for value in values:
        assert "http://" not in value and "https://" not in value
        assert "git.local" not in value
        assert "\\\\server\\share" not in value
        assert "C:\\repo" not in value
        assert "/mnt/data/release-notes.md" not in value
        assert "see commit" not in value.lower()
        assert "results:" not in value.lower()
        assert "plan:" not in value.lower()
        assert "update:" not in value.lower()
        sentences = [s for s in value.replace("?", ".").replace("!", ".").split(".") if s.strip()]
        assert len(sentences) <= 2
        assert len(value.split()) <= 48


def test_parse_vacations_excel_sheet_format(tmp_path: Path):
    path = tmp_path / "vac.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Vacations2026"
    ws.cell(row=3, column=2).value = "Name"
    ws.cell(row=3, column=6).value = datetime(2026, 3, 1)
    ws.cell(row=3, column=7).value = datetime(2026, 3, 2)
    ws.cell(row=3, column=8).value = datetime(2026, 3, 3)
    ws.cell(row=3, column=9).value = datetime(2026, 4, 20)  # outside horizon
    ws.cell(row=5, column=2).value = "Denis Mazur"
    ws.cell(row=5, column=6).value = "p"
    ws.cell(row=5, column=7).value = "P"
    ws.cell(row=5, column=8).value = "p"
    ws.cell(row=5, column=9).value = "p"
    wb.save(path)

    lines = parse_vacations_excel(
        path,
        sheet="Vacations2026",
        markers={"p", "P"},
        horizon_start=date(2026, 2, 20),
        horizon_days=30,
    )
    assert lines == ["Denis Mazur vacation 01.03.2026 - 03.03.2026"]


def test_parse_vacations_excel_supports_multiple_config_markers(tmp_path: Path):
    path = tmp_path / "vac_multi.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Vacations2026"
    ws.cell(row=3, column=2).value = "Name"
    ws.cell(row=3, column=6).value = datetime(2026, 3, 1)
    ws.cell(row=3, column=7).value = datetime(2026, 3, 2)
    ws.cell(row=3, column=8).value = datetime(2026, 3, 3)
    ws.cell(row=5, column=2).value = "Denis Mazur"
    ws.cell(row=5, column=6).value = "a"
    ws.cell(row=5, column=7).value = "p"
    ws.cell(row=5, column=8).value = "a,p"
    wb.save(path)

    lines = parse_vacations_excel(
        path,
        sheet="Vacations2026",
        markers={"a", "p"},
        horizon_start=date(2026, 2, 20),
        horizon_days=30,
    )
    assert lines == ["Denis Mazur vacation 01.03.2026 - 03.03.2026"]


def test_parse_vacations_excel_accepts_string_dates_in_header(tmp_path: Path):
    path = tmp_path / "vac_string_dates.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Vacations2026"
    ws.cell(row=3, column=2).value = "Name"
    ws.cell(row=3, column=6).value = "01.03.2026"
    ws.cell(row=3, column=7).value = "02.03.2026"
    ws.cell(row=3, column=8).value = "03.03.2026"
    ws.cell(row=5, column=2).value = "Denis Mazur"
    ws.cell(row=5, column=6).value = "p"
    ws.cell(row=5, column=7).value = "P"
    ws.cell(row=5, column=8).value = "p"
    wb.save(path)

    lines = parse_vacations_excel(
        path,
        sheet="Vacations2026",
        markers={"p", "P"},
        horizon_start=date(2026, 2, 20),
        horizon_days=30,
    )
    assert lines == ["Denis Mazur vacation 01.03.2026 - 03.03.2026"]


def test_parse_vacations_excel_real_template_60_days_from_2026_02_18():
    path = Path(__file__).parent / "fixtures" / "vacation_template.xlsx"
    lines = parse_vacations_excel(
        path,
        sheet="Vacations2026",
        markers={"a", "p"},
        horizon_start=date(2026, 2, 18),
        horizon_days=60,
    )
    assert lines == [
        "Alexey Horaskin vacation 23.03.2026 - 29.03.2026",
        "Andrey Khudenkikh vacation 17.04.2026 - 30.04.2026",
        "Gleb Skroba vacation 30.03.2026 - 05.04.2026",
        "Mikhail Politov vacation 13.04.2026 - 26.04.2026",
        "Roman Evstigneev vacation 09.03.2026 - 15.03.2026",
        "Sergey Kovalev vacation 17.04.2026 - 30.04.2026",
        "Sergey Samarin vacation 08.04.2026 - 14.04.2026",
    ]


def test_parse_vacations_excel_does_not_truncate_range_that_starts_within_horizon(tmp_path: Path):
    path = tmp_path / "vac_cross_horizon.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Vacations2026"
    ws.cell(row=3, column=2).value = "Name"

    start_day = date(2026, 4, 17)
    end_day = date(2026, 4, 30)
    day = start_day
    col = 6
    while day <= end_day:
        ws.cell(row=3, column=col).value = datetime.combine(day, datetime.min.time())
        ws.cell(row=5, column=col).value = "p"
        day += timedelta(days=1)
        col += 1

    ws.cell(row=5, column=2).value = "Denis Mazur"
    wb.save(path)

    lines = parse_vacations_excel(
        path,
        sheet="Vacations2026",
        markers={"p"},
        horizon_start=date(2026, 2, 18),
        horizon_days=60,
    )
    assert lines == ["Denis Mazur vacation 17.04.2026 - 30.04.2026"]


def test_render_outlook_html_uses_configurable_meta_header_and_footer_html():
    payload = {
        "meta": {
            "project": "ABC",
            "week_key": "26'w08",
            "week_start": "2026-02-16",
            "week_end": "2026-02-22",
        },
        "highlights": [],
        "epics": [],
        "next_week_plans": [],
        "vacations": [],
        "titles": {
            "main": "TelmaST Weekly Report",
            "header_project_info": "Execution Summary",
            "header_banner_bg_color": "rgb(23,88,98)",
            "meta_active_iteration": "Active iteration",
            "meta_active_iteration_value": "Sprint 42",
            "meta_report_period": "Report Period",
            "meta_report_owner": "Report Owner",
            "meta_report_owner_value": "Denis Mazur",
            "meta_team_member": "Team Member",
            "meta_team_member_value": "Core Team",
            "footer_html": "<p><b>Footer</b> line</p>",
        },
    }
    text = render_outlook_html(payload)
    assert "Execution Summary" in text
    assert "sub-left" not in text
    assert "<td class='sub-banner'" in text
    assert "background:rgb(23,88,98);" in text
    assert "Active iteration" in text
    assert "Sprint 42" in text
    assert "Report Period" in text
    assert "2026/02/16 - 2026/02/22" in text
    assert "Report Owner" in text
    assert "Denis Mazur" in text
    assert "Team Member" in text
    assert "Core Team" in text
    assert "<p><b>Footer</b> line</p>" in text


def test_render_outlook_html_renders_config_titles_as_html_and_lvl3_without_indent():
    payload = {
        "meta": {
            "project": "ABC",
            "week_key": "26'w08",
            "week_start": "2026-02-16",
            "week_end": "2026-02-22",
        },
        "highlights": [],
        "epics": [],
        "next_week_plans": [],
        "vacations": [],
        "titles": {
            "main": "<span>Weekly <b>Report</b></span>",
            "highlights": "<i>Highlights</i>",
            "meta_report_owner": "<u>Owner</u>",
            "meta_report_owner_value": "<b>Denis Mazur</b>",
        },
    }
    text = render_outlook_html(payload)
    assert "<span>Weekly <b>Report</b></span>" in text
    assert "<i>Highlights</i>" in text
    assert "<u>Owner</u>" in text
    assert "<b>Denis Mazur</b>" in text
    assert "&lt;span&gt;Weekly" not in text
    assert ".lvl3{margin-left:36px;}" in text
    assert ".lvl4{margin-left:20px;}" in text
    assert ".content .sec-label{width:190px;font-weight:700;font-size:14pt;" in text


@patch("stats_core.reports.jira_weekly_email.JiraSource")
def test_jira_weekly_email_report_run_html_snapshot_and_diff(mock_jira_source_cls, tmp_path: Path, capsys):
    issues_week_10 = [
        _make_issue(
            "ABC-1",
            summary="Feature delivery",
            issue_type="Feature",
            status="Done",
            resolution="Done",
            labels=["shine", "reportx"],
            priority="High",
            comment_body="Feature finalized and merged.",
        ),
        _make_issue(
            "ABC-2",
            summary="Parent task",
            issue_type="Task",
            status="In Progress",
            resolution="",
            labels=[],
            priority="Highest",
            comment_body="Work in progress for next week.",
        ),
        _make_issue(
            "ABC-3",
            summary="Bug fix",
            issue_type="Bug",
            status="Done",
            resolution="Done",
            labels=[],
            priority="Medium",
            comment_body="Bug resolved.",
        ),
        _make_issue(
            "ABC-4",
            summary="Subtask done",
            issue_type="Sub-task",
            status="Done",
            resolution="Done",
            labels=[],
            priority="Medium",
            parent_key="ABC-2",
            comment_body="Subtask completed.",
        ),
        _make_issue(
            "ABC-6",
            summary="Other epic task",
            issue_type="Task",
            status="Done",
            resolution="Done",
            labels=[],
            priority="Medium",
            epic_link="EPIC-2",
            comment_body="Completed in non-report epic.",
        ),
    ]
    issues_week_11 = issues_week_10 + [
        _make_issue(
            "ABC-5",
            summary="New improvement",
            issue_type="Improvement",
            status="Done",
            resolution="Done",
            labels=["reportx"],
            priority="Medium",
            comment_created="2026-03-10T10:00:00.000+0000",
            comment_body="Improvement closed.",
        )
    ]

    def _fake_search_issues(jql, *args, **kwargs):
        if "issuekey in (" in str(jql):
            return [
                _make_epic_issue("EPIC-1", "Epic One", ["reportx"]),
                _make_epic_issue("EPIC-2", "Epic Two", []),
            ]
        if "updated < '2026-03-09'" in str(jql):
            return issues_week_10
        if "updated < '2026-03-16'" in str(jql):
            return issues_week_11
        return []

    fake_jira = Mock()
    fake_jira.search_issues.side_effect = _fake_search_issues

    fake_source = Mock()
    fake_source.jira = fake_jira
    fake_source.fetch_epic_names.return_value = {"EPIC-1": "Epic One", "EPIC-2": "Epic Two"}
    mock_jira_source_cls.return_value = fake_source

    config = ConfigParser()
    config.read_dict(
        {
            "jira": {"jira-url": "https://jira.example.com", "username": "u", "password": "p"},
            "reporting": {"output_dir": str(tmp_path)},
            "ollama": {"enabled": "false"},
            "jira_weekly_email": {
                "labels_highlights": "highlights",
                "labels_report": "report",
            },
        }
    )

    report = JiraWeeklyEmailReport()
    report.run(
        dataset={},
        config=config,
        output_formats=["excel"],  # report forces html output mode
        extra_params={
            "project": "ABC",
            "week_date": "2026-03-03",
            "labels_highlights": "shine",
            "labels_report": "reportx",
            "output_dir": str(tmp_path),
        },
    )
    report.run(
        dataset={},
        config=config,
        output_formats=["html"],
        extra_params={
            "project": "ABC",
            "week_date": "2026-03-10",
            "labels_highlights": "shine",
            "labels_report": "reportx",
            "output_dir": str(tmp_path),
        },
    )

    out = capsys.readouterr().out
    assert "[DIFF] ABC" in out
    assert "\x1b[32m" in out
    assert "\x1b[31m" in out or "\x1b[37m" in out

    week10_path = tmp_path / "jira_weekly_email_ABC_26'w10.html"
    assert week10_path.exists()
    week10_text = week10_path.read_text(encoding="utf-8")
    assert "(ABC-1)" in week10_text
    assert "Report items" not in week10_text
    assert "Feature delivery (ABC-1)" in week10_text
    assert "Finished" in week10_text
    assert "Task completion:" not in week10_text
    assert "Parent task (ABC-2):" in week10_text
    assert "Subtask done - Done (ABC-4)" in week10_text
    assert "Subtask completed." in week10_text
    assert "High priority focus:" not in week10_text
    assert "Work in progress for next week." in week10_text
    assert "Plan: Work in progress for next week." not in week10_text
    assert "Continue implementation:" not in week10_text
    assert "Other completed work" not in week10_text
    assert "Epic Two" not in week10_text

    html_path = tmp_path / "jira_weekly_email_ABC_26'w11.html"
    assert html_path.exists()
    text = html_path.read_text(encoding="utf-8")
    assert "Highlights" in text
    assert "Key Results and Achievements" in text
    assert "(ABC-5)" in text
    assert "[ABC-5]" not in text
    assert "Changes vs Previous Week" not in text

    snapshot_path = tmp_path / "jira_weekly_email_ABC_26'w11.json"
    assert snapshot_path.exists()


@patch("stats_core.reports.jira_weekly_email.JiraSource")
def test_jira_weekly_email_highlights_include_finished_progress_and_no_progress(mock_jira_source_cls, tmp_path: Path):
    issues = [
        _make_issue(
            "ABC-90",
            summary="Finished highlight",
            issue_type="Task",
            status="Done",
            resolution="Done",
            labels=["shine"],
            priority="Medium",
            comment_body="Done and closed.",
        ),
        _make_issue(
            "ABC-91",
            summary="Active highlight",
            issue_type="Task",
            status="In Progress",
            resolution="",
            labels=["shine"],
            priority="Medium",
            comment_body="Implemented API endpoint.",
        ),
        _make_issue(
            "ABC-911",
            summary="Active subtask in progress",
            issue_type="Sub-task",
            status="In Progress",
            resolution="",
            labels=[],
            priority="Medium",
            parent_key="ABC-91",
            comment_body="Backend integration in progress.",
        ),
        _make_issue(
            "ABC-912",
            summary="Active subtask done",
            issue_type="Sub-task",
            status="Done",
            resolution="Done",
            labels=[],
            priority="Medium",
            parent_key="ABC-91",
            comment_body="Closed.",
        ),
        _make_issue(
            "ABC-92",
            summary="Stalled highlight",
            issue_type="Task",
            status="In Progress",
            resolution="",
            labels=["shine"],
            priority="Medium",
            comment_created="2026-02-01T10:00:00.000+0000",
            comment_body="Old note.",
        ),
    ]

    def _fake_search_issues(jql, *args, **kwargs):
        if "issuekey in (" in str(jql):
            return [_make_epic_issue("EPIC-1", "Epic One", ["reportx"])]
        return issues

    fake_jira = Mock()
    fake_jira.search_issues.side_effect = _fake_search_issues
    fake_source = Mock()
    fake_source.jira = fake_jira
    fake_source.fetch_epic_names.return_value = {"EPIC-1": "Epic One"}
    mock_jira_source_cls.return_value = fake_source

    config = ConfigParser()
    config.read_dict(
        {
            "jira": {"jira-url": "https://jira.example.com", "username": "u", "password": "p"},
            "reporting": {"output_dir": str(tmp_path)},
            "ollama": {"enabled": "false"},
            "jira_weekly_email": {"labels_report": "reportx", "labels_highlights": "shine"},
        }
    )

    report = JiraWeeklyEmailReport()
    report.run(
        dataset={},
        config=config,
        output_formats=["html"],
        extra_params={
            "project": "ABC",
            "week_date": "2026-03-03",
            "labels_report": "reportx",
            "labels_highlights": "shine",
            "output_dir": str(tmp_path),
        },
    )

    html_path = tmp_path / "jira_weekly_email_ABC_26'w10.html"
    text = html_path.read_text(encoding="utf-8")
    assert "Finished highlight - Finished this week. (ABC-90)" in text
    assert (
        "Active highlight - Progress: Implemented API endpoint.; Active subtask in progress: "
        "Backend integration in progress.; Active subtask done (ABC-91)"
    ) in text
    assert "Stalled highlight - No progress this week. (ABC-92)" in text


@patch("stats_core.reports.jira_weekly_email.JiraSource")
def test_jira_weekly_email_excludes_unknown_epic_when_labels_report_scoped(mock_jira_source_cls, tmp_path: Path):
    issues = [
        _make_issue(
            "ABC-93",
            summary="Scoped without epic",
            issue_type="Task",
            status="Done",
            resolution="Done",
            labels=["reportx"],
            priority="Medium",
            epic_link="",
            comment_body="Done this week.",
        ),
        _make_issue(
            "ABC-94",
            summary="Scoped with epic",
            issue_type="Task",
            status="Done",
            resolution="Done",
            labels=["reportx"],
            priority="Medium",
            epic_link="EPIC-1",
            comment_body="Done this week.",
        ),
    ]

    def _fake_search_issues(jql, *args, **kwargs):
        if "issuekey in (" in str(jql):
            return [_make_epic_issue("EPIC-1", "Epic One", ["reportx"])]
        return issues

    fake_jira = Mock()
    fake_jira.search_issues.side_effect = _fake_search_issues
    fake_source = Mock()
    fake_source.jira = fake_jira
    fake_source.fetch_epic_names.return_value = {"EPIC-1": "Epic One"}
    mock_jira_source_cls.return_value = fake_source

    config = ConfigParser()
    config.read_dict(
        {
            "jira": {"jira-url": "https://jira.example.com", "username": "u", "password": "p"},
            "reporting": {"output_dir": str(tmp_path)},
            "ollama": {"enabled": "false"},
            "jira_weekly_email": {"labels_report": "reportx"},
        }
    )

    report = JiraWeeklyEmailReport()
    report.run(
        dataset={},
        config=config,
        output_formats=["html"],
        extra_params={
            "project": "ABC",
            "week_date": "2026-03-03",
            "labels_report": "reportx",
            "output_dir": str(tmp_path),
        },
    )

    html_path = tmp_path / "jira_weekly_email_ABC_26'w10.html"
    text = html_path.read_text(encoding="utf-8")
    assert "Unknown Epic" not in text
    assert "Epic One (EPIC-1)" in text


@patch("stats_core.reports.jira_weekly_email.JiraSource")
def test_jira_weekly_email_omits_empty_high_priority_and_bugs(mock_jira_source_cls, tmp_path: Path):
    issues = [
        _make_issue(
            "ABC-10",
            summary="Report task done",
            issue_type="Task",
            status="Done",
            resolution="Done",
            labels=["reportx"],
            priority="Medium",
            comment_body="Done this week.",
        )
    ]

    fake_jira = Mock()
    fake_jira.search_issues.return_value = issues
    fake_source = Mock()
    fake_source.jira = fake_jira
    fake_source.fetch_epic_names.return_value = {"EPIC-1": "Epic One"}
    mock_jira_source_cls.return_value = fake_source

    config = ConfigParser()
    config.read_dict(
        {
            "jira": {"jira-url": "https://jira.example.com", "username": "u", "password": "p"},
            "reporting": {"output_dir": str(tmp_path)},
            "ollama": {"enabled": "false"},
            "jira_weekly_email": {"labels_report": "reportx"},
        }
    )

    report = JiraWeeklyEmailReport()
    report.run(
        dataset={},
        config=config,
        output_formats=["html"],
        extra_params={
            "project": "ABC",
            "week_date": "2026-03-03",
            "labels_report": "reportx",
            "output_dir": str(tmp_path),
        },
    )

    html_path = tmp_path / "jira_weekly_email_ABC_26'w10.html"
    text = html_path.read_text(encoding="utf-8")
    assert "High priority items" not in text
    assert "Bugs summary" not in text


@patch("stats_core.reports.jira_weekly_email.JiraSource")
def test_jira_weekly_email_bugs_summary_uses_project_level_in_progress_and_open(mock_jira_source_cls, tmp_path: Path):
    weekly_issues = [
        _make_issue(
            "ABC-15",
            summary="Closed high bug",
            issue_type="Bug",
            status="Done",
            resolution="Done",
            labels=["reportx"],
            priority="High",
            comment_body="Fixed this week.",
        )
    ]
    in_progress_bug = _make_issue(
        "ABC-16",
        summary="Background bug in progress",
        issue_type="Bug",
        status="In Progress",
        resolution="",
        labels=[],
        priority="Low",
        comment_body="Still in progress.",
    )
    open_bug = _make_issue(
        "ABC-17",
        summary="Open bug backlog item",
        issue_type="Bug",
        status="To Do",
        resolution="",
        labels=[],
        priority="Low",
        comment_body="Open.",
    )

    fake_jira = Mock()

    def _fake_search_issues(jql, *args, **kwargs):
        jql_text = str(jql)
        if "issuekey in (EPIC-1)" in jql_text:
            return [_make_epic_issue("EPIC-1", "Epic One", ["reportx"])]
        if "statusCategory = 'In Progress'" in jql_text:
            return [in_progress_bug]
        if "resolution = Unresolved" in jql_text:
            return [in_progress_bug, open_bug]
        return weekly_issues

    fake_jira.search_issues.side_effect = _fake_search_issues
    fake_source = Mock()
    fake_source.jira = fake_jira
    fake_source.fetch_epic_names.return_value = {"EPIC-1": "Epic One"}
    mock_jira_source_cls.return_value = fake_source

    config = ConfigParser()
    config.read_dict(
        {
            "jira": {"jira-url": "https://jira.example.com", "username": "u", "password": "p"},
            "reporting": {"output_dir": str(tmp_path)},
            "ollama": {"enabled": "false"},
            "jira_weekly_email": {"labels_report": "reportx"},
        }
    )

    report = JiraWeeklyEmailReport()
    report.run(
        dataset={},
        config=config,
        output_formats=["html"],
        extra_params={
            "project": "ABC",
            "week_date": "2026-03-03",
            "labels_report": "reportx",
            "output_dir": str(tmp_path),
        },
    )

    html_path = tmp_path / "jira_weekly_email_ABC_26'w10.html"
    text = html_path.read_text(encoding="utf-8")
    assert "Bugs summary" in text
    assert "1 trouble reports/issues are analyzed and closed, 1 currently in progress, 2 open in project." in text


@patch("stats_core.reports.jira_weekly_email.JiraSource")
def test_jira_weekly_email_high_priority_has_comment_and_no_duplicate_in_results(mock_jira_source_cls, tmp_path: Path):
    issues = [
        _make_issue(
            "ABC-74",
            summary="High priority task",
            issue_type="Task",
            status="Done",
            resolution="Done",
            labels=["reportx"],
            priority="High",
            epic_link="EPIC-1",
            comment_body="Done.",
        )
    ]
    fake_jira = Mock()

    def _fake_search_issues(jql, *args, **kwargs):
        jql_text = str(jql)
        if "issuekey in (EPIC-1)" in jql_text:
            return [_make_epic_issue("EPIC-1", "Epic One", ["reportx"])]
        if "statusCategory = 'In Progress'" in jql_text:
            return []
        if "resolution = Unresolved" in jql_text:
            return []
        return issues

    fake_jira.search_issues.side_effect = _fake_search_issues
    fake_source = Mock()
    fake_source.jira = fake_jira
    fake_source.fetch_epic_names.return_value = {"EPIC-1": "Epic One"}
    mock_jira_source_cls.return_value = fake_source

    config = ConfigParser()
    config.read_dict(
        {
            "jira": {"jira-url": "https://jira.example.com", "username": "u", "password": "p"},
            "reporting": {"output_dir": str(tmp_path)},
            "ollama": {"enabled": "false"},
            "jira_weekly_email": {"labels_report": "reportx"},
        }
    )

    report = JiraWeeklyEmailReport()
    report.run(
        dataset={},
        config=config,
        output_formats=["html"],
        extra_params={
            "project": "ABC",
            "week_date": "2026-03-03",
            "labels_report": "reportx",
            "output_dir": str(tmp_path),
        },
    )

    html_path = tmp_path / "jira_weekly_email_ABC_26'w10.html"
    text = html_path.read_text(encoding="utf-8")
    assert "High priority items" in text
    assert text.count("High priority task (ABC-74)") == 1
    assert "Done." in text


@patch("stats_core.reports.jira_weekly_email.JiraSource")
def test_jira_weekly_email_high_priority_subtask_not_duplicated(mock_jira_source_cls, tmp_path: Path):
    issues = [
        _make_issue(
            "ABC-901",
            summary="High priority subtask",
            issue_type="Sub-task",
            status="Done",
            resolution="Done",
            labels=["reportx"],
            priority="High",
            epic_link="",
            parent_key="ABC-900",
            comment_body="Subtask comment.",
            issue_is_subtask=True,
        )
    ]
    fake_jira = Mock()

    def _fake_search_issues(jql, *args, **kwargs):
        jql_text = str(jql)
        if "issuekey in (EPIC-1)" in jql_text:
            return [_make_epic_issue("EPIC-1", "Epic One", ["reportx"])]
        if "issuekey in (ABC-900)" in jql_text:
            return [_make_parent_issue("ABC-900", summary="Parent task", status="Done", resolution="Done", epic_link="EPIC-1")]
        if "statusCategory = 'In Progress'" in jql_text:
            return []
        if "resolution = Unresolved" in jql_text:
            return []
        return issues

    fake_jira.search_issues.side_effect = _fake_search_issues
    fake_source = Mock()
    fake_source.jira = fake_jira
    fake_source.fetch_epic_names.return_value = {"EPIC-1": "Epic One"}
    mock_jira_source_cls.return_value = fake_source

    config = ConfigParser()
    config.read_dict(
        {
            "jira": {"jira-url": "https://jira.example.com", "username": "u", "password": "p"},
            "reporting": {"output_dir": str(tmp_path)},
            "ollama": {"enabled": "false"},
            "jira_weekly_email": {"labels_report": "reportx"},
        }
    )

    report = JiraWeeklyEmailReport()
    report.run(
        dataset={},
        config=config,
        output_formats=["html"],
        extra_params={
            "project": "ABC",
            "week_date": "2026-03-03",
            "labels_report": "reportx",
            "output_dir": str(tmp_path),
        },
    )

    html_path = tmp_path / "jira_weekly_email_ABC_26'w10.html"
    text = html_path.read_text(encoding="utf-8")
    assert text.count("High priority subtask (ABC-901)") == 1
    assert "Subtask comment." in text


@patch("stats_core.reports.jira_weekly_email.JiraSource")
def test_jira_weekly_email_high_priority_parent_not_duplicated_as_parent_group(mock_jira_source_cls, tmp_path: Path):
    issues = [
        _make_issue(
            "ABC-920",
            summary="High parent task",
            issue_type="Task",
            status="In Progress",
            resolution="",
            labels=["reportx"],
            priority="High",
            epic_link="EPIC-1",
            comment_body="Parent progress.",
        ),
        _make_issue(
            "ABC-921",
            summary="Regular child subtask",
            issue_type="Sub-task",
            status="In Progress",
            resolution="",
            labels=[],
            priority="Medium",
            epic_link="",
            parent_key="ABC-920",
            comment_body="Child progress.",
            issue_is_subtask=True,
        ),
    ]
    fake_jira = Mock()

    def _fake_search_issues(jql, *args, **kwargs):
        jql_text = str(jql)
        if "issuekey in (EPIC-1)" in jql_text:
            return [_make_epic_issue("EPIC-1", "Epic One", ["reportx"])]
        if "statusCategory = 'In Progress'" in jql_text:
            return []
        if "resolution = Unresolved" in jql_text:
            return []
        return issues

    fake_jira.search_issues.side_effect = _fake_search_issues
    fake_source = Mock()
    fake_source.jira = fake_jira
    fake_source.fetch_epic_names.return_value = {"EPIC-1": "Epic One"}
    mock_jira_source_cls.return_value = fake_source

    config = ConfigParser()
    config.read_dict(
        {
            "jira": {"jira-url": "https://jira.example.com", "username": "u", "password": "p"},
            "reporting": {"output_dir": str(tmp_path)},
            "ollama": {"enabled": "false"},
            "jira_weekly_email": {"labels_report": "reportx"},
        }
    )

    report = JiraWeeklyEmailReport()
    report.run(
        dataset={},
        config=config,
        output_formats=["html"],
        extra_params={
            "project": "ABC",
            "week_date": "2026-03-03",
            "labels_report": "reportx",
            "output_dir": str(tmp_path),
        },
    )

    html_path = tmp_path / "jira_weekly_email_ABC_26'w10.html"
    text = html_path.read_text(encoding="utf-8")
    results_idx = text.find("Key Results")
    plans_idx = text.find("Next Week Plans")
    assert results_idx != -1 and plans_idx != -1 and plans_idx > results_idx
    results_text = text[results_idx:plans_idx]
    assert results_text.count("<li>High parent task (ABC-920) - In Progress</li>") == 1


@patch("stats_core.reports.jira_weekly_email.JiraSource")
def test_jira_weekly_email_results_status_renders_inline_not_as_nested_line(mock_jira_source_cls, tmp_path: Path):
    issues = [
        _make_issue(
            "ABC-930",
            summary="Standalone progress task",
            issue_type="Task",
            status="In Progress",
            resolution="",
            labels=["reportx"],
            priority="Medium",
            epic_link="EPIC-1",
            comment_body="Progress note.",
        ),
    ]
    fake_jira = Mock()

    def _fake_search_issues(jql, *args, **kwargs):
        jql_text = str(jql)
        if "issuekey in (EPIC-1)" in jql_text:
            return [_make_epic_issue("EPIC-1", "Epic One", ["reportx"])]
        if "statusCategory = 'In Progress'" in jql_text:
            return []
        if "resolution = Unresolved" in jql_text:
            return []
        return issues

    fake_jira.search_issues.side_effect = _fake_search_issues
    fake_source = Mock()
    fake_source.jira = fake_jira
    fake_source.fetch_epic_names.return_value = {"EPIC-1": "Epic One"}
    mock_jira_source_cls.return_value = fake_source

    config = ConfigParser()
    config.read_dict(
        {
            "jira": {"jira-url": "https://jira.example.com", "username": "u", "password": "p"},
            "reporting": {"output_dir": str(tmp_path)},
            "ollama": {"enabled": "false"},
            "jira_weekly_email": {"labels_report": "reportx"},
        }
    )

    report = JiraWeeklyEmailReport()
    report.run(
        dataset={},
        config=config,
        output_formats=["html"],
        extra_params={
            "project": "ABC",
            "week_date": "2026-03-03",
            "labels_report": "reportx",
            "output_dir": str(tmp_path),
        },
    )

    html_path = tmp_path / "jira_weekly_email_ABC_26'w10.html"
    text = html_path.read_text(encoding="utf-8")
    results_idx = text.find("Key Results")
    plans_idx = text.find("Next Week Plans")
    assert results_idx != -1 and plans_idx != -1 and plans_idx > results_idx
    results_text = text[results_idx:plans_idx]
    assert "Standalone progress task (ABC-930) - In Progress" in results_text
    assert "<li>In Progress</li>" not in results_text
    assert "Progress note." in results_text


@patch("stats_core.reports.jira_weekly_email.JiraSource")
def test_jira_weekly_email_labels_report_all_includes_any_labels(mock_jira_source_cls, tmp_path: Path):
    issues = [
        _make_issue(
            "ABC-20",
            summary="Done in epic one",
            issue_type="Task",
            status="Done",
            resolution="Done",
            labels=["foo"],
            priority="Medium",
            epic_link="EPIC-1",
            comment_body="Done.",
        ),
        _make_issue(
            "ABC-21",
            summary="Done in epic two",
            issue_type="Task",
            status="Done",
            resolution="Done",
            labels=[],
            priority="Medium",
            epic_link="EPIC-2",
            comment_body="Done.",
        ),
    ]

    fake_jira = Mock()
    fake_jira.search_issues.return_value = issues
    fake_source = Mock()
    fake_source.jira = fake_jira
    fake_source.fetch_epic_names.return_value = {"EPIC-1": "Epic One", "EPIC-2": "Epic Two"}
    mock_jira_source_cls.return_value = fake_source

    config = ConfigParser()
    config.read_dict(
        {
            "jira": {"jira-url": "https://jira.example.com", "username": "u", "password": "p"},
            "reporting": {"output_dir": str(tmp_path)},
            "ollama": {"enabled": "false"},
            "jira_weekly_email": {"labels_report": "@all"},
        }
    )

    report = JiraWeeklyEmailReport()
    report.run(
        dataset={},
        config=config,
        output_formats=["html"],
        extra_params={
            "project": "ABC",
            "week_date": "2026-03-03",
            "labels_report": "@all",
            "output_dir": str(tmp_path),
        },
    )

    html_path = tmp_path / "jira_weekly_email_ABC_26'w10.html"
    text = html_path.read_text(encoding="utf-8")
    assert "Epic One (EPIC-1)" in text
    assert "Epic Two (EPIC-2)" in text


@patch("stats_core.reports.jira_weekly_email.JiraSource")
def test_jira_weekly_email_includes_epic_by_epic_label_not_issue_label(mock_jira_source_cls, tmp_path: Path):
    issues = [
        _make_issue(
            "ABC-30",
            summary="Done work",
            issue_type="Task",
            status="Done",
            resolution="Done",
            labels=[],
            priority="Medium",
            epic_link="EPIC-9",
            comment_body="Done.",
        ),
    ]

    fake_jira = Mock()

    def _fake_search_issues(jql, *args, **kwargs):
        if "issuekey in (" in str(jql):
            return [_make_epic_issue("EPIC-9", "Epic Nine", ["reportx"])]
        return issues

    fake_jira.search_issues.side_effect = _fake_search_issues
    fake_source = Mock()
    fake_source.jira = fake_jira
    fake_source.fetch_epic_names.return_value = {"EPIC-9": "Epic Nine"}
    mock_jira_source_cls.return_value = fake_source

    config = ConfigParser()
    config.read_dict(
        {
            "jira": {"jira-url": "https://jira.example.com", "username": "u", "password": "p"},
            "reporting": {"output_dir": str(tmp_path)},
            "ollama": {"enabled": "false"},
            "jira_weekly_email": {"labels_report": "reportx"},
        }
    )

    report = JiraWeeklyEmailReport()
    report.run(
        dataset={},
        config=config,
        output_formats=["html"],
        extra_params={
            "project": "ABC",
            "week_date": "2026-03-03",
            "labels_report": "reportx",
            "output_dir": str(tmp_path),
        },
    )

    html_path = tmp_path / "jira_weekly_email_ABC_26'w10.html"
    text = html_path.read_text(encoding="utf-8")
    assert "Epic Nine (EPIC-9)" in text


@patch("stats_core.reports.jira_weekly_email.JiraSource")
def test_jira_weekly_email_uses_parent_summary_when_parent_not_in_main_query(mock_jira_source_cls, tmp_path: Path):
    issues = [
        _make_issue(
            "ABC-40",
            summary="Subtask delivered",
            issue_type="Sub-task",
            status="Done",
            resolution="Done",
            labels=[],
            priority="Medium",
            epic_link="",
            parent_key="ABC-41",
            comment_body="Subtask delivered.",
        ),
    ]

    fake_jira = Mock()

    def _fake_search_issues(jql, *args, **kwargs):
        jql_text = str(jql)
        if "issuekey in (ABC-41)" in jql_text:
            return [_make_parent_issue("ABC-41", summary="Parent headline", epic_link="EPIC-1")]
        if "issuekey in (EPIC-1)" in jql_text:
            return [_make_epic_issue("EPIC-1", "Epic One", ["reportx"])]
        return issues

    fake_jira.search_issues.side_effect = _fake_search_issues
    fake_source = Mock()
    fake_source.jira = fake_jira
    fake_source.fetch_epic_names.return_value = {"EPIC-1": "Epic One"}
    mock_jira_source_cls.return_value = fake_source

    config = ConfigParser()
    config.read_dict(
        {
            "jira": {"jira-url": "https://jira.example.com", "username": "u", "password": "p"},
            "reporting": {"output_dir": str(tmp_path)},
            "ollama": {"enabled": "false"},
            "jira_weekly_email": {"labels_report": "reportx"},
        }
    )

    report = JiraWeeklyEmailReport()
    report.run(
        dataset={},
        config=config,
        output_formats=["html"],
        extra_params={
            "project": "ABC",
            "week_date": "2026-03-03",
            "labels_report": "reportx",
            "output_dir": str(tmp_path),
        },
    )

    html_path = tmp_path / "jira_weekly_email_ABC_26'w10.html"
    text = html_path.read_text(encoding="utf-8")
    assert "Parent headline (ABC-41):" in text
    assert "Parent task (ABC-41):" not in text


@patch("stats_core.reports.jira_weekly_email.JiraSource")
def test_jira_weekly_email_vacation_file_accepts_quoted_absolute_path(mock_jira_source_cls, tmp_path: Path):
    path = tmp_path / "vacations.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Vacations2026"
    ws.cell(row=3, column=2).value = "Name"
    ws.cell(row=3, column=6).value = datetime(2026, 3, 1)
    ws.cell(row=3, column=7).value = datetime(2026, 3, 2)
    ws.cell(row=5, column=2).value = "Denis Mazur"
    ws.cell(row=5, column=6).value = "p"
    ws.cell(row=5, column=7).value = "p"
    wb.save(path)

    issues = [
        _make_issue(
            "ABC-50",
            summary="Done report item",
            issue_type="Task",
            status="Done",
            resolution="Done",
            labels=[],
            priority="Medium",
            epic_link="EPIC-1",
            comment_body="Done.",
        ),
    ]
    fake_jira = Mock()

    def _fake_search_issues(jql, *args, **kwargs):
        if "issuekey in (EPIC-1)" in str(jql):
            return [_make_epic_issue("EPIC-1", "Epic One", ["reportx"])]
        return issues

    fake_jira.search_issues.side_effect = _fake_search_issues
    fake_source = Mock()
    fake_source.jira = fake_jira
    fake_source.fetch_epic_names.return_value = {"EPIC-1": "Epic One"}
    mock_jira_source_cls.return_value = fake_source

    config = ConfigParser()
    config.read_dict(
        {
            "jira": {"jira-url": "https://jira.example.com", "username": "u", "password": "p"},
            "reporting": {"output_dir": str(tmp_path)},
            "ollama": {"enabled": "false"},
            "jira_weekly_email": {"labels_report": "reportx"},
        }
    )

    report = JiraWeeklyEmailReport()
    report.run(
        dataset={},
        config=config,
        output_formats=["html"],
        extra_params={
            "project": "ABC",
            "week_date": "2026-03-03",
            "labels_report": "reportx",
            "vacation_file": f"\"{path}\"",
            "vacation_sheet": "Vacations2026",
            "vacation_horizon_days": "30",
            "vacation_horizon_anchor": "week_start",
            "output_dir": str(tmp_path),
        },
    )

    html_path = tmp_path / "jira_weekly_email_ABC_26'w10.html"
    text = html_path.read_text(encoding="utf-8")
    assert "Denis Mazur vacation 01.03.2026 - 02.03.2026" in text


@patch("stats_core.reports.jira_weekly_email.JiraSource")
def test_jira_weekly_email_closed_subtask_not_added_to_plans(mock_jira_source_cls, tmp_path: Path):
    issues = [
        _make_issue(
            "ABC-60",
            summary="Closed subtask",
            issue_type="Sub-task",
            status="Done",
            resolution="Done",
            labels=[],
            priority="Medium",
            epic_link="",
            parent_key="ABC-61",
            comment_body="Subtask is done.",
        ),
    ]
    fake_jira = Mock()

    def _fake_search_issues(jql, *args, **kwargs):
        jql_text = str(jql)
        if "issuekey in (ABC-61)" in jql_text:
            return [_make_parent_issue("ABC-61", summary="Open parent", status="In Progress", epic_link="EPIC-1")]
        if "issuekey in (EPIC-1)" in jql_text:
            return [_make_epic_issue("EPIC-1", "Epic One", ["reportx"])]
        return issues

    fake_jira.search_issues.side_effect = _fake_search_issues
    fake_source = Mock()
    fake_source.jira = fake_jira
    fake_source.fetch_epic_names.return_value = {"EPIC-1": "Epic One"}
    mock_jira_source_cls.return_value = fake_source

    config = ConfigParser()
    config.read_dict(
        {
            "jira": {"jira-url": "https://jira.example.com", "username": "u", "password": "p"},
            "reporting": {"output_dir": str(tmp_path)},
            "ollama": {"enabled": "false"},
            "jira_weekly_email": {"labels_report": "reportx"},
        }
    )

    report = JiraWeeklyEmailReport()
    report.run(
        dataset={},
        config=config,
        output_formats=["html"],
        extra_params={
            "project": "ABC",
            "week_date": "2026-03-03",
            "labels_report": "reportx",
            "output_dir": str(tmp_path),
        },
    )

    html_path = tmp_path / "jira_weekly_email_ABC_26'w10.html"
    text = html_path.read_text(encoding="utf-8")
    assert "Closed subtask - Done (ABC-60)" in text
    assert "No in-progress plans collected for next week." in text


@patch("stats_core.reports.jira_weekly_email.JiraSource")
def test_jira_weekly_email_plans_include_report_tasks_and_in_progress_subtasks(mock_jira_source_cls, tmp_path: Path):
    issues = [
        _make_issue(
            "ABC-70",
            summary="Parent report task",
            issue_type="Task",
            status="In Progress",
            resolution="",
            labels=["reportx"],
            priority="Medium",
            epic_link="EPIC-1",
            comment_body="Parent work continues.",
        ),
        _make_issue(
            "ABC-71",
            summary="Subtask in progress",
            issue_type="Sub-task",
            status="In Progress",
            resolution="",
            labels=[],
            priority="Medium",
            epic_link="",
            parent_key="ABC-70",
            comment_body="Subtask work continues.",
        ),
        _make_issue(
            "ABC-72",
            summary="Subtask done",
            issue_type="Sub-task",
            status="Done",
            resolution="Done",
            labels=[],
            priority="Medium",
            epic_link="",
            parent_key="ABC-70",
            comment_body="Subtask delivered.",
        ),
    ]
    fake_jira = Mock()

    def _fake_search_issues(jql, *args, **kwargs):
        if "issuekey in (ABC-70)" in str(jql):
            return [_make_parent_issue("ABC-70", summary="Parent report task", status="In Progress", epic_link="EPIC-1")]
        if "issuekey in (EPIC-1)" in str(jql):
            return [_make_epic_issue("EPIC-1", "Epic One", ["reportx"])]
        return issues

    fake_jira.search_issues.side_effect = _fake_search_issues
    fake_source = Mock()
    fake_source.jira = fake_jira
    fake_source.fetch_epic_names.return_value = {"EPIC-1": "Epic One"}
    mock_jira_source_cls.return_value = fake_source

    config = ConfigParser()
    config.read_dict(
        {
            "jira": {"jira-url": "https://jira.example.com", "username": "u", "password": "p"},
            "reporting": {"output_dir": str(tmp_path)},
            "ollama": {"enabled": "false"},
            "jira_weekly_email": {"labels_report": "reportx"},
        }
    )

    report = JiraWeeklyEmailReport()
    report.run(
        dataset={},
        config=config,
        output_formats=["html"],
        extra_params={
            "project": "ABC",
            "week_date": "2026-03-03",
            "labels_report": "reportx",
            "output_dir": str(tmp_path),
        },
    )

    html_path = tmp_path / "jira_weekly_email_ABC_26'w10.html"
    text = html_path.read_text(encoding="utf-8")
    assert "Parent report task (ABC-70)" in text
    assert "Subtask in progress - In Progress (ABC-71)" in text
    assert "Subtask done - Done (ABC-72)" in text
    results_idx = text.index("Key Results and Achievements")
    plans_idx = text.index("Next Week Plans")
    results_text = text[results_idx:plans_idx]
    assert results_text.count("Parent report task (ABC-70)") == 1


@patch("stats_core.reports.jira_weekly_email.JiraSource")
def test_jira_weekly_email_excludes_non_in_progress_items_from_plans_and_progress_results(
    mock_jira_source_cls,
    tmp_path: Path,
):
    issues = [
        _make_issue(
            "ABC-90",
            summary="Feature in todo",
            issue_type="Feature",
            status="To Do",
            resolution="",
            labels=[],
            priority="Medium",
            epic_link="EPIC-1",
            comment_body="Todo stage update.",
        ),
        _make_issue(
            "ABC-91",
            summary="Subtask in todo",
            issue_type="Sub-task",
            status="To Do",
            resolution="",
            labels=[],
            priority="Medium",
            epic_link="",
            parent_key="ABC-92",
            comment_body="Subtask todo stage update.",
        ),
    ]
    fake_jira = Mock()

    def _fake_search_issues(jql, *args, **kwargs):
        jql_text = str(jql)
        if "issuekey in (ABC-92)" in jql_text:
            return [_make_parent_issue("ABC-92", summary="Parent task", status="In Progress", epic_link="EPIC-1")]
        if "issuekey in (EPIC-1)" in jql_text:
            return [_make_epic_issue("EPIC-1", "Epic One", ["reportx"])]
        return issues

    fake_jira.search_issues.side_effect = _fake_search_issues
    fake_source = Mock()
    fake_source.jira = fake_jira
    fake_source.fetch_epic_names.return_value = {"EPIC-1": "Epic One"}
    mock_jira_source_cls.return_value = fake_source

    config = ConfigParser()
    config.read_dict(
        {
            "jira": {"jira-url": "https://jira.example.com", "username": "u", "password": "p"},
            "reporting": {"output_dir": str(tmp_path)},
            "ollama": {"enabled": "false"},
            "jira_weekly_email": {"labels_report": "reportx"},
        }
    )

    report = JiraWeeklyEmailReport()
    report.run(
        dataset={},
        config=config,
        output_formats=["html"],
        extra_params={
            "project": "ABC",
            "week_date": "2026-03-03",
            "labels_report": "reportx",
            "output_dir": str(tmp_path),
        },
    )

    html_path = tmp_path / "jira_weekly_email_ABC_26'w10.html"
    text = html_path.read_text(encoding="utf-8")
    assert "Feature in todo" not in text
    assert "Subtask in todo" not in text
    assert "Todo stage update." not in text
    assert "Subtask todo stage update." not in text


@patch("stats_core.reports.jira_weekly_email.JiraSource")
def test_jira_weekly_email_plans_include_in_progress_subtask_when_only_epic_is_report_scoped(
    mock_jira_source_cls,
    tmp_path: Path,
):
    issues = [
        _make_issue(
            "ABC-82",
            summary="Scoped by epic subtask",
            issue_type="Sub-task",
            status="In Progress",
            resolution="",
            labels=[],
            priority="Medium",
            epic_link="",
            parent_key="ABC-81",
            comment_body="Subtask progress this week.",
        ),
    ]
    fake_jira = Mock()

    def _fake_search_issues(jql, *args, **kwargs):
        jql_text = str(jql)
        if "issuekey in (ABC-81)" in jql_text:
            return [_make_parent_issue("ABC-81", summary="Parent without report label", status="In Progress", epic_link="EPIC-1")]
        if "issuekey in (EPIC-1)" in jql_text:
            return [_make_epic_issue("EPIC-1", "Epic One", ["reportx"])]
        return issues

    fake_jira.search_issues.side_effect = _fake_search_issues
    fake_source = Mock()
    fake_source.jira = fake_jira
    fake_source.fetch_epic_names.return_value = {"EPIC-1": "Epic One"}
    mock_jira_source_cls.return_value = fake_source

    config = ConfigParser()
    config.read_dict(
        {
            "jira": {"jira-url": "https://jira.example.com", "username": "u", "password": "p"},
            "reporting": {"output_dir": str(tmp_path)},
            "ollama": {"enabled": "false"},
            "jira_weekly_email": {"labels_report": "reportx"},
        }
    )

    report = JiraWeeklyEmailReport()
    report.run(
        dataset={},
        config=config,
        output_formats=["html"],
        extra_params={
            "project": "ABC",
            "week_date": "2026-03-03",
            "labels_report": "reportx",
            "output_dir": str(tmp_path),
        },
    )

    html_path = tmp_path / "jira_weekly_email_ABC_26'w10.html"
    text = html_path.read_text(encoding="utf-8")
    assert "Epic One (EPIC-1)" in text
    assert "Parent without report label (ABC-81)" in text
    assert "Scoped by epic subtask - In Progress (ABC-82)" in text
    assert "Subtask progress this week." in text
    results_idx = text.index("Key Results and Achievements")
    plans_idx = text.index("Next Week Plans")
    subtask_result_idx = text.find("Scoped by epic subtask - In Progress (ABC-82)", results_idx, plans_idx)
    assert subtask_result_idx != -1


@patch("stats_core.reports.jira_weekly_email.JiraSource")
def test_jira_weekly_email_resolves_epic_from_parent_chain_for_updated_subtask(
    mock_jira_source_cls,
    tmp_path: Path,
):
    issues = [
        _make_issue(
            "ABC-84",
            summary="Chain subtask",
            issue_type="Sub-task",
            status="In Progress",
            resolution="",
            labels=[],
            priority="Medium",
            epic_link="",
            parent_key="ABC-83",
            comment_body="Subtask chain progress.",
        ),
    ]
    fake_jira = Mock()

    def _fake_search_issues(jql, *args, **kwargs):
        jql_text = str(jql)
        if "issuekey in (ABC-83)" in jql_text:
            return [
                _make_parent_issue(
                    "ABC-83",
                    summary="Parent in chain",
                    status="In Progress",
                    epic_link="",
                    parent_key="EPIC-1",
                    issue_type="Story",
                    labels=[],
                )
            ]
        if "issuekey in (EPIC-1)" in jql_text:
            return [_make_epic_issue("EPIC-1", "Epic One", ["reportx"])]
        return issues

    fake_jira.search_issues.side_effect = _fake_search_issues
    fake_source = Mock()
    fake_source.jira = fake_jira
    fake_source.fetch_epic_names.return_value = {"EPIC-1": "Epic One"}
    mock_jira_source_cls.return_value = fake_source

    config = ConfigParser()
    config.read_dict(
        {
            "jira": {"jira-url": "https://jira.example.com", "username": "u", "password": "p"},
            "reporting": {"output_dir": str(tmp_path)},
            "ollama": {"enabled": "false"},
            "jira_weekly_email": {"labels_report": "reportx"},
        }
    )

    report = JiraWeeklyEmailReport()
    report.run(
        dataset={},
        config=config,
        output_formats=["html"],
        extra_params={
            "project": "ABC",
            "week_date": "2026-03-03",
            "labels_report": "reportx",
            "output_dir": str(tmp_path),
        },
    )

    html_path = tmp_path / "jira_weekly_email_ABC_26'w10.html"
    text = html_path.read_text(encoding="utf-8")
    assert "Epic One (EPIC-1)" in text
    assert "Parent in chain (ABC-83)" in text
    assert "Chain subtask - In Progress (ABC-84)" in text
    assert "Subtask chain progress." in text


@patch("stats_core.reports.jira_weekly_email.JiraSource")
def test_jira_weekly_email_includes_in_progress_subtask_even_if_parent_finished(
    mock_jira_source_cls,
    tmp_path: Path,
):
    issues = [
        _make_issue(
            "ABC-86",
            summary="In-progress subtask under finished parent",
            issue_type="Sub-task",
            status="In Progress",
            resolution="",
            labels=[],
            priority="Medium",
            epic_link="",
            parent_key="ABC-85",
            comment_body="Subtask moved forward this week.",
        ),
    ]
    fake_jira = Mock()

    def _fake_search_issues(jql, *args, **kwargs):
        jql_text = str(jql)
        if "issuekey in (ABC-85)" in jql_text:
            return [
                _make_parent_issue(
                    "ABC-85",
                    summary="Finished parent task",
                    status="Done",
                    resolution="Done",
                    epic_link="EPIC-1",
                    labels=[],
                )
            ]
        if "issuekey in (EPIC-1)" in jql_text:
            return [_make_epic_issue("EPIC-1", "Epic One", ["reportx"])]
        return issues

    fake_jira.search_issues.side_effect = _fake_search_issues
    fake_source = Mock()
    fake_source.jira = fake_jira
    fake_source.fetch_epic_names.return_value = {"EPIC-1": "Epic One"}
    mock_jira_source_cls.return_value = fake_source

    config = ConfigParser()
    config.read_dict(
        {
            "jira": {"jira-url": "https://jira.example.com", "username": "u", "password": "p"},
            "reporting": {"output_dir": str(tmp_path)},
            "ollama": {"enabled": "false"},
            "jira_weekly_email": {"labels_report": "reportx"},
        }
    )

    report = JiraWeeklyEmailReport()
    report.run(
        dataset={},
        config=config,
        output_formats=["html"],
        extra_params={
            "project": "ABC",
            "week_date": "2026-03-03",
            "labels_report": "reportx",
            "output_dir": str(tmp_path),
        },
    )

    html_path = tmp_path / "jira_weekly_email_ABC_26'w10.html"
    text = html_path.read_text(encoding="utf-8")
    assert "Epic One (EPIC-1)" in text
    assert "Finished parent task (ABC-85)" in text
    assert "In-progress subtask under finished parent - In Progress (ABC-86)" in text
    assert "Subtask moved forward this week." in text


@patch("stats_core.reports.jira_weekly_email.JiraSource")
def test_jira_weekly_email_includes_custom_subtask_type_by_subtask_flag(
    mock_jira_source_cls,
    tmp_path: Path,
):
    issues = [
        _make_issue(
            "ABC-88",
            summary="Custom subtask type in progress",
            issue_type="QA Subtask",
            issue_is_subtask=True,
            status="In Progress",
            resolution="",
            labels=[],
            priority="Medium",
            epic_link="",
            parent_key="ABC-87",
            comment_body="Custom subtask progress.",
        ),
    ]
    fake_jira = Mock()

    def _fake_search_issues(jql, *args, **kwargs):
        jql_text = str(jql)
        if "issuekey in (ABC-87)" in jql_text:
            return [_make_parent_issue("ABC-87", summary="Parent for custom subtask", status="In Progress", epic_link="EPIC-1")]
        if "issuekey in (EPIC-1)" in jql_text:
            return [_make_epic_issue("EPIC-1", "Epic One", ["reportx"])]
        return issues

    fake_jira.search_issues.side_effect = _fake_search_issues
    fake_source = Mock()
    fake_source.jira = fake_jira
    fake_source.fetch_epic_names.return_value = {"EPIC-1": "Epic One"}
    mock_jira_source_cls.return_value = fake_source

    config = ConfigParser()
    config.read_dict(
        {
            "jira": {"jira-url": "https://jira.example.com", "username": "u", "password": "p"},
            "reporting": {"output_dir": str(tmp_path)},
            "ollama": {"enabled": "false"},
            "jira_weekly_email": {"labels_report": "reportx"},
        }
    )

    report = JiraWeeklyEmailReport()
    report.run(
        dataset={},
        config=config,
        output_formats=["html"],
        extra_params={
            "project": "ABC",
            "week_date": "2026-03-03",
            "labels_report": "reportx",
            "output_dir": str(tmp_path),
        },
    )

    html_path = tmp_path / "jira_weekly_email_ABC_26'w10.html"
    text = html_path.read_text(encoding="utf-8")
    assert "Epic One (EPIC-1)" in text
    assert "Parent for custom subtask (ABC-87)" in text
    assert "Custom subtask type in progress - In Progress (ABC-88)" in text
    assert "Custom subtask progress." in text


@patch("stats_core.reports.jira_weekly_email.JiraSource")
def test_jira_weekly_email_plans_include_epic_scoped_in_progress_task_without_issue_label(
    mock_jira_source_cls,
    tmp_path: Path,
):
    issues = [
        _make_issue(
            "ABC-73",
            summary="Epic scoped task",
            issue_type="New feature",
            status="In Progress",
            resolution="",
            labels=[],
            priority="Medium",
            epic_link="EPIC-1",
            comment_body="Ongoing implementation update.",
        ),
    ]
    fake_jira = Mock()

    def _fake_search_issues(jql, *args, **kwargs):
        if "issuekey in (EPIC-1)" in str(jql):
            return [_make_epic_issue("EPIC-1", "Epic One", ["reportx"])]
        return issues

    fake_jira.search_issues.side_effect = _fake_search_issues
    fake_source = Mock()
    fake_source.jira = fake_jira
    fake_source.fetch_epic_names.return_value = {"EPIC-1": "Epic One"}
    mock_jira_source_cls.return_value = fake_source

    config = ConfigParser()
    config.read_dict(
        {
            "jira": {"jira-url": "https://jira.example.com", "username": "u", "password": "p"},
            "reporting": {"output_dir": str(tmp_path)},
            "ollama": {"enabled": "false"},
            "jira_weekly_email": {"labels_report": "reportx"},
        }
    )

    report = JiraWeeklyEmailReport()
    report.run(
        dataset={},
        config=config,
        output_formats=["html"],
        extra_params={
            "project": "ABC",
            "week_date": "2026-03-03",
            "labels_report": "reportx",
            "output_dir": str(tmp_path),
        },
    )

    html_path = tmp_path / "jira_weekly_email_ABC_26'w10.html"
    text = html_path.read_text(encoding="utf-8")
    assert "Epic One (EPIC-1)" in text
    assert "Epic scoped task (ABC-73)" in text
    assert "Ongoing implementation update." in text
    assert "In Progress" in text
    results_idx = text.index("Key Results and Achievements")
    plans_idx = text.index("Next Week Plans")
    issue_idx = text.find("Epic scoped task (ABC-73)", results_idx, plans_idx)
    assert issue_idx != -1


@patch("stats_core.reports.jira_weekly_email.JiraSource")
def test_jira_weekly_email_plans_include_arbitrary_non_bug_type(mock_jira_source_cls, tmp_path: Path):
    issues = [
        _make_issue(
            "ABC-89",
            summary="Research stream item",
            issue_type="Research",
            status="In Progress",
            resolution="",
            labels=[],
            priority="Medium",
            epic_link="EPIC-1",
            comment_body="Research progressed this week.",
        ),
    ]
    fake_jira = Mock()

    def _fake_search_issues(jql, *args, **kwargs):
        if "issuekey in (EPIC-1)" in str(jql):
            return [_make_epic_issue("EPIC-1", "Epic One", ["reportx"])]
        return issues

    fake_jira.search_issues.side_effect = _fake_search_issues
    fake_source = Mock()
    fake_source.jira = fake_jira
    fake_source.fetch_epic_names.return_value = {"EPIC-1": "Epic One"}
    mock_jira_source_cls.return_value = fake_source

    config = ConfigParser()
    config.read_dict(
        {
            "jira": {"jira-url": "https://jira.example.com", "username": "u", "password": "p"},
            "reporting": {"output_dir": str(tmp_path)},
            "ollama": {"enabled": "false"},
            "jira_weekly_email": {"labels_report": "reportx"},
        }
    )

    report = JiraWeeklyEmailReport()
    report.run(
        dataset={},
        config=config,
        output_formats=["html"],
        extra_params={
            "project": "ABC",
            "week_date": "2026-03-03",
            "labels_report": "reportx",
            "output_dir": str(tmp_path),
        },
    )

    html_path = tmp_path / "jira_weekly_email_ABC_26'w10.html"
    text = html_path.read_text(encoding="utf-8")
    assert "Epic One (EPIC-1)" in text
    assert "Research stream item (ABC-89)" in text
    assert "Research progressed this week." in text


@patch("stats_core.reports.jira_weekly_email.JiraSource")
def test_jira_weekly_email_excludes_closed_items_with_non_done_resolution(mock_jira_source_cls, tmp_path: Path):
    issues = [
        _make_issue(
            "ABC-97",
            summary="Completed valid",
            issue_type="Task",
            status="Done",
            resolution="Done",
            labels=[],
            priority="Medium",
            epic_link="EPIC-1",
            comment_body="Delivered as planned.",
        ),
        _make_issue(
            "ABC-98",
            summary="Closed not for report",
            issue_type="Task",
            status="Done",
            resolution="Won't Do",
            labels=[],
            priority="Medium",
            epic_link="EPIC-1",
            comment_body="Out of scope.",
        ),
    ]
    fake_jira = Mock()

    def _fake_search_issues(jql, *args, **kwargs):
        if "issuekey in (EPIC-1)" in str(jql):
            return [_make_epic_issue("EPIC-1", "Epic One", ["reportx"])]
        return issues

    fake_jira.search_issues.side_effect = _fake_search_issues
    fake_source = Mock()
    fake_source.jira = fake_jira
    fake_source.fetch_epic_names.return_value = {"EPIC-1": "Epic One"}
    mock_jira_source_cls.return_value = fake_source

    config = ConfigParser()
    config.read_dict(
        {
            "jira": {"jira-url": "https://jira.example.com", "username": "u", "password": "p"},
            "reporting": {"output_dir": str(tmp_path)},
            "ollama": {"enabled": "false"},
            "jira_weekly_email": {"labels_report": "reportx"},
        }
    )

    report = JiraWeeklyEmailReport()
    report.run(
        dataset={},
        config=config,
        output_formats=["html"],
        extra_params={
            "project": "ABC",
            "week_date": "2026-03-03",
            "labels_report": "reportx",
            "output_dir": str(tmp_path),
        },
    )

    html_path = tmp_path / "jira_weekly_email_ABC_26'w10.html"
    text = html_path.read_text(encoding="utf-8")
    assert "Completed valid (ABC-97)" in text
    assert "Finished" in text
    assert "Closed not for report" not in text


@patch("stats_core.reports.jira_weekly_email.JiraSource")
def test_jira_weekly_email_priority_high_values_respects_config_exactly(mock_jira_source_cls, tmp_path: Path):
    issues = [
        _make_issue(
            "ABC-74",
            summary="High priority task",
            issue_type="Task",
            status="Done",
            resolution="Done",
            labels=["reportx"],
            priority="High",
            epic_link="EPIC-1",
            comment_body="Done.",
        ),
        _make_issue(
            "ABC-75",
            summary="Highest priority task",
            issue_type="Task",
            status="Done",
            resolution="Done",
            labels=["reportx"],
            priority="Highest",
            epic_link="EPIC-1",
            comment_body="Done.",
        ),
    ]
    fake_jira = Mock()

    def _fake_search_issues(jql, *args, **kwargs):
        if "issuekey in (EPIC-1)" in str(jql):
            return [_make_epic_issue("EPIC-1", "Epic One", ["reportx"])]
        return issues

    fake_jira.search_issues.side_effect = _fake_search_issues
    fake_source = Mock()
    fake_source.jira = fake_jira
    fake_source.fetch_epic_names.return_value = {"EPIC-1": "Epic One"}
    mock_jira_source_cls.return_value = fake_source

    config = ConfigParser()
    config.read_dict(
        {
            "jira": {"jira-url": "https://jira.example.com", "username": "u", "password": "p"},
            "reporting": {"output_dir": str(tmp_path)},
            "ollama": {"enabled": "false"},
            "jira_weekly_email": {
                "labels_report": "reportx",
                "priority_high_values": "High",
            },
        }
    )

    report = JiraWeeklyEmailReport()
    report.run(
        dataset={},
        config=config,
        output_formats=["html"],
        extra_params={
            "project": "ABC",
            "week_date": "2026-03-03",
            "labels_report": "reportx",
            "output_dir": str(tmp_path),
        },
    )

    html_path = tmp_path / "jira_weekly_email_ABC_26'w10.html"
    text = html_path.read_text(encoding="utf-8")
    assert "High priority items" in text
    assert "High priority task (ABC-74)" in text
    assert text.count("High priority task (ABC-74)") == 1
    assert "Done." in text
    assert "Highest priority task (ABC-75)" in text
    high_section_start = text.index("<li><b>High priority items</b></li>")
    high_section_end = text.find("</ul><ul class='lvl2'>", high_section_start)
    assert high_section_end != -1
    high_section = text[high_section_start:high_section_end]
    assert "High priority task (ABC-74)" in high_section
    assert "Highest priority task (ABC-75)" not in high_section


@patch("stats_core.reports.jira_weekly_email.JiraSource")
def test_jira_weekly_email_uses_only_last_week_comment_for_progress(mock_jira_source_cls, tmp_path: Path):
    issue = _make_issue(
        "ABC-76",
        summary="Task with two updates",
        issue_type="Task",
        status="In Progress",
        resolution="",
        labels=["reportx"],
        priority="Medium",
        epic_link="EPIC-1",
        comment_created="2026-03-07T10:00:00.000+0000",
        comment_body="Final result delivered.",
    )
    issue.fields.comment.comments.append(
        SimpleNamespace(
            body="Started implementation.",
            created="2026-03-03T09:00:00.000+0000",
            id="c-ABC-76-2",
        )
    )
    issue.fields.comment.comments.append(
        SimpleNamespace(
            body="Second update completed.",
            created="2026-03-04T09:00:00.000+0000",
            id="c-ABC-76-3",
        )
    )
    issue.fields.comment.comments.append(
        SimpleNamespace(
            body="Old comment from previous week.",
            created="2026-02-28T09:00:00.000+0000",
            id="c-ABC-76-4",
        )
    )
    issues = [issue]
    fake_jira = Mock()

    def _fake_search_issues(jql, *args, **kwargs):
        if "issuekey in (EPIC-1)" in str(jql):
            return [_make_epic_issue("EPIC-1", "Epic One", ["reportx"])]
        return issues

    fake_jira.search_issues.side_effect = _fake_search_issues
    fake_source = Mock()
    fake_source.jira = fake_jira
    fake_source.fetch_epic_names.return_value = {"EPIC-1": "Epic One"}
    mock_jira_source_cls.return_value = fake_source

    config = ConfigParser()
    config.read_dict(
        {
            "jira": {"jira-url": "https://jira.example.com", "username": "u", "password": "p"},
            "reporting": {"output_dir": str(tmp_path)},
            "ollama": {"enabled": "false"},
            "jira_weekly_email": {"labels_report": "reportx"},
        }
    )

    report = JiraWeeklyEmailReport()
    report.run(
        dataset={},
        config=config,
        output_formats=["html"],
        extra_params={
            "project": "ABC",
            "week_date": "2026-03-03",
            "labels_report": "reportx",
            "output_dir": str(tmp_path),
        },
    )

    html_path = tmp_path / "jira_weekly_email_ABC_26'w10.html"
    text = html_path.read_text(encoding="utf-8")
    assert "Task with two updates (ABC-76)" in text
    assert "Final result delivered." in text
    assert "Started implementation." not in text
    assert "Second update completed." not in text
    assert "Old comment from previous week." not in text


@patch("stats_core.reports.jira_weekly_email.JiraSource")
def test_jira_weekly_email_diff_fallback_to_latest_previous_snapshot(mock_jira_source_cls, tmp_path: Path, capsys):
    week10_issues = [
        _make_issue(
            "ABC-80",
            summary="Week10 task",
            issue_type="Task",
            status="Done",
            resolution="Done",
            labels=["reportx"],
            priority="Medium",
            epic_link="EPIC-1",
            comment_body="Week10 complete.",
        ),
    ]
    week12_issues = week10_issues + [
        _make_issue(
            "ABC-81",
            summary="Week12 task",
            issue_type="Task",
            status="Done",
            resolution="Done",
            labels=["reportx"],
            priority="Medium",
            epic_link="EPIC-1",
            comment_created="2026-03-17T10:00:00.000+0000",
            comment_body="Week12 complete.",
        ),
    ]

    fake_jira = Mock()

    def _fake_search_issues(jql, *args, **kwargs):
        jql_text = str(jql)
        if "issuekey in (EPIC-1)" in jql_text:
            return [_make_epic_issue("EPIC-1", "Epic One", ["reportx"])]
        if "updated < '2026-03-09'" in jql_text:
            return week10_issues
        if "updated < '2026-03-23'" in jql_text:
            return week12_issues
        return []

    fake_jira.search_issues.side_effect = _fake_search_issues
    fake_source = Mock()
    fake_source.jira = fake_jira
    fake_source.fetch_epic_names.return_value = {"EPIC-1": "Epic One"}
    mock_jira_source_cls.return_value = fake_source

    config = ConfigParser()
    config.read_dict(
        {
            "jira": {"jira-url": "https://jira.example.com", "username": "u", "password": "p"},
            "reporting": {"output_dir": str(tmp_path)},
            "ollama": {"enabled": "false"},
            "jira_weekly_email": {"labels_report": "reportx"},
        }
    )

    report = JiraWeeklyEmailReport()
    report.run(
        dataset={},
        config=config,
        output_formats=["html"],
        extra_params={
            "project": "ABC",
            "week_date": "2026-03-03",  # w10
            "labels_report": "reportx",
            "output_dir": str(tmp_path),
        },
    )
    report.run(
        dataset={},
        config=config,
        output_formats=["html"],
        extra_params={
            "project": "ABC",
            "week_date": "2026-03-17",  # w12, no w11 snapshot
            "labels_report": "reportx",
            "output_dir": str(tmp_path),
        },
    )

    out = capsys.readouterr().out
    assert "[DIFF] ABC" in out


@patch("stats_core.reports.jira_weekly_email.JiraSource")
def test_jira_weekly_email_diff_uses_previous_snapshot_with_week_key_only(mock_jira_source_cls, tmp_path: Path, capsys):
    previous_payload = {
        "meta": {"project": "ABC", "week_key": "26'w10"},
        "highlights": [],
        "epics": [],
        "next_week_plans": [],
        "vacations": [],
        "titles": {},
    }
    previous_snapshot = {
        "meta": {"project": "ABC", "week_key": "26'w10"},
        "payload": previous_payload,
    }
    previous_snapshot_path = tmp_path / "jira_weekly_email_ABC_26'w10.json"
    previous_snapshot_path.write_text(json.dumps(previous_snapshot, ensure_ascii=False, indent=2), encoding="utf-8")

    issues = [
        _make_issue(
            "ABC-95",
            summary="Current week task",
            issue_type="Task",
            status="Done",
            resolution="Done",
            labels=["reportx"],
            priority="Medium",
            comment_body="Done this week.",
        ),
    ]

    def _fake_search_issues(jql, *args, **kwargs):
        if "issuekey in (" in str(jql):
            return [_make_epic_issue("EPIC-1", "Epic One", ["reportx"])]
        return issues

    fake_jira = Mock()
    fake_jira.search_issues.side_effect = _fake_search_issues
    fake_source = Mock()
    fake_source.jira = fake_jira
    fake_source.fetch_epic_names.return_value = {"EPIC-1": "Epic One"}
    mock_jira_source_cls.return_value = fake_source

    config = ConfigParser()
    config.read_dict(
        {
            "jira": {"jira-url": "https://jira.example.com", "username": "u", "password": "p"},
            "reporting": {"output_dir": str(tmp_path)},
            "ollama": {"enabled": "false"},
            "jira_weekly_email": {"labels_report": "reportx"},
        }
    )

    report = JiraWeeklyEmailReport()
    report.run(
        dataset={},
        config=config,
        output_formats=["html"],
        extra_params={
            "project": "ABC",
            "week_date": "2026-03-10",
            "labels_report": "reportx",
            "output_dir": str(tmp_path),
        },
    )

    out = capsys.readouterr().out
    assert "[DIFF] ABC" in out


def test_load_previous_snapshot_supports_utf8_sig_and_recursive_search(tmp_path: Path):
    snapshot_base = tmp_path / "snapshots_root"
    nested = snapshot_base / "archive" / "nested"
    nested.mkdir(parents=True)

    previous_snapshot = {
        "meta": {"project": "OHOSUI", "week_key": "26'w07"},
        "payload": {"meta": {"project": "OHOSUI", "week_key": "26'w07"}},
    }
    previous_path = nested / "jira_weekly_email_OHOSUI_26'w07.json"
    previous_path.write_text(json.dumps(previous_snapshot, ensure_ascii=False), encoding="utf-8-sig")

    current_week = resolve_week_window({"week_date": "2026-02-18"})
    loaded = load_previous_snapshot(snapshot_base, "OHOSUI", current_week)
    assert loaded is not None
    assert (loaded.get("meta") or {}).get("week_key") == "26'w07"


def test_load_previous_snapshot_accepts_compact_week_filename(tmp_path: Path):
    snapshot_base = tmp_path / "snapshots_root"
    snapshot_base.mkdir(parents=True)
    previous_snapshot = {
        "meta": {"project": "OHOSUI", "week_key": "26w07"},
        "payload": {"meta": {"project": "OHOSUI", "week_key": "26w07"}},
    }
    previous_path = snapshot_base / "jira_weekly_email_OHOSUI_26w07.json"
    previous_path.write_text(json.dumps(previous_snapshot, ensure_ascii=False), encoding="utf-8")

    current_week = resolve_week_window({"week_date": "2026-02-18"})
    loaded = load_previous_snapshot(snapshot_base, "OHOSUI", current_week)
    assert loaded is not None
    assert (loaded.get("meta") or {}).get("week_key") == "26w07"
