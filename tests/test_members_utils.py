from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook

from stats_core.utils.members import read_member_list


def _write_workbook(tmp_path: Path) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.append(["Login", "Display Name", "Email"])
    ws.append(["jsmith", "John Smith", "john@example.com"])
    ws.append(["adoe", "Alice Doe", "alice@example.com"])
    ws.append(["jsmith", "John Smith", "john@example.com"])  # duplicate row
    file_path = tmp_path / "members.xlsx"
    wb.save(file_path)
    return file_path


def test_read_member_list_prefers_names(tmp_path):
    path = _write_workbook(tmp_path)
    members = read_member_list(path)
    assert members == ["John Smith", "Alice Doe"]


def test_read_member_list_can_return_logins(tmp_path):
    path = _write_workbook(tmp_path)
    members = read_member_list(path, prefer="login")
    assert members == ["jsmith", "adoe"]


def test_read_member_list_fallback_column(tmp_path):
    # workbook without headers - data starts in column E
    wb = Workbook()
    ws = wb.active
    ws["E1"] = "placeholder"
    ws["E2"] = "Engineer A"
    ws["E3"] = "Engineer B"
    file_path = tmp_path / "legacy_members.xlsx"
    wb.save(file_path)

    members = read_member_list(file_path)
    assert members == ["Engineer A", "Engineer B"]

