"""
Tests for Jira weekly report covering various scenarios.
"""

import pytest
from datetime import datetime, timedelta
from configparser import ConfigParser
from unittest.mock import Mock, MagicMock, patch
import pandas as pd
from docx import Document
from openpyxl import load_workbook

from stats_core.reports.jira_utils import (
    fetch_jira_data,
    build_resolved_issues_snapshot,
    mark_reassigned_tasks,
    fill_missing_weeks,
    norm_name,
    is_empty_task,
)
import stats_core.reports.jira_weekly as jira_weekly_module
from stats_core.reports.jira_weekly import JiraWeeklyReport
from stats_core.reports.jira_epic_report import (
    generate_epic_resolved_hierarchy,
    add_resolved_tasks_section,
)
from stats_core.sources.jira import JiraSource


@pytest.fixture
def mock_jira_source():
    """Create a mock JiraSource."""
    source = Mock(spec=JiraSource)
    source.jira_url = "https://test-jira.com"
    return source


@pytest.fixture
def sample_issues():
    """Create sample Jira issues for testing."""
    issues = []
    
    # Issue 1: Task closed in the same week it was taken
    issue1 = Mock()
    issue1.key = "TEST-1"
    issue1.fields.summary = "Task closed same week"
    issue1.fields.assignee = Mock(displayName="John Doe")
    issue1.fields.resolutiondate = "2025-01-15T10:00:00.000+0000"  # Wednesday of week 2025-W03
    issue1.fields.created = "2025-01-13T09:00:00.000+0000"  # Monday of week 2025-W03
    issue1.fields.customfield_10000 = None  # No epic
    issue1.fields.parent = None
    issue1.fields.issuetype = Mock(name="Task")
    issues.append(issue1)
    
    # Issue 2: Task stretched across multiple weeks
    issue2 = Mock()
    issue2.key = "TEST-2"
    issue2.fields.summary = "Task across weeks"
    issue2.fields.assignee = Mock(displayName="Jane Smith")
    issue2.fields.resolutiondate = "2025-01-25T10:00:00.000+0000"  # Saturday of week 2025-W04
    issue2.fields.created = "2025-01-06T09:00:00.000+0000"  # Monday of week 2025-W02
    issue2.fields.customfield_10000 = None
    issue2.fields.parent = None
    issue2.fields.issuetype = Mock(name="Task")
    issues.append(issue2)
    
    # Issue 3: Task with no worklogs (no activity)
    issue3 = Mock()
    issue3.key = "TEST-3"
    issue3.fields.summary = "No worklogs task"
    issue3.fields.assignee = Mock(displayName="Bob Wilson")
    issue3.fields.resolutiondate = None  # Not resolved
    issue3.fields.created = "2025-01-20T09:00:00.000+0000"  # Monday of week 2025-W04
    issue3.fields.customfield_10000 = None
    issue3.fields.parent = None
    issue3.fields.issuetype = Mock(name="Task")
    issues.append(issue3)
    
    # Issue 4: Task reassigned (worklogs by one person, final assignee different)
    issue4 = Mock()
    issue4.key = "TEST-4"
    issue4.fields.summary = "Reassigned task"
    issue4.fields.assignee = Mock(displayName="Alice Brown")  # Final assignee
    issue4.fields.resolutiondate = "2025-01-22T10:00:00.000+0000"  # Wednesday of week 2025-W04
    issue4.fields.created = "2025-01-13T09:00:00.000+0000"  # Monday of week 2025-W03
    issue4.fields.customfield_10000 = None
    issue4.fields.parent = None
    issue4.fields.issuetype = Mock(name="Task")
    issues.append(issue4)
    
    return issues


@pytest.fixture
def sample_worklogs():
    """Create sample worklogs for testing."""
    worklogs = {
        "TEST-1": [
            {
                "author": {"displayName": "John Doe"},
                "started": "2025-01-13T09:00:00.000+0000",  # Monday week 2025-W03
            },
            {
                "author": {"displayName": "John Doe"},
                "started": "2025-01-15T10:00:00.000+0000",  # Wednesday week 2025-W03 (resolved same day)
            },
        ],
        "TEST-2": [
            {
                "author": {"displayName": "Jane Smith"},
                "started": "2025-01-06T09:00:00.000+0000",  # Monday week 2025-W02
            },
            {
                "author": {"displayName": "Jane Smith"},
                "started": "2025-01-13T09:00:00.000+0000",  # Monday week 2025-W03
            },
            {
                "author": {"displayName": "Jane Smith"},
                "started": "2025-01-20T09:00:00.000+0000",  # Monday week 2025-W04
            },
            {
                "author": {"displayName": "Jane Smith"},
                "started": "2025-01-25T10:00:00.000+0000",  # Saturday week 2025-W04 (resolved)
            },
        ],
        "TEST-3": [],  # No worklogs
        "TEST-4": [
            {
                "author": {"displayName": "John Doe"},  # Different from final assignee
                "started": "2025-01-13T09:00:00.000+0000",  # Monday week 2025-W03
            },
            {
                "author": {"displayName": "John Doe"},
                "started": "2025-01-15T09:00:00.000+0000",  # Wednesday week 2025-W03
            },
            # Note: Final assignee is Alice Brown, but worklogs are by John Doe
        ],
    }
    return worklogs


def test_task_closed_same_week(mock_jira_source, sample_issues, sample_worklogs):
    """Test: Engineer has a task and closed it in the same week it was taken."""
    # Mock JiraSource methods
    mock_jira_source.fetch_issues = Mock(return_value=[sample_issues[0]])
    mock_jira_source.get_all_worklogs = Mock(return_value=sample_worklogs["TEST-1"])
    mock_jira_source.fetch_epic_names = Mock(return_value={})
    
    start_date = "2025-01-13"
    end_date = "2025-01-19"
    
    data = fetch_jira_data(mock_jira_source, "TEST", start_date, end_date)
    
    # Should have one row with Status="Resolved" and Week="2025-W03"
    assert not data.empty
    resolved_rows = data[data["Status"] == "Resolved"]
    assert len(resolved_rows) == 1
    assert resolved_rows.iloc[0]["Week"] == "2025-W03"
    assert resolved_rows.iloc[0]["Assignee"] == "John Doe"
    assert resolved_rows.iloc[0]["Issue_key"] == "TEST-1"


def test_task_stretched_multiple_weeks(mock_jira_source, sample_issues, sample_worklogs):
    """Test: Task stretched across multiple weeks."""
    # Mock JiraSource methods
    mock_jira_source.fetch_issues = Mock(return_value=[sample_issues[1]])
    mock_jira_source.get_all_worklogs = Mock(return_value=sample_worklogs["TEST-2"])
    mock_jira_source.fetch_epic_names = Mock(return_value={})
    
    start_date = "2025-01-06"
    end_date = "2025-01-26"
    
    data = fetch_jira_data(mock_jira_source, "TEST", start_date, end_date)
    
    # Should have rows for multiple weeks
    assert not data.empty
    test2_rows = data[data["Issue_key"] == "TEST-2"]
    
    # Should have "In progress" rows for weeks 2025-W02, 2025-W03, 2025-W04
    # and "Resolved" row for week 2025-W04
    weeks = set(test2_rows["Week"].unique())
    assert "2025-W02" in weeks
    assert "2025-W03" in weeks
    assert "2025-W04" in weeks
    
    # Should have one "Resolved" status
    resolved = test2_rows[test2_rows["Status"] == "Resolved"]
    assert len(resolved) == 1
    assert resolved.iloc[0]["Week"] == "2025-W04"
    
    # Should have "In progress" for earlier weeks
    in_progress = test2_rows[test2_rows["Status"] == "In progress"]
    assert len(in_progress) >= 2  # At least W02 and W03


def test_unresolved_task_with_worklog_counts_as_in_progress(mock_jira_source):
    """Test: A task without resolution but with worklog still appears as in-progress."""
    issue = Mock()
    issue.key = "TEST-5"
    issue.fields.summary = "Unresolved but active"
    issue.fields.assignee = Mock(displayName="Dev User")
    issue.fields.resolutiondate = None
    issue.fields.created = "2025-01-10T09:00:00.000+0000"
    issue.fields.customfield_10000 = None
    issue.fields.parent = None
    issue.fields.issuetype = Mock(name="Task")

    mock_jira_source.fetch_issues = Mock(return_value=[issue])
    mock_jira_source.get_all_worklogs = Mock(return_value=[
        {
            "author": {"displayName": "Dev User"},
            "started": "2025-01-15T12:00:00.000+0000",  # Week 2025-W03
        }
    ])
    mock_jira_source.fetch_epic_names = Mock(return_value={})

    start_date = "2025-01-13"
    end_date = "2025-01-19"
    data = fetch_jira_data(mock_jira_source, "TEST", start_date, end_date)

    expected_week = datetime.strptime("2025-01-15", "%Y-%m-%d").strftime("%G-W%V")
    active_rows = data[(data["Issue_key"] == "TEST-5") & (data["Status"] == "In progress")]
    assert len(active_rows) == 1
    assert active_rows.iloc[0]["Week"] == expected_week
    assert active_rows.iloc[0]["Assignee"] == "Dev User"


def test_no_tasks_no_logs(mock_jira_source, sample_issues, sample_worklogs):
    """Test: No tasks and no logs (empty data)."""
    # Mock JiraSource methods - return empty list
    mock_jira_source.fetch_issues = Mock(return_value=[])
    mock_jira_source.get_all_worklogs = Mock(return_value=[])
    mock_jira_source.fetch_epic_names = Mock(return_value={})
    
    start_date = "2025-01-13"
    end_date = "2025-01-19"
    
    data = fetch_jira_data(mock_jira_source, "TEST", start_date, end_date)
    
    # Should return empty DataFrame
    assert data.empty


def test_task_reassigned(mock_jira_source, sample_issues, sample_worklogs):
    """Test: Task had worklogs but was reassigned (final assignee not in worklog authors)."""
    # Mock JiraSource methods
    mock_jira_source.fetch_issues = Mock(return_value=[sample_issues[3]])
    mock_jira_source.get_all_worklogs = Mock(return_value=sample_worklogs["TEST-4"])
    mock_jira_source.fetch_epic_names = Mock(return_value={})
    
    start_date = "2025-01-13"
    end_date = "2025-01-26"
    
    data = fetch_jira_data(mock_jira_source, "TEST", start_date, end_date)
    
    # Mark reassigned tasks
    data = mark_reassigned_tasks(data)
    
    # Should have rows with worklog author (John Doe) and reassigned flag
    test4_rows = data[data["Issue_key"] == "TEST-4"]
    assert not test4_rows.empty
    
    # Check that reassigned flag is set correctly
    # The worklogs are by John Doe, but final assignee is Alice Brown
    # So rows with John Doe as Assignee should exist, and reassigned should be True
    # (because final assignee Alice Brown is not in worklog authors)
    john_rows = test4_rows[test4_rows["Assignee"] == "John Doe"]
    if not john_rows.empty:
        # The reassigned flag indicates that the final assignee (Alice) is not in worklog authors
        # So for John's rows, reassigned should be True
        assert john_rows["Reassigned"].any() or john_rows["Reassigned"].all()


def test_fill_missing_weeks():
    """Test: Fill missing weeks for assignees with no activity."""
    data = pd.DataFrame([
        {
            "Issue_key": "TEST-1",
            "Summary": "Task 1",
            "Assignee": "John Doe",
            "Status": "Resolved",
            "Week": "2025-W03",
            "Assignee_norm": "john doe",
        }
    ])
    
    valid_weeks = ["2025-W02", "2025-W03", "2025-W04"]
    required_assignees = ["John Doe", "Jane Smith"]
    
    filled = fill_missing_weeks(data, valid_weeks, required_assignees)
    
    # Should have rows for all assignees and weeks
    assert len(filled) > len(data)
    
    # Check that Jane Smith has filler rows for all weeks
    jane_rows = filled[filled["Assignee"] == "Jane Smith"]
    assert len(jane_rows) == 3  # One for each week
    
    # Check that John Doe has filler rows for missing weeks
    john_rows = filled[filled["Assignee"] == "John Doe"]
    assert len(john_rows) >= 3  # Original + fillers


def test_is_empty_task():
    """Test: Check if task is empty (no summary and no status)."""
    assert is_empty_task("", "") is True
    assert is_empty_task("   ", "   ") is True
    assert is_empty_task("Task summary", "Resolved") is False
    assert is_empty_task("Task summary", "") is False
    assert is_empty_task(None, None) is True


def test_norm_name():
    """Test: Name normalization."""
    assert norm_name("John Doe") == "john doe"
    assert norm_name("  Jane  Smith  ") == "jane smith"
    assert norm_name("Bob\tWilson\n") == "bob wilson"
    assert norm_name("") == ""
    assert norm_name(None) == ""


@patch('stats_core.reports.jira_weekly.JiraSource')
def test_jira_weekly_report_run(mock_jira_class, tmp_path):
    """Test: JiraWeeklyReport.run generates reports correctly."""
    # Create mock config
    from configparser import ConfigParser
    config = ConfigParser()
    config.add_section("jira")
    config.set("jira", "jira-url", "https://test-jira.com")
    config.set("jira", "username", "testuser")
    config.set("jira", "password", "testpass")
    
    # Create mock JiraSource instance
    mock_jira_source = Mock(spec=JiraSource)
    mock_jira_source.jira_url = "https://test-jira.com"
    mock_jira_source.fetch_issues = Mock(return_value=[])
    mock_jira_source.get_all_worklogs = Mock(return_value=[])
    mock_jira_source.fetch_epic_names = Mock(return_value={})
    mock_jira_class.return_value = mock_jira_source
    
    # Create report instance
    report = JiraWeeklyReport()
    
    # Run report
    extra_params = {
        "project": "TEST",
        "start": "2025-01-13",
        "end": "2025-01-19",
    }
    
    # Should not raise exception
    report.run(
        dataset={},
        config=config,
        output_formats=["excel", "word"],
        extra_params=extra_params,
    )



def test_build_resolved_issues_snapshot_filters_and_hierarchy(mock_jira_source):
    """Test: Resolved snapshot includes in-period issues and inherits epic from parent."""
    parent_issue = Mock()
    parent_issue.key = "TEST-10"
    parent_issue.fields.summary = "Parent task"
    parent_issue.fields.resolutiondate = "2025-01-15T10:00:00.000+0000"
    parent_issue.fields.customfield_10000 = "EPIC-1"
    parent_issue.fields.parent = None
    parent_issue.fields.issuetype = Mock(name="Task")

    subtask_issue = Mock()
    subtask_issue.key = "TEST-11"
    subtask_issue.fields.summary = "Subtask in period"
    subtask_issue.fields.resolutiondate = "2025-01-16T10:00:00.000+0000"
    subtask_issue.fields.customfield_10000 = None
    subtask_issue.fields.parent = parent_issue
    subtask_issue.fields.issuetype = Mock(name="Sub-task")

    out_of_range = Mock()
    out_of_range.key = "TEST-12"
    out_of_range.fields.summary = "Subtask out of range"
    out_of_range.fields.resolutiondate = "2025-01-25T10:00:00.000+0000"
    out_of_range.fields.customfield_10000 = None
    out_of_range.fields.parent = parent_issue
    out_of_range.fields.issuetype = Mock(name="Sub-task")

    mock_jira_source.fetch_issues = Mock(return_value=[parent_issue, subtask_issue, out_of_range])
    mock_jira_source.fetch_epic_names = Mock(return_value={"EPIC-1": "Epic One"})

    data = build_resolved_issues_snapshot(
        mock_jira_source,
        "TEST",
        "2025-01-13",
        "2025-01-19",
    )

    keys = set(data["Issue_key"].tolist())
    assert "TEST-10" in keys
    assert "TEST-11" in keys
    assert "TEST-12" not in keys

    subtask_row = data[data["Issue_key"] == "TEST-11"].iloc[0]
    assert subtask_row["Epic_Link"] == "EPIC-1"
    assert subtask_row["Epic_Name"] == "Epic One"
    assert subtask_row["Parent_Key"] == "TEST-10"


def test_generate_epic_resolved_hierarchy_includes_parent_for_subtask_only():
    """Test: Resolved hierarchy renders parent bucket even if only subtask resolved."""
    resolved_df = pd.DataFrame([
        {
            "Issue_key": "SUB-1",
            "Summary": "Subtask",
            "Resolution_Date": "2025-01-15",
            "Resolution_Week": "2025-W03",
            "Epic_Link": "EPIC-1",
            "Epic_Name": "Epic One",
            "Parent_Key": "PARENT-1",
            "Parent_Summary": "Parent",
            "Type": "Sub-task",
        }
    ])

    summary = generate_epic_resolved_hierarchy(resolved_df)
    assert len(summary) == 1
    parents = summary[0].get("Parents", [])
    assert len(parents) == 1
    assert parents[0]["Parent_Key"] == "PARENT-1"
    assert parents[0]["Subtasks"][0]["Task_Key"] == "SUB-1"


def test_add_resolved_tasks_section_lists_subtask_under_parent():
    """Test: Resolved Tasks section nests subtask under parent for the week."""
    resolved_df = pd.DataFrame([
        {
            "Issue_key": "SUB-1",
            "Summary": "Subtask",
            "Resolution_Date": "2025-01-15",
            "Resolution_Week": "2025-W03",
            "Epic_Link": "EPIC-1",
            "Epic_Name": "Epic One",
            "Parent_Key": "PARENT-1",
            "Parent_Summary": "Parent",
            "Type": "Sub-task",
        }
    ])

    document = Document()
    add_resolved_tasks_section(document, resolved_df)

    texts = [p.text for p in document.paragraphs]
    assert any("PARENT-1: Parent" in text for text in texts)
    assert any("SUB-1: Subtask" in text for text in texts)


def test_build_weekly_epic_summary_df_recovers_epic_from_open_parent_and_groups_subtasks():
    config = ConfigParser()

    resolved_df = pd.DataFrame(
        [
            {
                "Issue_key": "SUB-1",
                "Summary": "Finalize lifecycle cleanup",
                "Status": "Done",
                "Resolution": "Done",
                "Resolution_Date": "2025-01-15",
                "Epic_Link": "",
                "Epic_Name": "",
                "Parent": "FEATURE-1",
                "Parent_Key": "FEATURE-1",
                "Parent_Summary": "",
                "Type": "Sub-task",
                "Description": "Cleanup final implementation.",
                "Last_Comment": "",
            },
            {
                "Issue_key": "SUB-2",
                "Summary": "Add teardown regression coverage",
                "Status": "Done",
                "Resolution": "Done",
                "Resolution_Date": "2025-01-17",
                "Epic_Link": "",
                "Epic_Name": "",
                "Parent": "FEATURE-1",
                "Parent_Key": "FEATURE-1",
                "Parent_Summary": "",
                "Type": "Sub-task",
                "Description": "Coverage for repeated mount and unmount.",
                "Last_Comment": "",
            },
        ]
    )
    comments_df = pd.DataFrame(
        [
            {
                "Issue_key": "FEATURE-1",
                "CommentBody": (
                    "Implemented cleanup for detached nodes. "
                    "See https://example.com/report and \\\\server\\share\\weekly\\result.txt"
                ),
                "CommentDate": datetime(2025, 1, 15).date(),
                "CommentId": "10",
                "Is_Worklog_Comment": False,
            },
            {
                "Issue_key": "SUB-2",
                "CommentBody": (
                    "Regression fixed, attachment build.log uploaded and C:\\temp\\trace.txt checked. "
                    "{code:java}\nSystem.out.println(\"debug\");\n{code} ![image](https://example.com/img.png) [! leftover"
                ),
                "CommentDate": datetime(2025, 1, 17).date(),
                "CommentId": "11",
                "Is_Worklog_Comment": False,
            },
        ]
    )

    jira_source = Mock()
    jira_source.fetch_issue_details = Mock(
        return_value={
            "FEATURE-1": {
                "Issue_Key": "FEATURE-1",
                "Summary": "Component teardown stability",
                "Type": "Story",
                "Status": "In Progress",
                "Description": "Stabilize teardown and cleanup behavior.",
                "Epic_Link": "EPIC-1",
            }
        }
    )
    jira_source.fetch_epic_names = Mock(return_value={"EPIC-1": "Epic One"})

    summary_df = jira_weekly_module.build_weekly_epic_summary_df(
        jira_source,
        resolved_df,
        comments_df,
        "2025-01-13",
        "2025-01-19",
        config,
        {"ollama_enabled": False, "webui_enabled": False},
    )

    assert len(summary_df) == 1
    row = summary_df.iloc[0]
    assert row["Epic_Link"] == "EPIC-1"
    assert row["Epic_Name"] == "Epic One"
    assert row["Planned_Tasks_Resolved"] == 2
    assert row["Reported_Issues_Resolved"] == 0

    summary_text = str(row["Summary"])
    assert summary_text.count("- ") == 1
    assert "Component teardown stability" in summary_text
    assert "Resolved 2 planned tasks on time." in summary_text
    assert "https://" not in summary_text
    assert "\\\\server\\share" not in summary_text
    assert "build.log" not in summary_text
    assert "C:\\temp\\trace.txt" not in summary_text
    assert "{code" not in summary_text
    assert "System.out" not in summary_text
    assert "![" not in summary_text


@patch("stats_core.reports.jira_weekly.generate_file_suffix", return_value="")
@patch("stats_core.reports.jira_weekly.build_resolved_issues_snapshot")
@patch("stats_core.reports.jira_weekly.fetch_jira_activity_data")
@patch("stats_core.reports.jira_weekly.fetch_jira_data")
@patch("stats_core.reports.jira_weekly.JiraSource")
def test_jira_weekly_report_excel_contains_only_weekly_grid_sheet(
    mock_jira_source_cls,
    mock_fetch_jira_data,
    mock_fetch_jira_activity_data,
    mock_build_resolved,
    mock_suffix,
    tmp_path,
):
    config = ConfigParser()
    config.add_section("jira")
    config.set("jira", "jira-url", "https://test-jira.com")
    config.set("jira", "username", "testuser")
    config.set("jira", "password", "testpass")
    config.add_section("reporting")
    config.set("reporting", "output_dir", str(tmp_path))

    mock_jira_source = Mock(spec=JiraSource)
    mock_jira_source.jira_url = "https://test-jira.com"
    mock_jira_source_cls.return_value = mock_jira_source

    mock_fetch_jira_data.return_value = pd.DataFrame(
        [
            {
                "Issue_key": "ABC-1",
                "Summary": "First task",
                "Assignee": "Alice Dev",
                "Final_Assignee": "Alice Dev",
                "Status": "In progress",
                "Resolution_Date": "",
                "Created_Date": "2025-01-13",
                "Week": "2025-W03",
                "Epic_Link": "",
                "Epic_Name": "",
                "Parent_Key": "",
                "Parent_Summary": "",
                "Type": "Task",
            }
        ]
    )
    mock_fetch_jira_activity_data.return_value = (pd.DataFrame(), pd.DataFrame())
    mock_build_resolved.return_value = pd.DataFrame()

    report = JiraWeeklyReport()
    report.run(
        dataset={},
        config=config,
        output_formats=["excel"],
        extra_params={
            "project": "TEST",
            "start": "2025-01-13",
            "end": "2025-01-19",
            "output_dir": str(tmp_path),
        },
    )

    excel_path = tmp_path / "jira_report_TEST_2025-01-13-2025-01-19.xlsx"
    assert excel_path.exists()

    wb = load_workbook(excel_path)
    assert wb.sheetnames == ["Weekly_Grid"]
    assert "Developer_Activity" not in wb.sheetnames
