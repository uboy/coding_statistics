"""
Tests for Jira comprehensive report.
"""

from __future__ import annotations

from configparser import ConfigParser
from pathlib import Path
import re
from types import SimpleNamespace
from unittest.mock import Mock, patch

import pandas as pd
from openpyxl import load_workbook

from stats_core.reports.jira_comprehensive import (
    JiraComprehensiveReport,
    build_comments_period_df,
    build_jql_query,
    build_monthly_summary_df,
    calculate_engineer_metrics,
    fetch_jira_data,
    _build_comment_summary_prompt,
    rewrite_comment_items_with_ai,
    _extract_results_hint,
    _format_ai_comment_summary,
)


def test_build_jql_query_project_dates():
    params = {"project": "ABC", "start_date": "2025-01-01", "end_date": "2025-01-31"}
    assert (
        build_jql_query(params)
        == "project = ABC AND resolved >= '2025-01-01' AND resolved < '2025-02-01' ORDER BY created DESC"
    )


def _make_issue(
    key: str,
    *,
    summary: str,
    issue_type: str,
    status: str,
    assignee_display: str,
    assignee_username: str | None,
    reporter_display: str,
    reporter_username: str | None,
    description: str = "",
    comment_body: str | None = None,
    comment_id: str | None = None,
    labels: list[str] | None = None,
    resolution_name: str = "Done",
    resolved_at: str = "2025-01-10T10:00:00.000+0000",
    epic_link: str | None = "EPIC-1",
    parent_key: str | None = None,
    attachments: list[SimpleNamespace] | None = None,
):
    comment = SimpleNamespace(
        comments=(
            [
                SimpleNamespace(
                    body=comment_body,
                    author=SimpleNamespace(displayName="Commenter"),
                    created="2025-01-02T10:00:00.000+0000",
                    id=comment_id,
                )
            ]
            if comment_body
            else []
        )
    )
    parent = SimpleNamespace(key=parent_key) if parent_key else None
    fields = SimpleNamespace(
        summary=summary,
        assignee=SimpleNamespace(displayName=assignee_display, name=assignee_username),
        reporter=SimpleNamespace(displayName=reporter_display, name=reporter_username),
        resolutiondate=resolved_at,
        created="2025-01-01T10:00:00.000+0000",
        updated="2025-01-11T10:00:00.000+0000",
        description=description,
        comment=comment,
        labels=labels or [],
        priority=SimpleNamespace(name="P1"),
        status=SimpleNamespace(name=status),
        resolution=SimpleNamespace(name=resolution_name),
        issuetype=SimpleNamespace(name=issue_type),
        timeestimate=0,
        timespent=0,
        timeoriginalestimate=0,
        customfield_10000=epic_link,
        parent=parent,
        attachment=attachments or [],
    )
    return SimpleNamespace(key=key, fields=fields)


@patch("stats_core.reports.jira_comprehensive.rewrite_comment_items_with_ai")
@patch("stats_core.reports.jira_comprehensive.JiraSource")
def test_jira_comprehensive_report_run_writes_excel(
    mock_jira_source_cls,
    mock_rewrite_comments,
    tmp_path: Path,
):
    mock_rewrite_comments.return_value = {
        "ABC-1": "Сделано: фиксы. Планы: нет данных. Риски: нет данных. Зависимости: нет данных. Примечания: нет данных.",
    }
    issues = [
        _make_issue(
            "ABC-1",
            summary="Bug fix",
            issue_type="Bug",
            status="Released",
            assignee_display="Alice",
            assignee_username=None,
            reporter_display="Bob",
            reporter_username=None,
            description="See\x0b https://example.com",
            comment_body={
                "type": "doc",
                "content": [
                    {
                        "type": "paragraph",
                        "content": [
                            {
                                "type": "text",
                                "text": " *Results:* fixed issue https://result.example/one and https://result.example/two",
                            }
                        ],
                    }
                ],
            },
            comment_id="101",
            labels=["documentation"],
            resolved_at="2025-01-09T10:00:00.000+0000",
        ),
        _make_issue(
            "ABC-2",
            summary="QA task",
            issue_type="Task",
            status="In QA",
            assignee_display="Bob",
            assignee_username=None,
            reporter_display="Carol",
            reporter_username="carol",
            comment_body=(
                "TT_tdev_APIs - number of developed and executed tasks = 2\n"
                "TT_tested_APIs - number of executed tests = 3\n"
                "TT_tested_perf - number of executed performance tests = 1\n"
                "TT_tdev_perf - number of developed benchmark tests = 4"
            ),
            comment_id="201",
            labels=["documentation", "arkoala_perf"],
            resolved_at="2025-01-11T10:00:00.000+0000",
        ),
        _make_issue(
            "ABC-4",
            summary="Excluded QA task",
            issue_type="Task",
            status="Done",
            assignee_display="Bob",
            assignee_username=None,
            reporter_display="Bob",
            reporter_username=None,
            comment_body="TT_tdev_APIs: 200\nTT_tested_APIs: 300\nTT_tested_perf: 100\nTT_tdev_perf: 400",
            labels=["documentation"],
            resolution_name="Won't Do",
            resolved_at="2025-01-12T10:00:00.000+0000",
        ),
        _make_issue(
            "ABC-5",
            summary="Excluded invalid resolution",
            issue_type="Bug",
            status="Done",
            assignee_display="Alice",
            assignee_username=None,
            reporter_display="Bob",
            reporter_username=None,
            labels=["documentation"],
            resolution_name="Invalid",
            resolved_at="2025-01-13T10:00:00.000+0000",
        ),
        _make_issue(
            "ABC-3",
            summary="Epic",
            issue_type="Epic",
            status="Released",
            assignee_display="Carol",
            assignee_username="carol",
            reporter_display="Carol",
            reporter_username="carol",
            comment_body="h1. +Result+\n_epic shipped_",
            comment_id="301",
            resolved_at="2025-01-08T10:00:00.000+0000",
        ),
        _make_issue(
            "ABC-6",
            summary="Parent Task",
            issue_type="Task",
            status="Released",
            assignee_display="Alice",
            assignee_username="alice",
            reporter_display="Bob",
            reporter_username=None,
            epic_link="EPIC-2",
            resolved_at="2025-01-07T10:00:00.000+0000",
        ),
        _make_issue(
            "ABC-7",
            summary="Subtask Task",
            issue_type="Sub-task",
            status="Released",
            assignee_display="Alice",
            assignee_username="alice",
            reporter_display="Bob",
            reporter_username=None,
            epic_link=None,
            parent_key="ABC-6",
            resolved_at="2025-01-06T10:00:00.000+0000",
        ),
    ]
    issues[1].fields.comment.comments.append(
        SimpleNamespace(
            body="Follow-up QA note",
            author=SimpleNamespace(displayName="Commenter"),
            created="2025-01-03T10:00:00.000+0000",
            id="202",
        )
    )

    fake_jira = Mock()
    epic_issues = [
        _make_issue(
            "EPIC-1",
            summary="Epic One",
            issue_type="Epic",
            status="Released",
            assignee_display="Carol",
            assignee_username="carol",
            reporter_display="Carol",
            reporter_username="carol",
            labels=["report"],
        ),
        _make_issue(
            "EPIC-2",
            summary="Epic Two",
            issue_type="Epic",
            status="Released",
            assignee_display="Carol",
            assignee_username="carol",
            reporter_display="Carol",
            reporter_username="carol",
            labels=["report"],
        ),
    ]

    def _search_issues_side_effect(jql_query, *args, **kwargs):
        if "issuekey in" in jql_query:
            return epic_issues
        return issues

    fake_jira.search_issues.side_effect = _search_issues_side_effect

    fake_source = Mock()
    fake_source.jira = fake_jira
    fake_source.get_all_worklogs = Mock(return_value=[])
    mock_jira_source_cls.return_value = fake_source

    members_file = tmp_path / "members.xlsx"
    pd.DataFrame(
        [
            {"name": "Alice", "username": "alice", "role": "engineer"},
            {"name": "Bob", "username": "bob", "role": "test engineer"},
            {"name": "Carol", "username": "carol", "role": "project manager"},
        ]
    ).to_excel(members_file, index=False)

    code_volume_file = tmp_path / "code_volume.xlsx"
    pd.DataFrame([{"username": "alice", "code_volume": 123}]).to_excel(code_volume_file, index=False)

    config = ConfigParser()
    config.read_dict(
        {
            "jira": {"jira-url": "https://jira.example.com", "username": "u", "password": "p"},
            "reporting": {"output_dir": str(tmp_path)},
        }
    )

    report = JiraComprehensiveReport()
    report.run(
        dataset={},
        config=config,
        output_formats=["excel"],
        extra_params={
            "project": "ABC",
            "start": "2025-01-01",
            "end": "2025-01-31",
            "member_list_file": str(members_file),
            "code_volume_file": str(code_volume_file),
            "output": "out.xlsx",
        },
    )

    expected_jql = (
        "project = ABC AND resolved >= '2025-01-01' AND resolved < '2025-02-01' ORDER BY created DESC"
    )
    assert any(call.args[0] == expected_jql for call in fake_jira.search_issues.call_args_list)
    expected_comments_jql = (
        "project = ABC AND updated >= '2025-01-01' AND updated < '2025-02-01' ORDER BY created DESC"
    )
    assert any(call.args[0] == expected_comments_jql for call in fake_jira.search_issues.call_args_list)

    out_path = tmp_path / "out.xlsx"
    assert out_path.exists()

    wb = load_workbook(out_path)
    assert "Issues" in wb.sheetnames
    assert "Links" in wb.sheetnames
    assert "Results" in wb.sheetnames
    assert "Summary" in wb.sheetnames
    assert "Engineer_Performance" in wb.sheetnames
    assert "QA_Performance" in wb.sheetnames
    assert "PM_Performance" in wb.sheetnames
    assert "Comments_Period" in wb.sheetnames

    comments_sheet = wb["Comments_Period"]
    comments_headers = [cell.value for cell in comments_sheet[1]]
    for header in (
        "Issue_Key",
        "Summary",
        "Type",
        "Status",
        "Priority",
        "Assignee",
        "Created",
        "Epic_Name",
        "Parent",
        "Description",
        "Comments",
        "AI_Comments",
        "Comments_In_Period",
    ):
        assert header in comments_headers

    issues_sheet = wb["Issues"]
    headers = [cell.value for cell in issues_sheet[1]]
    desc_col = headers.index("Description") + 1
    comments_col = headers.index("Comments") + 1
    parent_col = headers.index("Parent") + 1
    epic_link_col = headers.index("Epic_Link") + 1
    epic_name_col = headers.index("Epic_Name") + 1
    assert "\x0b" not in str(issues_sheet.cell(row=2, column=desc_col).value)
    assert "\x00" not in str(issues_sheet.cell(row=2, column=comments_col).value)
    issue_key_col = headers.index("Issue_Key") + 1
    subtask_row = None
    for row_idx in range(2, issues_sheet.max_row + 1):
        if issues_sheet.cell(row=row_idx, column=issue_key_col).value == "ABC-7":
            subtask_row = row_idx
            break
    assert subtask_row is not None
    assert issues_sheet.cell(row=subtask_row, column=parent_col).value == "ABC-6"
    assert issues_sheet.cell(row=subtask_row, column=epic_link_col).value == "EPIC-2"
    assert issues_sheet.cell(row=subtask_row, column=epic_name_col).value == "Epic Two"
    resolved_col = headers.index("Resolved") + 1
    issues_sort_pairs = [
        (
            str(issues_sheet.cell(row=row_idx, column=epic_name_col).value or ""),
            str(issues_sheet.cell(row=row_idx, column=resolved_col).value or ""),
        )
        for row_idx in range(2, issues_sheet.max_row + 1)
    ]
    assert issues_sort_pairs == sorted(issues_sort_pairs, key=lambda item: (item[0], item[1]))

    links_sheet = wb["Links"]
    # header + at least 2 links (description + comment)
    assert links_sheet.max_row >= 2

    results_sheet = wb["Results"]
    results_headers = [cell.value for cell in results_sheet[1]]
    assert "Epic_Name" in results_headers
    assert "Resolved" in results_headers
    issue_key_col = results_headers.index("Issue_Key") + 1
    epic_col = results_headers.index("Epic_Name") + 1
    resolved_col = results_headers.index("Resolved") + 1
    result_col = results_headers.index("Result") + 1
    result_links_col = results_headers.index("Result_Links") + 1
    rows = []
    for row_idx in range(2, results_sheet.max_row + 1):
        rows.append(
            {
                "Issue_Key": results_sheet.cell(row=row_idx, column=issue_key_col).value,
                "Epic_Name": results_sheet.cell(row=row_idx, column=epic_col).value,
                "Resolved": results_sheet.cell(row=row_idx, column=resolved_col).value,
                "Result": results_sheet.cell(row=row_idx, column=result_col).value,
                "Result_Links": results_sheet.cell(row=row_idx, column=result_links_col).value,
            }
        )

    result_sort_pairs = [
        (str(row["Epic_Name"] or ""), str(row["Resolved"] or ""))
        for row in rows
    ]
    assert result_sort_pairs == sorted(result_sort_pairs, key=lambda item: (item[0], item[1]))

    abc1_rows = [row for row in rows if row["Issue_Key"] == "ABC-1"]
    assert abc1_rows
    assert abc1_rows[0]["Result"] == "fixed issue https://result.example/one and https://result.example/two"
    assert "https://result.example/one" in str(abc1_rows[0]["Result_Links"])
    assert "https://result.example/two" in str(abc1_rows[0]["Result_Links"])

    abc3_rows = [row for row in rows if row["Issue_Key"] == "ABC-3"]
    assert abc3_rows
    assert abc3_rows[0]["Result"] == "_epic shipped_"

    abc2_rows = [row for row in rows if row["Issue_Key"] == "ABC-2"]
    assert abc2_rows
    assert str(abc2_rows[0]["Result"]).casefold() == "no results"
    assert "focusedCommentId=202" in str(abc2_rows[0]["Result_Links"])

    summary_sheet = wb["Summary"]
    summary_headers = [cell.value for cell in summary_sheet[1]]
    epic_link_col = summary_headers.index("Epic_Link") + 1
    epic_name_col = summary_headers.index("Epic_Name") + 1
    summary_col = summary_headers.index("Summary") + 1
    planned_col = summary_headers.index("Planned_Tasks_Resolved") + 1
    bug_col = summary_headers.index("Reported_Issues_Resolved") + 1
    summary_rows: list[dict[str, object]] = []
    for row_idx in range(2, summary_sheet.max_row + 1):
        summary_rows.append(
            {
                "Epic_Link": summary_sheet.cell(row=row_idx, column=epic_link_col).value,
                "Epic_Name": summary_sheet.cell(row=row_idx, column=epic_name_col).value,
                "Summary": summary_sheet.cell(row=row_idx, column=summary_col).value,
                "Planned_Tasks_Resolved": summary_sheet.cell(row=row_idx, column=planned_col).value,
                "Reported_Issues_Resolved": summary_sheet.cell(row=row_idx, column=bug_col).value,
            }
        )

    epic_one_summary = next(row for row in summary_rows if row["Epic_Name"] == "Epic One")
    epic_two_summary = next(row for row in summary_rows if row["Epic_Name"] == "Epic Two")

    assert int(epic_one_summary["Planned_Tasks_Resolved"]) == 1
    assert int(epic_one_summary["Reported_Issues_Resolved"]) == 1
    assert "Resolved 1 planned tasks on time." in str(epic_one_summary["Summary"])
    assert "Resolved 1 reported issues." in str(epic_one_summary["Summary"])
    assert str(epic_one_summary["Summary"]).count("- ") == 1

    assert int(epic_two_summary["Planned_Tasks_Resolved"]) == 2
    assert int(epic_two_summary["Reported_Issues_Resolved"]) == 0
    assert "Resolved 2 planned tasks on time." in str(epic_two_summary["Summary"])
    assert "reported issues" not in str(epic_two_summary["Summary"])
    assert str(epic_two_summary["Summary"]).count("- ") == 1

    def _sheet_row(sheet_name: str) -> dict[str, object]:
        sheet = wb[sheet_name]
        headers = [cell.value for cell in sheet[1]]
        values = [sheet.cell(row=2, column=idx + 1).value for idx in range(len(headers))]
        return dict(zip(headers, values))

    engineer_row = _sheet_row("Engineer_Performance")
    assert engineer_row["Bugs"] == 1
    assert engineer_row["Documentation_Tasks"] == 1
    assert engineer_row["Total_Resolved_Issues"] == 3

    qa_row = _sheet_row("QA_Performance")
    assert qa_row["Test_Scenarios_Executed"] == 5
    assert qa_row["Issues_Raised"] == 1
    assert qa_row["Performance_Benchmarks"] == 5
    assert qa_row["Documentation_Tasks"] == 1
    assert "Outstanding_Contribution" in qa_row
    assert qa_row["Outstanding_Contribution"] == 0
    assert qa_row["TT_tdev_APIs"] == 2
    assert qa_row["TT_tested_APIs"] == 3
    assert qa_row["TT_tested_perf"] == 1
    assert qa_row["TT_tdev_perf"] == 4
    assert qa_row["Total_Resolved_Issues"] == 1


def _make_comment(
    body: str,
    *,
    created: str,
    updated: str | None = None,
    author: str = "Commenter",
):
    return SimpleNamespace(
        body=body,
        author=SimpleNamespace(displayName=author),
        created=created,
        updated=updated or created,
        id="1",
    )


def _make_issue_with_comments(key: str, comments: list[SimpleNamespace]):
    comment = SimpleNamespace(comments=comments)
    fields = SimpleNamespace(
        summary="Test task",
        assignee=SimpleNamespace(displayName="Alice", name="alice"),
        created="2025-01-01T10:00:00.000+0000",
        description="Description",
        comment=comment,
        priority=SimpleNamespace(name="P1"),
        status=SimpleNamespace(name="In Progress"),
        issuetype=SimpleNamespace(name="Task"),
        customfield_10000=None,
        parent=None,
    )
    return SimpleNamespace(key=key, fields=fields)


def test_build_comments_period_df_filters_created_and_updated():
    issues = [
        _make_issue_with_comments(
            "ABC-1",
            [
                _make_comment(
                    "created in period",
                    created="2025-01-10T10:00:00.000+0000",
                ),
                _make_comment(
                    "updated in period",
                    created="2024-12-10T10:00:00.000+0000",
                    updated="2025-01-15T10:00:00.000+0000",
                ),
            ],
        )
    ]
    fake_jira = SimpleNamespace(
        _options={"server": "https://jira.example.com"},
        search_issues=Mock(side_effect=[issues]),
    )
    config = ConfigParser()
    with patch(
        "stats_core.reports.jira_comprehensive.rewrite_comment_items_with_ai",
        return_value={},
    ):
        df = build_comments_period_df(
            fake_jira,
            "project = ABC AND updated >= '2025-01-01' AND updated < '2025-02-01'",
            "2025-01-01",
            "2025-01-31",
            config,
            {},
        )
    assert not df.empty
    row = df.iloc[0]
    assert "created in period" in str(row["Comments_In_Period"])
    assert "updated in period" in str(row["Comments_In_Period"])


def test_build_comments_period_df_without_dates_includes_all_comments():
    issues = [
        _make_issue_with_comments(
            "ABC-1",
            [
                _make_comment(
                    "first",
                    created="2024-12-01T10:00:00.000+0000",
                ),
                _make_comment(
                    "second",
                    created="2025-02-01T10:00:00.000+0000",
                ),
            ],
        )
    ]
    fake_jira = SimpleNamespace(
        _options={"server": "https://jira.example.com"},
        search_issues=Mock(side_effect=[issues]),
    )
    config = ConfigParser()
    with patch(
        "stats_core.reports.jira_comprehensive.rewrite_comment_items_with_ai",
        return_value={},
    ):
        df = build_comments_period_df(
            fake_jira,
            "project = ABC ORDER BY created DESC",
            None,
            None,
            config,
            {},
        )
    row = df.iloc[0]
    assert "first" in str(row["Comments_In_Period"])
    assert "second" in str(row["Comments_In_Period"])


def test_build_comments_period_df_sorts_by_epic_then_parent():
    issue_parent = _make_issue_with_comments(
        "ABC-2",
        [_make_comment("parent", created="2025-02-01T10:00:00.000+0000")],
    )
    issue_parent.fields.customfield_10000 = "EPIC-1"
    issue_parent.fields.parent = SimpleNamespace(key="ABC-1")

    issue_child = _make_issue_with_comments(
        "ABC-1",
        [_make_comment("child", created="2025-02-01T10:00:00.000+0000")],
    )
    issue_child.fields.customfield_10000 = "EPIC-1"
    fake_jira = SimpleNamespace(
        _options={"server": "https://jira.example.com"},
        search_issues=Mock(side_effect=[[issue_parent, issue_child]]),
    )
    config = ConfigParser()
    with patch(
        "stats_core.reports.jira_comprehensive.rewrite_comment_items_with_ai",
        return_value={},
    ), patch(
        "stats_core.reports.jira_comprehensive._fetch_epic_metadata",
        return_value={"EPIC-1": {"name": "Epic One"}},
    ):
        df = build_comments_period_df(
            fake_jira,
            "project = ABC ORDER BY created DESC",
            None,
            None,
            config,
            {"ai_comments_enabled": "false"},
        )
    assert df.iloc[0]["Parent"] == ""
    assert df.iloc[1]["Parent"] == "ABC-1"


def test_build_comment_summary_prompt_has_required_fields():
    _, prompt = _build_comment_summary_prompt([{"id": "ABC-1", "comments": "text"}])
    assert "JSON" in prompt
    assert "done" in prompt
    assert "planned" in prompt
    assert "risks" in prompt
    assert "dependencies" in prompt
    assert "notes" in prompt
    assert "Insufficient data" in prompt


def test_extract_results_hint_from_link_only():
    hint = _extract_results_hint("results: https://example.com/result")
    assert hint == "Results provided (link removed)."


def test_format_ai_comment_summary_uses_results_hint():
    comments = "Results: https://example.com/result"
    summary = _format_ai_comment_summary({}, _extract_results_hint(comments), comments)
    assert summary == "Results provided (link removed)."


def test_format_ai_comment_summary_handles_timeout_error():
    summary = _format_ai_comment_summary({"__error__": "timeout"}, None, "some comment")
    assert summary == "AI request timeout."


def test_rewrite_comment_items_with_ai_respects_flag():
    config = ConfigParser()
    result = rewrite_comment_items_with_ai(
        [{"id": "ABC-1", "comments": "text"}],
        config,
        {"ai_comments_enabled": "false"},
    )
    assert result == {"ABC-1": {"__error__": "AI comments disabled"}}


def test_fetch_jira_data_results_convert_attachment_markers_to_links():
    attachment_url = "https://jira.example.com/secure/attachment/123/demo.pptx"
    issues = [
        _make_issue(
            "ABC-99",
            summary="Attachment result",
            issue_type="Task",
            status="Released",
            assignee_display="Alice",
            assignee_username="alice",
            reporter_display="Bob",
            reporter_username="bob",
            comment_body="Results: [^demo.pptx]",
            comment_id="991",
            resolved_at="2025-01-15T10:00:00.000+0000",
            attachments=[SimpleNamespace(filename="demo.pptx", content=attachment_url)],
        )
    ]
    epic_issues = [
        _make_issue(
            "EPIC-1",
            summary="Epic One",
            issue_type="Epic",
            status="Released",
            assignee_display="Carol",
            assignee_username="carol",
            reporter_display="Carol",
            reporter_username="carol",
        )
    ]

    fake_jira = Mock()
    fake_jira._options = {"server": "https://jira.example.com"}

    def _search_issues_side_effect(jql_query, *args, **kwargs):
        if "issuekey in" in jql_query:
            return epic_issues
        return issues

    fake_jira.search_issues.side_effect = _search_issues_side_effect

    _, _, results_df = fetch_jira_data(fake_jira, "project = ABC ORDER BY created DESC")

    row = results_df[results_df["Issue_Key"] == "ABC-99"].iloc[0]
    assert row["Result"] == attachment_url
    assert attachment_url in str(row["Result_Links"])


def test_calculate_engineer_metrics_uses_members_jira_column():
    issues_df = pd.DataFrame(
        [
            {
                "Issue_Key": "ABC-1",
                "Summary": "Bug fix",
                "Type": "Bug",
                "Status": "Released",
                "Resolution": "Done",
                "Assignee": "Does Not Matter",
                "Assignee_Username": "eWX1025804",
                "Reporter": "Someone",
                "Reporter_Username": "someone",
                "Created": "2025-01-01",
                "Resolved": "2025-01-10",
                "Labels": "",
            }
        ]
    )
    members_df = pd.DataFrame(
        [
            {
                "name": "Evstigneev Roman",
                "username": "wx1025804",
                "Jira": "eWX1025804",
                "role": "engineer",
            }
        ]
    )
    code_volume_df = pd.DataFrame(columns=["username", "code_volume"])

    engineer_metrics = calculate_engineer_metrics(
        issues_df,
        members_df,
        code_volume_df,
        pd.DataFrame(columns=["Issue_Key", "Assignee", "Worklog_Author"]),
    )
    assert engineer_metrics.shape[0] == 1
    row = engineer_metrics.iloc[0].to_dict()
    assert row["Bugs"] == 1
    assert row["Total_Resolved_Issues"] == 1


def test_calculate_engineer_metrics_counts_assistance_from_unique_foreign_worklogs():
    issues_df = pd.DataFrame(
        [
            {
                "Issue_Key": "ABC-1",
                "Summary": "Own task",
                "Type": "Task",
                "Status": "In Progress",
                "Resolution": "",
                "Assignee": "Alice Dev",
                "Assignee_Username": "alice",
                "Reporter": "Someone",
                "Reporter_Username": "someone",
                "Created": "2025-01-01",
                "Resolved": "",
                "Labels": "",
            },
            {
                "Issue_Key": "ABC-2",
                "Summary": "Foreign task one",
                "Type": "Task",
                "Status": "In Progress",
                "Resolution": "",
                "Assignee": "Bob Dev",
                "Assignee_Username": "bob",
                "Reporter": "Someone",
                "Reporter_Username": "someone",
                "Created": "2025-01-01",
                "Resolved": "",
                "Labels": "",
            },
            {
                "Issue_Key": "ABC-3",
                "Summary": "Foreign task two",
                "Type": "Task",
                "Status": "In Progress",
                "Resolution": "",
                "Assignee": "Charlie Dev",
                "Assignee_Username": "charlie",
                "Reporter": "Someone",
                "Reporter_Username": "someone",
                "Created": "2025-01-01",
                "Resolved": "",
                "Labels": "",
            },
        ]
    )
    members_df = pd.DataFrame(
        [
            {
                "name": "Alice Dev",
                "username": "alice",
                "Jira": "alice",
                "role": "engineer",
            }
        ]
    )
    code_volume_df = pd.DataFrame(columns=["username", "code_volume"])
    worklog_entries_df = pd.DataFrame(
        [
            {"Issue_Key": "ABC-1", "Assignee": "Alice Dev", "Worklog_Author": "alice"},
            {"Issue_Key": "ABC-2", "Assignee": "Bob Dev", "Worklog_Author": "alice"},
            {"Issue_Key": "ABC-2", "Assignee": "Bob Dev", "Worklog_Author": "alice"},
            {"Issue_Key": "ABC-3", "Assignee": "Charlie Dev", "Worklog_Author": "alice"},
        ]
    )

    engineer_metrics = calculate_engineer_metrics(
        issues_df,
        members_df,
        code_volume_df,
        worklog_entries_df,
    )

    assert engineer_metrics.shape[0] == 1
    row = engineer_metrics.iloc[0].to_dict()
    assert row["Assistance_Provided"] == 2


def test_build_monthly_summary_df_uses_ai_rewrite_map_when_available():
    issues_df = pd.DataFrame(
        [
            {
                "Issue_Key": "EPIC-1",
                "Summary": "Epic one",
                "Type": "Epic",
                "Status": "In Progress",
                "Resolution": "",
                "Resolved": "",
                "Labels": "report",
                "Epic_Link": "",
                "Epic_Name": "Epic One",
            },
            {
                "Issue_Key": "ABC-1",
                "Summary": "Task one",
                "Type": "Task",
                "Status": "Done",
                "Resolution": "Done",
                "Resolved": "2025-01-02",
                "Description": "Initial implementation.",
                "Last_Comment": "Feature completed.",
                "Labels": "",
                "Epic_Link": "EPIC-1",
                "Epic_Name": "Epic One",
            },
            {
                "Issue_Key": "ABC-2",
                "Summary": "Bug one",
                "Type": "Bug",
                "Status": "Done",
                "Resolution": "Done",
                "Resolved": "2025-01-03",
                "Description": "",
                "Last_Comment": "Bug fixed.",
                "Labels": "",
                "Epic_Link": "EPIC-1",
                "Epic_Name": "Epic One",
            },
        ]
    )
    config = ConfigParser()

    with patch("stats_core.reports.jira_comprehensive.rewrite_summary_items_with_ai") as mock_rewrite:
        mock_rewrite.return_value = {"EPIC-1::ABC-1": "Delivered ArkUI workflow and stabilized behavior."}
        summary_df = build_monthly_summary_df(issues_df, config, extra_params={})

    assert summary_df.shape[0] == 1
    row = summary_df.iloc[0].to_dict()
    assert row["Epic_Name"] == "Epic One"
    assert row["Planned_Tasks_Resolved"] == 1
    assert row["Reported_Issues_Resolved"] == 1
    summary_text = str(row["Summary"])
    assert "Delivered ArkUI workflow and stabilized behavior." in summary_text
    assert "Resolved 1 planned tasks on time." in summary_text
    assert "Resolved 1 reported issues." in summary_text
    assert "- ABC-1:" not in summary_text
    assert re.search(r"\b[A-Z]+-\d+\b", summary_text) is None


def test_build_monthly_summary_df_filters_report_epics_and_groups_subtasks():
    issues_df = pd.DataFrame(
        [
            {
                "Issue_Key": "EPIC-REPORT",
                "Summary": "Report epic",
                "Type": "Epic",
                "Status": "In Progress",
                "Resolution": "",
                "Resolved": "",
                "Labels": "report",
                "Epic_Link": "",
                "Epic_Name": "Report Epic",
            },
            {
                "Issue_Key": "EPIC-NOREPORT",
                "Summary": "Other epic",
                "Type": "Epic",
                "Status": "In Progress",
                "Resolution": "",
                "Resolved": "",
                "Labels": "other",
                "Epic_Link": "",
                "Epic_Name": "Other Epic",
            },
            {
                "Issue_Key": "EPIC-CLOSED-OLD",
                "Summary": "Closed old epic",
                "Type": "Epic",
                "Status": "Done",
                "Resolution": "Done",
                "Resolved": "2024-12-01",
                "Labels": "report",
                "Epic_Link": "",
                "Epic_Name": "Closed Old Epic",
            },
            {
                "Issue_Key": "ABC-10",
                "Summary": "Parent feature",
                "Type": "Task",
                "Status": "Done",
                "Resolution": "Done",
                "Resolved": "2025-01-10",
                "Description": "Parent implementation",
                "Last_Comment": "Parent delivered",
                "Labels": "",
                "Epic_Link": "EPIC-REPORT",
                "Epic_Name": "Report Epic",
                "Parent": "",
                "Parent_Summary": "",
            },
            {
                "Issue_Key": "ABC-11",
                "Summary": "Subtask one",
                "Type": "Sub-task",
                "Status": "Done",
                "Resolution": "Done",
                "Resolved": "2025-01-11",
                "Description": "Subtask part one",
                "Last_Comment": "Part one done",
                "Labels": "",
                "Epic_Link": "EPIC-REPORT",
                "Epic_Name": "Report Epic",
                "Parent": "ABC-10",
                "Parent_Summary": "Parent feature",
            },
            {
                "Issue_Key": "ABC-12",
                "Summary": "Subtask two",
                "Type": "Sub-task",
                "Status": "Done",
                "Resolution": "Done",
                "Resolved": "2025-01-12",
                "Description": "Subtask part two",
                "Last_Comment": "Part two done",
                "Labels": "",
                "Epic_Link": "EPIC-REPORT",
                "Epic_Name": "Report Epic",
                "Parent": "ABC-10",
                "Parent_Summary": "Parent feature",
            },
            {
                "Issue_Key": "ABC-20",
                "Summary": "Excluded by epic label",
                "Type": "Task",
                "Status": "Done",
                "Resolution": "Done",
                "Resolved": "2025-01-15",
                "Description": "Should be excluded",
                "Last_Comment": "",
                "Labels": "",
                "Epic_Link": "EPIC-NOREPORT",
                "Epic_Name": "Other Epic",
            },
            {
                "Issue_Key": "ABC-30",
                "Summary": "Excluded by old closed epic",
                "Type": "Task",
                "Status": "Done",
                "Resolution": "Done",
                "Resolved": "2025-01-16",
                "Description": "Should be excluded",
                "Last_Comment": "",
                "Labels": "",
                "Epic_Link": "EPIC-CLOSED-OLD",
                "Epic_Name": "Closed Old Epic",
            },
        ]
    )
    config = ConfigParser()

    with patch("stats_core.reports.jira_comprehensive.rewrite_summary_items_with_ai") as mock_rewrite:
        mock_rewrite.side_effect = lambda items, *_: {item["id"]: f"Grouped summary for {item['id']}" for item in items}
        summary_df = build_monthly_summary_df(
            issues_df,
            config,
            extra_params={"start": "2025-01-01", "end": "2025-01-31"},
        )

    ai_inputs = mock_rewrite.call_args.args[0]
    assert len(ai_inputs) == 1
    assert "subtask" in ai_inputs[0]["description"].casefold()

    assert summary_df.shape[0] == 1
    row = summary_df.iloc[0].to_dict()
    assert row["Epic_Link"] == "EPIC-REPORT"
    assert row["Planned_Tasks_Resolved"] == 3
    summary_text = str(row["Summary"])
    assert summary_text.count("- ") == 1


@patch("stats_core.reports.jira_comprehensive.rewrite_comment_items_with_ai")
@patch("stats_core.reports.jira_comprehensive.JiraSource")
def test_jira_comprehensive_report_adds_developer_activity_sheet(
    mock_jira_source_cls,
    mock_rewrite_comments,
    tmp_path: Path,
):
    mock_rewrite_comments.return_value = {}

    issue_one = _make_issue(
        "ABC-1",
        summary="First task",
        issue_type="Task",
        status="Released",
        assignee_display="Alice Dev",
        assignee_username="alice.dev",
        reporter_display="Bob",
        reporter_username="bob",
        resolved_at="2025-01-18T10:00:00.000+0000",
    )
    issue_one.fields.comment.comments = [
        SimpleNamespace(
            body="Implemented API changes",
            author=SimpleNamespace(displayName="Alice Dev", name="alice.dev"),
            created="2025-01-14T10:00:00.000+0000",
            updated="2025-01-14T10:00:00.000+0000",
            id="101",
        ),
        SimpleNamespace(
            body="Added tests",
            author=SimpleNamespace(displayName="Alice Dev", name="alice.dev"),
            created="2025-01-15T12:00:00.000+0000",
            updated="2025-01-15T12:00:00.000+0000",
            id="102",
        ),
    ]

    issue_two = _make_issue(
        "ABC-2",
        summary="Second task",
        issue_type="Task",
        status="In QA",
        assignee_display="Bob Dev",
        assignee_username="bob.dev",
        reporter_display="Carol",
        reporter_username="carol",
        resolved_at="2025-01-19T10:00:00.000+0000",
    )
    issue_two.fields.comment.comments = [
        SimpleNamespace(
            body="Waiting for review",
            author=SimpleNamespace(displayName="Bob Dev", name="bob.dev"),
            created="2025-01-16T09:00:00.000+0000",
            updated="2025-01-16T09:00:00.000+0000",
            id="201",
        )
    ]

    issue_three = _make_issue(
        "ABC-3",
        summary="Only worklog task",
        issue_type="Task",
        status="In Progress",
        assignee_display="Carol Dev",
        assignee_username="carol.dev",
        reporter_display="Dan",
        reporter_username="dan",
        resolved_at="2025-01-20T10:00:00.000+0000",
    )
    issue_three.fields.comment.comments = []

    issues = [issue_one, issue_two, issue_three]
    epic_issues: list[SimpleNamespace] = []

    fake_jira = Mock()
    fake_jira._options = {"server": "https://jira.example.com"}

    def _search_issues_side_effect(jql_query, *args, **kwargs):
        if "issuekey in" in jql_query:
            return epic_issues
        return issues

    fake_jira.search_issues.side_effect = _search_issues_side_effect

    fake_source = Mock()
    fake_source.jira = fake_jira
    fake_source.jira_url = "https://jira.example.com"

    def _worklogs(issue_key):
        if issue_key == "ABC-1":
            return [
                {
                    "author": {"displayName": "Alice Dev", "name": "alice.dev"},
                    "started": "2025-01-14T10:00:00.000+0000",
                    "timeSpentSeconds": 3600,
                    "comment": "Investigated root cause",
                },
                {
                    "author": {"displayName": "Alice Dev", "name": "alice.dev"},
                    "started": "2025-01-15T11:00:00.000+0000",
                    "timeSpentSeconds": 1800,
                    "comment": "",
                },
                {
                    "author": {"displayName": "Reviewer", "name": "reviewer"},
                    "started": "2025-01-15T12:00:00.000+0000",
                    "timeSpentSeconds": 1200,
                    "comment": "Review support",
                },
            ]
        if issue_key == "ABC-2":
            return []
        if issue_key == "ABC-3":
            return [
                {
                    "author": {"displayName": "Carol Dev", "name": "carol.dev"},
                    "started": "2025-01-16T10:00:00.000+0000",
                    "timeSpentSeconds": 1200,
                    "comment": "Logged time only",
                }
            ]
        return []

    fake_source.get_all_worklogs.side_effect = _worklogs
    mock_jira_source_cls.return_value = fake_source

    members_file = tmp_path / "members.xlsx"
    pd.DataFrame(
        [
            {"name": "Alice Dev", "username": "alice.dev", "role": "engineer"},
            {"name": "Bob Dev", "username": "bob.dev", "role": "engineer"},
            {"name": "Carol Dev", "username": "carol.dev", "role": "engineer"},
        ]
    ).to_excel(members_file, index=False)

    config = ConfigParser()
    config.read_dict(
        {
            "jira": {"jira-url": "https://jira.example.com", "username": "u", "password": "p"},
            "reporting": {"output_dir": str(tmp_path)},
        }
    )

    report = JiraComprehensiveReport()
    report.run(
        dataset={},
        config=config,
        output_formats=["excel"],
        extra_params={
            "project": "ABC",
            "start": "2025-01-01",
            "end": "2025-01-31",
            "member_list_file": str(members_file),
            "output": "developer_activity.xlsx",
        },
    )

    workbook_path = tmp_path / "developer_activity.xlsx"
    assert workbook_path.exists()

    wb = load_workbook(workbook_path)
    assert "Developer_Activity" in wb.sheetnames

    ws = wb["Developer_Activity"]
    headers = [cell.value for cell in ws[1]]
    assert headers == ["Developer", "Issue", "Title", "Logged_Hours", "Worklog", "Comments"]

    rows = list(ws.iter_rows(min_row=2, values_only=False))
    assert len(rows) == 2

    data_rows = []
    for row in rows:
        row_map = {headers[idx]: cell.value for idx, cell in enumerate(row)}
        issue_cell = row[headers.index("Issue")]
        row_map["IssueHyperlink"] = issue_cell.hyperlink.target if issue_cell.hyperlink else None
        data_rows.append(row_map)

    alice_row = next(row for row in data_rows if row["Developer"] == "Alice Dev")
    assert alice_row["Issue"] == "ABC-1"
    assert alice_row["IssueHyperlink"] == "https://jira.example.com/browse/ABC-1"
    assert alice_row["Title"] == "First task"
    assert alice_row["Logged_Hours"] == 1.5
    assert "Investigated root cause" in str(alice_row["Worklog"])
    assert "Implemented API changes" in str(alice_row["Comments"])
    assert "Added tests" in str(alice_row["Comments"])

    bob_row = next(row for row in data_rows if row["Developer"] == "Bob Dev")
    assert bob_row["Issue"] == "ABC-2"
    assert bob_row["Logged_Hours"] in (0, 0.0)
    assert bob_row["Worklog"] in ("", None)
    assert "Waiting for review" in str(bob_row["Comments"])

    developers = {row["Developer"] for row in data_rows}
    assert "Carol Dev" not in developers
