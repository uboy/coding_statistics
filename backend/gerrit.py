# -*- coding: utf-8 -*-
import sys                                          # print etc
# import requests                                   #
import csv                                          # working with csv files
import codecs                                       # used for text encoding in config parser
import argparse                                     # good argument parser
import logging                                      #
# import re                                         # for parsing ticket no from description
from openpyxl import load_workbook                  # working with Excel files
from configparser import ConfigParser               # able to read configuration file
# from datetime import datetime                     # used for manipulations with dates
from pygerrit2 import GerritRestAPI, HTTPBasicAuth  # for gerrit and basic authentication

CONFIG_POINT_LOCAL = "gerrit"
CONFIG_POINT_GLOBAL = "global"
CONFIG_FILE = "config.ini"
CONFIG_BASE_URL = "gerrit-url"
CONFIG_MEMBER_LIST = "member-list"
CONFIG_USERNAME = "username"
CONFIG_PASSWORD = "password"


class OutputChangeItem:
    def __init__(self, change_link, change_added, change_deleted, change_total, change_created, change_merged,
                 change_owner, change_owner_name, msg_author, msg_date, comments_cnt):
        self.change_link = change_link
        self.change_owner = change_owner
        self.change_owner_name = change_owner_name
        self.msg_author = msg_author
        self.msg_date = msg_date
        self.change_added = change_added
        self.change_deleted = change_deleted
        self.change_total = change_total
        self.change_created = change_created
        self.change_merged = change_merged
        self.comments_cnt = comments_cnt

    # Format the time 2017-12-11 02:07:27.000000000 to 2017-12-11
    def get_display_time(self):
        return self.msg_date[0:19]

    def get_created_date(self):
        return self.msg_date[0:19]


def main():
    #   argument parsing section
    tool_description = 'Track comments left in code by using Gerrit API'
    parser = argparse.ArgumentParser(description=tool_description,
                                     formatter_class=argparse.ArgumentDefaultsHelpFormatter)
    parser.add_argument('-u', '--username', dest='username', help='username')
    parser.add_argument('-p', '--password', dest='password', help='password')
    parser.add_argument('-l', '--url-link', dest='base_url', help='the gerrit url')
    parser.add_argument('-f', '--file', dest='member_list_file', help='Group member list file')
    parser.add_argument('-d', '--date', dest='date_start', required=True,
                        help='query date string start, eg. 2019-06 or 2019-06-25')
    parser.add_argument('-v', '--verbose', dest='verbose', action='store_true', help='enable verbose (debug) logging')

    options = parser.parse_args()
    level = logging.DEBUG if options.verbose else logging.INFO
    logging.basicConfig(format='%(asctime)s %(levelname)s %(message)s', level=level)

    base_url = options.base_url
    username = options.username
    password = options.password
    date_start = options.date_start
    member_list_file = options.member_list_file
    config = ConfigParser()
    config.read_file(codecs.open(CONFIG_FILE, 'r', encoding='utf-8-sig'))

    if base_url is None:
        base_url = config.get(CONFIG_POINT_LOCAL, CONFIG_BASE_URL)
    if username is None:
        username = config.get(CONFIG_POINT_LOCAL, CONFIG_USERNAME)
    if password is None:
        password = config.get(CONFIG_POINT_LOCAL, CONFIG_PASSWORD)
    if member_list_file is None:
        member_list_file = config.get(CONFIG_POINT_GLOBAL, CONFIG_MEMBER_LIST)

    if base_url is None or username is None or password is None or member_list_file is None:
        raise ValueError("url or username or password or file is invalid")

    member_list = read_member_list("members.xlsx")
    # Create a report file with headlines
    file_name = "gerrit-comments-" + date_start + ".csv"
    create_csv_file(file_name)
    # create dict with members
    member_dict = {}
    for member in member_list:
        member_id = member["Username"]
        if member_id not in member_dict:
            member_dict[member_id] = member

    auth = HTTPBasicAuth(username, password)
    rest = GerritRestAPI(url=base_url, auth=auth)
    #   FULLLIST=$(ssh -p $GERRIT_PORT -l $GERRIT_USER $GERRIT_URL gerrit query --format=JSON
    #   ${STARTQUERY} ${FINISHQUERY}
    #   ${PROJQUERY} $BRQUERY --patch-sets --comments --all-approvals)
    query_limit = 1000
    query_start = 0
    changes_details = []
    gerrit_pagination = True
    while gerrit_pagination:
        changes_details_link = "/changes/?q=" \
                               "after:" + date_start + "%20project:project" \
                               "&o=MESSAGES&o=DETAILED_ACCOUNTS&n=" + str(query_limit) + "&start=" + str(query_start)
        query_response = rest.get(changes_details_link)
        changes_details = changes_details + query_response
        if "_more_changes" in query_response[len(query_response)-1]:
            gerrit_pagination = True
        else:
            gerrit_pagination = False
        query_start = query_start + min(query_limit, len(query_response))
    result_changes = []
    for change in changes_details:
        change_id = change["_number"]
        change_link = base_url + "/" + str(change_id)
        change_added = change["insertions"]
        change_deleted = change["deletions"]
        change_total = change_added + change_deleted
        change_created = change["created"][0:10]
        change_merged = ""
        if change["status"] == 'MERGED':
            change_merged = change["updated"][0:10]
        change_owner = change["owner"]["username"]
        change_owner_name = change["owner"]["name"]
        messages = change["messages"]
        comment_stat = {}
        for message in messages:
            if "author" in message:
                msg_author = message["author"]["username"]
                if msg_author != change_owner and msg_author in member_dict:
                    if msg_author in comment_stat:
                        msg_date = message["date"][0:10]
                        comments_cnt = comment_stat[msg_author][0] + 1
                        comment_stat[msg_author] = [comments_cnt, msg_date]
                    else:
                        msg_date = message["date"][0:10]
                        comment_stat[msg_author] = [1, msg_date]
        for comment_author in comment_stat:
            result_changes.append(OutputChangeItem(
                change_link, change_added, change_deleted,
                change_total, change_created, change_merged, change_owner,
                change_owner_name, comment_author, comment_stat[comment_author][1], comment_stat[comment_author][0]
            ))

    with open(file_name, "a", encoding='utf-8-sig', newline='') as csvfile:
        writer = csv.writer(csvfile)
        for change in result_changes:
            writer.writerow([change.change_link,
                             change.change_added,
                             change.change_deleted,
                             change.change_total,
                             change.change_owner,
                             change.change_created,
                             change.change_merged,
                             change.msg_author,
                             change.comments_cnt,
                             change.msg_date,
                             change.change_owner_name
                             ])
    print("All done!")
    return 0


def read_member_list(member_list_file):
    wb = load_workbook(member_list_file)
    sheet = wb.active
    has_name = "name" == sheet.cell(row=1, column=1).value
    has_mail = "email" == sheet.cell(row=1, column=2).value
    has_username = "username" == sheet.cell(row=1, column=3).value

    member_list = []
    if has_name and has_mail and has_username:
        for rx in range(2, sheet.max_row + 1):
            name = sheet.cell(row=rx, column=1).value
            mail = sheet.cell(row=rx, column=2).value
            username = sheet.cell(row=rx, column=3).value
            member_list.append({'Name': name, 'Email': mail, 'Username': username})
    else:
        raise ValueError("The table format is incorrect")
    return member_list


def create_csv_file(file_name):
    # Write Title
    with open(file_name, "w", encoding='utf-8-sig', newline='') as csvfile:
        writer = csv.writer(csvfile)
        writer.writerow(
            ["Review", "Added", "Deleted", "Total", "Owner", "Created", "Merged", "Reviewer", "Number", "Last Date"])


if __name__ == '__main__':
    # TODO check exit code
    sys.exit(main())
