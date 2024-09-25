# coding:utf-8
import sys
import argparse
import logging
import json
import re
import csv
import codecs
import functools
import time
import typing
from typing import Pattern

from pygerrit2 import GerritRestAPI, HTTPBasicAuth
from requests.exceptions import RequestException, HTTPError
from openpyxl import load_workbook
from configparser import ConfigParser
from concurrent.futures import ThreadPoolExecutor


class MemberInfo:
    def __init__(self, name: str, mail: str = '', full_name: str = '', account_id: int = 0, url: str = ''):
        self.__name = name
        self.__mail = mail
        self.__full_name = full_name
        self.__account_id = account_id
        self.__url = url

    def get_name(self):
        return self.__name

    def get_mail(self):
        return self.__mail

    def get_full_name(self):
        return self.__full_name

    def set_url(self, url: str):
        self.__url = url

    def get_url(self):
        return self.__url

    def set_account_id(self, account_id: int):
        self.__account_id = account_id

    def get_account_id(self):
        return self.__account_id

    def __eq__(self, other):
        return self.__url == other.__url and self.__account_id == other.__account_id

    def __hash__(self) -> int:
        return hash("%s%s" % (self.__url, self.__account_id))


class OutputChangeItem:
    def __init__(self, change_team, change_type, change_rejected, change_id, change_owner, change_project,
                 change_created, change_time, change_link, change_branch, change_reason,
                 change_description, change_code_count, change_inner_group_comments_count=0,
                 change_outer_group_comments_count=0, message_member=""):
        self.change_team = change_team
        self.change_type = change_type
        self.change_rejected = change_rejected
        self.change_id = change_id
        self.change_owner = change_owner
        self.change_project = change_project
        self.change_created = change_created
        self.change_time = change_time
        self.change_link = change_link
        self.change_branch = change_branch
        self.change_reason = change_reason
        self.change_description = change_description
        self.change_code_count = change_code_count
        self.change_inner_group_comments_count = change_inner_group_comments_count
        self.change_outer_group_comments_count = change_outer_group_comments_count
        self.message_member = message_member

    # Format the time 2017-12-11 02:07:27.000000000 to 2017-12-11
    def get_display_time(self):
        return self.change_time[0:19]

    def get_created_date(self):
        return self.change_created[0:19]


CONFIG_FILE = "config.ini"
CONFIG_POINT_LOCAL = "gerrit"
CONFIG_POINT_GLOBAL = "global"
CONFIG_BASE_URL = "gerrit-url"
CONFIG_USERNAME = "username"
CONFIG_PASSWORD = "password"
CONFIG_MEMBER_LIST = "member-list"

MEMBER_CACHE_FILE = "member_cache.json"

KEY_UPDATED = "updated"
KEY_CREATED = "created"
KEY_PROJECT = "project"
KEY_BRANCH = "branch"
KEY_OWNER = "owner"
KEY_ACCOUNT_ID = "_account_id"
KEY_NUMBER = "_number"
KEY_SUBJECT = "subject"
KEY_INSERT = "insertions"
KEY_DELETE = "deletions"
KEY_REVISIONS = "revisions"
KEY_REVISION_NUMBER = "_number"
KEY_MSG_MEMBER = "author"
KEY_MSG_MEMBER_ID = "_account_id"
KEY_TEAM = "team"


def cache_member_list(cache_file):
    def wrapper(original_get_member_id):
        try:
            cache = json.load(open(cache_file, 'r'))
        except (IOError, ValueError):
            cache = {}

        @functools.wraps(original_get_member_id)
        def new_get_member_id(rest_instance, member: MemberInfo):
            base_url = member.get_url()
            mail = member.get_mail()
            if base_url not in cache:
                print("url not in cache: %s" % base_url)
                member_id = original_get_member_id(rest_instance, member)
                if member_id is not None:
                    cache[base_url] = {mail: member_id}
                    json.dump(cache, open(cache_file, 'w'), indent=4)
                    return member_id
            else:
                member_dict = cache[base_url]
                if mail not in member_dict:
                    print("mail not in cache: %s" % mail)
                    member_id = original_get_member_id(rest_instance, member)
                    if member_id is not None:
                        member_dict[mail] = member_id
                        json.dump(cache, open(cache_file, 'w'), indent=4)
                        return member_id
                else:
                    return cache[base_url][mail]

        return new_get_member_id

    return wrapper


@cache_member_list("member_cache.json")
def get_member_id(rest_instance, member: MemberInfo):
    mail = member.get_mail()
    full_name = member.get_full_name()
    if mail is not None:
        account_link = 'accounts/?q=%s' % member.get_mail()
    elif full_name is not None:
        account_link = 'accounts/?q=%s' % member.get_mail()
    else:
        raise ValueError("Should at least input mail or full name")

    try:
        accounts = rest_instance.get(account_link)
    except HTTPError as e:
        if rest_instance.url.startswith('http://exception'):
            payload = {
                'jsonrpc': '2.0',
                'method': 'suggestAccount',
                'params': [member.get_mail(), "true", 20]
            }
            payload_header = {
                'Accept': 'application/json,application/json,application/jsonrequest',
                'Content-Type': 'application/json; charset=UTF-8',
            }
            response = rest_instance.session.post(
                'http://exception', data=json.dumps(payload),
                headers=payload_header)
            response_json = json.loads(response.content)
            if 'result' in response_json:
                response_members = response_json['result']
                for response_member in response_members:
                    # If multiple search results are found, there is a problem.
                    # Theoretically, only one search result is found by email or FullName
                    print(response_member)
                    return response_member['id']['id']
        print('The server could not fulfill the request.')
        print('Error code: ', e.args)
        return None
    for account_info in accounts:
        # If multiple search results are found, there is a problem. Theoretically,
        # only one search result is found by email or FullName
        print(account_info)
        return account_info[KEY_ACCOUNT_ID]


def read_member_list(member_list_file):
    wb = load_workbook(member_list_file)
    sheet = wb.active
    has_name = "name" == sheet.cell(row=1, column=1).value
    has_mail = "email" == sheet.cell(row=1, column=2).value

    member_list = []
    if has_name and has_mail:
        for rx in range(2, sheet.max_row + 1):
            name = sheet.cell(row=rx, column=1).value
            mail = sheet.cell(row=rx, column=2).value
            member_list.append(MemberInfo(name, mail))
    else:
        raise ValueError("The table format is incorrect")
    return member_list


def transfer_member_list(rest: GerritRestAPI, base_url: str, member_list: typing.List[MemberInfo]):
    member_dict = {}
    for member in member_list:
        member.set_url(base_url)
        member_id = get_member_id(rest, member)
        if member_id is not None:
            member.set_account_id(member_id)
            member_dict[member_id] = member
    return member_dict


def get_reason(subject):
    #   NOTE: subject can be just first line from commit message
    #   That happens when user send commit message with empty string after header
    #    print(subject)
    #    print("dts:"+";".join(dts_number)+" ar:"+";".join(ar_number)+" result:"+reason)
    description_pattern = re.compile(r'(?<=Description:)(.*?)(?=|Team:|Change-Id)', re.S)
    description_all = re.search(description_pattern, subject)
    if description_all:
        description = description_all.group(1).replace("\n", " ")
    else:
        description = ""
    reason = ""
    return reason, description


def create_csv_file(file_name):
    # Write Title     
    print(file_name)
    with open(file_name, "w", encoding='utf-8-sig', newline='') as csvfile:
        writer = csv.writer(csvfile)
        writer.writerow(["Team", "gerrit id", "description", "Created", "Merged", "LOC",
                         "change_inner_group_comments_count", "Author", "AR", "Type",
                         "Rejected", "Project", "branch", "Reviewed by", "change_outer_group_comments_count"])


def fetch_change(url, rest: GerritRestAPI, member: MemberInfo, member_dict, date_filter_string: str):
    #   https://gerrit-review.googlesource.com/Documentation/rest-api-changes.html
    date_filter_patten = re.compile(date_filter_string)
    comment_pattern = re.compile(r'(\d+?)(?=\scomments?\))')
    account_id = member.get_account_id()
    account_name = member.get_name()
    change_link = "/changes/?q=owner:" + \
                  str(account_id) + \
                  "%20status:merged&o=ALL_COMMITS&o=DETAILED_ACCOUNTS&o=CURRENT_REVISION"
    changes = []
    try:
        changes = rest.get(change_link)
    except HTTPError as e:
        print('The server could not fulfill the request.')
        print('Error code: ', e.args)
    filtered_changes = []
    for change in changes:
        if KEY_UPDATED in change:
            if re.search(date_filter_patten, str(change[KEY_UPDATED])):
                filtered_changes.append(change)
    print("change count:\t" + str(len(filtered_changes)) + "\t" + str(account_name) + "\n")
    result_changes = []
    for change in filtered_changes:
        # Gets review comments for each submission
        change_id = change[KEY_NUMBER]
        change_team = KEY_TEAM
        change_owner = member_dict[change[KEY_OWNER][KEY_ACCOUNT_ID]].get_name()
        change_owner_email = change["owner"]["email"]
        change_project = change[KEY_PROJECT]
        change_type = ""
        change_rejected = ""
        change_time = change[KEY_UPDATED]
        change_created = change[KEY_CREATED]
        change_link = url + "/" + str(change_id)
        change_branch = change[KEY_BRANCH]
        #       change_reason, change_description = get_reason(change[KEY_SUBJECT])
        change_reason, change_description = \
            get_reason(change["revisions"][list(change["revisions"])[0]]["commit"]["message"])
        change_code_count = change[KEY_INSERT] + change[KEY_DELETE]
        detail_link = "/changes/%s/detail" % change_id
        change_outer_group_comments_count = 0
        change_inner_group_comments_count = 0
        message_member = ""
        detail = {}
        try:
            detail = rest.get(detail_link)
        except HTTPError as e:
            print('The server could not fulfill the request.')
            print('Error code: ', e.args)
        if "messages" in detail:
            change_messages = detail["messages"]
            for change_message in change_messages:
                comment_message = re.search(comment_pattern, change_message["message"])
                if comment_message:
                    comment_member_id = change_message["author"][KEY_ACCOUNT_ID]
                    comment_member_email = change_message["author"]["email"]
                    #                    print(change_owner_email)
                    if change_owner_email != comment_member_email:
                        if comment_member_id not in member_dict:
                            change_outer_group_comments_count += int(comment_message.group(1))
                        else:
                            change_inner_group_comments_count += int(comment_message.group(1))
        if "labels" in detail:
            all_reviewer = detail["labels"]["Code-Review"]["all"]
            all_reviewer_name = []
            for reviewer in all_reviewer:
                if reviewer["value"] == 1 and change_owner_email != reviewer["email"]:
                    reviewer_id = reviewer[KEY_ACCOUNT_ID]
                    if reviewer_id in member_dict:
                        member = member_dict[reviewer_id]
                        all_reviewer_name += [member.get_name()]
            message_member = ";".join(all_reviewer_name)
        result_changes.append(
            OutputChangeItem(change_team, change_type, change_rejected, change_id, change_owner, change_project,
                             change_created, change_time, change_link, change_branch,
                             change_reason, change_description, change_code_count, change_inner_group_comments_count,
                             change_outer_group_comments_count, message_member=message_member))
    return result_changes


def _main():
    tool_description = 'Track commit by using Gerrit API'
    parser = argparse.ArgumentParser(description=tool_description,
                                     formatter_class=argparse.ArgumentDefaultsHelpFormatter)
    parser.add_argument('-u', '--username', dest='username', help='username')
    parser.add_argument('-p', '--password', dest='password', help='password')
    parser.add_argument('-l', '--url-link', dest='base_url', help='the gerrit url')
    parser.add_argument('-f', '--file', dest='member_list_file', help='Group member list file')
    parser.add_argument('-d', '--date', dest='date_start', required=True,
                        help='query date string start, eg. 2019-06 or 2019-06-25')
    parser.add_argument('-v', '--verbose', dest='verbose', action='store_true', help='enable verbose (debug) logging')

    options = parser.parse_args()
    level = logging.DEBUG if options.verbose else logging.INFO
    logging.basicConfig(format='%(asctime)s %(levelname)s %(message)s', level=level)

    base_url = options.base_url
    username = options.username
    password = options.password
    member_list_file = options.member_list_file

    config = ConfigParser()
    config.read_file(codecs.open(CONFIG_FILE, 'r', encoding='utf-8-sig'))

    if base_url is None:
        base_url = config.get(CONFIG_POINT_LOCAL, CONFIG_BASE_URL)
    if username is None:
        username = config.get(CONFIG_POINT_LOCAL, CONFIG_USERNAME)
    if password is None:
        password = config.get(CONFIG_POINT_LOCAL, CONFIG_PASSWORD)
    if member_list_file is None:
        member_list_file = config.get(CONFIG_POINT_GLOBAL, CONFIG_MEMBER_LIST)

    if base_url is None or username is None or password is None or member_list_file is None:
        raise ValueError("url or username or password or file is invalid")

    try:
        auth = HTTPBasicAuth(username, password)
        date_filter_string = options.date_start

        member_list = read_member_list(member_list_file)

        # Create a file
        file_name = "gerrit_review_tracking-" + date_filter_string + ".csv"
        create_csv_file(file_name)

        with open(file_name, "a", encoding='utf-8-sig', newline='') as csvfile:
            writer = csv.writer(csvfile)
            urls = str(base_url).split(';')
            all_change: typing.List[OutputChangeItem] = []

            # Use multiple threads to obtain the submission of each member,
            thread_list = []
            start_time = time.time()
            for url in urls:
                rest = GerritRestAPI(url=url, auth=auth)
                member_dict = transfer_member_list(rest, url, member_list)
                pool_size = min(10, len(member_dict))
                pool = ThreadPoolExecutor(pool_size)
                for owner in member_dict.values():
                    task = pool.submit(fetch_change, url, rest, owner, member_dict, date_filter_string)
                    thread_list.append(task)
            for thread in thread_list:
                if thread.result() is not None:
                    all_change += thread.result()

            all_change = sorted(all_change, key=lambda c: c.change_time)

            end_time = time.time()
            print("fetch all changes spent time: %s seconds" % (round(end_time - start_time, 2)))
            for result in all_change:
                writer.writerow([
                    result.change_team,
                    result.change_link,
                    result.change_description,
                    result.get_created_date(),
                    result.get_display_time(),
                    result.change_code_count,
                    "" if result.change_inner_group_comments_count == "0" else result.change_inner_group_comments_count,
                    result.change_owner,
                    result.change_reason,
                    result.change_type,
                    result.change_rejected,
                    result.change_project,
                    result.change_branch,
                    result.message_member,
                    "" if result.change_outer_group_comments_count == "0" else result.change_outer_group_comments_count,
                ])
        print("finished")

    except RequestException as err:
        logging.error("Error: %s", str(err))


if __name__ == '__main__':
    sys.exit(_main())
