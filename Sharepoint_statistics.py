# -*- coding: utf-8 -*-
#import sys                                          # print etc
from configparser import ConfigParser               # able to read configuration file
import argparse                                     # good argument parser
from openpyxl import load_workbook                  # working with Excel files
import logging                                      #
import codecs                                       # used for text encoding in config parser
import csv                                          # working with csv files
import requests
from datetime import date              # used for manipulations with dates


CONFIG_POINT_LOCAL = "sharepoint"
CONFIG_POINT_GLOBAL = "global"
CONFIG_FILE = "config.ini"
CONFIG_BASE_URL = "sharepoint-url"
CONFIG_MEMBER_LIST = "member-list"
CONFIG_USERNAME = "username"
CONFIG_PASSWORD = "password"
CONFIG_POINT_LIBRARY = "library"


class DisplayItem:
    def __init__(self, issue_type=""):
        self.issue_type = [issue_type, "Issue Type"]


def main():
    #   parse arguments and options and get dict
    options = parse_arguments_and_options()
    #   set log level according to options or argument
    level = logging.DEBUG if options['log_level'] else logging.INFO
    logging.basicConfig(format='%(asctime)s %(levelname)s %(message)s', level=level)
    #   read list of members
    #member_list = read_member_list("members.xlsx")
    date_start = date.fromisoformat(options['date_start'])
    sp_doc_library = options['library']
    #   connect to SHAREPOINT
    s = requests.Session()
    # generate token by echo -n 'user:password' | base64
    auth_secret = 'Bearer {}'.format(options['password'])
    s.headers = {'accept': "application/json;odata=verbose", "Authorization": auth_secret}
    # get list of files within given directory on SP
    list_files_url = '{}/_api/web/GetFolderByServerRelativeUrl(\'{}\')/Files' \
        .format(options['base_url'], sp_doc_library)
    list_files_resp = s.get(list_files_url).json()
    # iterating through found files to find where modification date more than report date
    for file in list_files_resp['d']['results']:
        last_modified = date.fromisoformat(file['TimeLastModified'][0:10])
        if date_start <= last_modified:
            # get details of a file recently modified
            file_versions_url = '{}/_api/Web/GetFileByServerRelativeUrl(\'{}\')/Versions'\
                .format(options['base_url'], file['ServerRelativeUrl'])
            file_versions_resp = s.get(file_versions_url).json()
            # store file size of previous version for statistics compare
            previous_version_size = 0
            # iterate through versions to find recently modified versions
            for version in file_versions_resp['d']['results']:
                version_created = date.fromisoformat(version['Created'][0:10])
                if date_start <= version_created:
                    version_author_url = '{}/Versions/GetById(\'{}\')/CreatedBy'\
                        .format(file['__metadata']['uri'], version['ID'])
                    version_author_resp = s.get(version_author_url).json()
                    size_diff = abs(version['Size'] - previous_version_size)
                    print(file['Name'] + " | " + version['VersionLabel'] + " | " +
                          version['Created'] + " | " + version_author_resp['d']['Email'] + " | " +
                          str(size_diff) + " | " + version['CheckInComment'] + " | " + str(version['IsCurrentVersion']))
                previous_version_size = version['Size']
            # get original author of the document
            #file_versions_url = '{}/_api/Web/GetFileByServerRelativeUrl(\'{}\')/Author'\
            #    .format(options['base_url'], file['ServerRelativeUrl'])


def read_member_list(member_list_file):
    wb = load_workbook(member_list_file)
    sheet = wb.active
    has_name = "name" == sheet.cell(row=1, column=1).value
    has_mail = "email" == sheet.cell(row=1, column=2).value
    has_username = "username" == sheet.cell(row=1, column=3).value

    member_list = []
    if has_name and has_mail and has_username:
        for rx in range(2, sheet.max_row + 1):
            name = sheet.cell(row=rx, column=1).value
            mail = sheet.cell(row=rx, column=2).value
            username = sheet.cell(row=rx, column=3).value
            member_list.append({'Name': name, 'Email': mail, 'Username': username})
    else:
        raise ValueError("The table format is incorrect")
    return member_list


def parse_arguments_and_options():
    #   argument parsing section
    tool_description = 'Track comments left in code by using Gerrit API'
    parser = argparse.ArgumentParser(description=tool_description,
                                     formatter_class=argparse.ArgumentDefaultsHelpFormatter)
    parser.add_argument('-u', '--username', dest='username', help='username')
    parser.add_argument('-p', '--password', dest='password', help='password')
    parser.add_argument('-l', '--url-link', dest='base_url', help='the gerrit url')
    parser.add_argument('-f', '--file', dest='member_list_file', help='Group member list file')
    parser.add_argument('-d', '--date', dest='date_start', required=False,
                        help='query date string start, eg. 2019-06 or 2019-06-25')
    parser.add_argument('-v', '--verbose', dest='verbose', action='store_true', help='enable verbose (debug) logging')
    options = parser.parse_args()

    base_url = options.base_url
    username = options.username
    password = options.password
    date_start = options.date_start
    member_list_file = options.member_list_file
    log_level = options.verbose
    config = ConfigParser()
    config.read_file(codecs.open(CONFIG_FILE, 'r', encoding='utf-8-sig'))

    if base_url is None:
        base_url = config.get(CONFIG_POINT_LOCAL, CONFIG_BASE_URL)
    if username is None:
        username = config.get(CONFIG_POINT_LOCAL, CONFIG_USERNAME)
    if password is None:
        password = config.get(CONFIG_POINT_LOCAL, CONFIG_PASSWORD)
    if member_list_file is None:
        member_list_file = config.get(CONFIG_POINT_GLOBAL, CONFIG_MEMBER_LIST)

    if base_url is None or username is None or password is None or member_list_file is None:
        raise ValueError("url or username or password or file is invalid")
    return {'base_url': base_url, 'username': username, 'password': password,
            'member_list_file': member_list_file, 'date_start': date_start, 'log_level': log_level}


def create_and_write_csv_file(file_name, data_to_write):
    with open(file_name, "w", encoding='utf-8-sig', newline='') as csvfile:
        writer = csv.writer(csvfile)
        # Write Title
        empty_class = DisplayItem()
        fields_name = []
        for field in empty_class.__dict__:
            fields_name.append(empty_class.__dict__[field][1])
        writer.writerow(fields_name)
        # Write data
        for item_to_write in data_to_write:
            row_to_write = []
            for field_to_write in item_to_write.__dict__:
                row_to_write.append(item_to_write.__dict__[field_to_write][0])
            writer.writerow(row_to_write)


if __name__ == '__main__':
    main()
