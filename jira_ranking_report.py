# -*- coding: utf-8 -*-
"""
Comprehensive Jira Report Generator
Generates detailed reports with task analysis, link extraction, and team performance ranking
"""

from jira import JIRA
from configparser import ConfigParser
import pandas as pd
from datetime import datetime
import argparse
import codecs
import os
import re
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

# Configuration constants
CONFIG_FILE = "config.ini"
CONFIG_SECTION = "jira"
CONFIG_URL = "jira-url"
CONFIG_USERNAME = "username"
CONFIG_PASSWORD = "password"

RESOLVED_STATUSES = set(['Done', 'Resolved', 'Closed'])


def parse_arguments_and_config():
    """Parse command-line arguments and configuration file."""
    parser = argparse.ArgumentParser(description="Generate comprehensive Jira report.")
    parser.add_argument("-c", "--config", default=CONFIG_FILE, help="Path to config file")
    parser.add_argument("-u", "--username", help="Jira username")
    parser.add_argument("-p", "--password", help="Jira password")
    parser.add_argument("-l", "--url", help="Jira base URL")
    parser.add_argument("--project", help="Jira project key (e.g., ABC)")
    parser.add_argument("--start-date", help="Start date in YYYY-MM-DD format")
    parser.add_argument("--end-date", help="End date in YYYY-MM-DD format")
    parser.add_argument("--version", help="Fix version name")
    parser.add_argument("--epic", help="Epic key")
    parser.add_argument("--jql", help="Custom JQL query")
    parser.add_argument("--member-list-file", default="members.xlsx", help="Path to member list Excel file")
    parser.add_argument("--code-volume-file", help="Path to code volume Excel file")
    args = parser.parse_args()

    # Load config
    config = ConfigParser(allow_no_value=False, comment_prefixes=('#', ';'))
    config.read_file(codecs.open(args.config, 'r', encoding='utf-8-sig'))

    jira_url = args.url or config.get(CONFIG_SECTION, CONFIG_URL, fallback=None)
    jira_username = args.username or config.get(CONFIG_SECTION, CONFIG_USERNAME, fallback=None)
    jira_password = args.password or config.get(CONFIG_SECTION, CONFIG_PASSWORD, fallback=None)

    if not jira_url or not jira_username or not jira_password:
        raise ValueError("Jira URL, username, and password must be specified")

    return {
        'jira_url': jira_url,
        'jira_username': jira_username,
        'jira_password': jira_password,
        'project': args.project,
        'start_date': args.start_date,
        'end_date': args.end_date,
        'version': args.version,
        'epic': args.epic,
        'jql': args.jql,
        'member_list_file': args.member_list_file,
        'code_volume_file': args.code_volume_file
    }


def build_jql_query(params):
    """Build JQL query based on provided parameters."""
    if params['jql']:
        return params['jql']

    conditions = []

    if params['project']:
        conditions.append(f"project = {params['project']}")

    if params['start_date'] and params['end_date']:
        conditions.append(f"resolved >= '{params['start_date']}' AND resolved <= '{params['end_date']}'")

    if params['version']:
        conditions.append(f"fixVersion = '{params['version']}'")

    if params['epic']:
        conditions.append(f"'Epic Link' = {params['epic']}")

    if not conditions:
        raise ValueError("Must specify at least one of: project+dates, version, epic, or jql")

    return " AND ".join(conditions) + " ORDER BY created DESC"


def extract_urls_from_text(text):
    """Extract all URLs from text."""
    if not text:
        return []
    url_pattern = r'https?://[^\s<>"{}|\\^`\[\]]+'
    return re.findall(url_pattern, text)


# --- New helper: извлечение чисел по паттернам из текста комментариев
def sum_numbers_by_patterns(text, patterns):
    """
    Суммирует все целые числа, найденные в текстe по списку регэксп-паттернов (группирующие).
    patterns - список регулярных выражений с одной захватывающей группой для числа.
    Возвращает сумму (int).
    """
    if not text:
        return 0
    total = 0
    for pat in patterns:
        for m in re.finditer(pat, text, flags=re.IGNORECASE):
            try:
                total += int(m.group(1))
            except Exception:
                continue
    return total


def fetch_jira_data(jira, jql_query):
    """
    Fetch Jira issues with all details including comments.

    Returns:
        tuple: (issues_df, links_df)
    """
    start_at = 0
    max_results = 100
    all_issues = []

    print(f"Executing JQL: {jql_query}")

    while True:
        issues = jira.search_issues(
            jql_query,
            startAt=start_at,
            maxResults=max_results,
            fields=[
                "key", "summary", "assignee", "reporter", "resolutiondate",
                "created", "updated", "description", "comment", "labels",
                "priority", "status", "resolution", "issuetype",
                "timeestimate", "timespent", "timeoriginalestimate",
                "customfield_10000"  # Epic Link
            ],
            expand="changelog"
        )

        all_issues.extend(issues)

        if len(issues) < max_results:
            break
        start_at += max_results

    print(f"Fetched {len(all_issues)} issues")

    data = []
    all_links = []

    for issue in all_issues:
        # Basic fields
        key = issue.key
        summary = getattr(issue.fields, 'summary', '') or ''
        assignee = issue.fields.assignee.displayName if getattr(issue.fields, 'assignee', None) else "Unassigned"
        # username fields may differ between on-prem/cloud. try to fetch stable value with fallback
        assignee_name = ''
        reporter_name = ''
        try:
            assignee_name = issue.fields.assignee.name if getattr(issue.fields, 'assignee', None) and hasattr(issue.fields.assignee, 'name') else (
                issue.fields.assignee.key if getattr(issue.fields, 'assignee', None) and hasattr(issue.fields.assignee, 'key') else '')
        except Exception:
            assignee_name = ''

        reporter = issue.fields.reporter.displayName if getattr(issue.fields, 'reporter', None) else ''
        try:
            reporter_name = issue.fields.reporter.name if getattr(issue.fields, 'reporter', None) and hasattr(issue.fields.reporter, 'name') else (
                issue.fields.reporter.key if getattr(issue.fields, 'reporter', None) and hasattr(issue.fields.reporter, 'key') else '')
        except Exception:
            reporter_name = ''

        # Dates
        created = issue.fields.created[:10] if getattr(issue.fields, 'created', None) else ""
        resolved = issue.fields.resolutiondate[:10] if getattr(issue.fields, 'resolutiondate', None) else ""

        # Time tracking (convert seconds to hours)
        original_estimate = (issue.fields.timeoriginalestimate or 0) / 3600 if getattr(issue.fields, 'timeoriginalestimate', None) else 0
        time_spent = (issue.fields.timespent or 0) / 3600 if getattr(issue.fields, 'timespent', None) else 0
        remaining = (issue.fields.timeestimate or 0) / 3600 if getattr(issue.fields, 'timeestimate', None) else 0

        # Description
        description = issue.fields.description or ""

        # Extract links from description
        desc_links = extract_urls_from_text(description)
        for link in desc_links:
            all_links.append({
                'Issue_Key': key,
                'Source': 'Description',
                'URL': link
            })

        # Comments
        comments = []
        if getattr(issue.fields, 'comment', None) and getattr(issue.fields.comment, 'comments', None):
            for comment in issue.fields.comment.comments:
                comment_text = comment.body or ''
                comment_author = comment.author.displayName if getattr(comment, 'author', None) else "Unknown"
                comment_created = comment.created[:10] if getattr(comment, 'created', None) else ""
                comments.append(f"[{comment_created}] {comment_author}: {comment_text}")

                # Extract links from comments
                links = extract_urls_from_text(comment_text)
                for link in links:
                    all_links.append({
                        'Issue_Key': key,
                        'Source': f'Comment by {comment_author}',
                        'URL': link
                    })

        all_comments = "\n---\n".join(comments)

        # Labels
        labels = ", ".join(issue.fields.labels) if getattr(issue.fields, 'labels', None) else ""

        # Other fields
        priority = issue.fields.priority.name if getattr(issue.fields, 'priority', None) else ""
        status = issue.fields.status.name if getattr(issue.fields, 'status', None) else ""
        resolution = issue.fields.resolution.name if getattr(issue.fields, 'resolution', None) else ""
        issue_type = issue.fields.issuetype.name if getattr(issue.fields, 'issuetype', None) else ""
        epic_link = getattr(issue.fields, "customfield_10000", "")

        data.append({
            'Issue_Key': key,
            'Summary': summary,
            'Type': issue_type,
            'Status': status,
            'Resolution': resolution,
            'Priority': priority,
            'Assignee': assignee,
            'Assignee_Username': assignee_name or '',
            'Reporter': reporter,
            'Reporter_Username': reporter_name or '',
            'Created': created,
            'Resolved': resolved,
            'Original_Estimate_Hours': original_estimate,
            'Time_Spent_Hours': time_spent,
            'Remaining_Hours': remaining,
            'Description': description,
            'Comments': all_comments,
            'Labels': labels,
            'Epic_Link': epic_link
        })

    return pd.DataFrame(data), pd.DataFrame(all_links)


def read_member_list(member_list_file):
    """Read team member details from Excel file and normalize roles."""
    if not os.path.exists(member_list_file):
        print(f"Warning: Member list file '{member_list_file}' not found")
        return pd.DataFrame(columns=['name', 'email', 'username', 'role'])

    df = pd.read_excel(member_list_file)
    required_columns = ['name', 'username', 'role']

    for col in required_columns:
        if col not in df.columns:
            raise ValueError(f"Member list file must contain '{col}' column")

    # Normalize role values to canonical set: 'Engineer', 'QA Engineer', 'Project Manager'
    def normalize_role(r):
        if not isinstance(r, str):
            return ''
        s = r.strip().lower()
        if 'qa' in s or 'test' in s:
            return 'QA Engineer'
        if 'engineer' in s and 'qa' not in s and 'test' not in s:
            return 'Engineer'
        if 'manager' in s or 'pm' in s:
            return 'Project Manager'
        return r.strip()

    df['role'] = df['role'].apply(normalize_role)

    # Ensure username and name columns are strings
    df['username'] = df['username'].astype(str)
    df['name'] = df['name'].astype(str)

    return df


def read_code_volume(code_volume_file):
    """Read code volume data from Excel file."""
    if not code_volume_file or not os.path.exists(code_volume_file):
        return pd.DataFrame(columns=['username', 'code_volume'])

    df = pd.read_excel(code_volume_file)
    return df


def calculate_engineer_metrics(issues_df, members_df, code_volume_df):
    """Calculate metrics for Engineers."""
    metrics = []

    if issues_df.empty or members_df.empty:
        return pd.DataFrame(metrics)

    # Preprocess issues: make columns safe to query
    issues_df['Assignee_Username'] = issues_df['Assignee_Username'].fillna('').astype(str)
    issues_df['Reporter_Username'] = issues_df['Reporter_Username'].fillna('').astype(str)
    issues_df['Labels'] = issues_df['Labels'].fillna('').astype(str)
    issues_df['Type'] = issues_df['Type'].fillna('').astype(str)
    issues_df['Status'] = issues_df['Status'].fillna('').astype(str)

    engineers = members_df[members_df['role'] == 'Engineer']

    for _, engineer in engineers.iterrows():
        username = str(engineer['username']).lower()
        name = engineer['name']

        # Get issues assigned to this engineer and resolved
        user_issues = issues_df[issues_df['Assignee_Username'].str.lower() == username]
        resolved_issues = user_issues[user_issues['Status'].isin(RESOLVED_STATUSES)]

        # Code Volume (from external file)
        code_volume = 0
        if not code_volume_df.empty and 'username' in code_volume_df.columns:
            cv_row = code_volume_df[code_volume_df['username'].astype(str).str.lower() == username]
            if not cv_row.empty:
                code_volume = cv_row.iloc[0].get('code_volume', 0)

        # Code Quality (Bugs / (Features + Improvements))
        bugs = resolved_issues[resolved_issues['Type'].str.lower() == 'bug'].shape[0]
        features = resolved_issues[resolved_issues['Type'].str.lower().isin(['story', 'new feature', 'improvement'])].shape[0]
        code_quality = bugs / features if features > 0 else 0

        # Quantity of documentation (labels)
        doc_tasks = resolved_issues[resolved_issues['Labels'].str.contains('documentation', case=False, na=False)].shape[0]

        # Outstanding contribution
        outstanding_contribution = resolved_issues[resolved_issues['Labels'].str.contains('outstanding_contribution', case=False, na=False)].shape[0]

        # Assistance provided (placeholder)
        assistance_provided = 0

        metrics.append({
            'Name': name,
            'Role': 'Engineer',
            'Code_Volume': code_volume,
            'Code_Quality_Score': round(1 / (1 + code_quality), 2) if features > 0 else 1.0,
            'Bugs': bugs,
            'Features': features,
            'Documentation_Tasks': doc_tasks,
            'Outstanding_Contribution': outstanding_contribution,
            'Assistance_Provided': assistance_provided,
            'Total_Resolved_Issues': resolved_issues.shape[0]
        })

    return pd.DataFrame(metrics)


def calculate_qa_metrics(issues_df, members_df):
    """Calculate metrics for QA Engineers using comments for TT patterns."""
    metrics = []

    if issues_df.empty or members_df.empty:
        return pd.DataFrame(metrics)

    issues_df['Assignee_Username'] = issues_df['Assignee_Username'].fillna('').astype(str)
    issues_df['Reporter_Username'] = issues_df['Reporter_Username'].fillna('').astype(str)
    issues_df['Labels'] = issues_df['Labels'].fillna('').astype(str)
    issues_df['Type'] = issues_df['Type'].fillna('').astype(str)
    issues_df['Status'] = issues_df['Status'].fillna('').astype(str)
    issues_df['Comments'] = issues_df['Comments'].fillna('').astype(str)

    qa_engineers = members_df[members_df['role'] == 'QA Engineer']

    # Patterns
    api_patterns = [r"\b(?:TT_tested_APIs|TT_tdev_APIs)\s*=\s*(\d+)\b"]
    perf_patterns = [r"\b(?:TT_tested_perf|TT_tdev_perf)\s*=\s*(\d+)\b"]

    for _, qa in qa_engineers.iterrows():
        username = str(qa['username']).lower()
        name = qa['name']

        # Resolved issues assigned to this QA
        user_issues = issues_df[issues_df['Assignee_Username'].str.lower() == username]
        resolved_issues = user_issues[user_issues['Status'].isin(RESOLVED_STATUSES)]

        # Issues raised (created bugs)
        bugs_created = issues_df[(issues_df['Reporter_Username'].str.lower() == username) & (issues_df['Type'].str.lower() == 'bug')].shape[0]

        # Extract metrics from comments in resolved issues
        api_tests = 0
        perf_tests = 0
        for _, issue in resolved_issues.iterrows():
            comments_text = issue.get('Comments', '') or ''
            api_tests += sum_numbers_by_patterns(comments_text, api_patterns)
            perf_tests += sum_numbers_by_patterns(comments_text, perf_patterns)

        # Documentation tasks (labels)
        doc_tasks = resolved_issues[resolved_issues['Labels'].str.contains('documentation', case=False, na=False)].shape[0]

        # Outstanding contribution
        outstanding_contribution = resolved_issues[resolved_issues['Labels'].str.contains('outstanding_contribution', case=False, na=False)].shape[0]

        metrics.append({
            'Name': name,
            'Role': 'QA Engineer',
            'Test_Scenarios_Executed': api_tests,
            'Performance_Benchmarks': perf_tests,
            'Issues_Raised': bugs_created,
            'Documentation_Tasks': doc_tasks,
            'Outstanding_Contribution': outstanding_contribution,
            'Total_Resolved_Issues': resolved_issues.shape[0]
        })

    return pd.DataFrame(metrics)


def calculate_pm_metrics(issues_df, members_df, jira, jql_query):
    """Calculate metrics for Project Managers."""
    metrics = []

    if issues_df.empty or members_df.empty:
        return pd.DataFrame(metrics)

    issues_df['Reporter_Username'] = issues_df['Reporter_Username'].fillna('').astype(str)
    issues_df['Labels'] = issues_df['Labels'].fillna('').astype(str)
    issues_df['Type'] = issues_df['Type'].fillna('').astype(str)
    issues_df['Status'] = issues_df['Status'].fillna('').astype(str)

    pms = members_df[members_df['role'] == 'Project Manager']

    for _, pm in pms.iterrows():
        username = str(pm['username']).lower()
        name = pm['name']

        # Total closed tasks (for the scope of query)
        total_closed = issues_df[issues_df['Status'].isin(RESOLVED_STATUSES)].shape[0]

        # Documentation tasks done (resolved)
        doc_tasks = issues_df[(issues_df['Status'].isin(RESOLVED_STATUSES)) & (issues_df['Labels'].str.contains('documentation', case=False, na=False))].shape[0]

        # Outstanding contribution (resolved)
        outstanding_contribution = issues_df[(issues_df['Status'].isin(RESOLVED_STATUSES)) & (issues_df['Labels'].str.contains('outstanding_contribution', case=False, na=False))].shape[0]

        # Epics created by this PM (by reporter) - считаем эпики, которые находятся в resolved статусах
        epics_created = issues_df[(issues_df['Type'].str.lower() == 'epic') & (issues_df['Status'].isin(RESOLVED_STATUSES))].shape[0]

        metrics.append({
            'Name': name,
            'Role': 'Project Manager',
            'Epics_Created': epics_created,
            'Total_Closed_Tasks': total_closed,
            'Documentation_Tasks': doc_tasks,
            'Outstanding_Contribution': outstanding_contribution
        })

    return pd.DataFrame(metrics)


def export_to_excel(issues_df, links_df, engineer_metrics, qa_metrics, pm_metrics, output_file):
    """Export all data to Excel file with multiple sheets."""
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        # Issues sheet
        issues_df.to_excel(writer, sheet_name='Issues', index=False)

        # Links sheet
        if not links_df.empty:
            links_df.to_excel(writer, sheet_name='Links', index=False)

        # Team Performance sheets
        if not engineer_metrics.empty:
            engineer_metrics.to_excel(writer, sheet_name='Engineer_Performance', index=False)

        if not qa_metrics.empty:
            qa_metrics.to_excel(writer, sheet_name='QA_Performance', index=False)

        if not pm_metrics.empty:
            pm_metrics.to_excel(writer, sheet_name='PM_Performance', index=False)

        # Format sheets
        workbook = writer.book

        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]

            # Format header
            for cell in worksheet[1]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

            # Auto-adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width

    print(f"Excel report created: {output_file}")


def main():
    """Main function."""
    params = parse_arguments_and_config()

    # Connect to Jira
    jira_options = {"verify": "bundle-ca"} if os.path.exists("bundle-ca") else {"verify": True}
    jira = JIRA(
        server=params['jira_url'],
        basic_auth=(params['jira_username'], params['jira_password']),
        options=jira_options
    )

    # Build and execute query
    jql_query = build_jql_query(params)
    issues_df, links_df = fetch_jira_data(jira, jql_query)

    if issues_df.empty:
        print("No issues found matching the query")
        return

    # Read member list
    members_df = read_member_list(params['member_list_file'])
    code_volume_df = read_code_volume(params['code_volume_file'])

    # Calculate metrics by role
    engineer_metrics = pd.DataFrame()
    qa_metrics = pd.DataFrame()
    pm_metrics = pd.DataFrame()

    if not members_df.empty:
        engineer_metrics = calculate_engineer_metrics(issues_df, members_df, code_volume_df)
        qa_metrics = calculate_qa_metrics(issues_df, members_df)
        pm_metrics = calculate_pm_metrics(issues_df, members_df, jira, jql_query)
    else:
        print("Warning: No member list found, skipping team performance calculations")

    # Generate output filename
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_file = f"jira_comprehensive_report_{timestamp}.xlsx"

    # Export to Excel
    export_to_excel(issues_df, links_df, engineer_metrics, qa_metrics, pm_metrics, output_file)

    # Print summary
    print("\n" + "=" * 60)
    print("REPORT SUMMARY")
    print("=" * 60)
    print(f"Total Issues: {len(issues_df)}")
    print(f"Total Links Found: {len(links_df)}")
    if not engineer_metrics.empty:
        print(f"Engineers Analyzed: {len(engineer_metrics)}")
    if not qa_metrics.empty:
        print(f"QA Engineers Analyzed: {len(qa_metrics)}")
    if not pm_metrics.empty:
        print(f"Project Managers Analyzed: {len(pm_metrics)}")
    print("=" * 60)


if __name__ == "__main__":
    main()
