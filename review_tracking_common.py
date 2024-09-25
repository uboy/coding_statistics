# -*- coding: utf-8 -*-
import csv  #
import codecs  # used for text encoding in config parser
import argparse  # good argument parser
import re  # for parsing ticket no from description
from enum import member
from idlelib.iomenu import encoding

from openpyxl import load_workbook  #
from configparser import ConfigParser  # able to read configuration file

from requests import options

from backend.gerrit import CONFIG_BASE_URL, CONFIG_MEMBER_LIST
from backend.gitlab import CONFIG_BRANCH
from codehub_review import CONFIG_PROJECT, CONFIG_POINT_GLOBAL, create_csv_file


def parse_arguments_and_options():
    #  argument parsing section
    tool_description = 'Track comments left in code by using Gerrit API'
    parser = argparse.ArgumentParser(description=tool_description,
                                     formatter_class=argparse.ArgumentDefaultsHelpFormatter)
    parser.add_argument('-t', '--token', dest='token', help='token')
    parser.add_argument('-l', '--url-link', dest='base_url', help='the gerrit url')
    parser.add_argument('-f', '--file', dest='member_list_file', help='Group member list file')
    parser.add_argument('-p', '--project', dest='project', help='Project where calculate statistics')
    parser.add_argument('-b', '--branch', dest='branch', help='Branch for statistics calculation')
    parser.add_argument('-d', '--date', dest='date_start', required=True,
                        help='query date string start, eg. 2019-06 or 2019-06-25')
    parser.add_argument('-v', '--verbose', dest='verbose', action='store_true', help='enable verbose (debug) logging')
    parser.add_argument('-u', '--until', dest='date_until', required=False, help='query date string until')
    options = parser.parse_args()

    base_url = options.base_url
    token = options.token
    branch = options.branch
    project = options.project
    date_start = options.date_start
    date_until = options.date_until
    member_list_file = options.member_list_file
    log_level = options.verbose
    config = ConfigParser()
    config.read_file(codecs.open(CONFIG_FILE, 'r', encoding='utf-8-sig'))

    if base_url is None:
        base_url = config.get(CONFI_POINT_LOCAL, CONFIG_BASE_URL)
    if token is None:
        token = config.get(CONFI_POINT_LOCAL, CONFIG_TOKEN)
    if branch is None:
        branch = config.get(CONFI_POINT_LOCAL, CONFIG_BRANCH)
    if project is None:
        project = config.get(CONFI_POINT_LOCAL, CONFIG_PROJECT)
    if member_list_file is None:
        member_list_file = config.get(CONFIG_POINT_GLOBAL, CONFIG_MEMBER_LIST)

    if base_url is None or token is None or member_list_file is None:
        raise ValueError("url or username or password of file is invalid")
    return {'base_url': base_url, 'token': token, 'branch': branch, 'project': project,
            'member_list_file': member_list_file, 'date_start': date_start, 'date_until': date_until,
            'log_level': log_level}

def read_member_list(member_list_file):
    wb = load_workbook(member_list_file)
    sheet = wb.active
    has_name = "name" == sheet.cell(row=1, column=1).value
    has_mail = "email" == sheet.cell(row=1, column=2).value
    has_username = "username" == sheet.cell(row=1, column=3).value
    has_accname = "accname" == sheet.cell(row=1, column=4).value
    #has_giteeacc = "giteeacc" == sheet.cell(row=1, column=5).value

    member_list = []
    if has_name and has_mail and has_username and has_accname:
        for rx in range(2, sheet.max_row + 1):
            name = sheet.cell(row=rx, column=1).value
            mail = sheet.cell(row=rx, column=2).value
            username = sheet.cell(row=rx, column=3).value
            accname = sheet.cell(row=rx, column=4).value
            giteeacc = sheet.cell(row=rx, column=5).value
            member_list.append({'Name': name, 'Email': mail, 'Username': username, 'AccName': accname, 'GiteeAcc': giteeacc})
    else:
        raise ValueError("The table format is incorrect")
    return member_list

def create_csv_file(file_name):
    # Write Title
    with open(file_name, "w", encoding='utf-8-sig', newline='') as csvfile:
        writer = csv.writer(csvfile)
        writer.writerow(["Name", "Project", "Date", "Gitlab id", "branch", "description", "LOC", "Reviewed By", "AR",
                         "change_inner_group_comments_count", "change_outer_group_comments_count"])


def get_ticketno(description):
    ticket_pattern = re.compile(r'(?<=Description:)(.*?)(?=TicketNo:|Team:|Change-Id)')
    ticket_number = re.findall(ticket_pattern, description)
    if ticket_number:
        ticket_number = ticket_number[0]
    else:
        ticket_number = ""
    return ticket_number